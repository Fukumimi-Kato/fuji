import datetime as dt
import logging
import os

import openpyxl as excel
from openpyxl.styles.borders import Side
from openpyxl.worksheet.datavalidation import DataValidation

from django.conf import settings
from django.core.management import call_command
from django.db.models import Max

from accounts.models import User, UserManager
from .excel import ExcelHellper
from .models import ImportUnit, UnitMaster, MealDisplay, MealMaster, MenuDisplay, MenuMaster, \
    AllergenMaster, AllergenDisplay, InvoiceException, Order, UserCreationInput, UserCreationMenuInput, \
    UserCreationMealInput, UserCreationUnitInput, UserCreationAllergenInput, UserCreationDefaultOrdersInput



class UnitImporter:
    def __init__(self):
        self.input_user = None

        self.company_name = 'インポートテスト'
        self.unit_names = ['追加テスト_ユニット']
        self.short_names = ['追加テス1']
        self.enable_day_str = '2023-12-30'
        self.enable_day = dt.datetime.strptime(self.enable_day_str, '%Y-%m-%d')
        self.address = 'hoge@test.com'
        self.menu_name = '基本食・嚥下'
        self.meals = [
            [100, '有', '有'],
            [150, '無', '有'],
            [150, '無', '無'],
        ]
        self.dry_cold_type = '乾燥'

        self.allergens = ['肉', '魚']

        self.enable_saturday = True
        self.enable_sunday = False

    def _read_karte_file(self, ws, unit_import_data: ImportUnit):
        error_list = []
        needs_confirm = False
        require_company = None
        require_facility = None

        # 会社名
        company_name = ws.cell(5, 2).value
        facility_name = ws.cell(8, 2).value
        if company_name:
            if User.objects.filter(company_name=company_name, is_parent=True).exists():
                # 既に親会社がある場合は問い合わせを表示しない
                pass
            elif User.objects.filter(company_name=company_name).exclude(facility_name=facility_name).exists():
                require_company = company_name.strip()
                require_facility = facility_name.strip()
        else:
            error_list.append('会社名の入力がありません。')

        # 施設名
        if not facility_name:
            error_list.append('施設名の入力がありません。')

        # 開始日
        start_day = None
        try:
            start_day = dt.datetime(
                year=int(ws.cell(11, 2).value), month=int(ws.cell(11, 4).value), day=int(ws.cell(11, 6).value))
        except:
            error_list.append('開始時期の入力内容が不正です。')

        # 開始日期限チェック
        today = dt.datetime.today().date()
        if today > start_day.date():
            error_list.append('運用開始日を本日より前に設定することはできません。')

        # メールアドレス
        e_mail = ws.cell(12, 2).value

        # 献立種類
        menu_list = []
        menu_col = 2
        for i in range(4):
            if ws.cell(14, menu_col + i).value == 'あり':
                menu_list.append(ws.cell(13, menu_col + i).value)

        if not menu_list:
            error_list.append('献立種類は1つ以上に「あり」を入力する必要があります。')

        # 食数人数
        orders_list = []
        for i in range(3):
            meal_orders_list = []
            for j in range(4):
                try:
                    quantity = int(ws.cell(15 + i, 2 + j).value)
                except Exception as e:
                    quantity = 0

                meal_orders_list.append(quantity)
            orders_list.append(meal_orders_list)

        # 食事区分
        meal_list = []
        meal_col = 2
        for i in range(3):
            basic_price = ws.cell(19, meal_col + i).value or 0
            enge_price = ws.cell(20, meal_col + i).value or 0
            has_soup = ws.cell(21, meal_col + i).value == '有'
            has_filling = ws.cell(22, meal_col + i).value == '有'
            meal_list.append((basic_price, enge_price, has_soup, has_filling))

        if not sum(x[0] + x[1] for x in meal_list):
            error_list.append('単価は1項目以上の入力が必要です。')

        if [x for x in meal_list if x[2] and (not x[3])]:
            error_list.append('汁あり、具なしは設定できません。')

        # 乾燥冷凍区分
        dry_cold_type = ws.cell(23, 2).value
        if dry_cold_type:
            if dry_cold_type == '冷凍(直送)':
                dry_cold_type = '冷凍'
            elif dry_cold_type == '冷凍(談から送る)':
                dry_cold_type = '冷凍_談'
        else:
            error_list.append('麺・ネギ・錦糸(乾燥冷凍区分)の入力がありません。')

        # 土曜受取
        ng_saturday = ws.cell(25, 3).value
        adjust_saturday = 0
        if ng_saturday == '不可':
            adjust_saturday = ws.cell(25, 5).value
            if not adjst_saturday:
                error_list.append('土曜日受取不可ですが、調整日数の入力がありません。')
        elif ng_saturday == '可':
            pass
        else:
            error_list.append('土曜日受取不可を入力してください。')

        # 日曜受取
        ng_sunday = ws.cell(26, 3).value
        adjust_sunday = 0
        if ng_sunday == '不可':
            adjust_sunday = ws.cell(26, 5).value
            if not adjust_sunday:
                error_list.append('日曜日受取不可ですが、調整日数の入力がありません。')
        elif ng_sunday == '可':
            pass
        else:
            error_list.append('日曜日受取不可を入力してください。')

        # 祝日受取
        ng_holyday = ws.cell(27, 3).value
        adjust_holyday = 0
        if ng_holyday == '不可':
            adjust_holyday = ws.cell(27, 5).value
            if not adjust_holyday:
                error_list.append('祝日受取不可ですが、調整日数の入力がありません。')
        elif ng_holyday == '可':
            pass
        else:
            error_list.append('祝日受取不可を入力してください。')

        # 金曜受取調整有無
        adjust_friday_value = ws.cell(28, 2).value
        if adjust_friday_value:
            if adjust_friday_value == '有':
                adjust_friday = True
            else:
                adjust_friday = False
        else:
            error_list.append('金曜受取調整有無の入力がありません。')

        # 業務委託有無
        is_reduced_value = ws.cell(29, 2).value
        if is_reduced_value:
            if is_reduced_value == '業務委託である':
                is_reduced = True
            else:
                is_reduced = False
        else:
            error_list.append('業務委託有無の入力がありません。')

        # ユニット情報
        unit_list = []
        for i in range(5):
            unit_name = ws.cell(34 + i, 2).value
            short_name = ws.cell(34 + i, 4).value
            if unit_name:
                if len(short_name) > 5:
                    error_list.append(f'ユニットNo.{i+1}の短縮名の文字数が多すぎます。')
                else:
                    unit_list.append((unit_name, short_name))
        if len(unit_list) == 0:
            error_list.append(f'ユニットは1件以上入力してください。')

        # アレルギー情報
        allergen_list = []
        if ws.cell(24, 2).value == '有':
            allergen_row = 44
            allergen_name = ws.cell(allergen_row, 1).value
            selected = ws.cell(allergen_row, 2).value
            while allergen_name:
                if selected == '〇':
                    if AllergenMaster.objects.filter(allergen_name=allergen_name).exists():
                        allergen_list.append(allergen_name)
                    else:
                        error_list.append(f'入力されたアレルギー情報({allergen_name})が存在しません。最新のカルテをダウンロードしてください。')
                        break

                allergen_row += 1
                allergen_name = ws.cell(allergen_row, 1).value
                selected = ws.cell(allergen_row, 2).value

        # 読み取ったデータを登録
        user_id = None
        if not error_list:
            user = UserCreationInput(
                import_file=unit_import_data,
                company_name=company_name.strip(),
                facility_name=facility_name.strip(),
                enable_start_day=start_day,
                e_mail=e_mail,
                dry_cold_type=dry_cold_type,
                adjust_saturday=adjust_saturday,
                adjust_sunday=adjust_sunday,
                adjust_holyday=adjust_holyday,
                has_adjust_friday=adjust_friday,
                is_reduced=is_reduced
            )
            user.save()
            user_id = user.id

            # 献立種類の登録
            for menu in menu_list:
                user_menu = UserCreationMenuInput(
                    user_creation=user,
                    menu_name=menu
                )
                user_menu.save()

            # 基本食数の登録
            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='朝食', menu_name='基本食',
                quantity=orders_list[0][0]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='朝食', menu_name='ソフト',
                quantity=orders_list[0][1]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='朝食', menu_name='ゼリー',
                quantity=orders_list[0][2]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='朝食', menu_name='ミキサー',
                quantity=orders_list[0][3]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='昼食', menu_name='基本食',
                quantity=orders_list[1][0]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='昼食', menu_name='ソフト',
                quantity=orders_list[1][1]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='昼食', menu_name='ゼリー',
                quantity=orders_list[1][2]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='昼食', menu_name='ミキサー',
                quantity=orders_list[1][3]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='夕食', menu_name='基本食',
                quantity=orders_list[2][0]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='夕食', menu_name='ソフト',
                quantity=orders_list[2][1]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='夕食', menu_name='ゼリー',
                quantity=orders_list[2][2]
            )
            orders_input.save()

            orders_input = UserCreationDefaultOrdersInput(
                user_creation=user,
                meal_name='夕食', menu_name='ミキサー',
                quantity=orders_list[2][3]
            )
            orders_input.save()

            # 食事区分の登録
            meal = meal_list[0]
            user_meal_breakfast = UserCreationMealInput(
                user_creation=user,
                basic_price=meal[0],
                enge_price=meal[1],
                has_soup=meal[2],
                has_filling=meal[3]
            )
            user_meal_breakfast.save()

            meal = meal_list[1]
            user_meal_lunch = UserCreationMealInput(
                user_creation=user,
                basic_price=meal[0],
                enge_price=meal[1],
                has_soup=meal[2],
                has_filling=meal[3]
            )
            user_meal_lunch.save()

            meal = meal_list[2]
            user_meal_dinner = UserCreationMealInput(
                user_creation=user,
                basic_price=meal[0],
                enge_price=meal[1],
                has_soup=meal[2],
                has_filling=meal[3]
            )
            user_meal_dinner.save()

            # ユニットの登録
            for unit in unit_list:
                user_unit = UserCreationUnitInput(
                    user_creation=user,
                    unit_name=unit[0],
                    short_name=unit[1]
                )
                user_unit.save()

            # アレルギーの登録
            for allergen in allergen_list:
                user_allergen = UserCreationAllergenInput(
                    user_creation=user,
                    allergen_name=allergen
                )
                user_allergen.save()

        return require_company, require_facility, user_id, error_list

    def pre_read(self, unit_import_data: ImportUnit):
        """
        アップロード済みファイル読込(第一段階)
        """
        # アップロード済みファイル読み込み
        path = unit_import_data.document_file.path
        wb = excel.load_workbook(path)
        ws = wb['input']

        # ファイルからの情報の取得
        company_name, facility_name, user_id, error_list = self._read_karte_file(ws, unit_import_data)

        return company_name, facility_name, user_id, error_list

    def read(self, user_id: int):
        # アップロードファイルから読込んだデータを取り込む
        input_data = UserCreationInput.objects.get(id=user_id)
        self.input_user = input_data

    def _create_user(self, code: int):
        user = User.objects.create_user(str(code), self.input_user.e_mail, 'dan' + str(code))
        user.company_name = self.input_user.company_name
        user.facility_name = self.input_user.facility_name
        user.seq_order = 10
        user.dry_cold_type = self.input_user.dry_cold_type
        user.save()

        return user

    def _create_parent_user(self, code: int):
        user = User.objects.create_user(str(code), self.input_user.e_mail, 'dan' + str(code))
        user.company_name = self.input_user.company_name
        user.facility_name = self.input_user.company_name
        user.seq_order = 10
        user.dry_cold_type = self.input_user.dry_cold_type
        user.is_parent = True
        user.invoice_pass = f'D{code}ivc'
        user.save()

        return user

    def _update_user(self, user):
        user.company_name = self.input_user.company_name
        user.facility_name = self.input_user.facility_name
        if self.input_user.e_mail:
            user.email = self.input_user.e_mail
        user.seq_order = 10
        user.dry_cold_type = self.input_user.dry_cold_type
        user.save()

        return user

    def register(self):
        # 既存確認
        registerd_user = User.objects.filter(
            company_name=self.input_user.company_name, facility_name=self.input_user.facility_name).first()
        if registerd_user:
            # 存在していれば、注文データ、施設毎データ削除
            Order.objects.filter(unit_name__username=registerd_user).delete()
            MealDisplay.objects.filter(username=registerd_user).delete()
            MenuDisplay.objects.filter(username=registerd_user).delete()
            AllergenDisplay.objects.filter(username=registerd_user).delete()
            InvoiceException.objects.filter(unit_name__username=registerd_user).delete()

            # 削除前と同じ番号を使用する
            first_unit = UnitMaster.objects.filter(username=registerd_user).order_by('unit_number').first()
            max_number = first_unit.unit_number
            max_code = first_unit.unit_code

            # ユニット削除
            UnitMaster.objects.filter(username=registerd_user).delete()

            # 親会社は再登録で追加されるので削除しない

            # 施設更新
            user = self._update_user(registerd_user)
        else:
            max_number_dict = UnitMaster.objects.filter(unit_code__lt=50000).aggregate(Max('unit_number'))
            max_number = max_number_dict['unit_number__max'] + 1
            max_code_dict = UnitMaster.objects.filter(unit_code__lt=50000).aggregate(Max('unit_code'))
            max_code = max_code_dict['unit_code__max'] + 1

            # 施設登録
            user = self._create_user(max_code)

        # ユニット登録
        unit_list = []
        for input_unit in self.input_user.usercreationunitinput_set.all():
            unit = UnitMaster(
                unit_name=input_unit.unit_name,
                group='なし',
                seq_order='10',
                is_active=True,
                username=user,
                unit_code=max_code,
                unit_number=max_number,
                calc_name=input_unit.unit_name,
                short_name=input_unit.short_name
            )
            unit.save()

            max_code += 1
            max_number += 1

            unit_list.append(str(unit.id))

        # -食事区分登録(朝、昼、夕の順で入力データは登録されている)
        input_meal_list = list(self.input_user.usercreationmealinput_set.all().order_by('id'))
        meal_dict = {
            '朝食': input_meal_list[0],
            '昼食': input_meal_list[1],
            '夕食': input_meal_list[2],
        }
        for meal_name in ['朝食', '昼食', '夕食']:
            if not (meal_dict[meal_name].basic_price or meal_dict[meal_name].enge_price):
                # 基本食、嚥下に単価がないものは登録しない。
                continue

            meal_master = MealMaster.objects.filter(
                meal_name=meal_name, soup=meal_dict[meal_name].has_soup, filling=meal_dict[meal_name].has_filling)
            tmp_meal = MealDisplay(
                username=user,
                meal_name=meal_master.first(),
            )
            tmp_meal.save()

        # -献立種類登録
        basic_prices = [meal_dict['朝食'].basic_price, meal_dict['昼食'].basic_price, meal_dict['夕食'].basic_price]
        enge_prices = [meal_dict['朝食'].enge_price, meal_dict['昼食'].enge_price, meal_dict['夕食'].enge_price]
        is_registerd_enge = False
        for menu in self.input_user.usercreationmenuinput_set.all().order_by('id'):
            if menu.menu_name == '基本食':
                menu_master = MenuMaster.objects.filter(menu_name='常食').first()
                tmp_menu = MenuDisplay(
                    username=user,
                    menu_name=menu_master,
                    price_breakfast=basic_prices[0],
                    price_lunch=basic_prices[1],
                    price_dinner=basic_prices[2],
                )
                tmp_menu.save()
            elif not is_registerd_enge:
                is_registerd_enge = True
                for enge_menu in ['ゼリー', 'ミキサー', 'ソフト']:
                    menu_master = MenuMaster.objects.filter(menu_name=enge_menu).first()
                    tmp_menu = MenuDisplay(
                        username=user,
                        menu_name=menu_master,
                        price_breakfast=enge_prices[0],
                        price_lunch=enge_prices[1],
                        price_dinner=enge_prices[2],
                    )
                    tmp_menu.save()

        # -アレルギー登録
        for allergen in self.input_user.usercreationallergeninput_set.all().order_by('id'):
            qs = AllergenMaster.objects.filter(allergen_name=allergen.allergen_name)
            if qs.exists():
                tmp = AllergenDisplay(
                    username=user, allergen_name=qs.first()
                )
                tmp.save()

        # 顧客別_売上日調整日数登録
        if self.input_user.adjust_saturday or \
                self.input_user.adjust_sunday or \
                self.input_user.adjust_holyday or \
                self.input_user.has_adjust_friday or \
                self.input_user.is_reduced:
            ex = InvoiceException(
                unit_name=unit,
                ng_saturday=self.input_user.adjust_saturday or 0,
                ng_sunday=self.input_user.adjust_sunday or 0,
                ng_holiday=self.input_user.adjust_holyday or 0,
                reduced_rate=self.input_user.is_reduced,
                is_far=self.input_user.has_adjust_friday,
            )
            ex.save()

        # 親会社を登録する
        parent_user_qs = User.objects.filter(company_name=user.company_name, is_parent=False).exclude(id=user.id)
        if parent_user_qs.exists():
            if User.objects.filter(company_name=user.company_name, is_parent=True).exists():
                # 既に親会社がある場合は登録しない
                pass
            else:
                parent_user_code = None
                for user in parent_user_qs.order_by('username'):
                    # もっとも若い得意先コードを代表にする。
                    tmp_user_code = f'9{user.username}'
                    if not User.objects.filter(username=tmp_user_code).exists():
                        # 念のため、得意先コードが使われていないことをチェック
                        parent_user_code = tmp_user_code
                        break

                if parent_user_code:
                    # 親会社の登録
                    self._create_parent_user(parent_user_code)
                else:
                    messages.warn(request, 'データ不整合により、親会社のログイン情報が登録できませんでした。システム管理者にお問い合わせください。')
        else:
            user.invoice_pass = f'D{user.username}ivc'
            user.save()

        # -注文データ登録
        enable_day_str = self.input_user.enable_start_day.strftime('%Y-%m-%d')
        if unit_list:
            call_command('weekly_for_call', enable_day_str, ",".join(unit_list))
        else:
            call_command('weekly_for_call', enable_day_str, str(unit.id))
        call_command('generate_qr')


class KarteWriter:
    def __init__(self, path, filename):
        self.filepath = os.path.join(path, filename)
        self.wb = excel.load_workbook(self.filepath)

        self.path = path

    def refreesh(self):
        now = dt.datetime.now()

        # 既存のファイルをバックアップ
        backup_path = os.path.join(self.path, f'{now.strftime("%Y%m%d%H%M%S")}_{settings.KARTE_FORMAT_DOWNLOAD_NAME}')
        self.wb.save(backup_path)

        # アレルギーの更新
        ws = self.wb['input']
        for row in ws.iter_rows(min_row=43):
            # A列、B列をクリア
            row[0].value = ''

        row = 44
        ignore_allergen_names = ['なし', '個食', 'フリーズ', 'あり']
        qs = AllergenMaster.objects.filter(
            ignore_karte=False).exclude(allergen_name__in=ignore_allergen_names).order_by('seq_order', '-id')
        for index, allergen in enumerate(qs):
            # 名称の設定
            ws.cell(row, 1, allergen.allergen_name)

            # 入力規則の設定
            dv = DataValidation(type="list", formula1='"〇"')
            dv.add(ws.cell(row, 2))
            ws.add_data_validation(dv)

            row += 1

        # 枠線の引き直し
        ExcelHellper.set_grid_border(
            ws, f'A43:B{row}', Side(style='medium', color="000000"), Side(style='thin', color="000000"))

        # ファイルの更新
        self.wb.save(self.filepath)
        self.wb.close()

