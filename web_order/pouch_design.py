from functools import cmp_to_key

from itertools import groupby
import logging
import math
import os
import re

import openpyxl as excel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

from django.conf import settings
from django.db.models import Sum

from .cooking_direction_plates import PlateNameAnalizeUtil
from .excel import ExcelHellper
from .models import AllergenPlateRelations, CookingDirectionPlate, Order, PackageMaster, UnitPackage, UnitMaster, \
    CommonAllergen, UncommonAllergenHistory
from .p7 import P7Util
from .picking import EatingManagement

logger = logging.getLogger(__name__)


class PouchMeal:
    """
    パウチ設計図で管理する食事のクラス。
    """

    def __init__(self, eating_day, meal_name):
        self.eating_day = eating_day
        self.meal_name = meal_name

        # 集計情報
        self.bacic_aggregate = None
        self.bacic_5p_aggregate = None
        self.unit_aggregate = None
        self.unit_mix_rice_aggregate = None
        self.soup_aggregate = None
        self.filling_aggregate = None

        self.soft_aggregate = None
        self.mixer_aggregate = None
        self.jelly_aggregate = None

        self.allergen_aggregate = None

        self.unit_for_dry_aggregate = None

    def exists_units(self):
        return self.unit_aggregate.exists()

    def exists_mix_rice(self):
        return self.unit_mix_rice_aggregate.exists()

    def exists_soup(self):
        return self.filling_aggregate.exists()

    def exists_units_for_dry(self):
        return self.unit_for_dry_aggregate

    def exists_5p(self):
        return self.bacic_5p_aggregate.exists()

    def __lt__(self, other):
        cmp_date = self.eating_day - other.eating_day
        if cmp_date:
            return self.eating_day < other.eating_day
        else:
            seq_dict = {
                '朝食': 0,
                '昼食': 1,
                '夕食': 2,
            }

            return seq_dict[self.meal_name] < seq_dict[other.meal_name]


class DesignedPouch:
    """
    設計図出力用のパウチ情報
    """

    def __init__(self, name: str):
        self.name = name

        # 1人用
        self.package_count_1p = 0

        # 施設
        self.package_units = []

        # 保存用
        self.package_preserve = 0

        # 針刺し用
        self.package_needle = 0

    def add_unit(self, unit_number: int, unit_name: str, count: int):
        """
        施設毎の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.package_units.append((unit_qs.first(), count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class DesignedPouchBasic(DesignedPouch):
    """
    設計図パウチ情報(基本食)
    """
    def __init__(self, name: str):
        super(DesignedPouchBasic, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.preserve_1p = 1
        self.photo = 1 + 1

        # 10人用
        self.package_count_10p = 0

    def add_unit(self, unit_number: int, unit_name: str, count: int, ko_units):
        """
        施設毎の袋情報を追加する(個食対応を追加)
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            first_unit = unit_qs.first()
            logger.debug(f'ユニット情報登録:{first_unit.unit_name}-{count}')
            # 個食の対応
            if ko_units:
                logger.debug(f'個食元{unit_number}-{count}')
                for ko_unit in ko_units:
                    if '個食' in ko_unit.calc_name:
                        logger.debug(f'個食add{unit_number}')
                        self.package_units.append((ko_unit, 1))
                    else:
                        logger.debug(f'個食元add{unit_number}')
                        self.package_units.append((ko_unit, count - 1))
            else:
                logger.debug(f'個食なしadd{unit_number}')
                self.package_units.append((first_unit, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class DesignedPouchBasic5p(DesignedPouch):
    """
    設計図パウチ情報(基本食5人用)
    """
    def __init__(self, name: str):
        super(DesignedPouchBasic5p, self).__init__(name)

        # 針刺し用・保存用の設定(10人用の倍になる)
        self.package_needle = 4
        self.package_preserve = 10

        self.preserve_1p = 1
        self.photo = 1 + 1

        # 5人用
        self.package_count_5p = 0
        self.plate_name_5p = None

    def add_unit(self, unit_number: int, unit_name: str, count: int, ko_units):
        """
        施設毎の袋情報を追加する(個食対応を追加)
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            first_unit = unit_qs.first()
            logger.debug(f'ユニット情報登録:{first_unit.unit_name}-{count}')
            # 個食の対応
            if ko_units:
                logger.debug(f'個食元{unit_number}-{count}')
                for ko_unit in ko_units:
                    if '個食' in ko_unit.calc_name:
                        self.package_units.append((ko_unit, 1))
                    else:
                        self.package_units.append((ko_unit, count - 1))
            else:
                self.package_units.append((first_unit, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def exists(self):
        """
        データの有無を返却する
        """
        units_len = len(self.package_units)
        if (units_len > 0) or self.package_count_5p:
            return True
        else:
            return False


class DesignedPouchUnit(DesignedPouch):
    """
    設計図パウチ情報(施設毎)
    """
    def __init__(self, name: str):
        super(DesignedPouchUnit, self).__init__(name)

        self.preserve_1p = 1
        self.photo = 1 + 1

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

    def exists(self):
        """
        データの有無を返却する
        """
        units_len = len(self.package_units)
        if units_len > 0:
            return True
        else:
            return False

    def add_unit(self, unit_number: int, unit_name: str, count: int, has_koshoku: bool=False):
        """
        施設毎の袋情報を追加する(個食対応を追加-基本食と同じ内容)
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            # 個食の対応
            if has_koshoku:
                self.package_units.append((unit_qs.first(), count - 1))
                ko_qs = UnitMaster.objects.filter(unit_number=unit_number).exclude(calc_name=unit_name)
                if ko_qs.exists():
                    self.package_units.append((ko_qs.first(), 1))
            else:
                self.package_units.append((unit_qs.first(), count))

        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')



class DesignedPouchMixrice(DesignedPouchUnit):
    """
    設計図パウチ情報(施設毎-混ぜご飯)
    """
    def __init__(self, name: str):
        super(DesignedPouchMixrice, self).__init__(name)

        # 混ぜご飯パーツの袋数の有無(メイン料理がg以外の単位の場合)
        self.has_parts = False

    def exists(self):
        """
        データの有無を返却する
        """
        count_1p = self.package_count_1p
        units_len = len(self.package_units)
        if count_1p > 0:
            return True
        elif units_len > 0:
            return True
        else:
            return self.has_parts

    def add_unit(self, unit_number: int, unit_name: str, count: int, has_koshoku: bool=False):
        """
        施設毎の袋情報を追加する(個食対応を追加-基本食と同じ内容)
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.package_units.append((unit_qs.first(), count, False))

            # 個食の対応
            if has_koshoku:
                ko_qs = UnitMaster.objects.filter(unit_number=unit_number).exclude(calc_name=unit_name)
                if ko_qs.exists():
                    self.package_units.append((ko_qs.first(), 1, False))
        else:
            aggregate_first_units = [x[0] for x in settings.MIX_RICE_AGGREGATE_UNITS]
            if unit_number in aggregate_first_units:
                aggregate_qs = UnitMaster.objects.filter(unit_number=unit_number).exclude(calc_name=unit_name)
                self.package_units.append((aggregate_qs.first(), count, True))
            else:
                logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

class DesignedPouchSoup(DesignedPouch):
    """
    設計図パウチ情報(汁)
    """
    def __init__(self, name: str):
        super(DesignedPouchSoup, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.preserve_1p = 1
        self.photo = 2

        self.plate_name = ''

        self.designed_soft = DesignedPouchEngeSoup('(ソフト)')
        self.designed_soft.preserve_1p = 4
        self.designed_mixer = DesignedPouchEngeSoup('(ミキサー)')
        self.designed_jelly = DesignedPouchEngeSoup('(ゼリー)')

        self.is_unit_pouch_use = False

    def add_unit(self, unit: UnitMaster, size: int, order_count: int, package_count: int):
        self.package_units.append((unit, size, order_count, package_count))

    def set_soft_1p(self, count: int):
        self.designed_soft.package_count_1p = count

    def set_mixer_1p(self, count: int):
        self.designed_mixer.package_count_1p = count

    def set_jellry_1p(self, count: int):
        self.designed_jelly.package_count_1p = count

    def set_soft_2p(self, count: int):
        self.designed_soft.package_count_2p = count

    def set_mixer_2p(self, count: int):
        self.designed_mixer.package_count_2p = count

    def set_jellry_2p(self, count: int):
        self.designed_jelly.package_count_2p = count

    def add_soft_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ソフト食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_soft.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_jelly_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ゼリー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_jelly.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_mixer_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ミキサー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_mixer.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class DesignedPouchFilling(DesignedPouch):
    """
    設計図パウチ情報(汁具)
    """
    def __init__(self, name: str):
        super(DesignedPouchFilling, self).__init__(name)

        self.plate_name = ''

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.preserve_1p = 1
        self.photo = 1 + 1

        self.designed_soft = DesignedPouchEngeFilling('(ソフト)')
        self.designed_soft.preserve_1p = 4
        self.designed_mixer = DesignedPouchEngeFilling('(ミキサー)')
        self.designed_jelly = DesignedPouchEngeFilling('(ゼリー)')

    def exists(self):
        if self.designed_soft.package_count_1p or self.designed_soft.package_count_2p or self.designed_soft.package_units:
            return True
        elif self.designed_mixer.package_count_1p or self.designed_mixer.package_count_2p or self.designed_mixer.package_units:
            return True
        elif self.designed_jelly.package_count_1p or self.designed_jelly.package_count_2p or self.designed_jelly.package_units:
            return True
        elif self.package_units:
            return True
        else:
            return False

    def set_soft_1p(self, count: int):
        self.designed_soft.package_count_1p = count

    def set_mixer_1p(self, count: int):
        self.designed_mixer.package_count_1p = count

    def set_jellry_1p(self, count: int):
        self.designed_jelly.package_count_1p = count

    def set_soft_2p(self, count: int):
        self.designed_soft.package_count_2p = count

    def set_mixer_2p(self, count: int):
        self.designed_mixer.package_count_2p = count

    def set_jellry_2p(self, count: int):
        self.designed_jelly.package_count_2p = count

    def add_soft_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ソフト食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_soft.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_jelly_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ゼリー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_jelly.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_mixer_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ミキサー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_mixer.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class DesignedPouchEngeFilling(DesignedPouch):
    """
    設計図パウチ情報(嚥下)
    """
    def __init__(self, name: str):
        super(DesignedPouchEngeFilling, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.package_count_2p = 0

        self.preserve_1p = 2


class DesignedPouchEngeSoup(DesignedPouch):
    """
    設計図パウチ情報(嚥下)-汁用
    """
    def __init__(self, name: str):
        super(DesignedPouchEngeSoup, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.package_count_2p = 0

        self.preserve_1p = 2


class DesignedPouchEnge(DesignedPouch):
    """
    設計図パウチ情報(嚥下)
    """
    def __init__(self, name: str):
        super(DesignedPouchEnge, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.preserve_1p = 2

        self.package_count_2p = 0

    def set_1p(self, count: int):
        self.package_count_1p = count

    def set_2p(self, count: int):
        self.package_count_2p = count

    def add_unit(self, unit: UnitMaster, size: int, order_count: int, package_count: int):
        self.package_units.append((unit, size, order_count, package_count))


class DesignedPouchAllergen(DesignedPouch):
    """
    設計図パウチ情報(アレルギー)
    """
    def __init__(self, name: str):
        super(DesignedPouchAllergen, self).__init__(name)

        # 針刺し用・保存用の設定
        self.package_preserve = 1

        self.allergen_plates = []

    def add_allergen_plate(self, unit_number: str, unit_name: str, plate_index: int, menu_name: str, count):
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.allergen_plates.append((unit_qs.first(), plate_index, menu_name, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class DesignedUnitForDryRawPlate(DesignedPouch):
    """
    設計図パウチ情報(乾燥原体製品の特別出力用)
    """
    def __init__(self, name: str):
        super(DesignedUnitForDryRawPlate, self).__init__(name)

        self.plate_name = ''

        # 針刺し用・保存用の設定
        self.package_needle = 2
        self.package_preserve = 5

        self.preserve_1p = 1
        self.photo = 1 + 1

        self.designed_soft = DesignedPouchEngeFilling('(ソフト)')
        self.designed_soft.preserve_1p = 4
        self.designed_mixer = DesignedPouchEngeFilling('(ミキサー)')
        self.designed_jelly = DesignedPouchEngeFilling('(ゼリー)')

    def set_soft_1p(self, count: int):
        self.designed_soft.package_count_1p = count

    def set_mixer_1p(self, count: int):
        self.designed_mixer.package_count_1p = count

    def set_jellry_1p(self, count: int):
        self.designed_jelly.package_count_1p = count

    def set_soft_2p(self, count: int):
        self.designed_soft.package_count_2p = count

    def set_mixer_2p(self, count: int):
        self.designed_mixer.package_count_2p = count

    def set_jellry_2p(self, count: int):
        self.designed_jelly.package_count_2p = count

    def add_soft_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ソフト食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_soft.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_jelly_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ゼリー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_jelly.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')

    def add_mixer_unit(self, unit_number: int, unit_name: str, size: int, order_count: int, count: int):
        """
        ミキサー食の1人以上の袋情報を追加する
        """
        unit_qs = UnitMaster.objects.filter(unit_number=unit_number, calc_name=unit_name)
        if unit_qs.exists():
            self.designed_mixer.package_units.append((unit_qs.first(), size, order_count, count))
        else:
            logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')


class PouchAggregate:
    """
    パウチ設計図用集計クラス
    """
    ENGE_PACKAGE_ID_LIST = [
        settings.PICKING_PACKAGES['ENGE_2'],
        settings.PICKING_PACKAGES['ENGE_7'],
        settings.PICKING_PACKAGES['ENGE_14'],
        settings.PICKING_PACKAGES['ENGE_20']
    ]

    def __init__(self, cooking_day):
        self.cooking_day = cooking_day
        self.meal_list = []

    def compare_allergen_unit(self, unit_order1, unit_order2):
        seq_dict = {
            '常食': 0,
            'ソフト': 1,
            'ゼリー': 2,
            'ミキサー': 3
        }

        cmp_seq = seq_dict[unit_order1.menu_name.menu_name] - seq_dict[unit_order2.menu_name.menu_name]
        if cmp_seq:
            return cmp_seq
        else:
            cmp_allergen_id = unit_order1.allergen_id - unit_order2.allergen_id
            if cmp_allergen_id:
                return cmp_allergen_id
            return unit_order1.unit_name.unit_number - unit_order2.unit_name.unit_number

    def read_eating_time(self):
        """
        喫食日に紐づく食事を読み込む
        """
        self.meal_list.clear()

        # 製造日に紐づく喫食日を取得する
        tmp_list = []
        eating_dict = EatingManagement.get_meals_dict_by_cooking_day(self.cooking_day)
        for key, meal_name_list in eating_dict.items():
            for meal_name in meal_name_list:
                tmp_list.append(PouchMeal(key, meal_name))

        # ソート
        self.meal_list = sorted(tmp_list)

    def aggreate(self):
        """
        読込済みの全ての食事毎に、集計を実施する
        """
        logger.info('パウチ設計図-集計開始')

        # 基本食(10人袋)の集計
        self.aggregate_basic()

        # 基本食(5人袋)の集計
        self.aggregate_basic_5p()

        # 施設毎の集計
        self.aggregate_unit()

        # 施設毎(混ぜご飯)の集計
        self.aggregate_mix_rice()

        # 汁の集計
        self.aggregate_soup()

        # 汁具の集計
        self.aggregate_filling()

        # 嚥下の集計
        self.aggregate_enge()

        # アレルギーの集計
        self.aggregate_allergen()

        # 乾燥出力用施設の集計
        self.aggregate_unit_for_dry_plate()

    def aggregate_basic(self):
        """
        基本食(10人袋)の集計
        """
        logger.debug('パウチ設計図-基本食集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchBasic('基本食')

            # 5人袋の集計(10人用の有無確認のため、先に取得)
            # ユニット毎の袋数情報取得
            unit_package_10p_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_10'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none').order_by('unit_number')
            first_unit_package = None
            if unit_package_10p_qs.exists():
                # 以降で料理名を条件に加えるため、先頭を取得
                first_unit_package = unit_package_10p_qs.first()
            else:
                # 10人用袋適用の料理が存在しない
                pouch_meal.bacic_aggregate = dp
                continue

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_1'], is_basic_plate=True, menu_name='常食',
                plate_name=first_unit_package.plate_name,
                mix_rice_type='none', soup_type='none').order_by('unit_number')

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_1p += 1

            # 各ユニットの10人用袋の集計
            for key, group in groupby(unit_package_10p_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_10p += unit_package.count

            # 端数の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_FRACTION'], is_basic_plate=True, menu_name='常食',
                plate_name=first_unit_package.plate_name,
                mix_rice_type='none', soup_type='none').order_by('unit_number', '-count')

            # ユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)

                # 個食の対応
                has_koshoku = False
                ko_unit_list = []
                ko_qs = Order.objects.filter(unit_name__unit_number=unit_package.unit_number, quantity__gt=0,
                                             eating_day=pouch_meal.eating_day,
                                             meal_name__meal_name=pouch_meal.meal_name).select_related('unit_name')
                if ko_qs.exists():
                    for order in ko_qs:
                        if (order.quantity % 10) > 1:
                            ko_unit_list.append(order.unit_name)
                            if '個食' in order.unit_name.unit_name:
                                has_koshoku = True
                    if not has_koshoku:
                        ko_unit_list.clear()

                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count, ko_unit_list)

            pouch_meal.bacic_aggregate = dp


    def aggregate_basic_5p(self):
        """
        基本食(5人袋)の集計
        """
        logger.debug('パウチ設計図-基本食集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchBasic5p('基本食')

            # 5人袋の集計(5人用の有無確認のため、先に取得)
            # ユニット毎の袋数情報取得
            unit_package_5p_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_5'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none').order_by('unit_number')
            first_unit_package = None
            if unit_package_5p_qs.exists():
                # 以降で料理名を条件に加えるため、先頭を取得
                first_unit_package = unit_package_5p_qs.first()
                dp.plate_name_5p = first_unit_package.plate_name
            else:
                # 5人用袋適用の料理が存在しない
                pouch_meal.bacic_5p_aggregate = dp
                continue

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_1'], is_basic_plate=True, menu_name='常食',
                plate_name=first_unit_package.plate_name,
                mix_rice_type='none', soup_type='none').order_by('unit_number')

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_1p += 1

            # 各ユニットの5人用袋の集計
            for key, group in groupby(unit_package_5p_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_5p += unit_package.count

            # 端数の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_FRACTION'], is_basic_plate=True, menu_name='常食',
                plate_name=first_unit_package.plate_name,
                mix_rice_type='none', soup_type='none').order_by('unit_number', '-count')

            # ユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)

                # 個食の対応
                has_koshoku = False
                ko_unit_list = []
                ko_qs = Order.objects.filter(unit_name__unit_number=unit_package.unit_number, quantity__gt=0,
                                             eating_day=pouch_meal.eating_day,
                                             meal_name__meal_name=pouch_meal.meal_name).select_related('unit_name')
                if ko_qs.exists():
                    for order in ko_qs:
                        if (order.quantity % 5) > 0:
                            ko_unit_list.append(order.unit_name)
                            if '個食' in order.unit_name.unit_name:
                                has_koshoku = True
                    if not has_koshoku:
                        ko_unit_list.clear()

                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count, ko_unit_list)

            pouch_meal.bacic_5p_aggregate = dp

    def aggregate_unit(self):
        """
        施設毎の集計
        """
        logger.debug('パウチ設計図-施設毎集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchUnit('施設毎')

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            # ユニット毎以外の料理を参照しないように、ユニット毎の袋を抽出
            pre_unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none'
            ).order_by('unit_number')
            if pre_unit_package_qs.exists():
                unit_plate_name = pre_unit_package_qs.first().plate_name

                unit_package_qs = UnitPackage.objects.filter(
                    cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                    package_id=settings.PICKING_PACKAGES['BASIC_1'], is_basic_plate=True, menu_name='常食',
                    mix_rice_type='none', soup_type='none', plate_name=unit_plate_name
                ).order_by('unit_number')

                # 各ユニットの1人用袋の集計
                for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                    unit_package = next(group)
                    # サンシティあい検食用(施設番号=3)を含める
                    if (key == 3) or unit_package.count:
                        dp.package_count_1p += 1

            # 施設毎の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none').order_by('unit_number', '-count')

            # 端数のあるユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)

                # 個食の対応
                has_koshoku = False
                ko_qs = Order.objects.filter(unit_name__unit_number=unit_package.unit_number, quantity__gt=0,
                                             eating_day=pouch_meal.eating_day,
                                             meal_name__meal_name=pouch_meal.meal_name).exclude(
                    unit_name__calc_name=unit_package.unit_name).values('unit_name__calc_name')
                if ko_qs.exists():
                    ko_unit = ko_qs.first()
                    if '個食' in ko_unit['unit_name__calc_name']:
                        has_koshoku = True

                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count, has_koshoku)

            pouch_meal.unit_aggregate = dp

    def aggregate_mix_rice(self):
        """
        施設毎(混ぜご飯)の集計
        """
        logger.debug('パウチ設計図-施設毎(混ぜご飯)集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchMixrice('施設毎(混ぜご飯)')

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_1'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='main', soup_type='none'
            ).order_by('unit_number')

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_1p += 1

            # 混ぜご飯パーツの有無の判断
            parts_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='parts', soup_type='none'
            )
            dp.has_parts = parts_package_qs.exists()

            # 施設毎の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='main', soup_type='none').order_by('unit_number', '-count')

            # 端数のあるユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)

                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count)

            pouch_meal.unit_mix_rice_aggregate = dp

    def aggregate_soup(self):
        """
        汁(スープ)の集計
        """
        logger.debug('パウチ設計図-汁集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchSoup('汁')

            # 料理名の出力
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='soup'
            )
            if unit_package_qs.exists():
                dp.plate_name = unit_package_qs.first().plate_name

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['SOUP_1'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='soup'
            ).order_by('unit_number', '-count')


            if unit_package_qs.exists():
                dp.plate_name = unit_package_qs.first().plate_name

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_1p += 1

            # 施設毎の集計
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['SOUP_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='soup').order_by('unit_number', 'package_id', '-count')
            if unit_package_qs.exists():
                # ユニット毎の袋数情報取得(施設毎)
                dp.is_unit_pouch_use = True
            else:
                # 端数形式の場合の袋数情報取得(施設毎)
                package_ids = [settings.PICKING_PACKAGES['SOUP_FRACTION'], settings.PICKING_PACKAGES['SOUP_10']]
                unit_package_qs = UnitPackage.objects.filter(
                    cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                    package_id__in=package_ids, is_basic_plate=True, menu_name='常食',
                    mix_rice_type='none', soup_type='soup').order_by('unit_number', 'package_id', '-count')

            # 施設毎の袋数の登録
            for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.package_id)):
                number, package_id = key
                package = PackageMaster.objects.get(id=package_id)
                unit_package = next(group)
                if unit_package.count:
                    unit_qs = UnitMaster.objects.filter(unit_number=number, calc_name=unit_package.unit_name)
                    if unit_qs.exists():
                        unit = unit_qs.first()
                        order_aggreagte = Order.objects.filter(
                            unit_name=unit, eating_day=unit_package.eating_day,
                            meal_name__meal_name=unit_package.meal_name, menu_name__menu_name='常食'
                        ).aggregate(total=Sum("quantity"))

                        dp.add_unit(unit, package.quantity, order_aggreagte['total'], unit_package.count)

            # 嚥下食の集計
            if not PlateNameAnalizeUtil.is_miso_soup(dp.plate_name):
                self.aggregate_enge_soup(pouch_meal, 'ソフト', dp.set_soft_1p, dp.set_soft_2p, dp.add_soft_unit)
                self.aggregate_enge_soup(pouch_meal, 'ミキサー', dp.set_mixer_1p, dp.set_mixer_2p, dp.add_mixer_unit)
                self.aggregate_enge_soup(pouch_meal, 'ゼリー', dp.set_jellry_1p, dp.set_jellry_2p, dp.add_jelly_unit)

            pouch_meal.soup_aggregate = dp

    def aggregate_filling(self):
        """
        汁具の集計
        """
        logger.debug('パウチ設計図-汁具集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchFilling('汁具')

            # 料理名の出力
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='filling'
            ).values('plate_name').distinct()
            if unit_package_qs.exists():
                if len(unit_package_qs) == 1:
                    dp.plate_name = unit_package_qs.first()['plate_name']
                else:
                    dp.plate_name = f'{unit_package_qs.first()["plate_name"]}・{unit_package_qs.last()["plate_name"]}'

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['SOUP_1'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='filling'
            ).order_by('unit_number', 'count')

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key= lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.package_count_1p += 1

            # 施設毎の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['SOUP_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='filling').order_by('unit_number', '-count')

            # 端数のあるユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count)

            # 嚥下食の集計
            self.aggregate_enge_filling(pouch_meal, 'ソフト', dp.set_soft_1p, dp.set_soft_2p, dp.add_soft_unit)
            self.aggregate_enge_filling(pouch_meal, 'ミキサー', dp.set_mixer_1p, dp.set_mixer_2p, dp.add_mixer_unit)
            self.aggregate_enge_filling(pouch_meal, 'ゼリー', dp.set_jellry_1p, dp.set_jellry_2p, dp.add_jelly_unit)

            pouch_meal.filling_aggregate = dp

    def aggregate_enge_filling(self, pouch_meal, menu_name, func_for_1p, func_for_2p, func_for_unit):
        # 1人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_1'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='filling').order_by('plate_name', 'unit_number')

        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_1p(len(list(group)))
                break

        # 2人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_2'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='filling').order_by('unit_number')
        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_2p(len(list(group)))
                break

        # 可変サイズ
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id__in=self.ENGE_PACKAGE_ID_LIST, is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='filling').order_by('unit_number', 'package_id', '-count')

        for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.package_id)):
            number, package_id = key
            package = PackageMaster.objects.get(id=package_id)
            unit_package = next(group)
            if unit_package.count:
                unit_qs = UnitMaster.objects.filter(unit_number=number, calc_name=unit_package.unit_name)
                if unit_qs.exists():
                    unit = unit_qs.first()
                    order_aggreagte = Order.objects.filter(
                        unit_name=unit, eating_day=unit_package.eating_day,
                        meal_name__meal_name=unit_package.meal_name, menu_name__menu_name=menu_name
                    ).aggregate(total=Sum("quantity"))

                    func_for_unit(unit_package.unit_number, unit_package.unit_name, package.quantity,
                                  order_aggreagte['total'], unit_package.count)

    def aggregate_enge_soup(self, pouch_meal, menu_name, func_for_1p, func_for_2p, func_for_unit):
        # 1人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_1'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='soup').order_by('unit_number')

        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_1p(len(list(group)))
                break

        # 2人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_2'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='soup').order_by('unit_number')
        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_2p(len(list(group)))
                break

        # 可変サイズ
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id__in=self.ENGE_PACKAGE_ID_LIST, is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='soup').order_by('unit_number', 'package_id', '-count')

        for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.package_id)):
            number, package_id = key
            package = PackageMaster.objects.get(id=package_id)
            unit_package = next(group)
            if unit_package.count:
                unit_qs = UnitMaster.objects.filter(unit_number=number, calc_name=unit_package.unit_name)
                if unit_qs.exists():
                    unit = unit_qs.first()
                    order_aggreagte = Order.objects.filter(
                        unit_name=unit, eating_day=unit_package.eating_day,
                        meal_name__meal_name=unit_package.meal_name, menu_name__menu_name=menu_name
                    ).aggregate(total=Sum("quantity"))

                    func_for_unit(unit_package.unit_number, unit_package.unit_name, package.quantity,
                                  order_aggreagte['total'], unit_package.count)

    def aggregate_enge_dry_units(self, pouch_meal, menu_name, func_for_1p, func_for_2p, func_for_unit, unit_numbers, unit_names, plate_name):
        # 1人袋
        unit_package_qs = UnitPackage.objects.filter(
            unit_number__in=unit_numbers, unit_name__in=unit_names, plate_name=plate_name,
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_1'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('unit_number')

        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                logger.debug(f'錦糸卵-嚥下{menu_name}1p:{key}')
                func_for_1p(len(list(group)))
                break

        # 2人袋
        unit_package_qs = UnitPackage.objects.filter(
            unit_number__in=unit_numbers, unit_name__in=unit_names, plate_name=plate_name,
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_2'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('unit_number')
        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_2p(len(list(group)))
                break

        # 可変サイズ
        unit_package_qs = UnitPackage.objects.filter(
            unit_number__in=unit_numbers, unit_name__in=unit_names, plate_name=plate_name,
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id__in=self.ENGE_PACKAGE_ID_LIST, is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('unit_number', 'package_id', '-count')

        for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.package_id)):
            number, package_id = key
            package = PackageMaster.objects.get(id=package_id)
            unit_package = next(group)
            if unit_package.count:
                unit_qs = UnitMaster.objects.filter(unit_number=number, calc_name=unit_package.unit_name)
                if unit_qs.exists():
                    unit = unit_qs.first()
                    order_aggreagte = Order.objects.filter(
                        unit_name=unit, eating_day=unit_package.eating_day,
                        meal_name__meal_name=unit_package.meal_name, menu_name__menu_name=menu_name
                    ).aggregate(total=Sum("quantity"))

                    func_for_unit(unit_package.unit_number, unit_package.unit_name, package.quantity,
                                  order_aggreagte['total'], unit_package.count)

    def aggregate_enge(self):
        """
        嚥下の集計
        """
        logger.info('パウチ設計図-嚥下集計開始')

        for pouch_meal in self.meal_list:
            dp_soft = DesignedPouchEnge('(ソフト)')
            dp_soft.preserve_1p = 4
            dp_jelly = DesignedPouchEnge('(ゼリー)')
            dp_mixer = DesignedPouchEnge('(ミキサー)')

            self.aggregate_enge_detail(pouch_meal, 'ソフト', dp_soft.set_1p, dp_soft.set_2p, dp_soft.add_unit)
            self.aggregate_enge_detail(pouch_meal, 'ゼリー', dp_jelly.set_1p, dp_jelly.set_2p, dp_jelly.add_unit)
            self.aggregate_enge_detail(pouch_meal, 'ミキサー', dp_mixer.set_1p, dp_mixer.set_2p, dp_mixer.add_unit)

            pouch_meal.soft_aggregate = dp_soft
            pouch_meal.jelly_aggregate = dp_jelly
            pouch_meal.mixer_aggregate = dp_mixer

    def aggregate_enge_detail(self, pouch_meal, menu_name, func_for_1p, func_for_2p, func_for_unit):
        logger.info(f'嚥下詳細:{menu_name}-{pouch_meal.eating_day}-{pouch_meal.meal_name}')
        # 1人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_1'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('plate_name')

        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                if self._is_indivisual_dry_plate(key):
                    continue
                p1_list = list(group)
                logger.info(f'嚥下1人袋リスト:{p1_list}')
                func_for_1p(len(p1_list))
                break

        # 2人袋
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id=settings.PICKING_PACKAGES['ENGE_2'], is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('unit_number')
        if unit_package_qs.exists():
            for key, group in groupby(unit_package_qs, key=lambda x: x.plate_name):
                func_for_2p(len(list(group)))
                break

        # 可変サイズ
        unit_package_qs = UnitPackage.objects.filter(
            cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
            package_id__in=self.ENGE_PACKAGE_ID_LIST, is_basic_plate=True, menu_name=menu_name,
            mix_rice_type='none', soup_type='none').order_by('unit_number', 'package_id')

        for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.package_id)):
            number, package_id = key
            logger.debug(f'{number}-{package_id}')
            package = PackageMaster.objects.get(id=package_id)
            unit_package = next(group)
            if unit_package.count:
                # 注文数の取得
                unit_qs = UnitMaster.objects.filter(unit_number=number, calc_name=unit_package.unit_name)
                if unit_qs.exists():
                    unit = unit_qs.first()
                    order_aggreagte = Order.objects.filter(
                        unit_name=unit, eating_day=unit_package.eating_day,
                        meal_name__meal_name=unit_package.meal_name, menu_name__menu_name=menu_name
                    ).aggregate(total=Sum("quantity"))

                    # 施設情報の登録
                    logger.debug(f'嚥下施設読込{unit.unit_number}.{unit.short_name}-{package.quantity}-{menu_name}')
                    func_for_unit(
                        unit, package.quantity, order_aggreagte['total'],
                        unit_package.count)
                else:
                    logger.warning(f'ユニットマスタ該当なし：{unit_number}.{unit_name}')
            else:
                logger.warning(f'ユニットマスタ-パッケージ該当なし：{number}-{package_id}')

    def aggregate_allergen(self):
        """
        アレルギーの集計
        """
        logger.debug('パウチ設計図-アレルギー集計開始')

        for pouch_meal in self.meal_list:
            dp = DesignedPouchAllergen('アレルギー')

            unit_package_qs = UnitPackage.objects.filter(
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                is_basic_plate=False).order_by('unit_number', 'menu_name', 'index')

            for key, group in groupby(unit_package_qs, key=lambda x: (x.unit_number, x.menu_name, x.index)):
                unit_package = next(group)
                number, menu, plate_index = key
                if unit_package.count:
                    dp.add_allergen_plate(number, unit_package.unit_name, plate_index, menu, unit_package.count)

            pouch_meal.allergen_aggregate = dp

    def _get_units_for_dry_plate(self):
        qs = UnitMaster.objects.filter(unit_code__gt=0, username__dry_cold_type='乾燥').values('unit_number', 'calc_name').\
            exclude(unit_code__range=[80001, 80008]).exclude(calc_name=None)

        return list(qs)

    def _is_indivisual_dry_plate(self, plate_name: str) -> bool:
        """
        別ページに個別出力(方式は施設毎)する料理かどうかを判定する。現在は乾燥錦糸卵のみ。
        ※乾燥施設かどうかの判定はここでは行わない。
        """
        if ('錦糸卵' in plate_name) or ('きんしたまご' in plate_name):
            return True
        else:
            return False

    def aggregate_unit_for_dry_plate(self):
        """
        乾燥原体特別表示用の集計
        """
        logger.debug('乾燥原体施設毎-集計開始')

        units = self._get_units_for_dry_plate()
        unit_numbers = [x['unit_number'] for x in units]
        unit_names = [x['calc_name'] for x in units]
        for pouch_meal in self.meal_list:
            dp = DesignedUnitForDryRawPlate('乾燥原体')

            # 料理名の出力
            unit_package_qs = UnitPackage.objects.filter(
                unit_number__in=unit_numbers, unit_name__in=unit_names,
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none'
            ).values('plate_name').distinct()
            for up in unit_package_qs:
                if self._is_indivisual_dry_plate(up['plate_name']):
                    dp.plate_name = up['plate_name']
            if not dp.plate_name:
                continue

            # 1人袋の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                unit_number__in=unit_numbers, unit_name__in=unit_names, plate_name=dp.plate_name,
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_1'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none'
            ).order_by('unit_number', 'count')

            # 各ユニットの1人用袋の集計
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    logger.debug(f'錦糸卵-基本食1p:{key}')
                    dp.package_count_1p += 1

            # 施設毎の集計
            # ユニット毎の袋数情報取得
            unit_package_qs = UnitPackage.objects.filter(
                unit_number__in=unit_numbers, unit_name__in=unit_names, plate_name=dp.plate_name,
                cooking_day=self.cooking_day, eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name,
                package_id=settings.PICKING_PACKAGES['BASIC_UNIT'], is_basic_plate=True, menu_name='常食',
                mix_rice_type='none', soup_type='none').order_by('unit_number', '-count')

            # 端数のあるユニットの登録
            for key, group in groupby(unit_package_qs, key=lambda x: x.unit_number):
                unit_package = next(group)
                if unit_package.count:
                    dp.add_unit(unit_package.unit_number, unit_package.unit_name, unit_package.count)

            # 嚥下食の集計
            self.aggregate_enge_dry_units(pouch_meal, 'ソフト', dp.set_soft_1p, dp.set_soft_2p, dp.add_soft_unit, unit_numbers, unit_names, dp.plate_name)
            self.aggregate_enge_dry_units(pouch_meal, 'ミキサー', dp.set_mixer_1p, dp.set_mixer_2p, dp.add_mixer_unit, unit_numbers, unit_names, dp.plate_name)
            self.aggregate_enge_dry_units(pouch_meal, 'ゼリー', dp.set_jellry_1p, dp.set_jellry_2p, dp.add_jelly_unit, unit_numbers, unit_names, dp.plate_name)

            pouch_meal.unit_for_dry_aggregate = dp



class PouchDesignWriter:
    """
    パウチ設計図を出力するクラス
    """
    DEFAULT_KOBACHI_GRAM = 40.0
    PAGE_MAX_ROW = 42
    PAGE_MAX_ROW_ENGE = 36
    NEXT_PAGE_ROW = 42
    NEXT_PAGE_ROW_ENGE = 36

    plate_title_fill = PatternFill(patternType='solid', fgColor='A9D08E')
    enge_plate_titele_fill = PatternFill(patternType='solid', fgColor='D9E1F2')
    desigined_basic_table_fill = PatternFill(patternType='solid', fgColor='FFF2CC')
    desigined_unit_table_fill = PatternFill(patternType='solid', fgColor='E2EFDA')
    desigined_soup_table_fill = PatternFill(patternType='solid', fgColor='F5E8FE')

    plate_soft_fill = PatternFill(patternType='solid', fgColor='FFCCFF')
    plate_jelly_fill = PatternFill(patternType='solid', fgColor='F8CBAD')
    plate_mixer_fill = PatternFill(patternType='solid', fgColor='FFFF99')

    plate_breakfast_fill = PatternFill(patternType='solid', fgColor='FCE4D6')
    plate_lunch_fill = PatternFill(patternType='solid', fgColor='FFE699')
    plate_dinner_fill = PatternFill(patternType='solid', fgColor='DDEBF7')

    designed_basic_outer_side = Side(style='thick', color="000000")
    designed_soup_outer_side = Side(style='thick', color="FF0000")
    designed_filling_outer_side = Side(style='thick', color="0000FF")
    designed_unit_table_outer_side = Side(style='medium', color="000000")
    designed_unit_table_inner_side = Side(style='thin', color="000000")

    eating_font = Font(name='游ゴシック', size=36, bold=True)
    plate_number_font = Font(name='游ゴシック', size=28, bold=True)
    plate_font = Font(name='游ゴシック', size=28, bold=True)
    plate_warning_font = Font(name='游ゴシック', size=28, bold=True, color="FF0000")
    plate_small_font = Font(name='游ゴシック', size=20, bold=True)
    allergen_small_font = Font(name='游ゴシック', size=16, bold=True)
    allergen_alter_font = Font(name='游ゴシック', size=28, bold=True, u='single')
    allergen_small_alter_font = Font(name='游ゴシック', size=16, bold=True, u='single')
    allergen_unit_font = Font(name='游ゴシック', size=28, bold=True)

    design_title_font = Font(name='游ゴシック', size=36, bold=True, u='single')
    design_soup_title_font = Font(name='游ゴシック', size=36, bold=True, u='single', color="FF0000")
    design_filling_title_font = Font(name='游ゴシック', size=36, bold=True, u='single', color="0000FF")
    design_font = Font(name='游ゴシック', size=28, bold=True)
    design_emphasis_font = Font(name='游ゴシック', size=28, bold=True, color="FF0000")

    center_alignment = Alignment(horizontal="center")

    def __init__(self, aggregate: PouchAggregate):
        self.aggreagte = aggregate

        self.cooking_day = self.aggreagte.cooking_day
        self.plates_qs = None
        self.plate_index = None

        self.miso_fill_dict = False
        self.miso_soup_dict = False

        self.enge_package_list = []
        self.food_count = 0

    def open_template(self):
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/pouch_design.xlsx')  # 書き起こし票のテンプレート
        return excel.load_workbook(template_path)

    def remove_template_sheet(self, wb):
        wb.remove(wb["basic"])
        wb.remove(wb["plate"])
        wb.remove(wb["enge"])

    def save(self, wb):
        """
        チラー毎の指示書エクセルファイルを保存する。
        """
        output_dir = os.path.join(settings.OUTPUT_DIR, 'pouch_design')
        os.makedirs(output_dir, exist_ok=True)  # 上書きOK

        save_path = os.path.join(output_dir, f'パウチ設計図_{str(self.cooking_day)}_製造.xlsx')

        for ws in wb.worksheets:
            ws.sheet_view.tabSelected = True
            footer_text = '&P / &Nページ'
            ws.oddFooter.center.text = footer_text
            ws.HeaderFooter.oddFooter.center.size = 48
        wb.save(save_path)
        wb.close()

    def prpare_work_sheet(self, wb, src_sheet_name: str, dst_sheet_name: str):
        """
        テンプレートのワークシートをコピーし、作業量のワークシートを準備する。
        """
        ws = wb.copy_worksheet(wb[src_sheet_name])
        ws.title = dst_sheet_name
        ws.print_title_rows = '1:2'

        return ws

    def _get_weekday_for_excel(self, date):
        weekday = date.weekday()
        if weekday == 6:
            return '(日Sun.)'
        elif weekday == 0:
            return '(月Mon.)'
        elif weekday == 1:
            return '(火Tue.)'
        elif weekday == 2:
            return '(水Wed.)'
        elif weekday == 3:
            return '(木Thu.)'
        elif weekday == 4:
            return '(金Fri.)'
        else:
            return '(土Sat.)'

    def _get_meal_name_for_excel(self, meal_name: str):
        if '朝' in meal_name:
            return '△朝(あさ)'
        elif '昼' in meal_name:
            return '〇昼(ひる)'
        elif '夕' in meal_name:
            return '□夕(ゆう)'
        else:
            return meal_name

    def write_eating_time_for_plate(self, ws, eating_day, pouch_meal: PouchMeal, start_row: int):
        """
        喫食タイミングの出力(料理シート用)
        """

        meal = pouch_meal.meal_name
        if '朝' in meal:
            col = 2
            p7_fill = self.plate_breakfast_fill
        elif '昼' in meal:
            col = 10
            p7_fill = self.plate_lunch_fill
        elif '夕' in meal:
            col = 18
            p7_fill = self.plate_dinner_fill
        else:
            raise RuntimeError('食事区分に想定外の内容')

        current_row = start_row
        if '昼' in pouch_meal.meal_name:
            # 昼食料理の出力は、見間違いを防ぐため、1行下げる
            current_row += 1

        # 喫食日の出力
        ws.cell(current_row, col).font = self.eating_font
        ws.cell(current_row, col, f'{eating_day.strftime("%Y/%m/%d")}{self._get_weekday_for_excel(eating_day)}')
        for tmp_i in range(7):
            ws.cell(current_row, col + tmp_i).fill = p7_fill

        # 食事区分の出力
        ws.cell(current_row + 1, col).font = self.eating_font
        ws.cell(current_row + 1, col, self._get_meal_name_for_excel(pouch_meal.meal_name))
        for tmp_i in range(7):
            ws.cell(current_row + 1, col + tmp_i).fill = p7_fill

    def write_eating_time(self, ws, pouch_meal: PouchMeal, start_row: int):

        # 喫食日の出力
        eating_day = pouch_meal.eating_day
        ws.cell(start_row + 1, 2).font = self.eating_font
        ws.cell(start_row + 1, 2, f'{eating_day.strftime("%Y/%m/%d")}{self._get_weekday_for_excel(eating_day)}')

        # 食事区分の出力
        ws.cell(start_row + 1, 10).font = self.eating_font
        ws.cell(start_row + 1, 10, self._get_meal_name_for_excel(pouch_meal.meal_name))

    def _get_plate_gram(self, plate_name: str):
        """
        料理のg数を取得する。gでない単位の場合は小鉢のデフォルト値(中サイズ対象となるg数)を返却する。
        """
        # 全角半角変換(+は全角に、gは半角に揃える)
        replaced = plate_name.replace('+', '＋').replace('ｇ', 'g')
        plus_index = replaced.find('＋')
        if plus_index != -1:
            replaced = replaced[:plus_index]
        res = re.findall('(\d+|\d+\.\d+)g', replaced)
        if res:
            return float(res[0])

        return self.DEFAULT_KOBACHI_GRAM

    def _get_plate_total_gram(self, plate_name: str):
        """
        料理のg数を取得する。gでない単位の場合は小鉢のデフォルト値(中サイズ対象となるg数)を返却する。
        """
        # 全角半角変換(+は全角に、gは半角に揃える)
        replaced = plate_name.replace('+', '＋').replace('%', '％').replace('ｇ', 'g').replace('（', '(').replace('）', ')')

        enable_units = settings.MEASURE_ENABLE_UNITS + ['個', '丁']
        not_gram_unit_list = "|".join(enable_units)
        res = re.findall(f'({not_gram_unit_list})\((\d+|\d+\.\d+)g\)＋(\D*)(\d+|\d+\.\d+)％', replaced)
        if res:
            src = float(res[0][1])
            return src + src * float(res[0][3]) / 100

        res = re.findall('(\d+|\d+\.\d+)g＋(\D*)(\d+|\d+\.\d+)g', replaced)
        if res:
            return float(res[0][0]) + float(res[0][2])
        else:
            res = re.findall('(\d+|\d+\.\d+)g＋(\D*)(\d+|\d+\.\d+)％', replaced)
            if res:
                src = float(res[0][0])
                return src + src * float(res[0][2]) / 100
            else:
                plus_index = replaced.find('＋')
                if plus_index != -1:
                    replaced = replaced[:plus_index]
                res = re.findall('(\d+|\d+\.\d+)g', replaced)
                if res:
                    return float(res[0])

        return self.DEFAULT_KOBACHI_GRAM

    def _get_sub_plate_package(self, plate_name):
        """
        副菜の袋サイズ判断
        """
        logger.info(f'袋サイズ判断:{plate_name}')
        enge_size = 7 if ('△' in plate_name) else 14
        total_gram = self._get_plate_total_gram(plate_name)
        if (total_gram < 20.0) or ('▼' in plate_name):
            if '△' in plate_name:
                self._add_enge_package(7)
                return 'しせつごと(嚥下7人用)'
            else:
                self._add_enge_package(20)
                return 'しせつごと(嚥下20人用)'
        elif total_gram >= 50.0:
            self._add_enge_package(enge_size)
            return f'大(嚥下{enge_size}人用)'
        elif total_gram >= 40.0:
            self._add_enge_package(enge_size)
            return f'中(嚥下{enge_size}人用)'
        else:
            self._add_enge_package(enge_size)
            return f'小(嚥下{enge_size}人用)'

    def _add_enge_package(self, size):
        logger.info(f'{size}人用追加-{self.enge_package_list}')
        if not (size in self.enge_package_list):
            self.enge_package_list.append(size)

    def _get_package(self, plate, prev_plate_name):
        """
        袋サイズを取得する。書き起こし票と同じ内容。
        """
        logger.info(f'袋サイズ判断2:{prev_plate_name}')
        number = plate.plate_name[0]
        plate_name = plate.plate_name

        if number == '⑩':
            self._add_enge_package(7)
            return '中(嚥下7人用)'
        if number == '①':
            if PlateNameAnalizeUtil.is_sansyokudon_plate(plate_name):
                self._add_enge_package(14)
                return '小(嚥下14人用)'
            elif plate.is_mix_rice:
                total_gram = self._get_plate_total_gram(plate.plate_name)
                is_not_big, is_unit = PlateNameAnalizeUtil.is_sub_package_size_enge_mix_rice(plate.plate_name, total_gram)
                if is_not_big:
                    if is_unit:
                        self._add_enge_package(20)
                    else:
                        self._add_enge_package(14)

                    return '中(嚥下14人用)'
                else:
                    self._add_enge_package(7)
                    return '大(嚥下7人用)'
            else:
                self._add_enge_package(7)
                return '大(嚥下7人用)'
        elif (number == '②') or (number == '③'):
            return self._get_sub_plate_package(plate_name)
        elif number == '④':
            if '錦糸卵' in plate_name:
                self._add_enge_package(20)
            if plate.is_mix_rice:
                #　具材量の判定が必要になるかも?
                #self._add_enge_package(20)
                return '中'
            else:
                is_special_enge = prev_plate_name and ('①小松菜と団子の煮物' in prev_plate_name) and ('④里芋煮物2個' in plate.plate_name)
                is_7p_size = is_special_enge or ('△' in plate_name)
                size = 7 if is_7p_size else 14
                gram = self._get_plate_gram(plate_name)
                if (gram < 20.0) or ('▼' in plate_name):
                    total_gram = self._get_plate_total_gram(plate_name)
                    if (total_gram < 20.0) or ('▼' in plate_name):
                        if '△' in plate_name:
                            # △と▼が同時指定は運用外とする
                            self._add_enge_package(7)
                            return 'しせつごと(嚥下7人用)'
                        else:
                            self._add_enge_package(20)
                            return 'しせつごと(嚥下20人用)'
                    elif total_gram >= 50.0:
                        self._add_enge_package(size)
                        return f'大(嚥下{size}人用)'
                    elif total_gram >= 40.0:
                        self._add_enge_package(size)
                        return f'中(嚥下{size}人用)'
                    else:
                        self._add_enge_package(size)
                        return f'小(嚥下{size}人用)'
                else:
                    if gram >= 50:
                        self._add_enge_package(size)
                        return f'大(嚥下{size}人用)'
                    elif gram >= 40:
                        self._add_enge_package(size)
                        return f'中(嚥下{size}人用)'
                    else:
                        self._add_enge_package(size)
                        return f'小(嚥下{size}人用)'
        elif number == '⑤':
            if not PlateNameAnalizeUtil.is_miso_soup(plate_name):
                if PlateNameAnalizeUtil.is_soup_liquid(plate_name):
                    # 味噌汁以外の汁・スープの液の料理は、小で出力する(年に1,2回程度中のケースがあるがそちらは運用対応)
                    logger.info(f'汁サイズ-小:{self.enge_package_list}')
                    return '小'

            return '中'
        else:
            # 本来この内容は出てはならない
            return '未定'

    def _is_need_adjust_index(self, plate):
        if '◆' in plate.plate_name:
            res = re.findall('具(\d+|\d+\.\d+)[g|ｇ]\s*\D液(\d+|\d+\.\d+)[g|ｇ]', plate.plate_name)
            if res and (res[0][0] and res[0][1]):
                return True
        if ('カレーライス' in plate.plate_name) or ('シチュー' in plate.plate_name):
            return True
        else:
            return False

    def _is_indivisual_display_plate(self, plate_name: str) -> bool:
        """
        別ページに個別出力(方式は施設毎)する料理かどうかを判定する。現在は乾燥錦糸卵のみ。
        ※乾燥施設かどうかの判定はここでは行わない。
        """
        if ('錦糸卵' in plate_name) or ('きんしたまご' in plate_name):
            return True
        else:
            return False

    def write_plate_name(self, ws, row, plate, package, is_adjusted, eating_day, is_enge: bool=False):
        """
        パウチ設計図に料理名を出力する。
        """
        # スタイルの設定
        # P7番号は色を付けない

        # 料理名
        ws.cell(row+1, 2).fill = self.plate_title_fill
        ws.cell(row+1, 3).fill = self.plate_title_fill
        ws.cell(row+1, 4).fill = self.plate_title_fill
        ws.cell(row+1, 5).fill = self.plate_title_fill
        ws.cell(row+1, 6).fill = self.plate_title_fill
        ws.cell(row+1, 7).fill = self.plate_title_fill
        ws.cell(row+1, 8).fill = self.plate_title_fill
        ws.cell(row+1, 2).font = self.plate_font

        # P7番号の作成
        food_count = self.food_count
        is_write_warning = False
        if self._is_need_adjust_index(plate):
            basic_number_index = P7Util.get_number_index(food_count)
            enge_number_index = P7Util.get_number_index(food_count)
            enge_append_index = P7Util.get_number_index(food_count + 1)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, self.cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_number_index}'
            soft_append_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_append_index}'

            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_number_index}'
            jelly_append_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_append_index}'

            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_number_index}'
            mixer_append_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_append_index}'

            p7_numbers = "、".join([
                basic_number,
                soft_number, soft_append_number,
                jelly_number, jelly_append_number,
                mixer_number, mixer_append_number
            ])

            is_write_warning = True
            result = food_count + 3
        elif PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
            p7_numbers = P7Util.get_number_miso_soup()
            result = food_count
        else:
            basic_number_index = P7Util.get_number_index(food_count)
            if is_adjusted:
                enge_number_index = P7Util.get_number_index(food_count + 1)
            else:
                enge_number_index = P7Util.get_number_index(food_count)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, self.cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_number_index}'
            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_number_index}'
            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_number_index}'
            p7_numbers = "、".join([basic_number, soft_number, jelly_number, mixer_number])

            result = food_count + 1

        # P7番号の出力
        ws.cell(row, 2).font = self.plate_font
        ws.cell(row, 2, p7_numbers)
        if len(p7_numbers) >= 26:
            ws.cell(row, 2).font = self.plate_small_font

        # 料理名の出力
        if PlateNameAnalizeUtil.is_raw_plate(plate):
            if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate.plate_name, eating_day)[0]:
                if PlateNameAnalizeUtil.is_required_dry_notice(plate):
                    write_name = f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)'
                else:
                    write_name = f'☆{plate.plate_name}　{package}'
            else:
                if PlateNameAnalizeUtil.is_required_dry_notice(plate):
                    write_name = f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)'
                else:
                    write_name = f'☆{plate.plate_name}'
            ws.cell(row + 1, 2, write_name)
            if len(write_name) >= 26:
                ws.cell(row + 1, 2).font = self.plate_small_font
        else:
            write_name = f'{plate.plate_name}　{package}'
            ws.cell(row + 1, 2, write_name)
            if len(write_name) >= 26:
                ws.cell(row + 1, 2).font = self.plate_small_font
        if is_write_warning and is_enge:
            ws.cell(row + 2, 2, '※嚥下食は具と液の袋2つ')
            ws.cell(row + 2, 2).font = self.plate_warning_font

        self.food_count = result

    def write_plate_name_for_plate_sheet(self, ws, row, col, plate, package, is_adjusted, eating_day, meal):
        """
        パウチ設計図の料理名シートに料理名を出力する。
        """
        # スタイルの設定

        # 料理名
        ws.cell(row+1, col).fill = self.plate_title_fill
        ws.cell(row+1, col + 1).fill = self.plate_title_fill
        ws.cell(row+1, col + 2).fill = self.plate_title_fill
        ws.cell(row+1, col + 3).fill = self.plate_title_fill
        ws.cell(row+1, col + 4).fill = self.plate_title_fill
        ws.cell(row+1, col + 5).fill = self.plate_title_fill
        ws.cell(row+1, col + 6).fill = self.plate_title_fill
        ws.cell(row+1, col).font = self.plate_font

        # P7番号の作成
        food_count = self.food_count
        is_write_warning = False
        if self._is_need_adjust_index(plate):
            basic_number_index = P7Util.get_number_index(food_count)
            enge_number_index = P7Util.get_number_index(food_count)
            enge_append_index = P7Util.get_number_index(food_count + 1)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, self.cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_number_index}'
            soft_append_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_append_index}'

            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_number_index}'
            jelly_append_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_append_index}'

            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_number_index}'
            mixer_append_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_append_index}'

            p7_numbers = "、".join([
                basic_number,
                soft_number, soft_append_number,
                jelly_number, jelly_append_number,
                mixer_number, mixer_append_number
            ])

            is_write_warning = True
            result = food_count + 3
        elif PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
            p7_numbers = P7Util.get_number_miso_soup()
            result = food_count
        else:
            basic_number_index = P7Util.get_number_index(food_count)
            if is_adjusted:
                enge_number_index = P7Util.get_number_index(food_count + 1)
            else:
                enge_number_index = P7Util.get_number_index(food_count)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, self.cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, self.cooking_day)}{enge_number_index}'
            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, self.cooking_day)}{enge_number_index}'
            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, self.cooking_day)}{enge_number_index}'
            p7_numbers = "、".join([basic_number, soft_number, jelly_number, mixer_number])

            result = food_count + 1

        # P7番号の出力
        ws.cell(row, col).font = self.plate_font
        if '朝' in meal:
            p7_fill = self.plate_breakfast_fill
        elif '昼' in meal:
            p7_fill = self.plate_lunch_fill
        elif '夕' in meal:
            p7_fill = self.plate_dinner_fill
        for tmp_i in range(7):
            ws.cell(row, col + tmp_i).fill = p7_fill
        ws.cell(row, col, p7_numbers)

        # 料理名の出力
        if PlateNameAnalizeUtil.is_raw_plate(plate):
            if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate.plate_name, eating_day)[0]:
                if PlateNameAnalizeUtil.is_required_dry_notice(plate):
                    write_name = f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)'
                else:
                    write_name = f'☆{plate.plate_name}　{package}'
            else:
                if PlateNameAnalizeUtil.is_required_dry_notice(plate):
                    write_name = f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)'
                else:
                    write_name = f'☆{plate.plate_name}'
            ws.cell(row + 1, col, write_name)
            if len(write_name) >= 26:
                ws.cell(row + 1, col).font = self.plate_small_font
        else:
            write_name = f'{plate.plate_name}　{package}'
            ws.cell(row + 1, col, write_name)
            if len(write_name) >= 26:
                ws.cell(row + 1, col).font = self.plate_small_font
        if is_write_warning:
            ws.cell(row + 2, col, '※嚥下食は具と液の袋2つ')
            ws.cell(row + 2, col).font = self.plate_warning_font

        self.food_count = result

    def read_plates(self, pouch_meal: PouchMeal, only_not_soup: bool = True):
        # 料理データの取得
        if only_not_soup:
            self.plates_qs = CookingDirectionPlate.objects.filter(
                cooking_day=self.cooking_day, is_basic_plate=True, is_soup=False,
                eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name
            ).order_by('index')
        else:
            self.plates_qs = CookingDirectionPlate.objects.filter(
                cooking_day=self.cooking_day, is_basic_plate=True,
                eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name
            ).order_by('index')

        self.plate_index = None

    def is_plate_writing_over(self, row: int, plate: CookingDirectionPlate, page: int, is_enge: bool = False):
        # 料理名出力分
        count = 1

        length = self.simulate_write_allergen(row, plate)
        count += length

        logger.debug(f'改ページ予想:{row + count}')
        if is_enge:
            return (row + count) >= self.PAGE_MAX_ROW_ENGE * page
        else:
            return (row + count) >= self.PAGE_MAX_ROW * page

    def write_plates_plate_sheet(self, ws, start_row: int, pouch):
        """
        料理情報の出力。味噌汁の汁は出力しない。
        """
        meal = pouch.meal_name
        if '朝' in meal:
            col = 2
        elif '昼' in meal:
            col = 10
        elif '夕' in meal:
            col = 18
        else:
            raise RuntimeError('食事区分に想定外の内容')

        if '昼' in meal:
            row = start_row + 3
        else:
            row = start_row + 2
        is_adjusted = False
        prev_plate = None

        for index, iter_value in enumerate(groupby(self.plates_qs, key=lambda x: x.index)):
            key, group = iter_value

            for plate in group:
                # 味噌汁の汁は出力しない
                if PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
                    continue

                # 料理名の出力
                self.write_plate_name_for_plate_sheet(
                    ws, row, col, plate, self._get_package(plate, prev_plate), is_adjusted, pouch.eating_day, pouch.meal_name)
                row += 2

                if self._is_need_adjust_index(plate):
                    is_adjusted = True

                # 次の料理のために空白行を入れる
                prev_plate = plate.plate_name
                row += 1

    def write_plates(self, ws, start_row: int, page: int, pouch, is_enge: bool = False):
        """
        料理情報の出力。汁・汁具の料理は出力しない。
        """
        row = start_row + 3
        is_adjusted = False
        is_first = True
        is_break = False
        is_complete = True
        prev_plate = None

        for index, iter_value in enumerate(groupby(self.plates_qs, key=lambda x: x.index)):
            key, group = iter_value

            # 改ページ後の料理を表示できるよう、表示済みをスキップ
            if self.plate_index:
                if index < self.plate_index:
                    continue

            if not is_break:
                self.plate_index = index
            for plate in group:
                if is_enge:
                    # 嚥下のシートの場合、味噌汁の汁は出力しない
                    if PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
                        continue

                if not is_first:
                    # ページ初回料理は全て出力する(運用上、1ページを超える出力はないはず。。。)
                    if self.is_plate_writing_over(row, plate, page, is_enge):
                        is_break = True

                if is_break:
                    # 嚥下登場袋サイズを読み込むため、袋サイズ判定だけは行う
                    self._get_package(plate, prev_plate)
                else:
                    # 料理名の出力
                    self.write_plate_name(
                        ws, row, plate, self._get_package(plate, prev_plate), is_adjusted, pouch.eating_day, is_enge)
                    row += 2

                    if self._is_need_adjust_index(plate):
                        is_adjusted = True

                    # アレルギーの出力
                    row += self.write_allergen(ws, row, plate)

                    # 次の料理のために空白行を入れる
                    prev_plate = plate.plate_name
                    row += 1

            if is_break:
                is_complete = False
            is_first = False

        return row, is_complete

    def write_plates_not_write(self, ws, start_row: int, page: int, pouch, is_enge: bool = False):
        """
        料理情報の出力。汁・汁具の料理は出力しない。
        """
        row = start_row + 3
        is_adjusted = False
        is_first = True
        is_break = False
        is_complete = True
        prev_plate = None

        for index, iter_value in enumerate(groupby(self.plates_qs, key=lambda x: x.index)):
            key, group = iter_value

            # 改ページ後の料理を表示できるよう、表示済みをスキップ
            if self.plate_index:
                if index < self.plate_index:
                    continue

            if not is_break:
                self.plate_index = index
            for plate in group:
                if is_enge:
                    # 嚥下のシートの場合、味噌汁の汁は出力しない
                    if PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
                        continue

                if not is_first:
                    # ページ初回料理は全て出力する(運用上、1ページを超える出力はないはず。。。)
                    if self.is_plate_writing_over(row, plate, page, is_enge):
                        is_break = True

                if is_break:
                    # 嚥下登場袋サイズを読み込むため、袋サイズ判定だけは行う
                    self._get_package(plate, prev_plate)
                else:
                    # 袋サイズ読込
                    self._get_package(plate, prev_plate)
                    row += 2

                    if self._is_need_adjust_index(plate):
                        is_adjusted = True

                    # アレルギーの出力
                    #row += self.write_allergen(ws, row, plate)

                    # 次の料理のために空白行を入れる
                    prev_plate = plate.plate_name
                    row += 1

            if is_break:
                is_complete = False
            is_first = False

        return row, is_complete

    def write_plates_soup(self, ws, pouch_meal: PouchMeal, start_row: int):
        """
        汁・汁具の料理情報の出力。運用上、1ページに収まるはず。。
        """
        # 料理データの取得
        self.plates_qs = CookingDirectionPlate.objects.filter(
            cooking_day=self.cooking_day, is_basic_plate=True, is_soup=True,
            eating_day=pouch_meal.eating_day, meal_name=pouch_meal.meal_name
        ).order_by('index')

        row = start_row
        is_adjusted = False
        food_count = 0

        for index, iter_value in enumerate(groupby(self.plates_qs, key=lambda x: x.index)):
            key, group = iter_value

            self.plate_index = index
            for plate in group:
                # 料理名の出力
                food_count = self.write_plate_name(ws, row, plate, self._get_package(plate, None), is_adjusted, food_count, pouch_meal.eating_day)
                row += 2

                if self._is_need_adjust_index(plate):
                    is_adjusted = True

                # アレルギーの出力
                row += self.write_allergen(ws, row, plate)

                # 次の料理のために空白行を入れる
                row += 1

    def get_kind_menu_name(self, code):
        """
        食種の献立種類を取得する
        """
        if '常' in code:
            return '常食'
        elif 'ソ' in code:
            return 'ソフト'
        elif 'ゼ' in code:
            return 'ゼリー'
        elif 'ミ' in code:
            return 'ミキサー'
        else:
            return ''

    def get_allergens_with_menu(self, code):
        if settings.KIZAWA_RAKUKON_CODE in code:
            if settings.KOSHOKU_UNIT_IDS:
                return ['個食'], '常食'
        if settings.FREEZE_RACKUKON_CODE in code:
            if settings.FREEZE_UNIT_IDS:
                return ['ﾌﾘｰｽﾞ'], '常食'

        # 献立種類名の取得
        menu_name = self.get_kind_menu_name(code)

        # 散発アレルギーから検索
        uncommon_qs = UncommonAllergenHistory.objects.filter(code=code, cooking_day=self.cooking_day, menu_name=menu_name)
        if uncommon_qs.exists():
            allergen = uncommon_qs.first().allergen
            return [allergen], menu_name
        else:
            common_qs = CommonAllergen.objects.filter(code=code, menu_name__menu_name=menu_name)
            if common_qs.exists():
                return [x.allergen for x in common_qs], menu_name
            else:
                return [], None

    def compare_allergen_unit(self, unit_order1, unit_order2):
        seq_dict = {
            '常食': 0,
            'ソフト': 1,
            'ゼリー': 2,
            'ミキサー': 3
        }

        cmp_seq = seq_dict[unit_order1.menu_name.menu_name] - seq_dict[unit_order2.menu_name.menu_name]
        if cmp_seq:
            return cmp_seq
        else:
            cmp_allergen_id = unit_order1.allergen_id - unit_order2.allergen_id
            if cmp_allergen_id:
                return cmp_allergen_id
            return unit_order1.unit_name.unit_number - unit_order2.unit_name.unit_number

    def is_soe_allergen(self, plate_unit_list, prev_unit_list):
        if not prev_unit_list:
            return False
        if len(plate_unit_list) != len(prev_unit_list):
            return False

        for current in plate_unit_list:
            is_find = False
            for prev in prev_unit_list:
                if current == prev:
                    is_find = True
                    break

            if not is_find:
                return False

        return True

    def _get_display_unit_name(self, unit):
        display_unit_number = ''
        if unit.unit_number < 10:
            # 1桁の場合は、半角空白を入れて、2桁の施設名とも並ぶようにするう
            display_unit_number = f' {unit.unit_number}'
        else:
            display_unit_number = f'{unit.unit_number}'

        display_unit_name = unit.short_name
        if len(unit.short_name) < 5:
            diff = 5 - len(unit.short_name)
            # 5文字になるまで全角空白を入れる
            for i in range(diff):
                display_unit_name = display_unit_name + '　'

        return f'{display_unit_number}.{display_unit_name}'

    def _get_enge_cell_color(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.plate_soft_fill
        elif menu_name == 'ゼリー':
            return self.plate_jelly_fill
        elif menu_name == 'ミキサー':
            return self.plate_mixer_fill
        else:
            return None

    def write_allergen(self, ws, row, plate):
        # アレルギー情報の取得
        relations = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')

        col = 2
        prev_unit_list = None
        current_row = row
        prev_row = None
        for key, group in groupby(relations, key=lambda x: x.plate):
            # 代替先料理名の表示
            ws.cell(current_row, 2).font = self.allergen_alter_font
            ws.cell(current_row, 2, f'{key.plate_name}')

            allergen_unit_list = []
            for allergen_plate in group:
                code = allergen_plate.code
                allergen_list, menu_name = self.get_allergens_with_menu(code)
                for allergen in allergen_list:
                    # 対象アレルギーを注文している施設を取得
                    if allergen == '個食':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.KOSHOKU_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    elif allergen == 'ﾌﾘｰｽﾞ':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.FREEZE_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    else:
                        if key.is_soup:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name,
                                meal_name__filling=True
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                        else:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    logger.debug(f'{plate.eating_day}-{allergen}-{plate.meal_name}-{menu_name}:')
                    logger.debug(f'{order_qs}:')
                    logger.debug('')

                    allergen_unit_list += list(order_qs)

            # 施設情報のソート
            allergen_orders = sorted(allergen_unit_list, key=cmp_to_key(self.compare_allergen_unit))

            if self.is_soe_allergen(allergen_unit_list, prev_unit_list):
                ws.cell(prev_row, 2).font = self.allergen_small_alter_font
                ws.cell(prev_row, 2).value = f'{ws.cell(prev_row, 2).value}、{key.plate_name}'
                ws.cell(current_row, 2).value = ''

                # 次の行にいかないように制御
                current_row -= 1
            else:
                prev_row = current_row

                current_row += 1

                prev_unit_list = []
                prev_unit_list += allergen_unit_list

                # 施設情報の出力
                total_count = 0
                prev_menu_name = None
                for order in allergen_orders:
                    unit_col = col
                    if order.menu_name.menu_name == '常食':
                        kana_name = order.allergen.kana_name
                        if not kana_name:
                            if '個食' in order.unit_name.unit_name:
                                kana_name = 'こしょく'
                            elif 'ﾌﾘｰｽﾞ' in order.unit_name.unit_name:
                                kana_name = 'ﾌﾘｰｽﾞ'
                        # 出力内容(常食はセル色を変えない)
                        label = f'{self._get_display_unit_name(order.unit_name)}　{kana_name}　{order.unit_quantity}'
                    else:
                        # セルの色付け
                        fill_color = self._get_enge_cell_color(order.menu_name.menu_name)
                        if fill_color:
                            for i in range(6):
                                ws.cell(current_row, unit_col + i).fill = fill_color

                        # 出力内容
                        label = f'{self._get_display_unit_name(order.unit_name)}　{order.allergen.kana_name}　{order.menu_name.menu_name} {order.unit_quantity}'

                    # アレルギー対象施設の出力
                    if len(label) > 16:
                        ws.cell(current_row, unit_col).font = self.allergen_small_font
                    else:
                        ws.cell(current_row, unit_col).font = self.allergen_unit_font
                    ws.cell(current_row, unit_col, f'{label}')

                    if prev_menu_name:
                        # 料理単位の合計数を表示
                        if prev_menu_name != order.menu_name.menu_name:
                            ws.cell(current_row - 1, 8).font = self.allergen_unit_font
                            ws.cell(current_row - 1, 8, f'合計:{total_count}、保存:1')
                            total_count = 0
                    prev_menu_name = order.menu_name.menu_name
                    total_count += math.ceil(order.unit_quantity / 10)
                    current_row += 1

                if total_count:
                    ws.cell(current_row - 1, 8).font = self.allergen_unit_font
                    ws.cell(current_row - 1, 8, f'合計:{total_count}、保存:1')

            current_row += 1

        return current_row - row

    def simulate_write_allergen(self, row, plate):
        # アレルギー情報の取得
        relations = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')

        col = 2
        index = 0
        prev_unit_list = None
        current_row = row
        for key, group in groupby(relations, key=lambda x: x.plate):
            allergen_unit_list = []
            for allergen_plate in group:
                code = allergen_plate.code
                allergen_list, menu_name = self.get_allergens_with_menu(code)
                for allergen in allergen_list:
                    # 対象アレルギーを注文している施設を取得
                    if allergen == '個食':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.KOSHOKU_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    elif allergen == 'ﾌﾘｰｽﾞ':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.FREEZE_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    else:
                        if key.is_soup:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name,
                                meal_name__filling=True
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                        else:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    allergen_unit_list += list(order_qs)

            # 施設情報のソート
            allergen_orders = sorted(allergen_unit_list, key=cmp_to_key(self.compare_allergen_unit))

            if self.is_soe_allergen(allergen_unit_list, prev_unit_list):
                pass
            else:
                current_row += 1

                prev_unit_list = []
                prev_unit_list += allergen_unit_list

                # 施設情報の出力
                for order in allergen_orders:
                    current_row += 1

        return current_row - row

    def write_basic_design(self, ws, pouch: PouchMeal, start_row):
        """
        基本食設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        basic_aggregate = pouch.bacic_aggregate

        row = start_row + 3
        # タイトルの表記
        ws.cell(row, 10).font = self.design_title_font
        ws.cell(row, 10, basic_aggregate.name)
        row += 1

        # 1人用袋の出力
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '1人用:')
        ws.cell(row, 11, basic_aggregate.package_count_1p + basic_aggregate.preserve_1p + basic_aggregate.photo)

        # 袋必要数(タイトル)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, '袋必要数:')
        row += 1

        # 10人用袋の出力
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '10人用:')
        count_for_10 = basic_aggregate.package_count_10p + basic_aggregate.package_preserve
        ws.cell(row, 11, count_for_10)

        # 袋必要数(値)
        units_len = len(basic_aggregate.package_units)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, f'{units_len+count_for_10}+{basic_aggregate.package_needle}')

        row += 1

        # 端数
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '端数:')
        ws.cell(row, 11, units_len)
        row += 1

        # 端数 施設の表
        col = 11
        table_start_row = row
        for index, value in enumerate(basic_aggregate.package_units):
            unit, count = value
            row_added = (index // 3)
            col_added = (index % 3)
            unit_row = row + row_added
            unit_col = col + col_added

            ws.cell(unit_row, unit_col).font = self.design_font
            if ((row_added % 2) == 0) and (col_added == 0):
                ws.cell(unit_row, unit_col).fill = self.desigined_basic_table_fill
                ws.cell(unit_row, unit_col+1).fill = self.desigined_basic_table_fill
                ws.cell(unit_row, unit_col+2).fill = self.desigined_basic_table_fill
            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')

        row += math.ceil(units_len / 3)
        table_end_row = row
        row += 1

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'K{table_start_row}:M{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.cell(row-1, 10).font = self.design_font
        ws.cell(row-1, 11).font = self.design_font
        ws.cell(row-1, 10, '保存用:')
        ws.cell(row-1, 11, basic_aggregate.package_preserve)

        # 針刺し
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '針刺し:')
        ws.cell(row, 11, basic_aggregate.package_needle)

        # 外枠
        ExcelHellper.set_outer_border(ws, f'J{start_row + 3}:N{row}', self.designed_basic_outer_side)

        return row

    def write_basic_5p_design(self, ws, pouch: PouchMeal, start_row):
        """
        基本食設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        basic_aggregate = pouch.bacic_5p_aggregate

        row = start_row + 3
        # タイトルの表記
        ws.cell(row-1, 10).font = self.design_title_font
        ws.cell(row, 10).font = self.design_title_font
        ws.cell(row-1, 10, f'{basic_aggregate.plate_name_5p})')
        ws.cell(row, 10, f'{basic_aggregate.name}')
        row += 1

        # 1人用袋の出力
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '1人用:')
        ws.cell(row, 11, basic_aggregate.package_count_1p + basic_aggregate.preserve_1p + basic_aggregate.photo)

        # 袋必要数(タイトル)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, '袋必要数:')
        row += 1

        # 5人用袋の出力
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '5人用:')
        count_for_5 = basic_aggregate.package_count_5p + basic_aggregate.package_preserve
        ws.cell(row, 11, count_for_5)

        # 袋必要数(値)
        units_len = len(basic_aggregate.package_units)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, f'{units_len+count_for_5}+{basic_aggregate.package_needle}')

        row += 1

        # 端数
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '端数:')
        ws.cell(row, 11, units_len)
        row += 1

        # 端数 施設の表
        col = 11
        table_start_row = row
        for index, value in enumerate(basic_aggregate.package_units):
            unit, count = value
            row_added = (index // 3)
            col_added = (index % 3)
            unit_row = row + row_added
            unit_col = col + col_added

            ws.cell(unit_row, unit_col).font = self.design_font
            if ((row_added % 2) == 0) and (col_added == 0):
                ws.cell(unit_row, unit_col).fill = self.desigined_basic_table_fill
                ws.cell(unit_row, unit_col+1).fill = self.desigined_basic_table_fill
                ws.cell(unit_row, unit_col+2).fill = self.desigined_basic_table_fill
            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')

        row += math.ceil(units_len / 3)
        table_end_row = row
        row += 1

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'K{table_start_row}:M{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.cell(row-1, 10).font = self.design_font
        ws.cell(row-1, 11).font = self.design_font
        ws.cell(row-1, 10, '保存用:')
        ws.cell(row-1, 11, basic_aggregate.package_preserve)

        # 針刺し
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '針刺し:')
        ws.cell(row, 11, basic_aggregate.package_needle)

        # 外枠
        ExcelHellper.set_outer_border(ws, f'J{start_row + 3}:N{row}', self.designed_basic_outer_side)

        return row

    def write_units_design(self, ws, pouch: PouchMeal, start_row):
        """
        基本食設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        unit_aggregate = pouch.unit_aggregate

        row = start_row + 3
        # タイトルの表記
        ws.cell(row, 16).font = self.design_title_font
        ws.cell(row, 16, unit_aggregate.name)
        row += 1

        # 1人用袋の出力
        ws.cell(row, 16).font = self.design_font
        ws.cell(row, 17).font = self.design_font
        ws.cell(row, 16, '1人用:')
        ws.cell(row, 17, unit_aggregate.package_count_1p + unit_aggregate.preserve_1p + unit_aggregate.photo)

        # 袋必要数(タイトル)
        units_len = len(unit_aggregate.package_units)
        units_quantity = sum([x[1] for x in unit_aggregate.package_units])
        ws.cell(row, 19).font = self.design_font
        ws.cell(row+1, 19).font = self.design_font
        ws.cell(row, 19, '袋必要数:')
        ws.cell(row+1, 19, f'{units_quantity+unit_aggregate.package_preserve}+{unit_aggregate.package_needle}')
        row += 1

        # 施設
        ws.cell(row, 16).font = self.design_font
        ws.cell(row, 17).font = self.design_font
        ws.cell(row, 16, '施設:')
        ws.cell(row, 17, units_len)
        row += 1

        # 端数 施設の表
        col = 17
        table_start_row = row
        for index, value in enumerate(unit_aggregate.package_units):
            unit, count = value
            row_added = (index // 3)
            col_added = (index % 3)
            unit_row = row + row_added
            unit_col = col + col_added * 2

            # 2件以上の施設を赤字に表示
            if count >= 2:
                ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col+1).font = self.design_emphasis_font
            else:
                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col+1).font = self.design_font

            # 行ごとの背景色設定
            if ((row_added % 2) == 0) and (col_added == 0):
                for i in range(6):
                    ws.cell(unit_row, unit_col+i).fill = self.desigined_unit_table_fill

            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
            ws.cell(unit_row, unit_col+1, f'{count}')

        row += math.ceil(units_len / 3)
        table_end_row = row
        row += 1

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'Q{table_start_row}:R{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'S{table_start_row}:T{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'U{table_start_row}:V{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.cell(row, 16).font = self.design_font
        ws.cell(row, 17).font = self.design_font
        ws.cell(row, 16, '保存用:')
        ws.cell(row, 17, unit_aggregate.package_preserve)
        row += 1

        # 針刺し
        ws.cell(row, 16).font = self.design_font
        ws.cell(row, 17).font = self.design_font
        ws.cell(row, 16, '針刺し:')
        ws.cell(row, 17, unit_aggregate.package_needle)

        # 外枠
        ExcelHellper.set_outer_border(ws, f'P{start_row + 3}:W{row}', self.designed_basic_outer_side)

        return row

    def write_dry_plate(self, ws, pouch: PouchMeal, start_row):
        """
        乾燥品施設毎出力エリアのページ先頭の料理名を出力
        """
        unit_aggregate = pouch.unit_for_dry_aggregate

        row = start_row + 3
        # タイトルの表記
        ws.cell(row, 10).font = self.design_title_font
        ws.cell(row, 10, f'{unit_aggregate.plate_name}(乾燥)　設計図')

    def write_dry_units_design(self, ws, pouch: PouchMeal, start_row):
        """
        乾燥品施設毎出力エリアの出力。戻り値として、エリアの最終行を返す。
        """
        unit_aggregate = pouch.unit_for_dry_aggregate

        row = start_row + 3
        # タイトルの表記
        ws.cell(row, 10).font = self.design_title_font
        ws.cell(row, 10, '基本食')
        row += 1

        # 1人用袋の出力
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '1人用:')
        ws.cell(row, 11, unit_aggregate.package_count_1p + unit_aggregate.preserve_1p + unit_aggregate.photo)

        # 袋必要数(タイトル)
        units_len = len(unit_aggregate.package_units)
        units_quantity = sum([x[1] for x in unit_aggregate.package_units])
        ws.cell(row, 13).font = self.design_font
        ws.cell(row+1, 13).font = self.design_font
        ws.cell(row, 13, '袋必要数:')
        ws.cell(row+1, 13, f'{units_quantity+unit_aggregate.package_preserve}+{unit_aggregate.package_needle}')
        row += 1

        # 施設
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '施設:')
        ws.cell(row, 11, units_len)
        row += 1

        # 端数 施設の表
        col = 11
        table_start_row = row
        for index, value in enumerate(unit_aggregate.package_units):
            unit, count = value
            row_added = (index // 3)
            col_added = (index % 3)
            unit_row = row + row_added
            unit_col = col + col_added * 2

            # 2件以上の施設を赤字に表示
            if count >= 2:
                ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col+1).font = self.design_emphasis_font
            else:
                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col+1).font = self.design_font

            # 行ごとの背景色設定
            if ((row_added % 2) == 0) and (col_added == 0):
                for i in range(6):
                    ws.cell(unit_row, unit_col+i).fill = self.desigined_unit_table_fill

            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
            ws.cell(unit_row, unit_col+1, f'{count}')

        row += math.ceil(units_len / 3)
        table_end_row = row
        row += 1

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'K{table_start_row}:L{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'M{table_start_row}:N{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'O{table_start_row}:P{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '保存用:')
        ws.cell(row, 11, unit_aggregate.package_preserve)
        row += 1

        # 針刺し
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '針刺し:')
        ws.cell(row, 11, unit_aggregate.package_needle)

        # 外枠
        ExcelHellper.set_outer_border(ws, f'J{start_row + 3}:R{row}', self.designed_basic_outer_side)

        return row

    def write_mix_rice_design(self, ws, pouch: PouchMeal, start_row):
        """
        混ぜご飯設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        mix_rice_aggregate = pouch.unit_mix_rice_aggregate

        row = start_row + 1
        # タイトルの表記
        ws.cell(row, 10).font = self.design_title_font
        ws.cell(row, 10, f'{mix_rice_aggregate.name}　※混ぜご飯計量表合数一覧参照')
        row += 1

        return row

    def write_soup_design(self, ws, pouch: PouchMeal, start_row: int, is_miso_soup: bool):
        """
        汁の設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        soup_aggregate = pouch.soup_aggregate
        if is_miso_soup:
            return self.write_miso_soup_design(ws, pouch, start_row, soup_aggregate)
        else:
            return self.write_other_soup_design(ws, pouch, start_row, soup_aggregate)

    def write_other_soup_design(self, ws, pouch: PouchMeal, start_row: int, soup_aggregate: DesignedPouchSoup):
        """
        汁の設計図エリアの出力。味噌汁以外の汁・スープとして出力する)。戻り値として、エリアの最終行を返す。
        """

        row = start_row
        # タイトルの表記
        eating_day = pouch.eating_day
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_soup_title_font
        ws.cell(row, 10, f'{soup_aggregate.plate_name or soup_aggregate.name}　{eating_day.strftime("%m/%d")}{pouch.meal_name}')
        row += 1

        # 1人用袋の出力
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '1人用:')
        ws.cell(row, 11, soup_aggregate.package_count_1p + soup_aggregate.preserve_1p + soup_aggregate.photo)

        # 袋必要数(タイトル)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, '袋必要数:')
        row += 1

        ws.row_dimensions[row].height = 45.75
        if soup_aggregate.is_unit_pouch_use:
            #　施設出力時は施設向けの袋のみ
            pack_10_sum = 0
            units_len = len(soup_aggregate.package_units)
        else:
            #　端数出力時は、10人用袋と端数に分ける
            pack_10_sum = sum([x[3] for x in soup_aggregate.package_units if x[1] == 10])
            units_fra = [x for x in soup_aggregate.package_units if x[1] != 10]
            units_len = len(units_fra)
        units_quantity = sum([x[3] for x in soup_aggregate.package_units])

        # 袋必要数(値)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, f'{units_quantity+soup_aggregate.package_preserve}+{soup_aggregate.package_needle}')

        if not soup_aggregate.is_unit_pouch_use:
            ws.cell(row, 10).font = self.design_font
            ws.cell(row, 11).font = self.design_font
            ws.cell(row, 10, '10人袋:')
            ws.cell(row, 11, pack_10_sum + soup_aggregate.package_preserve)
            row += 1

        # 施設
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '施設:')
        ws.cell(row, 11, units_len)

        row += 1

        # 端数 施設の表
        table_start_row = row
        if soup_aggregate.is_unit_pouch_use:
            col = 10
            for index, value in enumerate(soup_aggregate.package_units):
                ws.row_dimensions[table_start_row].height = 45.75
                unit, pack, orders, count = value
                row_added = (index // 2)
                col_added = (index % 2)
                unit_row = row + row_added
                unit_col = col + col_added * 3

                if count >= 2:
                    ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                    ws.cell(unit_row, unit_col + 1).font = self.design_emphasis_font
                    ws.cell(unit_row, unit_col + 2).font = self.design_emphasis_font
                else:
                    ws.cell(unit_row, unit_col).font = self.design_font
                    ws.cell(unit_row, unit_col + 1).font = self.design_font
                    ws.cell(unit_row, unit_col + 2).font = self.design_font

                if ((row_added % 2) == 0) and (col_added == 0):
                    for i in range(6):
                        ws.cell(unit_row, unit_col + i).fill = self.desigined_soup_table_fill
                ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
                ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
                ws.cell(unit_row, unit_col + 1, f'{orders}')
                ws.cell(unit_row, unit_col + 2).alignment = self.center_alignment
                ws.cell(unit_row, unit_col + 2, f'{count}')
        else:
            col = 11
            for index, value in enumerate(units_fra):
                ws.row_dimensions[table_start_row].height = 45.75
                unit, pack, orders, count = value
                row_added = (index // 2)
                col_added = (index % 2)
                unit_row = row + row_added
                unit_col = col + col_added * 2

                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col + 1).font = self.design_font
                if ((row_added % 2) == 0) and (col_added == 0):
                    ws.cell(unit_row, unit_col).fill = self.desigined_soup_table_fill
                    ws.cell(unit_row, unit_col + 1).fill = self.desigined_soup_table_fill
                    ws.cell(unit_row, unit_col + 2).fill = self.desigined_soup_table_fill
                    ws.cell(unit_row, unit_col + 3).fill = self.desigined_soup_table_fill
                ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
                ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
                ws.cell(unit_row, unit_col + 1, f'{count}')

        row += math.ceil(units_len / 2)
        table_end_row = row

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            if soup_aggregate.is_unit_pouch_use:
                ExcelHellper.set_grid_border(ws, f'J{table_start_row}:L{table_end_row - 1}',
                                              self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
                ExcelHellper.set_grid_border(ws, f'M{table_start_row}:O{table_end_row - 1}',
                                              self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            else:
                ExcelHellper.set_grid_border(ws, f'K{table_start_row}:L{table_end_row - 1}',
                                              self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
                ExcelHellper.set_grid_border(ws, f'M{table_start_row}:N{table_end_row - 1}',
                                              self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 施設毎の場合、端数の表に袋数は不要なので、施設名の表示を広げる
        if not soup_aggregate.is_unit_pouch_use:
            for i in range(table_end_row - table_start_row):
                ws.merge_cells(f'K{table_start_row + i}:L{table_start_row + i}')
                ws.merge_cells(f'M{table_start_row + i}:N{table_start_row + i}')

        # 外枠
        ExcelHellper.set_outer_border(ws, f'J{start_row}:O{row + 2}', self.designed_soup_outer_side)

        # 保存用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '保存用:')
        ws.cell(row, 11, soup_aggregate.package_preserve)
        row += 1

        # 針刺し
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '針刺し:')
        ws.cell(row, 11, soup_aggregate.package_needle)

        return row + 1

    def write_miso_soup_design(self, ws, pouch: PouchMeal, start_row: int, soup_aggregate: DesignedPouchSoup):
        """
        汁の設計図エリアを味噌汁として出力する(嚥下出力なし)。戻り値として、エリアの最終行を返す。
        """

        row = start_row
        # タイトルの表記
        eating_day = pouch.eating_day
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_soup_title_font
        ws.cell(row, 10, f'{soup_aggregate.plate_name or soup_aggregate.name}　{eating_day.strftime("%m/%d")}{pouch.meal_name}')
        row += 1

        # 1人用袋の出力
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '1人用:')
        if self.miso_soup_dict:
            ws.cell(row, 11, soup_aggregate.package_count_1p)
        else:
            # 味噌汁の汁の場合は、写真用が不要だが、その代わりに1件必要
            ws.cell(row, 11, soup_aggregate.package_count_1p + soup_aggregate.preserve_1p + 1)

        # 袋必要数(タイトル)
        ws.cell(row, 13).font = self.design_font
        ws.cell(row, 13, '袋必要数:')
        row += 1

        # 施設
        ws.row_dimensions[row].height = 45.75
        units_len = len(soup_aggregate.package_units)
        units_quantity = sum([x[3] for x in soup_aggregate.package_units])
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '施設:')
        ws.cell(row, 11, units_len)

        # 袋必要数(値)
        ws.cell(row, 13).font = self.design_font
        if self.miso_soup_dict:
            ws.cell(row, 13, f'{units_quantity}')
        else:
            ws.cell(row, 13, f'{units_quantity+soup_aggregate.package_preserve}+{soup_aggregate.package_needle}')

        row += 1

        # 端数 施設の表
        col = 11
        table_start_row = row
        for index, value in enumerate(soup_aggregate.package_units):
            ws.row_dimensions[table_start_row].height = 45.75
            unit, p, o, count = value
            row_added = (index // 2)
            col_added = (index % 2)
            unit_row = row + row_added
            unit_col = col + col_added * 2

            if count >= 2:
                ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col+1).font = self.design_emphasis_font
            else:
                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col+1).font = self.design_font
            if ((row_added % 2) == 0) and (col_added == 0):
                ws.cell(unit_row, unit_col).fill = self.desigined_soup_table_fill
                ws.cell(unit_row, unit_col+1).fill = self.desigined_soup_table_fill
                ws.cell(unit_row, unit_col+2).fill = self.desigined_soup_table_fill
                ws.cell(unit_row, unit_col+3).fill = self.desigined_soup_table_fill
            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
            ws.cell(unit_row, unit_col+1, f'{count}')

        row += math.ceil(units_len / 2)
        table_end_row = row

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'K{table_start_row}:L{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'M{table_start_row}:N{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '保存用:')
        if PlateNameAnalizeUtil.is_miso_soup(soup_aggregate.plate_name):
            if self.miso_soup_dict:
                ws.cell(row, 11, 0)
            else:
                ws.cell(row, 11, soup_aggregate.package_preserve)
                self.miso_soup_dict = True
        else:
            ws.cell(row, 11, soup_aggregate.package_preserve)
        row += 1

        # 針刺し
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_font
        ws.cell(row, 11).font = self.design_font
        ws.cell(row, 10, '針刺し:')
        if PlateNameAnalizeUtil.is_miso_soup(soup_aggregate.plate_name):
            if self.miso_fill_dict:
                ws.cell(row, 11, 0)
            else:
                ws.cell(row, 11, soup_aggregate.package_needle)
                self.miso_fill_dict = True
        else:
            ws.cell(row, 11, soup_aggregate.package_needle)

        # 外枠
        ExcelHellper.set_outer_border(ws, f'J{start_row}:O{row}', self.designed_soup_outer_side)

        return row

    def write_soup_enge_design(self, ws, pouch: PouchMeal, start_row: int):
        soup_aggregate = pouch.soup_aggregate

        # 嚥下
        # ソフト
        row_soft = self.write_filling_enge_detail(ws, start_row, 10, self.plate_soft_fill, soup_aggregate.designed_soft)

        # ゼリー
        row_jelly = self.write_filling_enge_detail(ws, start_row, 14, self.plate_jelly_fill, soup_aggregate.designed_jelly, width=5)

        # ミキサー
        row_mixer = self.write_filling_enge_detail(ws, start_row, 19, self.plate_mixer_fill, soup_aggregate.designed_mixer, adjust=0)

        # 外枠
        write_row = max([row_soft, row_jelly, row_mixer])
        ExcelHellper.set_outer_border(ws, f'J{start_row}:V{write_row}', self.designed_soup_outer_side)

        return write_row


    def write_filling_design(self, ws, pouch: PouchMeal, start_row: int):
        """
        汁・汁具食設計図エリアの出力。戻り値として、エリアの最終行を返す。
        """
        filling_aggregate = pouch.filling_aggregate

        row = start_row
        # タイトルの表記
        eating_day = pouch.eating_day
        ws.cell(row, 18).font = self.design_filling_title_font
        ws.cell(row, 18, f'{filling_aggregate.plate_name or filling_aggregate.name}　{eating_day.strftime("%m/%d")}{pouch.meal_name}')
        row += 1

        # 1人用袋の出力
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 18).font = self.design_font
        ws.cell(row, 20).font = self.design_font
        ws.cell(row, 18, '1人用:')
        ws.cell(row, 20, filling_aggregate.package_count_1p + filling_aggregate.preserve_1p + filling_aggregate.photo)

        # 袋必要数(タイトル)
        ws.cell(row, 23).font = self.design_font
        ws.cell(row, 23, '袋必要数:')
        row += 1

        # 施設
        ws.row_dimensions[row].height = 45.75
        units_len = len(filling_aggregate.package_units)
        ws.cell(row, 18).font = self.design_font
        ws.cell(row, 20).font = self.design_font
        ws.cell(row, 18, '施設:')
        ws.cell(row, 20, units_len)

        # 袋必要数(値)
        ws.cell(row, 23).font = self.design_font
        quantity = sum([x[1] for x in filling_aggregate.package_units])
        ws.cell(row, 23, f'{quantity+filling_aggregate.package_preserve}+{filling_aggregate.package_needle}')

        row += 1

        # 端数 施設の表
        col = 19
        table_start_row = row
        for index, value in enumerate(filling_aggregate.package_units):

            unit, count = value
            row_added = (index // 5)
            col_added = (index % 5)
            unit_row = row + row_added
            unit_col = col + col_added * 2
            ws.row_dimensions[unit_row].height = 45.75

            if count >= 2:
                ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col+1).font = self.design_emphasis_font
            else:
                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col+1).font = self.design_font
            if ((row_added % 2) == 0) and (col_added == 0):
                for i in range(10):
                    ws.cell(unit_row, unit_col + i).fill = self.desigined_soup_table_fill
            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
            ws.cell(unit_row, unit_col+1, f'{count}')

        row += math.ceil(units_len / 5)
        table_end_row = row

        # 表の枠
        if table_end_row > table_start_row:
            # 施設データがある場合のみ枠を記載
            ExcelHellper.set_grid_border(ws, f'S{table_start_row}:T{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'U{table_start_row}:V{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'W{table_start_row}:X{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'Y{table_start_row}:Z{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)
            ExcelHellper.set_grid_border(ws, f'AA{table_start_row}:AB{table_end_row - 1}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        # 保存用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 18).font = self.design_font
        ws.cell(row, 20).font = self.design_font
        ws.cell(row, 18, '保存用:')
        ws.cell(row, 20, filling_aggregate.package_preserve)
        row += 1

        # 針刺し
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 18).font = self.design_font
        ws.cell(row, 20).font = self.design_font
        ws.cell(row, 18, '針刺し:')
        ws.cell(row, 20, filling_aggregate.package_needle)
        row += 1

        # 嚥下情報
        ws.row_dimensions[row].height = 45.75

        # ソフト
        row_soft = self.write_filling_enge_detail(ws, row, 18, self.plate_soft_fill, filling_aggregate.designed_soft)

        # ゼリー
        row_jelly = self.write_filling_enge_detail(ws, row, 22, self.plate_jelly_fill, filling_aggregate.designed_jelly)

        # ミキサー
        row_mixer = self.write_filling_enge_detail(ws, row, 26, self.plate_mixer_fill, filling_aggregate.designed_mixer)

        # 外枠
        row = max([row_soft, row_jelly, row_mixer])
        ExcelHellper.set_outer_border(ws, f'R{start_row}:AD{row}', self.designed_filling_outer_side)

        return row

    def write_filling_enge_detail(self, ws, start_row: int, col: int, fill, dp, width: int = 4, adjust: int = 1):
        row = start_row

        ws.row_dimensions[row].height = 45.75
        ws.cell(row, col).font = self.design_font

        for i in range(width):
            ws.cell(row, col + i).fill = fill

        # 嚥下タイトル
        ws.cell(row, col, dp.name)
        row += 1

        # 1人用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, col).font = self.design_font
        ws.cell(row, col+2).font = self.design_font
        ws.cell(row, col, '1人用：')
        ws.cell(row, col + 2, dp.package_count_1p + dp.preserve_1p)
        row += 1

        # 袋数
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, col).font = self.design_font
        ws.cell(row, col+2).font = self.design_font
        ws.cell(row, col, '袋数：')
        units_len = len(dp.package_units)
        package_quantity = sum([x[3] for x in dp.package_units])
        ws.cell(row, col + 2, package_quantity)
        row += 1

        # 施設の表
        table_start_row = row
        adjusted_col = col + adjust
        for index, value in enumerate(dp.package_units):
            ws.row_dimensions[row].height = 45.75
            unit, quantity, orders, count = value
            if (index % 2) == 0:
                ws.cell(row, adjusted_col).fill = fill
                ws.cell(row, adjusted_col + 1).fill = fill
                ws.cell(row, adjusted_col + 2).fill = fill

            # 文字色の設定
            if count > 1:
                # 2袋以上の施設を赤字で出力
                ws.cell(row, adjusted_col).font = self.design_emphasis_font
                ws.cell(row, adjusted_col + 1).font = self.design_emphasis_font
                ws.cell(row, adjusted_col + 2).font = self.design_emphasis_font
            else:
                ws.cell(row, adjusted_col).font = self.design_font
                ws.cell(row, adjusted_col + 1).font = self.design_font
                ws.cell(row, adjusted_col + 2).font = self.design_font
            ws.cell(row, adjusted_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(row, adjusted_col + 1).alignment = self.center_alignment
            ws.cell(row, adjusted_col + 1, f'{orders}')
            ws.cell(row, adjusted_col + 2).alignment = self.center_alignment
            ws.cell(row, adjusted_col + 2, f'{count}')
            row += 1

        if units_len:
            start_cell = ws.cell(table_start_row, adjusted_col)
            end_cell = ws.cell(row - 1, adjusted_col + 2)
            ExcelHellper.set_grid_border(ws, f'{start_cell.coordinate}:{end_cell.coordinate}',
                                          self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        return row

    def is_basic_page_over(self, row: int, pouch: PouchMeal):
        """
        汁、汁具の設計図を出力したときに、ページを跨いでしまうかどうかを計算する。
        """
        # 汁具は嚥下も出力するため、通常汁具の方が出力行が多くなる
        filling = pouch.filling_aggregate

        basic_lines = 6 + (len(filling.package_units) // 5)

        # 嚥下領域の計算
        soft = filling.designed_soft
        jelly = filling.designed_jelly
        mixer = filling.designed_mixer

        soft_lines = 4 + len(soft.package_units)
        jelly_lines = 4 + len(jelly.package_units)
        mixer_lines = 4 + len(mixer.package_units)
        enge_lines = max([soft_lines, jelly_lines, mixer_lines])

        return (row + basic_lines + enge_lines) > self.PAGE_MAX_ROW

    def write_basic_sheet(self, wb, pouch: PouchMeal):
        """
        基本食・施設毎・汁・汁具のパウチ設計図を出力する。喫食日、食事区分の組み合わせで1シートとする。
        """
        ws = self.prpare_work_sheet(wb, 'basic', f'{pouch.eating_day.strftime("%m%d")}_{pouch.meal_name}_基本食・施設毎・汁・汁具')
        logger.info(f'{pouch.eating_day.strftime("%m%d")}_{pouch.meal_name}_基本食・施設毎・汁・汁具')

        # 最終データ更新日時出力
        last_modify = UnitPackage.objects.filter(cooking_day=self.cooking_day).order_by('-register_at').first()
        ws.cell(1, 23, f'最終データ更新日時：{last_modify.register_at.strftime("%Y/%m/%d %H:%M:%S")}')

        # 料理情報の読込
        self.read_plates(pouch)

        # 喫食日、食事区分の出力
        row = 0
        self.write_eating_time(ws, pouch, row)

        is_write_complete = False
        page = 1
        while not is_write_complete:
            # 料理情報の出力
            plate_row, is_write_complete = self.write_plates(ws, row, page, pouch)

            # 基本食パウチ設計の出力
            basic_end_row = self.write_basic_design(ws, pouch, row)

            # 施設毎パウチ設計の出力
            if pouch.exists_units():
                units_end_row = self.write_units_design(ws, pouch, row)
            else:
                units_end_row = basic_end_row

            # 施設毎(混ぜご飯)パウチ設計の出力
            if pouch.exists_mix_rice():
                rice_row = max([basic_end_row, units_end_row])

                self.write_mix_rice_design(ws, pouch, rice_row + 1)

            # 次のページに表示
            next_row = self.NEXT_PAGE_ROW * page - 2
            ws.row_breaks.append(Break(next_row))
            row = next_row
            page += 1

        soup_start_row = self.NEXT_PAGE_ROW * (page - 1)

        if pouch.exists_soup():
            # 汁・汁具の料理名出力(食数が存在する場合のみ出力)
            self.write_plates_soup(ws, pouch, soup_start_row)

            # ws.row_dimensions[soup_start_row].height = 57.75

            # 汁のパウチ設計の出力
            is_miso_soup = PlateNameAnalizeUtil.is_miso_soup(pouch.soup_aggregate.plate_name)
            soup_row = self.write_soup_design(ws, pouch, soup_start_row, is_miso_soup)

            # 汁具のパウチ設計の出力
            filling_row = self.write_filling_design(ws, pouch, soup_start_row)

            # 味噌汁以外の汁・スープの場合のみ、汁の嚥下を出力する
            if not is_miso_soup:
                enge_start_row = max([soup_row, filling_row]) + 2
                filling_row = self.write_soup_enge_design(ws, pouch, enge_start_row)

            dry_units_start_row = max([filling_row, soup_row])

            # 次のページに表示
            ws.row_breaks.append(Break(dry_units_start_row + 2))
            row = dry_units_start_row + 3
            page += 1
        else:
            dry_units_start_row = soup_start_row

        # 基本食(5人用)の出力
        if pouch.exists_5p():
            # 基本食(5人用)の設計図を記載
            basic_5p_row = self.write_basic_5p_design(ws, pouch, row)

            # 次のページに表示
            ws.row_breaks.append(Break(basic_5p_row + 2))
            row = basic_5p_row + 3
            page += 1

        # 乾燥特別集計対象品(現在は錦糸卵のみ)
        if pouch.exists_units_for_dry():
            # 料理名を出力
            self.write_dry_plate(ws, pouch, row - 2)

            # 施設毎の設計図を記載
            units_end_row = self.write_dry_units_design(ws, pouch, row)

            # 嚥下の設計図を出力
            self.write_dry_enge_design(ws, units_end_row + 2, pouch, 20)

    def is_enge_page_over(self, row: int, pouch, package_size: int):
        """
        嚥下の設計図を出力したときに、ページを跨いでしまうかどうかを計算する。
        """
        # 嚥下領域の計算
        soft = pouch.soft_aggregate
        jelly = pouch.jelly_aggregate
        mixer = pouch.mixer_aggregate

        soft_lines = 4 + len([x for x in soft.package_units if x[1] == package_size])
        jelly_lines = 4 + len([x for x in jelly.package_units if x[1] == package_size])
        mixer_lines = 4 + len([x for x in mixer.package_units if x[1] == package_size])
        enge_lines = max([soft_lines, jelly_lines, mixer_lines])

        return (row + enge_lines) > self.PAGE_MAX_ROW

    def compare_units(self, unit1, unit2):
        return unit1[0].unit_number - unit2[0].unit_number

    def write_enge_detail(self, ws, start_row: int, start_col: int, cols: int, fill, dp, package_size):
        row = start_row

        ws.row_dimensions[row].height = 45.75
        ws.cell(row, start_col).font = self.design_font

        ws.cell(row, start_col).fill = fill
        ws.cell(row, start_col + 1).fill = fill
        for i in range(cols):
            c = start_col + 2 + i * 3
            ws.cell(row, c).fill = fill
            ws.cell(row, c+1).fill = fill
            ws.cell(row, c+2).fill = fill

        # 運用上、0件になってしまうことはないが、テストデータ使用時のために中断を設定
        if not dp:
            logger.warning('dp is none.')
            return row

        # 嚥下タイトル
        ws.cell(row, start_col, dp.name)
        row += 1

        # 1人用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, start_col).font = self.design_font
        ws.cell(row, start_col+2).font = self.design_font
        ws.cell(row, start_col, '1人用：')
        ws.cell(row, start_col + 2, dp.package_count_1p + dp.preserve_1p)
        logger.info(f'{dp.package_count_1p}-{dp.preserve_1p}')
        row += 1

        # 施設用
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, start_col).font = self.design_font
        ws.cell(row, start_col+2).font = self.design_font
        ws.cell(row, start_col, '施設：')
        work_units = [x for x in dp.package_units if x[1] == package_size]
        if package_size in self.enge_package_list:
            work_units += [y for y in dp.package_units if y[1] == 2]
        work_units = sorted(work_units, key=cmp_to_key(self.compare_units))
        units_len = sum([w[3] for w in work_units]) if work_units else 0
        ws.cell(row, start_col + 2, units_len)
        row += 1

        # 施設の表
        table_start_row = row
        unit_row = row
        for index, value in enumerate(work_units):
            ws.row_dimensions[row].height = 45.75
            unit, quantity, orders, count = value
            row_added = (index // cols)
            col_added = (index % cols)
            unit_row = row + row_added
            unit_col = start_col + col_added * 3 + 1
            if ((row_added % 2) == 0) and (col_added == 0):
                for i in range(cols * 3):
                    ws.cell(unit_row, unit_col + i).fill = fill

            # 文字色の設定
            if count > 1:
                # 2袋以上の施設を赤字で出力
                ws.cell(unit_row, unit_col).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col + 1).font = self.design_emphasis_font
                ws.cell(unit_row, unit_col + 2).font = self.design_emphasis_font
            else:
                ws.cell(unit_row, unit_col).font = self.design_font
                ws.cell(unit_row, unit_col + 1).font = self.design_font
                ws.cell(unit_row, unit_col + 2).font = self.design_font

            ws.cell(unit_row, unit_col, f'{unit.unit_number}.{unit.short_name}')
            ws.cell(unit_row, unit_col + 1).alignment = self.center_alignment
            ws.cell(unit_row, unit_col + 1, f'{orders}')
            ws.cell(unit_row, unit_col + 2).alignment = self.center_alignment
            ws.cell(unit_row, unit_col + 2, f'{count}')
        row = unit_row + 1

        # 表の罫線を描画
        if units_len:
            table_start_col = start_col + 1
            for i in range(cols):
                start_cell = ws.cell(table_start_row, table_start_col + i * 3)
                end_cell = ws.cell(row - 1, start_col + (i + 1) * 3)
                ExcelHellper.set_grid_border(ws, f'{start_cell.coordinate}:{end_cell.coordinate}',
                                              self.designed_unit_table_outer_side, self.designed_unit_table_inner_side)

        return row

    def write_enge_design(self, ws, start_row: int, pouch: PouchMeal, package_size):
        row = start_row

        # 袋タイトルの出力
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 3).font = self.design_title_font
        for i in range(30):
            ws.cell(row, 3 + i).fill = self.enge_plate_titele_fill

        ws.cell(row, 3, f'{package_size}人袋')
        row += 1

        # ソフトの出力
        row_soft = self.write_enge_detail(ws, row, 3, 4, self.plate_soft_fill, pouch.soft_aggregate, package_size)

        # ゼリーの出力
        row_jelly = self.write_enge_detail(ws, row, 17, 2, self.plate_jelly_fill, pouch.jelly_aggregate, package_size)

        # ミキサーの出力
        row_mixer = self.write_enge_detail(ws, row, 25, 2, self.plate_mixer_fill, pouch.mixer_aggregate, package_size)

        # 外枠
        row = max([row_soft, row_jelly, row_mixer])
        ExcelHellper.set_outer_border(ws, f'C{start_row}:AF{row}', self.designed_basic_outer_side)

        return row

    def write_dry_enge_design(self, ws, start_row: int, pouch: PouchMeal, package_size):
        row = start_row

        # 袋タイトルの出力
        ws.row_dimensions[row].height = 45.75
        ws.cell(row, 10).font = self.design_title_font
        for i in range(15):
            ws.cell(row, 10 + i).fill = self.enge_plate_titele_fill

        ws.cell(row, 10, f'{package_size}人袋')
        row += 1

        # ソフトの出力
        row_soft = self.write_enge_detail(
            ws, row, 10, 1, self.plate_soft_fill, pouch.unit_for_dry_aggregate.designed_soft, package_size)

        # ゼリーの出力
        row_jelly = self.write_enge_detail(
            ws, row, 15, 1, self.plate_jelly_fill, pouch.unit_for_dry_aggregate.designed_jelly, package_size)

        # ミキサーの出力
        row_mixer = self.write_enge_detail(
            ws, row, 20, 1, self.plate_mixer_fill, pouch.unit_for_dry_aggregate.designed_mixer, package_size)

        # 外枠
        row = max([row_soft, row_jelly, row_mixer])
        ExcelHellper.set_outer_border(ws, f'J{start_row}:X{row}', self.designed_basic_outer_side)

        return row

    def write_enge_plate_sheet(self, wb, meal_list):
        """
        嚥下用料理名シート。1喫食日1ページで、アレルギー表記なしで料理情報を出力する。
        """
        ws = wb.copy_worksheet(wb['plate'])
        ws.title = '嚥下食_料理出力'

        logger.info('料理情報シート出力')
        current_eating_day = None
        max_page = 31
        start_row = 1
        row = 1
        for pouch in meal_list:
            if current_eating_day:
                # 喫食日の切り替え
                # 出力のベース行の更新
                if current_eating_day != pouch.eating_day:
                    current_eating_day = pouch.eating_day
                    start_row += max_page
            else:
                # ループ初回
                current_eating_day = pouch.eating_day
            row = start_row

            # 料理情報読込
            self.read_plates(pouch, False)

            # 喫食日、食事区分の出力
            self.write_eating_time_for_plate(ws, current_eating_day, pouch, row)

            # 料理情報の出力
            self.write_plates_plate_sheet(ws, row, pouch)


    def write_enge_sheet(self, wb, pouch: PouchMeal):
        """
        嚥下食のパウチ設計図を出力する。喫食日、食事区分の組み合わせで1シートとする。
        """
        ws = self.prpare_work_sheet(wb, 'enge', f'{pouch.eating_day.strftime("%m%d")}_{pouch.meal_name}_嚥下食')
        logger.info(f'{pouch.eating_day.strftime("%m%d")}_{pouch.meal_name}_嚥下食')
        self.enge_package_list.clear()

        # 最終データ更新日時出力
        last_modify = UnitPackage.objects.filter(cooking_day=self.cooking_day).order_by('-register_at').first()
        ws.cell(1, 33, f'最終データ更新日時：{last_modify.register_at.strftime("%Y/%m/%d %H:%M:%S")}')

        # 料理情報の読込
        self.read_plates(pouch, False)

        is_write_complete = False
        row = 0
        # 喫食日、食事区分の出力
        self.write_eating_time(ws, pouch, row)

        # 複数ページに葉ならない
        page = 1

        # 料理情報の出力
        # plate_row, is_write_complete = self.write_plates(ws, row, page, pouch, True)
        self.write_plates_not_write(ws, row, page, pouch, True)
        row += 3

        ws.cell(row, 3).font = self.design_title_font
        ws.cell(row, 3, '嚥下')
        row += 2

        # 7人用パウチ設計の出力
        row_7p = self.write_enge_design(ws, row, pouch, 7)

        # 14人用パウチ設計の出力
        row_14p = self.write_enge_design(ws, row_7p + 2, pouch, 14)

        # 20人用パウチ設計の出力
        logger.info(f'enge:{self.enge_package_list}')
        row_20p = self.write_enge_design(ws, row_14p + 2, pouch, 20)

    def write(self):
        """
        パウチ設計図を出力する。シートは基本食(喫食日・食事区分毎)->嚥下(同左)の順に出力
        """
        wb = self.open_template()

        for pouch_meal in self.aggreagte.meal_list:
            # 基本食他の出力
            self.write_basic_sheet(wb, pouch_meal)

        self.food_count = 0
        # 料理情報シートの出力
        self.write_enge_plate_sheet(wb, self.aggreagte.meal_list)

        self.food_count = 0
        for pouch_meal in self.aggreagte.meal_list:
            # 嚥下食の出力
            self.write_enge_sheet(wb, pouch_meal)

        self.remove_template_sheet(wb)
        self.save(wb)
        logger.info(f'帳票出力完了(パウチ設計図)-{self.cooking_day}製造')

