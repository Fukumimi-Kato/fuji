import datetime as dt
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from functools import cmp_to_key
from itertools import groupby
import logging
import math
import numpy as np
import os
import re

import openpyxl as excel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

from django.conf import settings
from django.utils.functional import cached_property

from .cooking_direction_plates import CookingDirectionPlatesManager, PlateNameAnalizeUtil
from .meal import MealUtil
from .models import RawPlatePackageMaster, UnitMaster, PickingRawPlatePackage, MealMaster, CookingDirectionPlate
from .models import Order, MealDisplay, AllergenPlateRelations, UnitPackage, PackageMaster
from .models import PickingResultRaw

from web_order.cooking_direction_plates import PlateNameAnalizeUtil
logger = logging.getLogger(__name__)


class PlatePackageRegister:
    @classmethod
    def get_package_master(cls, name):
        qs = PackageMaster.objects.filter(name=name)
        if qs.exists():
            return qs.first()
        else:
            return None

    @classmethod
    def register_unit_package(cls, cooking_day, eating_day, meal_name, index,
                              unit_number, unit_name, count, master_name, parts_name,
                              menu_name='常食', is_basic_plate=True, mix_rice_type='none', soup_type='none',
                              is_raw_to_enge: bool = False, is_soup_parts: bool = False):
        if ('原体' in parts_name) and (not is_raw_to_enge):
            return
        if index == -1:
            return

        package_master = cls.get_package_master(master_name)
        if not package_master:
            logger.warning(f'package is none:{master_name}')
            return

        cooking_direction_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, eating_day=eating_day,
                                                                meal_name=meal_name, index=index, is_basic_plate=is_basic_plate)
        if not cooking_direction_qs.exists():
            logger.warning(f'plate for cooking direction is none:{cooking_day}-{eating_day}-{meal_name}-{index}')
            return

        cooking_direction = cooking_direction_qs.first()
        if is_soup_parts:
            if '◆' in cooking_direction.plate_name:
                res = re.findall('具(\d+|\d+\.\d+)[g|ｇ]\s*\D液(\d+|\d+\.\d+)[g|ｇ]', cooking_direction.plate_name)
                if res and (res[0][0] and res[0][1]):
                    is_curry_soup = True
            elif ('カレーライス' in cooking_direction.plate_name) or ('シチュー' in cooking_direction.plate_name):
                is_curry_soup = True
            else:
                is_curry_soup = False
        else:
            is_curry_soup = False

        # 個食の数値は、元のユニットに加算する
        if count and package_master:
            if '個食' in unit_name:
                qs = UnitPackage.objects.filter(
                    cooking_day=cooking_day,
                    unit_number=unit_number,
                    plate_name=cooking_direction.plate_name,
                    eating_day=eating_day,
                    meal_name=meal_name,
                    index=index,
                    menu_name=menu_name,
                    is_basic_plate=is_basic_plate,
                    mix_rice_type=mix_rice_type,
                    soup_type=soup_type
                ).exclude(
                    unit_name=unit_name
                )
                if qs.exists():
                    # 「個食」でないユニットにパウチ数を加算。通常個食だけ注文はない
                    up = qs.filter(package=package_master).first()
                    if up:
                        # 対象のサイズの袋が存在する場合
                        up.count += count
                        up.save()
                    else:
                        # 対象のサイズの袋が存在しない場合
                        up = qs.first()

                        # 対象の袋サイズでデータ登録
                        unit_package = UnitPackage(
                            cooking_day=cooking_day,
                            unit_number=unit_number,
                            unit_name=up.unit_name,
                            plate_name=cooking_direction.plate_name,
                            eating_day=eating_day,
                            meal_name=meal_name,
                            index=index,
                            package=package_master,
                            count=count,
                            menu_name=menu_name,
                            is_basic_plate=is_basic_plate,
                            mix_rice_type=mix_rice_type,
                            soup_type=soup_type
                        )
                        unit_package.save()
                else:
                    logger.warning(f'{unit_name}のメイン側の袋情報未登録')
                    # 「個食」でないユニットにパウチ数を加算。通常個食だけ注文はない
                    up = UnitMaster.objects.filter(unit_number=unit_number).exclude(unit_name=unit_name).order_by('id').first()
                    unit_package = UnitPackage(
                        cooking_day=cooking_day,
                        unit_number=unit_number,
                        unit_name=up.unit_name,
                        plate_name=cooking_direction.plate_name,
                        eating_day=eating_day,
                        meal_name=meal_name,
                        index=index,
                        package=package_master,
                        count=1,
                        menu_name=menu_name,
                        is_basic_plate=is_basic_plate,
                        mix_rice_type=mix_rice_type,
                        soup_type=soup_type
                    )
                    unit_package.save()
            else:
                plate_name = f'{cooking_direction.plate_name}(ルー)' if is_curry_soup else cooking_direction.plate_name
                if soup_type == 'soup' and ('具' in plate_name):
                    pass
                else:
                    unit_package = UnitPackage(
                        cooking_day=cooking_day,
                        unit_number=unit_number,
                        unit_name=unit_name,
                        plate_name=plate_name,
                        eating_day=eating_day,
                        meal_name=meal_name,
                        index=index,
                        package=package_master,
                        count=count,
                        menu_name=menu_name,
                        is_basic_plate=is_basic_plate,
                        mix_rice_type=mix_rice_type,
                        soup_type=soup_type
                    )
                    unit_package.save()


class RawPlatePackageRegisterBase:
    def __init__(self, plate_name: str):
        self.plate_name = plate_name.strip()
        self.is_valid = True

    def register(self, df, cook_day, eating_day, meal_name: str):
        pass


class InvalidRawPlatePackageRegister(RawPlatePackageRegisterBase):
    """
    原体送り対象外料理のパッケージ情報登録クラス
    """
    def __init__(self, plate_name: str):
        super(InvalidRawPlatePackageRegister, self).__init__(plate_name)
        self.is_valid = False


class RawPlatePackageRegister(RawPlatePackageRegisterBase):
    """
    原体送り料理のパッケージ情報を登録するクラス
    """
    def __init__(self, plate_name: str, raw_quantity_dry=None, raw_quantity_cold=None, raw_quantity_both=None):
        super(RawPlatePackageRegister, self).__init__(plate_name)
        if raw_quantity_both:
            self.dry_raw_quantity = raw_quantity_both
            self.cold_raw_quantity = raw_quantity_both
        else:
            self.dry_raw_quantity = raw_quantity_dry
            self.cold_raw_quantity = raw_quantity_cold

    def _get_package_detail(self, raw_plate_package, dry_cold_type: str):
        """
        指定した区分の、1人前使用量、パッケージ分量、取得した区分を取得する。
        指定した区分がない場合は、乾燥、または冷凍の内容を返す。
        """
        if dry_cold_type == '乾燥':
            if raw_plate_package.dry_eneble_name_gram and self.dry_raw_quantity:
                return Decimal(self.dry_raw_quantity), raw_plate_package.dry_package_quantity, dry_cold_type
            elif raw_plate_package.dry_quantity:
                return raw_plate_package.dry_quantity, raw_plate_package.dry_package_quantity, dry_cold_type
            else:
                # 冷凍の原体を使用
                if raw_plate_package.cold_eneble_name_gram and self.cold_raw_quantity:
                    return Decimal(self.cold_raw_quantity), raw_plate_package.cold_package_quantity, '冷凍'
                else:
                    return raw_plate_package.cold_quantity, raw_plate_package.cold_package_quantity, '冷凍'
        elif '冷凍' in dry_cold_type:
            if raw_plate_package.cold_eneble_name_gram and self.cold_raw_quantity:
                return Decimal(self.cold_raw_quantity), raw_plate_package.cold_package_quantity, dry_cold_type
            elif raw_plate_package.cold_quantity:
                return raw_plate_package.cold_quantity, raw_plate_package.cold_package_quantity, dry_cold_type
            else:
                # 乾燥の原体を使用
                if raw_plate_package.dry_eneble_name_gram and self.dry_raw_quantity:
                    return Decimal(self.dry_raw_quantity), raw_plate_package.dry_package_quantity, '乾燥'
                else:
                    return raw_plate_package.dry_quantity, raw_plate_package.dry_package_quantity, '乾燥'
        else:
            # 冷蔵は今のところ運用予定なし
            logger.warning('乾燥・冷凍区分未定')
            return raw_plate_package.cold_quantity, raw_plate_package.cold_package_quantity, '冷凍'

    def _in_plate(self, plate_name):
        for raw_plate in RawPlatePackageMaster.objects.all():
            if raw_plate.base_name in plate_name:
                return raw_plate

        # 見つからなかった場合
        return None

    def register(self, df, cook_day, eating_day, meal_name: str):
        logger.info(f'原体送りパッケージ情報登録:{self.plate_name}')
        #df.to_csv('tmp/raw_insert_df.csv', encoding='cp932')
        is_raw_to_enge = False

        if PlateNameAnalizeUtil.is_raw_enge_plate_name(self.plate_name, eating_day)[0]:
            logger.info(f'嚥下製造対象')
            is_raw_to_enge = True

        # 登録処理を実施する
        raw_plate_package = self._in_plate(self.plate_name)
        if raw_plate_package:
            logger.info(f'原体マスタ情報:{raw_plate_package}')

            PickingRawPlatePackage.objects.filter(cooking_day=cook_day, eating_day=eating_day, meal_name=meal_name, package_master=raw_plate_package).delete()
            for index, row in df.iterrows():
                unit_qs = UnitMaster.objects.filter(unit_number=row['呼出番号'], calc_name=row['ユニット名'])
                if not unit_qs.exists():
                    logger.debug(f'ユニットなし:{row["呼出番号"]}')
                    continue

                if is_raw_to_enge and (not row['献立種類'] == '常食'):
                    logger.info(f'嚥下製造-スキップ:{row["呼出番号"]}-{row["献立種類"]}')
                    continue

                # 合算でも個別のユニットを判断する必要はないため、先頭を取得
                unit = unit_qs.first()
                package, is_create = PickingRawPlatePackage.objects.get_or_create(
                    cooking_day=cook_day,
                    eating_day=eating_day,
                    meal_name=meal_name,
                    package_master=raw_plate_package,
                    unit_name=unit
                )
                if is_create:
                    logger.debug(f'インスタンス作成前:{row["呼出番号"]}-{row["献立種類"]}')
                    package.quantity = Decimal(0.0)
                else:
                    logger.debug(f'インスタンス作成済み:{row["呼出番号"]}-{row["献立種類"]}-{package.quantity}')

                quantity, package_quantity, dc_type = self._get_package_detail(raw_plate_package, unit.username.dry_cold_type)
                if unit.username.dry_cold_type == '乾燥':
                    logger.info(f'乾燥数量:{row["呼出番号"]}-{row["献立種類"]}-{row["注文数"] * quantity}-{unit.username.dry_cold_type}/{dc_type}')
                    package.quantity += row["注文数"] * quantity
                    if package.package:
                        logger.debug('pass')
                    else:
                        if (not package.package) and (package_quantity > 0.0):
                            package.package = package_quantity
                        else:
                            package.package = 1.0
                        package.dry_cold_type = dc_type
                elif unit.username.dry_cold_type == '冷凍' or unit.username.dry_cold_type == '冷凍_談':
                    logger.info(f'冷凍数量:{row["呼出番号"]}-{row["献立種類"]}-{row["注文数"] * quantity}-{unit.username.dry_cold_type}/{dc_type}')
                    package.quantity += row["注文数"] * quantity
                    if package.package:
                        logger.debug('pass')
                    else:
                        if (not package.package) and (package_quantity > 0.0):
                            package.package = package_quantity
                        else:
                            package.package = 1.0
                        package.dry_cold_type = dc_type
                else:
                    logger.info(f'冷蔵数量:{row["呼出番号"]}-{row["献立種類"]}-{row["注文数"] * quantity}-{unit.username.dry_cold_type}/{dc_type}')
                    package.quantity += row["注文数"] * quantity
                    if package.package:
                        logger.debug('pass')
                    else:
                        if (not package.package) and (package_quantity > 0.0):
                            package.package = package_quantity
                        else:
                            package.package = 1.0
                        package.dry_cold_type = dc_type

                package.save()

        else:
            logger.warning('マスタ情報未登録のため、原体送りパッケージ情報登録を行いません。')


class QrCodeUtil:
    @classmethod
    def get_image_path_root(cls):
        """
        QRコード画像保存先のパスを取得する。
        """
        image_dir_path = os.path.join(settings.MEDIA_ROOT, 'qr')
        return image_dir_path

    @classmethod
    def get_value(cls, unit: UnitMaster, meal_value: str, picking_type_value: str):
        number = str(unit.unit_number).zfill(3)
        return f'{number}{meal_value}{picking_type_value}'

    @classmethod
    def get_value_v2(cls, unit: UnitMaster, meal_value: str, picking_type_value: str, day: int):
        number = str(unit.unit_number).zfill(3)
        day_str = str(day).zfill(2)
        return f'{number}{meal_value}{picking_type_value}{day_str}'

    @classmethod
    def get_all_in_value(cls, unit: UnitMaster, meal_value: str):
        number = str(unit.unit_number).zfill(3)
        return f'{number}{meal_value}01{number}{meal_value}02{number}{meal_value}03{number}{meal_value}04'

    @classmethod
    def get_all_in_value_v2(cls, unit: UnitMaster, meal_value: str, day: int):
        number = str(unit.unit_number).zfill(3)
        day_str = str(day).zfill(2)
        return f'{number}{meal_value}01{day_str}{number}{meal_value}02{day_str}{number}{meal_value}03{day_str}{number}{meal_value}04{day_str}'

    @classmethod
    def get_value_from_number(cls, number: int, meal_value: str, picking_type_value: str):
        qr_number = str(number).zfill(3)
        return f'{qr_number}{meal_value}{picking_type_value}'

    @classmethod
    def get_value_from_number_v2(cls, number: int, meal_value: str, picking_type_value: str, day: int):
        qr_number = str(number).zfill(3)
        day_str = str(day).zfill(2)
        return f'{qr_number}{meal_value}{picking_type_value}{day_str}'

    @classmethod
    def get_transfer_value(cls, unit: UnitMaster, meal: str):
        number = str(unit.unit_number).zfill(3)
        return f'{number}{meal}'

    @classmethod
    def get_file_name_by_value(cls, value: str):
        return f'{value}.png'

    @classmethod
    def get_file_name_by_prefix_all_value(cls, value: str):
        return f'{value[0:5]}_all.png'

    @classmethod
    def get_file_name_by_prefix_all_value_v2(cls, value: str, day: int):
        day_str = str(day).zfill(2)
        return f'{value[0:5]}_{day_str}_all.png'

    @classmethod
    def get_file_name(cls, unit: UnitMaster, meal_value: str, picking_type_value: str):
        value = cls.get_value(unit, meal_value, picking_type_value)
        return cls.get_file_name_by_value(value)

    @classmethod
    def get_file_name_v2(cls, unit: UnitMaster, meal_value: str, picking_type_value: str, date: int):
        value = cls.get_value_v2(unit, meal_value, picking_type_value, date)
        return cls.get_file_name_by_value(value)

    @classmethod
    def get_file_name_for_mixrice_aggregate(cls, meal_value: str, picking_type_value: str):
        value = cls.get_value_from_number(4, meal_value, picking_type_value)
        return cls.get_file_name_by_value(value)

    @classmethod
    def get_imege_path(cls, unit: UnitMaster, meal_value: str, picking_type_value: str):
        image_dir_path = cls.get_image_path_root()
        return os.path.join(image_dir_path, cls.get_file_name(unit, meal_value, picking_type_value))

    @classmethod
    def get_imege_path_v2(cls, unit: UnitMaster, meal_value: str, picking_type_value: str, day: int):
        image_dir_path = cls.get_image_path_root()
        return os.path.join(image_dir_path, cls.get_file_name_v2(unit, meal_value, picking_type_value, day))

    @classmethod
    def get_imege_path_for_mix_rice_agg(cls, meal_value: str, picking_type_value: str):
        image_dir_path = cls.get_image_path_root()
        return os.path.join(image_dir_path, cls.get_file_name_for_mixrice_aggregate(meal_value, picking_type_value))

    @classmethod
    def get_transfer_imege_path(cls, number: int, meal: str):
        image_dir_path = cls.get_image_path_root()
        number = str(number).zfill(3)
        meal_vaue = cls.convert_to_t_meal_value(meal)
        value = f'{number}{meal_vaue}'
        return os.path.join(image_dir_path, cls.get_file_name_by_prefix_all_value(value))

    @classmethod
    def get_transfer_imege_path_v2(cls, number: int, meal: str, date: int):
        image_dir_path = cls.get_image_path_root()
        number = str(number).zfill(3)
        date_str = str(date).zfill(2)
        meal_vaue = cls.convert_to_t_meal_value(meal)
        value = f'{number}{meal_vaue}'
        return os.path.join(image_dir_path, cls.get_file_name_by_prefix_all_value_v2(value, date_str))

    @classmethod
    def convert_to_meal_value(cls, meal_name: str):
        if meal_name == '朝食':
            return '01'
        elif meal_name == '昼食':
            return '02'
        elif meal_name == '夕食':
            return '03'

    @classmethod
    def convert_to_t_meal_value(cls, meal_name: str):
        if meal_name == '△ 朝':
            return '01'
        elif meal_name == '○ 昼':
            return '02'
        elif meal_name == '□ 夕':
            return '03'

    @classmethod
    def parse_meal(cls, value: str):
        if value == '01':
            return '朝食'
        elif value == '02':
            return '昼食'
        elif value == '03':
            return '夕食'
        else:
            return '不明'

    @classmethod
    def parse_type(cls, value: str):
        if value == '01':
            return '基本食'
        elif value == '02':
            return '嚥下食'
        elif value == '03':
            return '汁・汁具'
        elif value == '04':
            return '原体'
        else:
            return '不明'

    @classmethod
    def perse_qr_value(cls, value: str):
        """
        中袋用のQRコード解析処理。
        """
        number = int(value[:3])
        meal_name = cls.parse_meal(value[3:5])
        type_name = cls.parse_type(value[5:])

        return number, meal_name, type_name

    @classmethod
    def perse_qr_value_v2(cls, value: str):
        """
        中袋用のQRコード解析処理。
        """
        try:
            number = int(value[:3])
            meal_name = cls.parse_meal(value[3:5])
            type_value = value[5:7]
            day = int(value[7:])

            return number, meal_name, type_value, day
        except Exception as e:
            logger.error(f'QRコード値解析失敗:{value}')
            raise e

    @classmethod
    def perse_qr_value_transfer(cls, value: str):
        """
        配送用段ボール用のQRコード解析処理。中袋の種類は考慮しないため、ピッキング種類を返却しない(QRコード値にも入っていない)
        """
        number = int(value[:2])
        meal_name = cls.parse_meal(value[2:4])

        return number, meal_name


class RawPlatePackageRegisterFactory:
    @classmethod
    def create(cls, name: str):
        if '原体' in name:
            # 「原体」の前にある料理名を取得
            word_index = name.find('原体')
            search_name = name[:word_index]
            logger.info(f'調理表原体名:{search_name}')
            res = re.findall('(\D+)(\d|\s)+', search_name)
            if res:
                logger.info(f'調理表原体名抜き出し:{res}')
                base_name = res[0][0]
                if ('(' in base_name) or ('（' in base_name):
                    if (')' in base_name) or ('）' in base_name):
                        pass
                    else:
                        for str in res[0][1:]:
                            base_name = base_name + str
                        for str in res[1]:
                            base_name = base_name + str
                            if (')' in str) or ('）' in str):
                                break
                res_q = re.findall('(\d+|\d+.\d+)g', name)
                res_d = re.findall('乾燥(\d+|\d+.\d+)g', name)
                res_c = re.findall('冷凍(\d+|\d+.\d+)g', name)
                if res_d and res_c:
                    return RawPlatePackageRegister(base_name, raw_quantity_dry=res_d[0], raw_quantity_cold=res_c[0])
                elif res_q:
                    return RawPlatePackageRegister(base_name, raw_quantity_both=res_q[0])
                else:
                    return RawPlatePackageRegister(base_name)

            return InvalidRawPlatePackageRegister(name)
        else:
            logger.info('原体料理外')
            return InvalidRawPlatePackageRegister(name)


class ChillerPicking:
    """
    チラー情報
    """
    def __init__(self, no: int, start_unit_number: int, end_unit_number: int):
        self.no = no
        self.unit_number_range = (start_unit_number, end_unit_number)


class PlateUnitPackage:
    """
    料理・ユニット単位のパッケージ情報
    """
    def __init__(self, unit_package_list):
        self.unit_package_list = unit_package_list

    def get_package(self, package, menu_name=None):
        for unit_package in self.unit_package_list:
            if unit_package.package == package:
                if menu_name:
                    if unit_package.menu_name == menu_name:
                        return unit_package
                else:
                    return unit_package

        return None

    def get_packages_count(self, package, menu_name=None):
        packages = []
        for unit_package in self.unit_package_list:
            if unit_package.package == package:
                if menu_name:
                    if unit_package.menu_name == menu_name:
                        packages.append(unit_package)
                else:
                    packages.append(unit_package)

        if packages:
            return sum([x.count for x in packages])
        else:
            return None

    def is_valid(self):
        return self.unit_package_list

class PickingDirectionWriter:
    thin_Side = Side(style='thin')
    medium_Side = Side(style='medium')
    double_Side = Side(style='double')

    # 施設(ユニット用)
    unit_top_left_border = Border(top=medium_Side, left=medium_Side)
    unit_top_t_right_border = Border(top=medium_Side, right=thin_Side)
    unit_top_bose_b_left_border = Border(top=medium_Side, left=medium_Side, right=thin_Side)
    unit_top_bose_b_right_border = Border(top=medium_Side, left=thin_Side, right=medium_Side)
    unit_top_t_bottom_t_left_border = Border(top=medium_Side, bottom=thin_Side, left=thin_Side)
    unit_top_t_bottom_border = Border(top=medium_Side, bottom=thin_Side)
    unit_top_t_bottom_t_right_border = Border(top=medium_Side, bottom=thin_Side, right=thin_Side)
    unit_top_t_bottom_bose_d_right_border = Border(top=medium_Side, bottom=thin_Side, left=thin_Side, right=double_Side)
    unit_top_t_bottom_bose_d_left_border = Border(top=medium_Side, bottom=thin_Side, left=double_Side, right=thin_Side)
    unit_top_t_bottom_bose_border = Border(top=medium_Side, bottom=thin_Side, left=thin_Side, right=thin_Side)
    unit_top_t_bottom_bose_b_right_border = Border(top=medium_Side, bottom=thin_Side, left=thin_Side, right=medium_Side)
    unit_top_t_bottom_bose_b_left_border = Border(top=medium_Side, bottom=thin_Side, left=medium_Side, right=thin_Side)

    unit_bottom_left_border = Border(bottom=medium_Side, left=medium_Side)
    unit_bottom_t_right_border = Border(bottom=medium_Side, right=thin_Side)
    unit_bottom_bose_b_left_border = Border(bottom=medium_Side, right=thin_Side, left=medium_Side)
    unit_bottom_bose_b_right_border = Border(bottom=medium_Side, right=medium_Side, left=thin_Side)
    unit_bottom_t_top_t_left_border = Border(top=thin_Side, bottom=medium_Side, left=thin_Side)
    unit_bottom_t_top_border = Border(top=thin_Side, bottom=medium_Side)
    unit_bottom_t_top_t_right_border = Border(top=thin_Side, bottom=medium_Side, right=thin_Side)
    unit_bottom_t_top_bose_d_right_border = Border(top=thin_Side, bottom=medium_Side, left=thin_Side, right=double_Side)
    unit_bottom_t_top_bose_d_left_border = Border(top=thin_Side, bottom=medium_Side, left=double_Side, right=thin_Side)
    unit_bottom_t_top_bose_border = Border(top=thin_Side, bottom=medium_Side, left=thin_Side, right=thin_Side)
    unit_bottom_t_top_bose_b_right_border = Border(top=thin_Side, bottom=medium_Side, left=thin_Side, right=medium_Side)
    unit_bottom_t_top_bose_b_left_border = Border(top=thin_Side, bottom=medium_Side, left=medium_Side, right=thin_Side)

    unit_cell_b_left_border = Border(left=medium_Side)
    unit_cell_left_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side)
    unit_cell_border = Border(top=thin_Side, bottom=thin_Side)
    unit_cell_right_border = Border(top=thin_Side, bottom=thin_Side, right=thin_Side)
    unit_cell_bose_d_right_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side, right=double_Side)
    unit_cell_bose_d_left_border = Border(top=thin_Side, bottom=thin_Side, left=double_Side, right=thin_Side)
    unit_cell_bose_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side, right=thin_Side)
    unit_cell_bose_b_right_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side, right=medium_Side)

    # ヘッダ用
    header_top_bottom_left_border = Border(top=medium_Side, bottom=double_Side, left=medium_Side)
    header_top_bottom_t_left_border = Border(top=medium_Side, bottom=double_Side, left=thin_Side)
    header_top_bottom_t_right_border = Border(top=medium_Side, bottom=double_Side, right=thin_Side)
    header_top_bottom_bose_right_border = Border(top=medium_Side, bottom=double_Side, left=thin_Side, right=double_Side)
    header_top_bottom_bose_left_border = Border(top=medium_Side, bottom=double_Side, left=double_Side, right=thin_Side)
    header_top_bottom_border = Border(top=medium_Side, bottom=double_Side)
    header_top_bottom_t_bose_border = Border(top=medium_Side, bottom=double_Side, left=thin_Side, right=thin_Side)
    header_top_bottom_bose_right_b_border = Border(top=medium_Side, bottom=double_Side, left=thin_Side, right=medium_Side)
    header_top_bottom_bose_left_b_border = Border(top=medium_Side, bottom=double_Side, left=medium_Side, right=thin_Side)

    cell_top_left_border = Border(top=thin_Side, left=thin_Side)
    cell_top_right_border = Border(top=thin_Side, right=thin_Side)
    cell_bottom_left_border = Border(bottom=thin_Side, left=thin_Side)
    cell_bottom_right_border = Border(bottom=thin_Side, right=thin_Side)

    cell_top_bottom_left_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side)
    cell_top_bottom_right_border = Border(top=thin_Side, bottom=thin_Side, right=thin_Side)

    cell_top_border = Border(top=thin_Side)
    cell_bottom_border = Border(bottom=thin_Side)
    cell_top_bottom_border = Border(top=thin_Side, bottom=thin_Side)

    cell_left_border = Border(left=thin_Side)
    cell_right_border = Border(right=thin_Side)

    cell_all_border = Border(top=thin_Side, bottom=thin_Side, left=thin_Side, right=thin_Side)

    # セル色
    soft_fill = PatternFill(patternType='solid', fgColor='FFCCFF')
    jerry_fill = PatternFill(patternType='solid', fgColor='F8CBAD')
    mixer_fill = PatternFill(patternType='solid', fgColor='FFFF99')

    allergen_fill = PatternFill(patternType='solid', fgColor='C6E0B4')
    disable_fill = PatternFill(patternType='solid', fgColor='AAAAAA')

    dry_fill = PatternFill(patternType='solid', fgColor='FCE4D6')
    cold_fill = PatternFill(patternType='solid', fgColor='DDEBF7')

    # 文字そろえ
    top_alignment = Alignment(vertical='top')
    right_alignment = Alignment(horizontal='right')

    # 喫食日、食事区分
    eating_day_meal_font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True, u='single')

    # 汎用
    # 太字フォント
    bold_font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True)

    # 列最大行数ラベル
    max_column_label = 'A'

    # 最大出力行
    PAGE_MAX_ROW = 36

    # 食種ソート順
    meal_order_dict = {
        '朝食': 0,
        '昼食': 1,
        '夕食': 2,
    }

    def __init__(self, chillers):
        self.chillers = chillers
        self.workbook = None

    def write(self, eating_day, meal_list, cooking_eating_dict):
        pass

    def get_package_list(self, eating_day, meal, chiller):
        pass

    def get_plate_list(self, eating_day, meal):
        pass

    def write_excel(self):
        pass

    def get_qr_image(self, unit: UnitMaster, meal_name, picking_type_value, day: int):
        path = QrCodeUtil.get_imege_path_v2(unit, QrCodeUtil.convert_to_meal_value(meal_name), picking_type_value, day)
        qr_image = excel.drawing.image.Image(path)
        qr_image.width = 100
        qr_image.height = 100

        return qr_image

    def save(self, ws, cooking_day, type_name: str):
        output_dir = os.path.join(settings.OUTPUT_DIR, 'picking')
        os.makedirs(output_dir, exist_ok=True)  # 上書きOK
        save_path = os.path.join(output_dir, f'ピッキング指示書_{str(cooking_day)}_製造_{type_name}.xlsx')
        footer_text = '&P / &Nページ'

        ws.oddFooter.center.text = footer_text
        ws.page_margins.bottom = 1.0

        self.workbook.save(save_path)

        self.workbook.close()

    def write_chiller_no(self, ws, row, no: int):
        ws.cell(row, 11, 'チラーNo.')
        ws.cell(row, 12, f'{no}')

        ws.cell(row, 11).font = self.bold_font
        ws.cell(row, 12).font = self.bold_font
        ws.cell(row, 12).alignment = self.right_alignment

    def write_eating_day_with_meal(self, ws, row, eating_day, meal: str):
        display_date = eating_day.strftime('%Y年%m月%d日')
        ws.cell(row, 2, f'{display_date} {MealUtil.add_name_mark(meal)}')
        ws.cell(row, 2).font = self.eating_day_meal_font

    def write_unit_name_for_name(self, ws, row, unit_number, display_name):
        display_unit = f'{unit_number}.{display_name}'
        ws.cell(row, 2, display_unit)
        ws.cell(row, 2).font = self.bold_font

    def write_unit_name(self, ws, row, unit):
        self.write_unit_name_for_name(ws, row, unit.unit_number, unit.short_name)

    def get_unit_package(self, index, unit_package_list):
        package_list = []
        for up in unit_package_list:
            if up.index == index and up.is_basic_plate:
                package_list.append(up)

        return PlateUnitPackage(package_list)

    def get_mix_rice_aggregate_package_from_list(self, unit, unit_package_list):
        """
        混ぜご飯集約対象のユニットのリストから、指定されたユニットの袋情報を取得する
        """
        for up in unit_package_list:
            # 現状、通常の合算名称と混ぜご飯用の合算名称は異なるため、それを前提としている
            if up.unit_name != unit.calc_name and up.is_basic_plate:
                return PlateUnitPackage([up])

        return None

    def get_mix_rice_aggregate_package(self, unit, cooking_day, eating_day, meal, menu):
        """
        混ぜご飯集約対象のユニットのリストから、指定されたユニットの袋情報を取得する
        """
        unit_pacage_qs = UnitPackage.objects.filter(
            unit_number=unit.unit_number, cooking_day=cooking_day, eating_day=eating_day,
            meal_name=meal, menu_name=menu, is_basic_plate=True
        ).exclude(
            unit_name=unit.calc_name
        )

        if unit_pacage_qs.exists:
            return PlateUnitPackage([unit_pacage_qs.first()])

        return None

    def get_allergen_unit_package(self, instance, unit_package_list):
        package_list = []
        for up in unit_package_list:
            if up.cooking_direction == instance:
                package_list.append(up)

        return PlateUnitPackage(package_list)

    def adjust_print_area(self, ws, row: int, break_list):
        ws.print_area = f'A1:{self.max_column_label}{row}'
        for br in break_list:
            ws.row_breaks.append(br)

    def compare_meal(self, meal1, meal2):
        cmp_seq = self.meal_order_dict[meal1] - self.meal_order_dict[meal2]
        return cmp_seq

    def generate_cooking_eating_dict(self, cooking_eating_dict):
        eating_day_list = sorted([key for key in cooking_eating_dict.keys()])

        for eating_day in eating_day_list:
            meal_list = cooking_eating_dict[eating_day]

            yield eating_day, sorted(meal_list, key=cmp_to_key(self.compare_meal))

    def open_workbook(self):
        """
        ピッキング指示書テンプレートファイルを開く
        """
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/picking_basic.xlsx')
        self.workbook = excel.load_workbook(template_path)


class MixRicePlates:
    """
    ピッキング指示書出力に置ける混ぜご飯制御を行うクラス
    """

    def __init__(self, qs):
        # CookingDirectionPlateのquerysetを想定
        self.list = list(qs)

    def is_diplay_mixrice_aaggregate(self):
        if self.list:
            first_plate = self.list[0]
            if first_plate.is_mix_rice:
                # gの量を名称に含むかどうかを判断する
                return ('g' in first_plate.plate_name) or ('ｇ' in first_plate.plate_name)
            else:
                False

        return False

    def is_all_mixrice(self):
        if [x for x in self.list if not x.is_mix_rice]:
            return False
        else:
            return True


class BasicPickingDirectionWriter(PickingDirectionWriter):
    # 5人用袋セル色
    package_5_fill = PatternFill(patternType='solid', fgColor='FCE4D6')

    max_column_label = 'M'

    def __init__(self, chillers, chiller_no):
        super(BasicPickingDirectionWriter, self).__init__(chillers)
        self.chiller_no = chiller_no
        self.PACKAGE_FOR_10 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['BASIC_10'])
        self.PACKAGE_FOR_5 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['BASIC_5'])
        self.PACKAGE_FOR_FRACTION = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['BASIC_FRACTION'])
        self.PACKAGE_FOR_1 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['BASIC_1'])
        self.PACKAGE_FOR_UNIT = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['BASIC_UNIT'])

        # 混ぜご飯合算用の情報を集計する辞書
        self.agg_mix_rice_dict = {}

        self.eating_plates = {}

    def add_mix_rice_plate(self, unit_number: int, plate):
        # 複数セットを想定して記載をしているが、運用上単一セットのみのため、詳細検証していない。複数セット対応時は検討必要。
        if unit_number in self.agg_mix_rice_dict:
            plate_list = self.agg_mix_rice_dict[unit_number]
            if not (plate in plate_list):
                plate_list.append(plate)
        else:
            self.agg_mix_rice_dict[unit_number] = [plate]

    def write_table_header(self, ws, row):
        ws.cell(row, 2, '施設名')
        ws.cell(row, 4, '料理名')
        ws.cell(row, 7, 'アレルギー')
        ws.cell(row, 8, '10人用')
        ws.cell(row, 9, '5人用')
        ws.cell(row, 9).fill = self.package_5_fill
        ws.cell(row, 10, '端数')
        ws.cell(row, 11, '1人用')
        ws.cell(row, 12, '施設毎')

        # フォント
        for i in range(11):
            ws.cell(row, 2 + i).font = self.bold_font

        # 枠の記入
        ws.cell(row, 2).border = self.header_top_bottom_left_border
        ws.cell(row, 3).border = self.header_top_bottom_t_right_border

        ws.cell(row, 4).border = self.header_top_bottom_t_left_border
        ws.cell(row, 5).border = self.header_top_bottom_border
        ws.cell(row, 6).border = self.header_top_bottom_t_right_border

        ws.cell(row, 7).border = self.header_top_bottom_bose_right_border
        ws.cell(row, 8).border = self.header_top_bottom_bose_left_border
        ws.cell(row, 9).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 10).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 11).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 12).border = self.header_top_bottom_bose_right_b_border

    def write_plate(self, ws, row, plate, unit_package, allergen=None):
        ws.cell(row, 4, plate.plate_name)
        ws.cell(row, 4).font = self.bold_font

        if allergen:
            allergen_name = allergen.kana_name or 'あり'
            ws.cell(row, 7, allergen_name)
            ws.cell(row, 7).font = self.bold_font

            for i in range(12 - 4 + 1):
                ws.cell(row, 4 + i).fill = self.allergen_fill
        else:
            ws.cell(row, 7, '')
            ws.cell(row, 9).fill = self.package_5_fill

        ws.cell(row, 8).font = self.bold_font
        ws.cell(row, 9).font = self.bold_font
        ws.cell(row, 10).font = self.bold_font
        ws.cell(row, 11).font = self.bold_font
        ws.cell(row, 12).font = self.bold_font

        # 原体送り対応
        if '原体' in plate.plate_name:
            ws.cell(row, 8, '原体送り資料参照')
            # セル結合は別メソッドで対応
            return

        # 袋数の出力
        # 10人用
        package = unit_package.get_package(self.PACKAGE_FOR_10)
        if package:
            ws.cell(row, 8, package.count)

        # 5人用
        package = unit_package.get_package(self.PACKAGE_FOR_5)
        if package:
            ws.cell(row, 9, package.count)

        # 端数用
        if allergen:
            pc = unit_package.get_packages_count(self.PACKAGE_FOR_FRACTION)
            if pc:
                ws.cell(row, 10, pc)
        else:
            package = unit_package.get_package(self.PACKAGE_FOR_FRACTION)
            if package:
                ws.cell(row, 10, package.count)

        # 1人用
        if allergen:
            pc = unit_package.get_packages_count(self.PACKAGE_FOR_1)
            if pc:
                ws.cell(row, 11, pc)
        else:
            package = unit_package.get_package(self.PACKAGE_FOR_1)
            if package:
                ws.cell(row, 11, package.count)

        # 施設毎用
        package = unit_package.get_package(self.PACKAGE_FOR_UNIT)
        if package:
            ws.cell(row, 12, package.count)

    def write_unit_border(self, ws, start_row, end_row, is_first_unit):
        # ユニット枠上部
        ws.cell(start_row, 2).border = self.unit_top_left_border
        ws.cell(start_row, 3).border = self.unit_top_t_right_border
        ws.cell(start_row, 4).border = self.unit_top_t_bottom_t_left_border
        ws.cell(start_row, 5).border = self.unit_top_t_bottom_border
        ws.cell(start_row, 6).border = self.unit_top_t_bottom_t_right_border
        ws.cell(start_row, 7).border = self.unit_top_t_bottom_bose_d_right_border
        ws.cell(start_row, 8).border = self.unit_top_t_bottom_bose_d_left_border
        ws.cell(start_row, 9).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 10).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 11).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 12).border = self.unit_top_t_bottom_bose_b_right_border

        # 料理各行
        for i in range(end_row - start_row - 1):
            ws.cell(start_row + i + 1, 2).border = self.unit_cell_b_left_border
            ws.cell(start_row + i + 1, 4).border = self.unit_cell_left_border
            ws.cell(start_row + i + 1, 5).border = self.unit_cell_border
            ws.cell(start_row + i + 1, 6).border = self.unit_cell_right_border
            ws.cell(start_row + i + 1, 7).border = self.unit_cell_bose_d_right_border
            ws.cell(start_row + i + 1, 8).border = self.unit_cell_bose_d_left_border
            ws.cell(start_row + i + 1, 9).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 10).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 11).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 12).border = self.unit_cell_bose_b_right_border

        # ユニット枠下部
        ws.cell(end_row, 2).border = self.unit_bottom_left_border
        ws.cell(end_row, 3).border = self.unit_bottom_t_right_border
        ws.cell(end_row, 4).border = self.unit_bottom_t_top_t_left_border
        ws.cell(end_row, 5).border = self.unit_bottom_t_top_border
        ws.cell(end_row, 6).border = self.unit_bottom_t_top_t_right_border
        ws.cell(end_row, 7).border = self.unit_bottom_t_top_bose_d_right_border
        ws.cell(end_row, 8).border = self.unit_bottom_t_top_bose_d_left_border
        ws.cell(end_row, 9).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 10).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 11).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 12).border = self.unit_bottom_t_top_bose_b_right_border

        # 原体送りのセル結合
        for row_index in range(end_row - start_row + 1):
            current_row = start_row + row_index
            value = ws.cell(current_row, 4).value
            if value and ('原体' in value):
                ws.merge_cells(start_row=current_row, end_row=current_row, start_column=8, end_column=12)

    def _merge_plate_rows(self, ws, start_row, end_row):
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=4, end_column=6)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=7, end_column=7)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=8, end_column=8)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=9, end_column=9)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=10, end_column=10)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=11, end_column=11)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=12, end_column=12)

    def _contains_mix_rice_plate(self, plate_qs):
        """
        対象食事区分の料理の中に、混ぜご飯があるかどうかを確認する。plate_qsはCookingDirectionPlateを参照するquerysetを前提とする。
        """
        combine_ps = plate_qs.filter(is_mix_rice=True)
        if combine_ps.exists():
            return True
        else:
            return False

    def _is_mixrice_aggregate_unit(self, unint_number):
        """
        混ぜご飯限定で集約するユニットかどうかを判定する。
        """
        for unit_numbers in settings.MIX_RICE_AGGREGATE_UNITS:
            if unint_number in unit_numbers:
                return True, unit_numbers[0], unit_numbers[-1]

        return False, None, None

    def get_qr_image_for_mix_rice_agg(self, meal_name, picking_type_value):
        path = QrCodeUtil.get_imege_path_for_mix_rice_agg(QrCodeUtil.convert_to_meal_value(meal_name), picking_type_value)
        qr_image = excel.drawing.image.Image(path)
        qr_image.width = 100
        qr_image.height = 100

        return qr_image

    def write_mixrice_aggregate_unit(self, ws, meal, mix_rice_plates, unit_number, row, aggreate_unit_package, tmp_row, page_start_row):
        aup = [x for x in aggreate_unit_package.unit_package_list if (x is not None)]
        if not aup:
            return row, None, tmp_row
        # 施設名の出力
        unit_start_row = row
        self.write_unit_name_for_name(ws, row, unit_number, aup[0].unit_name)

        # QRコード画像の表示
        qr_image = self.get_qr_image_for_mix_rice_agg(meal, '01')
        ws.add_image(qr_image, ws.cell(row + 1, 3).coordinate)

        # 料理の取得
        for p in mix_rice_plates:
            self.write_plate(ws, row, p, PlateUnitPackage(aup))
            row += 1
        unit_end_row = row - 1

        # QRコード画像表示のため、出力行数が少ない場合は、行の高さを広げる
        row_diff = unit_end_row - unit_start_row
        if row_diff < 5:
            target_row = unit_start_row + 1
            added_height = 18.75 * (5 - row_diff) + 5
            ws.row_dimensions[target_row].height = added_height
            tmp_row += 5 - row_diff

        page_row_diff = unit_end_row + tmp_row - page_start_row
        braak_row = None
        if page_row_diff > self.PAGE_MAX_ROW:
            braak_row = unit_start_row - 1

        # ユニット単位の枠描画
        self.write_unit_border(ws, unit_start_row, unit_end_row + 1, False)

        return unit_end_row + 2, braak_row, tmp_row

    def save_by_chiller(self, ws, cooking_day, chiller_no: int):
        """
        チラー毎の指示書エクセルファイルを保存する。
        """
        output_dir = os.path.join(settings.OUTPUT_DIR, 'picking')
        os.makedirs(output_dir, exist_ok=True)  # 上書きOK
        save_path = os.path.join(output_dir, f'ピッキング指示書_{str(cooking_day)}_製造_基本食_チラー{chiller_no}.xlsx')
        footer_text = '&P / &Nページ'

        ws.oddFooter.center.text = footer_text
        ws.page_margins.bottom = 1.0

        self.workbook.save(save_path)
        self.workbook.close()

    def _get_plates(self, cooking_day, eating_day, meal_name):
        if (eating_day, meal_name) in self.eating_plates:
            return self.eating_plates[(eating_day, meal_name)]
        else:
            plates = CookingDirectionPlate.objects.filter(
                cooking_day=cooking_day,
                is_basic_plate=True, eating_day=eating_day, meal_name=meal_name,
                is_soup=False).order_by('seq_meal', 'index')
            self.eating_plates[(eating_day, meal_name)] = plates
            return plates

    def write(self, cooking_day, meal_list):
        logger.info('基本食のピッキング指示書を出力')

        cooking_eating_dict = EatingManagement.get_meals_dict_by_cooking_day(cooking_day)

        for chiller in self.chillers:
            if chiller.no != self.chiller_no:
                continue

            logger.info(f'チラー={chiller.no}')
            # 全体まとめて出力->種類ごと(基本食はさらにチラー毎)に出力になったため、個別でworkbookを扱う
            self.open_workbook()

            row = 3
            page_start_row = row
            ws = self.workbook['基本食']
            max_modified = UnitPackage.objects.filter(cooking_day=cooking_day).order_by('-register_at').first()
            ws.cell(1, 1, f"出力日時:{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(1, 12, f"最終データ更新日時:{max_modified.register_at.strftime('%Y-%m-%d %H:%M:%S')}")
            break_list = []
            write_units = []  # 処理済みのユニット
            unit_list = list(UnitMaster.objects.filter(
                unit_number__range=chiller.unit_number_range).order_by('unit_number', 'id'))
            for eating_day, cooking_meal_list in self.generate_cooking_eating_dict(cooking_eating_dict):

                # 対象製造日の全ての料理を取得する
                for meal in cooking_meal_list:
                    # 対象喫食日の朝食、昼食、夕食が製造対象でなければ次にする
                    meal_plats_qs = self._get_plates(cooking_day, eating_day, meal)
                    if not meal_plats_qs.exists():
                        continue

                    # 混ぜご飯を含むかどうかを取得
                    is_contains_mix_rice = self._contains_mix_rice_plate(meal_plats_qs)
                    agg_plates = MixRicePlates(meal_plats_qs)

                    write_units.clear()
                    aggreate_unit_package = None

                    # 喫食日、食事区分を出力する
                    self.write_eating_day_with_meal(ws, row, eating_day, meal)
                    row += 1

                    tmp_rows = 0

                    self.agg_mix_rice_dict.clear()

                    # チラー番号を出力する
                    self.write_chiller_no(ws, row, chiller.no)
                    row += 1

                    # ヘッダ出力
                    self.write_table_header(ws, row)
                    row += 1

                    is_first_unit = True
                    for unit in unit_list:
                        if (unit.unit_number, unit.calc_name) in write_units:
                            logger.info(f'処理済みユニット：{unit.unit_number}, {unit.calc_name}({unit.unit_name})')
                            continue

                        # 個食の出力はメインに含めるため、出力しない
                        if '個食' in unit.unit_name:
                            logger.info(f'対象外(個食)：{unit.unit_number}-{eating_day}-{meal}')
                            continue

                        if is_contains_mix_rice:
                            # 混ぜご飯集約対象のユニットかどうかを判定
                            is_aggregated, first_unit_number, last_unit_number = \
                                self._is_mixrice_aggregate_unit(unit.unit_number)
                        else:
                            is_aggregated = False
                            last_unit_number = None

                        unit_start_row = row

                        # 対象喫食日・食事区分にユニットの注文があるか確認
                        unit_order_qs = Order.objects.filter(
                            eating_day=eating_day, meal_name__meal_name=meal, quantity__gt=0,
                            unit_name=unit
                        )
                        is_kizawa_special = False
                        if not unit_order_qs.exists():
                            logger.info(f'注文無し：{unit.unit_number}-{eating_day}-{meal}')

                            # 自ユニットの注文は0でも、混ぜご飯集約の対象の可能性があるので、チェック
                            if is_aggregated:
                                # 混ぜご飯集約対象のユニットの場合

                                if unit.unit_number == last_unit_number:
                                    # 集約ユニット群の最後のユニットの場合、混ぜご飯集約情報の抽出を行う。

                                    mix_rice_plate_qs = meal_plats_qs.filter(is_mix_rice=True)
                                    if mix_rice_plate_qs.exists():
                                        for rice_plate in mix_rice_plate_qs:
                                            self.add_mix_rice_plate(first_unit_number, rice_plate)

                                        # 対象のユニットのみの情報を抽出
                                        if aggreate_unit_package and aggreate_unit_package.unit_package_list and (first_unit_number in self.agg_mix_rice_dict):
                                            plates = self.agg_mix_rice_dict[first_unit_number]
                                            row, break_row, tmp_rows = self.write_mixrice_aggregate_unit(
                                                ws, meal, plates, first_unit_number, row,
                                                aggreate_unit_package, tmp_rows, unit_start_row)

                                            # 改ページ制御
                                            if break_row:
                                                break_list.append(Break(break_row))
                                                logger.debug(f'{unit}[Break-Row]:{break_row}')
                                                page_start_row = row + 1
                                                tmp_rows = 0

                                    else:
                                        logger.warning(f'混ぜご飯対象plateの取得に失敗：{unit.unit_number}-{eating_day}-{meal}')

                                elif first_unit_number == unit.unit_number:
                                    # 集約ユニット群の最初のユニットの場合、混ぜご飯集約情報を取得する。

                                    aup = self.get_mix_rice_aggregate_package(
                                        unit, cooking_day, eating_day, meal, '常食')
                                    if aup:
                                        aggreate_unit_package = aup


                            # ユニットの注文がないならこれ以上の処理は不要
                            if unit.unit_number == 31:
                                if Order.objects.filter(eating_day=eating_day, meal_name__meal_name=meal, quantity__gt=0, unit_name__unit_number=unit.unit_number).exclude(unit_name=unit).exists():
                                    logger.info(f'木沢個特別対応')
                                    is_kizawa_special = True
                                else:
                                    continue
                            else:
                                continue

                        # アレルギー以外の注文がないかどうか
                        if is_kizawa_special:
                            is_only_allergen = False
                        else:
                            is_only_allergen = (not unit_order_qs.filter(allergen__allergen_name='なし').exists())

                        # 袋数情報の取得
                        unit_pacage_qs = UnitPackage.objects.filter(
                            unit_number=unit.unit_number, cooking_day=cooking_day, eating_day=eating_day,
                            meal_name=meal, menu_name='常食'
                        )
                        unit_package_list = list(unit_pacage_qs)

                        # 施設名の出力
                        self.write_unit_name(ws, row, unit)

                        # QRコード画像の表示
                        qr_image = self.get_qr_image(unit, meal, '01', eating_day.day)
                        ws.add_image(qr_image, ws.cell(row + 1, 3).coordinate)
                        del qr_image

                        # 料理の出力
                        for plate in agg_plates.list:
                            if is_aggregated:
                                if plate.is_mix_rice and (last_unit_number == unit.unit_number):
                                    self.add_mix_rice_plate(first_unit_number, plate)
                                elif first_unit_number == unit.unit_number:
                                    p_list = self.get_mix_rice_aggregate_package_from_list(unit, unit_package_list)
                                    if p_list:
                                        aggreate_unit_package = p_list

                            plate_unit_package = self.get_unit_package(plate.index, unit_package_list)
                            if is_aggregated:
                                if agg_plates.is_all_mixrice():
                                    # 料理が混ぜご飯のみの場合
                                    if not is_only_allergen:
                                        logger.info('料理名出力')
                                        self.write_plate(ws, row, plate, plate_unit_package)
                                else:
                                    # 混ぜご飯以外もある場合

                                    if agg_plates.is_diplay_mixrice_aaggregate():
                                        if plate.is_mix_rice:
                                            row -= 1
                                        else:
                                            if not is_only_allergen:
                                                logger.info('料理名出力')
                                                self.write_plate(ws, row, plate, plate_unit_package)
                                    else:
                                        if not is_only_allergen:
                                            logger.info('料理名出力')
                                            self.write_plate(ws, row, plate, plate_unit_package)
                            else:
                                if not is_only_allergen:
                                    # アレルギー以外の注文があれば出力、なければアレルギーを出力する場合に出力(処理は後述)
                                    logger.info('料理名出力')
                                    self.write_plate(ws, row, plate, plate_unit_package)

                            # アレルギー代替料理の出力対象かどうか判定
                            ar_relation_ps = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')
                            is_output = False
                            is_write_source = False
                            if ar_relation_ps.exists():
                                # アレルギーの元は出力する
                                if is_aggregated and plate.is_mix_rice:
                                    row += 1    # 減らした分を戻す
                                    self.write_plate(ws, row, plate, plate_unit_package)
                                for key, group in groupby(ar_relation_ps, key=lambda x: x.plate):
                                    display_allergens = []

                                    for relation in group:
                                        ar_list, ar_menu = \
                                            CookingDirectionPlatesManager.get_allergens_with_menu(relation.code, cooking_day)
                                        if ar_menu == '常食':
                                            # 嚥下の内容は、嚥下の指示書で出力する
                                            """
                                            if '個食' in ar_list:
                                                ar_list.remove('個食')
                                            """
                                            if 'ﾌﾘｰｽﾞ' in ar_list:
                                                ar_list.remove('ﾌﾘｰｽﾞ')
                                            if '個食' in [x.allergen_name for x in ar_list]:
                                                for unit_ar_order in Order.objects.filter(
                                                        eating_day=eating_day, meal_name__meal_name=meal,
                                                        quantity__gt=0,
                                                        unit_name__unit_number=unit.unit_number, allergen__allergen_name='なし'):
                                                    if not ('個食' in unit_ar_order.unit_name.unit_name):
                                                        break
                                                    if not (unit_ar_order.allergen in display_allergens):
                                                        display_allergens.append(unit_ar_order.allergen)

                                                        # アレルギー代替食を表示するために、アレルギー代替元を空で表示する
                                                        if is_only_allergen:
                                                            if is_write_source:
                                                                row += 1
                                                                self.write_plate(ws, row, plate, PlateUnitPackage([]))
                                                                is_write_source = True
                                            else:
                                                for unit_ar_order in Order.objects.filter(
                                                        eating_day=eating_day, meal_name__meal_name=meal,
                                                        quantity__gt=0,
                                                        unit_name__unit_number=unit.unit_number, allergen__in=ar_list):
                                                    if not (unit_ar_order.allergen in display_allergens):
                                                        display_allergens.append(unit_ar_order.allergen)

                                                        # アレルギー代替食を表示するために、アレルギー代替元を空で表示する
                                                        if is_only_allergen:
                                                            if is_write_source:
                                                                row += 1
                                                                self.write_plate(ws, row, plate, PlateUnitPackage([]))
                                                                is_write_source = True

                                    # アレルギー代替料理の出力
                                    logger.info(f'【{eating_day}-{meal}】{unit.unit_number}{unit.unit_name}:{display_allergens}')
                                    for display_allergen in display_allergens:
                                        is_output = True
                                        row += 1
                                        ar_plate_package = self.get_allergen_unit_package(key, unit_package_list)
                                        logger.info(
                                            f'【{key.plate_name}】{ar_plate_package.unit_package_list}')
                                        self.write_plate(ws, row, key, ar_plate_package, display_allergen)

                            if is_output or (not is_only_allergen):
                                row += 1

                        unit_package_list.clear()
                        unit_end_row = row - 1
                        logger.debug(f'{unit}[Unit-End-Row]:{unit_end_row}({row})')

                        # QRコード画像表示のため、出力行数が少ない場合は、行の高さを広げる
                        row_diff = unit_end_row - unit_start_row
                        if row_diff < 5:
                            if row_diff == 0:
                                row += 1
                                unit_end_row += 1
                                self._merge_plate_rows(ws, unit_start_row, unit_end_row)
                            target_row = unit_start_row + 1
                            added_height = 18.75 * (5 - row_diff) + 5
                            ws.row_dimensions[target_row].height = added_height

                            tmp_rows += 5 - row_diff

                        # 同一ユニットが別ページに分かれないようにする対応
                        page_row_diff = unit_end_row + tmp_rows - page_start_row
                        if page_row_diff > self.PAGE_MAX_ROW:
                            logger.debug(f'{unit}[Break-Page]:{page_row_diff}({unit_end_row}+{tmp_rows}-{page_start_row})')
                            break_list.append(Break(unit_start_row - 1))
                            logger.debug(f'{unit}[Break-Row]:{unit_start_row - 1}')
                            page_start_row = row + 1
                            tmp_rows = 0

                        # ユニット単位の枠描画
                        self.write_unit_border(ws, unit_start_row, unit_end_row, is_first_unit)
                        is_first_unit = False

                        # 混ぜご飯合算ユニットを出力する場合
                        if is_aggregated and self.agg_mix_rice_dict and agg_plates.is_diplay_mixrice_aaggregate() and (not agg_plates.is_all_mixrice()):
                            mix_rice_plates = self.agg_mix_rice_dict[first_unit_number]
                            if aggreate_unit_package and aggreate_unit_package.unit_package_list:
                                row, break_row, tmp_rows = self.write_mixrice_aggregate_unit(
                                    ws, meal, mix_rice_plates, first_unit_number, row, aggreate_unit_package, tmp_rows, unit_start_row)

                                # 改ページ制御
                                if break_row:
                                    break_list.append(Break(break_row))
                                    page_start_row = row + 1
                                    tmp_rows = 0

                        # 処理済みユニットの登録
                        write_units.append((unit.unit_number, unit.calc_name))

                    # 食事区分違いを別ページ開始とするように対応
                    break_list.append(Break(row))

                    # 食事区分の出力完了
                    row += 1
                    page_start_row = row

                cooking_meal_list.clear()

            # 印刷範囲の修正
            self.adjust_print_area(ws, row, break_list[:-1])

            # ファイルの保存
            self.save_by_chiller(ws, cooking_day, chiller.no)

            unit_list.clear()


class EngePickingDirectionWriter(PickingDirectionWriter):
    ENGE_MENU_NAME_LIST = ['ソフト', 'ゼリー', 'ミキサー']
    max_column_label = 'N'

    # 最大出力行
    PAGE_MAX_ROW = 34

    def __init__(self, chillers):
        super(EngePickingDirectionWriter, self).__init__(chillers)
        self.PACKAGE_FOR_7 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['ENGE_7'])
        self.PACKAGE_FOR_14 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['ENGE_14'])
        self.PACKAGE_FOR_20 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['ENGE_20'])
        self.PACKAGE_FOR_2 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['ENGE_2'])
        self.PACKAGE_FOR_1 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['ENGE_1'])
        self.reference_font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True, color="FF0000")

    def _is_miso_soup(self, plate):
        if plate.is_soup:
            plate_name = plate.plate_name
            if PlateNameAnalizeUtil.is_miso_soup(plate_name):
                if PlateNameAnalizeUtil.is_soup_liquid(plate_name):
                    return True
        return False

    # 書き込み先列が異なるため、オーバライド
    def write_chiller_no(self, ws, row, no: int):
        ws.cell(row, 12, 'チラーNo.')
        ws.cell(row, 13, f'{no}')

        ws.cell(row, 12).font = self.bold_font
        ws.cell(row, 13).font = self.bold_font
        ws.cell(row, 13).alignment = self.right_alignment

    def write_table_header(self, ws, row):
        ws.cell(row, 2, '施設名')
        ws.cell(row, 4, '料理名')
        ws.cell(row, 7, '献立種別')
        ws.cell(row, 8, 'アレルギー')
        ws.cell(row, 9, '主菜7人用')
        ws.cell(row, 10, '小鉢14人用')
        ws.cell(row, 11, '添え20人用')
        ws.cell(row, 12, '1人用')
        ws.cell(row, 13, '2人用')

        # フォント
        for i in range(12):
            ws.cell(row, 2 + i).font = self.bold_font

        # 枠の記入
        ws.cell(row, 2).border = self.header_top_bottom_left_border
        ws.cell(row, 3).border = self.header_top_bottom_t_right_border

        ws.cell(row, 4).border = self.header_top_bottom_t_left_border
        ws.cell(row, 5).border = self.header_top_bottom_border
        ws.cell(row, 6).border = self.header_top_bottom_t_right_border

        ws.cell(row, 7).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 8).border = self.header_top_bottom_bose_right_border
        ws.cell(row, 9).border = self.header_top_bottom_bose_left_border
        ws.cell(row, 10).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 11).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 12).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 13).border = self.header_top_bottom_bose_right_b_border

    def fill_menu_color(self, ws, row, fill):
        for i in range(13 - 7 + 1):
            ws.cell(row, 7 + i).fill = fill

    def write_plate(self, ws, row, plate, menu_name, unit_package, eating_day, unit, allergen=None):
        plate_name = plate.plate_name

        if PlateNameAnalizeUtil.is_raw_plate_name(plate_name):
            if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate_name, eating_day)[0]:
                ws.cell(row, 4, f'{plate_name}\n(嚥下食は袋あり)')
                ws.cell(row, 4).alignment = Alignment(wrapText=True)
                ws.row_dimensions[row].height = 18.75 * 3
            else:
                ws.cell(row, 4, plate_name)
        else:
            ws.cell(row, 4, plate_name)
        ws.cell(row, 4).font = self.bold_font

        ws.cell(row, 7, menu_name)
        ws.cell(row, 7).font = self.bold_font

        if allergen:
            allergen_name = allergen.kana_name or 'あり'
            ws.cell(row, 8, allergen_name)
            ws.cell(row, 8).font = self.bold_font
            self.fill_menu_color(ws, row, self.allergen_fill)
        else:
            ws.cell(row, 8, '')
            if menu_name == 'ソフト':
                self.fill_menu_color(ws, row, self.soft_fill)
            elif menu_name == 'ゼリー':
                self.fill_menu_color(ws, row, self.jerry_fill)
            elif menu_name == 'ミキサー':
                self.fill_menu_color(ws, row, self.mixer_fill)

        ws.cell(row, 9).font = self.bold_font
        ws.cell(row, 10).font = self.bold_font
        ws.cell(row, 11).font = self.bold_font
        ws.cell(row, 12).font = self.bold_font
        ws.cell(row, 13).font = self.bold_font

        # 原体送り対応
        required_write_reference = False
        if PlateNameAnalizeUtil.is_raw_plate_name(plate_name):
            if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate_name, eating_day)[0]:
                # 嚥下製造対象は通常通り出力する
                pass
            else:
                if (unit.username.dry_cold_type == '乾燥') and (PlateNameAnalizeUtil.is_required_reference(plate_name)):
                    required_write_reference = True
                    ws.row_dimensions[row].height = 18.75 * 2
                else:
                    ws.cell(row, 9, '原体送り資料参照')
                    return required_write_reference

        # 袋数出力
        # 7人用
        package = unit_package.get_package(self.PACKAGE_FOR_7, menu_name)
        if package:
            if required_write_reference:
                ws.cell(row, 9, f'{package.count}\n(原体参照)')
                ws.cell(row, 9).font = self.reference_font
                ws.cell(row, 9).alignment = Alignment(shrink_to_fit=True)
            else:
                ws.cell(row, 9, package.count)

        # 14人用
        package = unit_package.get_package(self.PACKAGE_FOR_14, menu_name)
        if package:
            if required_write_reference:
                ws.cell(row, 10, f'{package.count}\n(原体参照)')
                ws.cell(row, 10).font = self.reference_font
                ws.cell(row, 10).alignment = Alignment(shrink_to_fit=True)
            else:
                ws.cell(row, 10, package.count)

        # 20人用
        package = unit_package.get_package(self.PACKAGE_FOR_20, menu_name)
        if package:
            if required_write_reference:
                ws.cell(row, 11, f'{package.count}\n(原体参照)')
                ws.cell(row, 11).font = self.reference_font
                ws.cell(row, 11).alignment = Alignment(shrink_to_fit=True)
            else:
                ws.cell(row, 11, package.count)

        # 2人用
        package = unit_package.get_package(self.PACKAGE_FOR_2, menu_name)
        if package:
            if required_write_reference:
                ws.cell(row, 13, f'{package.count}\n(原体参照)')
                ws.cell(row, 13).font = self.reference_font
                ws.cell(row, 13).alignment = Alignment(shrink_to_fit=True)
            else:
                ws.cell(row, 13, package.count)

        # 1人用
        package = unit_package.get_package(self.PACKAGE_FOR_1, menu_name)
        if package:
            if required_write_reference:
                ws.cell(row, 12, f'{package.count}\n(原体参照)')
                ws.cell(row, 12).font = self.reference_font
                ws.cell(row, 12).alignment = Alignment(shrink_to_fit=True)
            else:
                ws.cell(row, 12, package.count)

        return required_write_reference

    def write_unit_border(self, ws, start_row, end_row, eating_day):
        # ユニット枠上部
        ws.cell(start_row, 2).border = self.unit_top_left_border
        ws.cell(start_row, 3).border = self.unit_top_t_right_border
        ws.cell(start_row, 4).border = self.unit_top_t_bottom_t_left_border
        ws.cell(start_row, 5).border = self.unit_top_t_bottom_border
        ws.cell(start_row, 6).border = self.unit_top_t_bottom_t_right_border
        ws.cell(start_row, 7).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 8).border = self.unit_top_t_bottom_bose_d_right_border
        ws.cell(start_row, 9).border = self.unit_top_t_bottom_bose_d_left_border
        ws.cell(start_row, 10).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 11).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 12).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 13).border = self.unit_top_t_bottom_bose_b_right_border

        # 料理各行
        for i in range(end_row - start_row - 1):
            ws.cell(start_row + i + 1, 2).border = self.unit_cell_b_left_border
            ws.cell(start_row + i + 1, 4).border = self.unit_cell_left_border
            ws.cell(start_row + i + 1, 5).border = self.unit_cell_border
            ws.cell(start_row + i + 1, 6).border = self.unit_cell_right_border
            ws.cell(start_row + i + 1, 7).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 8).border = self.unit_cell_bose_d_right_border
            ws.cell(start_row + i + 1, 9).border = self.unit_cell_bose_d_left_border
            ws.cell(start_row + i + 1, 10).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 11).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 12).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 13).border = self.unit_cell_bose_b_right_border

        # ユニット枠下部
        ws.cell(end_row, 2).border = self.unit_bottom_left_border
        ws.cell(end_row, 3).border = self.unit_bottom_t_right_border
        ws.cell(end_row, 4).border = self.unit_bottom_t_top_t_left_border
        ws.cell(end_row, 5).border = self.unit_bottom_t_top_border
        ws.cell(end_row, 6).border = self.unit_bottom_t_top_t_right_border
        ws.cell(end_row, 7).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 8).border = self.unit_bottom_t_top_bose_d_right_border
        ws.cell(end_row, 9).border = self.unit_bottom_t_top_bose_d_left_border
        ws.cell(end_row, 10).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 11).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 12).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 13).border = self.unit_bottom_t_top_bose_b_right_border

        # 原体送りのセル結合
        is_prev_row_raw = False
        for row_index in range(end_row - start_row + 1):
            current_row = start_row + row_index
            current_cell = ws.cell(current_row, 4)
            if isinstance(current_cell, excel.cell.cell.Cell):
                if current_cell.value and ('原体' in current_cell.value):
                    if PlateNameAnalizeUtil.is_raw_enge_plate_name(current_cell.value, eating_day)[0]:
                        is_prev_row_raw = False
                    else:
                        value_text = ws.cell(current_row, 9).value
                        if value_text == '原体送り資料参照':
                            ws.merge_cells(start_row=current_row, end_row=current_row, start_column=9, end_column=13)
                            is_prev_row_raw = True
                else:
                    is_prev_row_raw = False
            else:
                if is_prev_row_raw:
                    ws.merge_cells(start_row=current_row, end_row=current_row, start_column=9, end_column=13)

    def merge_same_plate(self, ws, start_row, end_row):
        """
        同一の料理名が続いた場合にマージする
        """
        prev_plate = None
        first_row = start_row
        for row_index in range(end_row - start_row):
            current_cell = ws.cell(start_row + row_index, 4)
            current_plate = current_cell.value
            if prev_plate != current_plate:
                # 料理名が切り替わった
                if prev_plate:
                    # 初回は除く。
                    # セルのマージ処理
                    ws.merge_cells(start_row=first_row, end_row=(start_row + row_index - 1), start_column=4, end_column=6)
                    diff = (start_row + row_index - 1) - first_row
                    if diff >= 2:
                        if '袋あり' in prev_plate:
                            for i in range(diff+1):
                                ws.row_dimensions[first_row+i].height = 18.75
                        else:
                            value_text = ws.cell(first_row, 9).value
                            if value_text == '原体送り資料参照':
                                ws.row_dimensions[first_row].height = 18.75
                    elif diff == 1:
                        if '袋あり' in prev_plate:
                            ws.row_dimensions[first_row].height = 18.75
                            ws.row_dimensions[first_row+1].height = 18.75
                    first_row = (start_row + row_index)
                prev_plate = current_plate

        # 残りのマージ
        if first_row < end_row:
            current_cell = ws.cell(end_row, 4)
            current_plate = current_cell.value
            if prev_plate == current_plate:
                # セルのマージ処理
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=4, end_column=6)

        # 1行のみの対応
        if first_row < end_row:
            next_menu_cell = ws.cell(end_row, 7)
            value = next_menu_cell.value
            if (not value) or (value == ""):
                # セルのマージ処理
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=4, end_column=6)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=7, end_column=7)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=8, end_column=8)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=9, end_column=9)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=10, end_column=10)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=11, end_column=11)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=12, end_column=12)
                ws.merge_cells(start_row=first_row, end_row=end_row, start_column=13, end_column=13)

    def open_workbook(self):
        """
        ピッキング指示書テンプレートファイルを開く
        """
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/picking_enge.xlsx')
        self.workbook = excel.load_workbook(template_path)

    def write(self, cooking_day, meal_list):
        logger.info('嚥下食のピッキング指示書を出力')

        # 全体まとめて出力->種類ごと(基本食はさらにチラー毎)に出力になったため、個別でworkbookを扱う
        self.open_workbook()
        cooking_eating_dict = EatingManagement.get_meals_dict_by_cooking_day(cooking_day)

        break_list = []
        row = 3
        page_start_row = row
        ws = self.workbook['嚥下食']
        max_modified = UnitPackage.objects.filter(cooking_day=cooking_day).order_by('-register_at').first()
        ws.cell(1, 1, f"出力日時:{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(1, 13, f"最終データ更新日時:{max_modified.register_at.strftime('%Y-%m-%d %H:%M:%S')}")
        for eating_day, cooking_meal_list in self.generate_cooking_eating_dict(cooking_eating_dict):

            # 対象製造日の全ての料理を取得する
            plates = CookingDirectionPlate.objects.filter(
                cooking_day=cooking_day,
                is_basic_plate=True, eating_day=eating_day, meal_name__in=cooking_meal_list).order_by('seq_meal', 'index')

            for meal in cooking_meal_list:
                # 対象喫食日の朝食、昼食、夕食が製造対象でなければ次にする
                meal_plats_qs = plates.filter(meal_name=meal)
                if not meal_plats_qs.exists():
                    continue

                # 喫食日、食事区分を出力する
                self.write_eating_day_with_meal(ws, row, eating_day, meal)
                row += 2

                tmp_rows = 0

                # ヘッダ出力
                self.write_table_header(ws, row)
                row += 1

                # ユニットの取得。過去の指示書出力時に、ユニットが無効化されていても出力できるように、is_active条件に含めない
                unit_qs = UnitMaster.objects.filter(unit_number__isnull=False).order_by('unit_number', 'id')
                is_first_unit = True
                add_row = 0
                for unit in unit_qs:
                    # 個食の出力はメインに含めるため、出力しない
                    if '個食' in unit.unit_name:
                        logger.debug(f'対象外(個食)：{unit.unit_number}-{eating_day}-{meal}')
                        continue

                    # 対象喫食日・食事区分にユニットの注文があるか確認
                    unit_order_qs = Order.objects.filter(
                        eating_day=eating_day, meal_name__meal_name=meal, quantity__gt=0,
                        unit_name=unit, menu_name__menu_name__in=self.ENGE_MENU_NAME_LIST
                    )
                    if not unit_order_qs.exists():
                        logger.debug(f'嚥下注文無し：{unit.unit_number}-{eating_day}-{meal}')
                        continue

                    # 施設名の出力
                    unit_start_row = row
                    self.write_unit_name(ws, row, unit)

                    # 袋数情報の取得
                    unit_pacage_qs = UnitPackage.objects.filter(
                        unit_number=unit.unit_number, cooking_day=cooking_day, eating_day=eating_day, meal_name=meal,
                        menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                    )
                    unit_package_list = list(unit_pacage_qs)

                    # QRコード画像の表示
                    qr_image = self.get_qr_image(unit, meal, '02', eating_day.day)
                    ws.add_image(qr_image, ws.cell(row + 1, 3).coordinate)

                    # 料理の出力
                    for plate in meal_plats_qs:
                        # 味噌汁の汁
                        if self._is_miso_soup(plate):
                            continue

                        # 汁なしの施設に汁の料理を表示しないように対応
                        if plate.is_soup and (not '具' in plate.plate_name):
                            if unit_order_qs.filter(meal_name__soup=True).exists():
                                pass
                            else:
                                continue

                        for enge_menu in self.ENGE_MENU_NAME_LIST:
                            if unit_order_qs.filter(menu_name__menu_name=enge_menu).exists():
                                plate_unit_package = self.get_unit_package(plate.index, unit_package_list)
                                is_add_row = self.write_plate(ws, row, plate, enge_menu, plate_unit_package, eating_day, unit)
                                if is_add_row:
                                    add_row += 1
                                # アレルギー代替料理の出力対象かどうか判定
                                ar_relation_ps = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')
                                if ar_relation_ps.exists():
                                    for key, group in groupby(ar_relation_ps, key=lambda x: x.plate):
                                        display_allergens = []

                                        for relation in group:
                                            ar_list, ar_menu = \
                                                CookingDirectionPlatesManager.get_allergens_with_menu(relation.code, cooking_day)
                                            if ar_menu == enge_menu:
                                                # 嚥下の内容は、嚥下の指示書で出力する
                                                unit_ar_qs = unit_order_qs.filter(
                                                    allergen__in=ar_list, menu_name__menu_name=ar_menu)
                                                for unit_ar_order in unit_ar_qs:
                                                    if not (unit_ar_order.allergen in display_allergens):
                                                        display_allergens.append((unit_ar_order.allergen, ar_menu))

                                        # アレルギー代替料理の出力
                                        for display_allergen, display_menu_name in display_allergens:
                                            row += 1
                                            ar_plate_package = self.get_allergen_unit_package(key, unit_package_list)
                                            is_add_row = self.write_plate(ws, row, key, display_menu_name, ar_plate_package, eating_day, unit, display_allergen)
                                            if is_add_row:
                                                add_row += 1

                                row += 1
                    unit_package_list.clear()
                    unit_end_row = row - 1
                    logger.debug(f'{unit}[Unit-End-Row]:{unit_end_row}({row})')

                    # QRコード画像表示のため、出力行数が少ない場合は、行の高さを広げる
                    row_diff = unit_end_row - unit_start_row + add_row
                    if row_diff == 0:
                        row_diff = 1
                        unit_end_row += 1
                        row += 1
                    if row_diff < 5:
                        added_height = 18.75 * (5 - row_diff) + 5
                        tmp_rows += 5 - row_diff
                        ws.row_dimensions[unit_start_row + 1].height = added_height

                    # 同一ユニットが別ページに分かれないようにする対応
                    page_row_diff = unit_end_row + tmp_rows - page_start_row + add_row
                    if page_row_diff >= self.PAGE_MAX_ROW:
                        logger.debug(f'{unit}[Break-Page]:{page_row_diff}({unit_end_row}+{tmp_rows}-{page_start_row})')
                        break_row = unit_start_row - 1
                        if is_first_unit:
                            break_row -= 3
                            is_first_unit = False
                        logger.debug(f'{unit}[Break-Row]:{break_row}')

                        break_list.append(Break(break_row))
                        page_start_row = row + 1
                        tmp_rows = 0
                        add_row = 0

                    # ユニット単位の枠描画
                    self.write_unit_border(ws, unit_start_row, unit_end_row, eating_day)

                    # 同一料理名のセル結合
                    self.merge_same_plate(ws, unit_start_row, unit_end_row)

                    is_first_unit = False
                # 食事区分単位の改ページ
                break_list.append(Break(row))
                row += 1
                page_start_row = row

        # 印刷範囲の修正
        self.adjust_print_area(ws, row, break_list[:-1])

        # ファイルの保存
        self.save(ws, cooking_day, '嚥下')


class SoupPickingDirectionWriter(PickingDirectionWriter):
    max_column_label = 'N'

    PAGE_MAX_ROW = 45

    def __init__(self, chillers):
        super(SoupPickingDirectionWriter, self).__init__(chillers)
        self.PACKAGE_FOR_10 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['SOUP_10'])
        self.PACKAGE_FOR_FRACTION = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['SOUP_FRACTION'])
        self.PACKAGE_FOR_UNIT = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['SOUP_UNIT'])
        self.PACKAGE_FOR_1 = PackageMaster.objects.get(id=settings.PICKING_PACKAGES['SOUP_1'])

    def _is_miso_soup(self, plate):
        """
        対象の料理が味噌汁の汁の料理かどうかを判定する。
        """
        if plate.is_soup:
            plate_name = plate.plate_name
            if PlateNameAnalizeUtil.is_miso_soup(plate_name):
                if PlateNameAnalizeUtil.is_soup_liquid(plate_name):
                    return True
        return False

    def write_table_header(self, ws, row):
        ws.cell(row, 2, '施設名')
        ws.cell(row, 4, '料理名')
        ws.cell(row, 7, '献立種別')
        ws.cell(row, 8, 'アレルギー')
        ws.cell(row, 9, '10人用')
        ws.cell(row, 10, '端数')
        ws.cell(row, 11, '施設毎')
        ws.cell(row, 12, '1人用')

        # フォント
        for i in range(11):
            ws.cell(row, 2 + i).font = self.bold_font

        # 枠の記入
        ws.cell(row, 2).border = self.header_top_bottom_left_border
        ws.cell(row, 3).border = self.header_top_bottom_t_right_border

        ws.cell(row, 4).border = self.header_top_bottom_t_left_border
        ws.cell(row, 5).border = self.header_top_bottom_border
        ws.cell(row, 6).border = self.header_top_bottom_t_right_border

        ws.cell(row, 7).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 8).border = self.header_top_bottom_bose_right_border
        ws.cell(row, 9).border = self.header_top_bottom_bose_left_border
        ws.cell(row, 10).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 11).border = self.header_top_bottom_t_bose_border
        ws.cell(row, 12).border = self.header_top_bottom_bose_right_b_border

    def write_plate(self, ws, row, plate, menu_name, unit_package, allergen=None):
        ws.cell(row, 4, plate.plate_name)
        ws.cell(row, 4).font = self.bold_font
        ws.cell(row, 7, menu_name)
        ws.cell(row, 7).font = self.bold_font

        if allergen:
            allergen_name = allergen.kana_name or 'あり'
            ws.cell(row, 8, allergen_name)
            ws.cell(row, 8).font = self.bold_font

            for i in range(12 - 4 + 1):
                ws.cell(row, 4 + i).fill = self.allergen_fill
        else:
            ws.cell(row, 8, '')

        # 原体送り対応
        if '原体' in plate.plate_name:
            ws.cell(row, 9, '原体送り資料参照')
            # セル結合は別メソッドで対応

        ws.cell(row, 9).font = self.bold_font
        ws.cell(row, 10).font = self.bold_font
        ws.cell(row, 11).font = self.bold_font
        ws.cell(row, 12).font = self.bold_font

        # 袋数の出力
        if unit_package.is_valid():
            # 10人用
            package = unit_package.get_package(self.PACKAGE_FOR_10)
            if package:
                ws.cell(row, 9, package.count)

            # 端数用
            package = unit_package.get_package(self.PACKAGE_FOR_FRACTION)
            if package:
                ws.cell(row, 10, package.count)

            # 施設毎用
            package = unit_package.get_package(self.PACKAGE_FOR_UNIT)
            if package:
                ws.cell(row, 11, package.count)

            # 1人用
            package = unit_package.get_package(self.PACKAGE_FOR_1)
            if package:
                ws.cell(row, 12, package.count)
        else:
            # 常食スープで1件のみの場合は袋を出力していないため、対象のunit_packageが見つからない
            logger.debug(f'1件のみの汁具のため、出力対象外：{plate.plate_name}')
            ws.cell(row, 9).fill = self.disable_fill
            ws.cell(row, 10).fill = self.disable_fill
            ws.cell(row, 11).fill = self.disable_fill
            ws.cell(row, 12).fill = self.disable_fill

    def write_unit_border(self, ws, start_row, end_row, is_first_unit=False):
        # ユニット枠上部
        ws.cell(start_row, 2).border = self.unit_top_left_border
        ws.cell(start_row, 3).border = self.unit_top_t_right_border
        ws.cell(start_row, 4).border = self.unit_top_t_bottom_t_left_border
        ws.cell(start_row, 5).border = self.unit_top_t_bottom_border
        ws.cell(start_row, 6).border = self.unit_top_t_bottom_t_right_border
        ws.cell(start_row, 7).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 8).border = self.unit_top_t_bottom_bose_d_right_border
        ws.cell(start_row, 9).border = self.unit_top_t_bottom_bose_d_left_border
        ws.cell(start_row, 10).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 11).border = self.unit_top_t_bottom_bose_border
        ws.cell(start_row, 12).border = self.unit_top_t_bottom_bose_b_right_border

        # 料理各行
        for i in range(end_row - start_row - 1):
            ws.cell(start_row + i + 1, 2).border = self.unit_cell_b_left_border
            ws.cell(start_row + i + 1, 4).border = self.unit_cell_left_border
            ws.cell(start_row + i + 1, 5).border = self.unit_cell_border
            ws.cell(start_row + i + 1, 6).border = self.unit_cell_right_border
            ws.cell(start_row + i + 1, 7).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 8).border = self.unit_cell_bose_d_right_border
            ws.cell(start_row + i + 1, 9).border = self.unit_cell_bose_d_left_border
            ws.cell(start_row + i + 1, 10).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 11).border = self.unit_cell_bose_border
            ws.cell(start_row + i + 1, 12).border = self.unit_cell_bose_b_right_border

        # ユニット枠下部
        ws.cell(end_row, 2).border = self.unit_bottom_left_border
        ws.cell(end_row, 3).border = self.unit_bottom_t_right_border
        ws.cell(end_row, 4).border = self.unit_bottom_t_top_t_left_border
        ws.cell(end_row, 5).border = self.unit_bottom_t_top_border
        ws.cell(end_row, 6).border = self.unit_bottom_t_top_t_right_border
        ws.cell(end_row, 7).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 8).border = self.unit_bottom_t_top_bose_d_right_border
        ws.cell(end_row, 9).border = self.unit_bottom_t_top_bose_d_left_border
        ws.cell(end_row, 10).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 11).border = self.unit_bottom_t_top_bose_border
        ws.cell(end_row, 12).border = self.unit_bottom_t_top_bose_b_right_border

        # 原体送りのセル結合
        for row_index in range(end_row - start_row + 1):
            current_row = start_row + row_index
            value = ws.cell(current_row, 4).value
            if value and ('原体' in value):
                ws.merge_cells(start_row=current_row, end_row=current_row, start_column=9, end_column=12)

    def merge_unit_plate(self, ws, row):
        ws.cell(row, 4).border = self.unit_bottom_t_top_t_left_border
        ws.cell(row, 5).border = self.unit_bottom_t_top_border
        ws.cell(row, 6).border = self.unit_bottom_t_top_t_right_border
        ws.cell(row, 7).border = self.unit_bottom_t_top_bose_border
        ws.cell(row, 8).border = self.unit_bottom_t_top_bose_d_right_border
        ws.cell(row, 9).border = self.unit_bottom_t_top_bose_d_left_border
        ws.cell(row, 10).border = self.unit_bottom_t_top_bose_border
        ws.cell(row, 11).border = self.unit_bottom_t_top_bose_border
        ws.cell(row, 12).border = self.unit_bottom_t_top_bose_b_right_border

        self._merge_plate_rows(ws, row, row + 1)

    def _merge_plate_rows(self, ws, start_row, end_row):
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=4, end_column=6)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=7, end_column=7)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=8, end_column=8)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=9, end_column=9)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=10, end_column=10)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=11, end_column=11)
        ws.merge_cells(start_row=start_row, end_row=end_row, start_column=12, end_column=12)

    def open_workbook(self):
        """
        ピッキング指示書テンプレートファイルを開く
        """
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/picking_soup.xlsx')
        self.workbook = excel.load_workbook(template_path)

    def write(self, cooking_day, meal_list):
        logger.info('汁・汁具のピッキング指示書を出力')

        # 全体まとめて出力->種類ごと(基本食はさらにチラー毎)に出力になったため、個別でworkbookを扱う
        self.open_workbook()
        cooking_eating_dict = EatingManagement.get_meals_dict_by_cooking_day(cooking_day)

        break_list = []
        row = 3
        page_start_row = row
        ws = self.workbook['汁・汁具']
        max_modified = UnitPackage.objects.filter(cooking_day=cooking_day).order_by('-register_at').first()
        ws.cell(1, 1, f"出力日時:{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(1, 12, f"最終データ更新日時:{max_modified.register_at.strftime('%Y-%m-%d %H:%M:%S')}")
        for eating_day, cooking_meal_list in self.generate_cooking_eating_dict(cooking_eating_dict):

            # 対象製造日の全ての料理を取得する
            plates = CookingDirectionPlate.objects.filter(
                cooking_day=cooking_day, is_soup=True,
                is_basic_plate=True, eating_day=eating_day, meal_name__in=cooking_meal_list).order_by('seq_meal', 'index')

            for meal in cooking_meal_list:
                # 対象喫食日の朝食、昼食、夕食が製造対象でなければ次にする
                meal_plats_qs = plates.filter(meal_name=meal)
                if not meal_plats_qs.exists():
                    continue

                # 喫食日、食事区分を出力する
                self.write_eating_day_with_meal(ws, row, eating_day, meal)
                row += 2

                tmp_rows = 0

                # ヘッダ出力
                self.write_table_header(ws, row)
                row += 1

                # ユニットの取得。過去の指示書出力時に、ユニットが無効化されていても出力できるように、is_active条件に含めない
                unit_qs = UnitMaster.objects.filter(unit_number__isnull=False).order_by('unit_number', 'id')
                is_first_unit = True
                for unit in unit_qs:
                    # 対象喫食日・食事区分にユニットの注文があるか確認(汁なしは対象外-具があるかどうかで判断)
                    unit_order_qs = Order.objects.filter(
                        eating_day=eating_day, meal_name__meal_name=meal, quantity__gt=0,
                        unit_name=unit, meal_name__filling=True
                    )
                    if not unit_order_qs.exists():
                        logger.debug(f'汁・汁具注文無し：{unit.unit_number}-{eating_day}-{meal}')
                        continue

                    # 個食の出力はメインに含めるため、出力しない
                    if '個食' in unit.unit_name:
                        logger.debug(f'対象外(個食)：{unit.unit_number}-{eating_day}-{meal}')
                        continue

                    # 施設名の出力
                    unit_start_row = row
                    self.write_unit_name(ws, row, unit)

                    # 袋数情報の取得
                    unit_pacage_qs = UnitPackage.objects.filter(
                        unit_number=unit.unit_number, cooking_day=cooking_day, eating_day=eating_day, meal_name=meal,
                        menu_name='常食'
                    )
                    unit_package_list = list(unit_pacage_qs)

                    # QRコード画像の表示
                    qr_image = self.get_qr_image(unit, meal, '03', eating_day.day)
                    ws.add_image(qr_image, ws.cell(row + 1, 3).coordinate)

                    # 料理の出力
                    for plate in meal_plats_qs:
                        if PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
                            if not unit_order_qs.filter(meal_name__soup=True).exists():
                                # スープの注文がないユニットには、スープ・汁の料理は出さない
                                continue
                        plate_unit_package = self.get_unit_package(plate.index, unit_package_list)
                        self.write_plate(ws, row, plate, '常食', plate_unit_package)
                        row += 1

                        # アレルギー代替料理の出力対象かどうか判定
                        ar_relation_ps = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')
                        if ar_relation_ps.exists():
                            for key, group in groupby(ar_relation_ps, key=lambda x: x.plate):
                                display_allergens = []

                                for relation in group:
                                    ar_list, ar_menu = \
                                        CookingDirectionPlatesManager.get_allergens_with_menu(relation.code, cooking_day)
                                    if ar_menu == '常食':
                                        # 嚥下の内容は、嚥下の指示書で出力する
                                        for unit_ar_order in unit_order_qs.filter(allergen__in=ar_list):
                                            if not (unit_ar_order.allergen in display_allergens):
                                                display_allergens.append(unit_ar_order.allergen)

                                # アレルギー代替料理の出力
                                for display_allergen in display_allergens:
                                    ar_plate_package = self.get_allergen_unit_package(key, unit_package_list)
                                    self.write_plate(ws, row, key, '常食', ar_plate_package, display_allergen)
                                    row += 1

                    unit_package_list.clear()
                    unit_end_row = row - 1

                    # QRコード画像表示のため、出力行数が少ない場合は、行の高さを広げる
                    row_diff = unit_end_row - unit_start_row
                    if row_diff < 5:
                        if row_diff == 0:
                            row += 1
                            unit_end_row += 1
                            self._merge_plate_rows(ws, unit_start_row, unit_end_row)
                        target_row = unit_start_row + 1
                        added_height = 18.75 * (5 - row_diff) + 5
                        ws.row_dimensions[target_row].height = added_height

                        tmp_rows += 5 - row_diff

                    # 同一ユニットが別ページに分かれないようにする対応
                    page_row_diff = unit_end_row + tmp_rows - page_start_row
                    if page_row_diff >= self.PAGE_MAX_ROW:
                        logger.debug(f'{unit}[Break-Page]:{page_row_diff}({unit_end_row}+{tmp_rows}-{page_start_row})')
                        break_row = unit_start_row - 1
                        if is_first_unit:
                            break_row -= 2
                            is_first_unit = False
                        logger.debug(f'{unit}[Break-Row]:{break_row}')
                        break_list.append(Break(break_row))
                        page_start_row = row + 1
                        tmp_rows = 0

                    # ユニット単位の枠描画
                    if row_diff == 0:
                        self.write_unit_border(ws, unit_start_row, unit_start_row + 1)
                        self.merge_unit_plate(ws, unit_start_row)
                    else:
                        self.write_unit_border(ws, unit_start_row, unit_end_row)

                    # 同一料理名のセル結合
                    # self.merge_same_plate(ws, unit_start_row, unit_end_row)

                    is_first_unit = False
                # 食事区分単位の改ページ
                break_list.append(Break(row))
                row += 1
                page_start_row = row

        # 改ページの設定
        self.adjust_print_area(ws, row, break_list[:-1])

        # ファイルの保存
        self.save(ws, cooking_day, '汁・汁具')


class RawPlatePickingDirectionWriter(PickingDirectionWriter):
    max_column_label = 'M'
    bold_font_count = Font(name='ＭＳ Ｐゴシック', size=20, bold=True)
    count_alignment = Alignment(vertical='center')

    PAGE_MAX_ROW = 23

    def __init__(self, chillers):
        super(RawPlatePickingDirectionWriter, self).__init__(chillers)

    def convert_plate_name(self, source: str) -> str:
        """
        計量表出力時の同様の料理名へ変換する。
        """

        # 先頭の番号以降を対象
        converted = source[1:]

        """
        # ※下記内容は、cooking_direction.pyの料理名変換処理に合わせること
        # 全角のｇを半角のgに揃える
        converted = converted.replace('ｇ', 'g')

        # 半角の%を全角の％に揃える
        converted = converted.replace('%', '％')

        # 半角の+を全角の＋に揃える
        converted = converted.replace('+', '＋')

        # 半角の()を半角スペースにする
        converted = converted.replace('(', ' ')
        converted = converted.replace(')', ' ')

        # 全角の（）を半角スペースにする
        converted = converted.replace('（', ' ')
        converted = converted.replace('）', ' ')

        converted = converted.replace('/', '÷')
        """

        word_index = converted.find('原体')
        search_name = converted[:word_index]
        res = re.findall('(\D+)(\d|\s)+', search_name)

        base_name = res[0][0]
        if ('(' in base_name) or ('（' in base_name):
            if (')' in base_name) or ('）' in base_name):
                pass
            else:
                for str in res[0][1:]:
                    base_name = base_name + str
                for str in res[1]:
                    base_name = base_name + str
                    if (')' in str) or ('）' in str):
                        break
        if res:
            return base_name.strip()
        else:
            raise ValueError('名称不正')

    def _in_plate(self, plate_name):
        for raw_plate in RawPlatePackageMaster.objects.all():
            if raw_plate.base_name in plate_name:
                return raw_plate

        # 見つからなかった場合
        return None

    def get_raw_plate_list(self, cooking_day, eating_day, meal_name):
        # 対象製造日の全ての料理を取得する
        plates = CookingDirectionPlate.objects.filter(
            cooking_day=cooking_day,
            is_basic_plate=True, eating_day=eating_day,
            meal_name=meal_name).order_by('seq_meal', 'index')
        plate_list = []
        is_none = True
        for plate in plates:
            is_none = False
            if '原体' in plate.plate_name:
                raw_plate = self._in_plate(self.convert_plate_name(plate.plate_name))
                plate_list.append((plate, raw_plate))

        return plate_list, is_none

    def write_table_header(self, ws, row, raw_plate_list):
        ws.cell(row, 2, '施設名')
        ws.cell(row, 2).font = self.bold_font

        # 表ヘッダ内容を出力
        col = 4
        for plate, raw_plate in raw_plate_list:
            if raw_plate is None:
                logger.warning(f'原体情報がないのでスキップ：{plate}')
                continue

            ws.cell(row, col, f'{raw_plate.dry_name}(乾燥)')
            ws.cell(row, col+1, f'{raw_plate.cold_name}(冷凍)')

            # フォント
            ws.cell(row, col).font = self.bold_font
            ws.cell(row, col+1).font = self.bold_font

            # セル色
            ws.cell(row, col).fill = self.dry_fill
            ws.cell(row, col+1).fill = self.cold_fill

            col += 2

        # 最後の増加分を補正
        col -= 2

        if col >= 4:
            # 枠の記入
            ws.cell(row, 2).border = self.header_top_bottom_left_border
            ws.cell(row, 3).border = self.header_top_bottom_t_right_border

            for i in range(len(raw_plate_list)):
                draw_col = 4 + i * 2
                ws.cell(row, draw_col).border = self.header_top_bottom_bose_left_b_border
                ws.cell(row, draw_col + 1).border = self.header_top_bottom_bose_right_b_border

    def adjust_print_area(self, ws, row: int, break_list, max_col):
        if max_col:
            ws.print_area = f'A1:{ws.cell(row, max_col).coordinate}'
            logger.debug(f'[max_col]:{max_col}')
            for br in break_list:
                ws.row_breaks.append(br)

    def write_unit_border(self, ws, start_row, raw_plate_list):
        col = len(raw_plate_list)
        if col == 0:
            return

        # 上段
        ws.cell(start_row, 2).border = self.unit_top_left_border
        ws.cell(start_row, 3).border = self.unit_top_t_right_border
        for col_index in range(col):
            ws.cell(start_row, 4 + col_index * 2).border = self.unit_top_bose_b_left_border
            ws.cell(start_row, 4 + col_index * 2 + 1).border = self.unit_top_bose_b_right_border

        # 下段
        end_row = start_row + 1
        ws.cell(end_row, 2).border = self.unit_bottom_left_border
        ws.cell(end_row, 3).border = self.unit_bottom_t_right_border
        for col_index in range(col):
            ws.cell(end_row, 4 + col_index * 2).border = self.unit_bottom_bose_b_left_border
            ws.cell(end_row, 4 + col_index * 2 + 1).border = self.unit_bottom_bose_b_right_border

        return (col * 2 + 1) if col else 0

    def open_workbook(self):
        """
        ピッキング指示書テンプレートファイルを開く
        """
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/picking_raw.xlsx')
        self.workbook = excel.load_workbook(template_path)

    def write(self, cooking_day, meal_list):
        logger.info('原体のピッキング指示書を出力')

        # 全体まとめて出力->種類ごと(基本食はさらにチラー毎)に出力になったため、個別でworkbookを扱う
        self.open_workbook()
        cooking_eating_dict = EatingManagement.get_meals_dict_by_cooking_day(cooking_day)

        break_list = []
        row = 3
        page_start_row = row
        ws = self.workbook['原体']
        max_modified = UnitPackage.objects.filter(cooking_day=cooking_day).order_by('-register_at').first()
        ws.cell(1, 1, f"出力日時:{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(1, 5, f"最終データ更新日時:{max_modified.register_at.strftime('%Y-%m-%d %H:%M:%S')}")
        max_col = 0
        for eating_day, cooking_meal_list in self.generate_cooking_eating_dict(cooking_eating_dict):

            for meal in cooking_meal_list:
                # 原体にチラーは関係ないため、参照しない

                # ヘッダに出力する料理情報
                plate_list, is_none = self.get_raw_plate_list(cooking_day, eating_day, meal)
                if is_none:
                    # 対象喫食日の朝食、昼食、夕食が製造対象でなければ次にする
                    continue

                has_raw_enge_only = False
                for plate in plate_list:
                    if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate[0].plate_name, eating_day)[0]:
                        if has_raw_enge_only:
                            has_raw_enge_only = False
                            break
                        else:
                            has_raw_enge_only = True
                    else:
                        has_raw_enge_only = False
                        break

                # 喫食日、食事区分を出力する
                self.write_eating_day_with_meal(ws, row, eating_day, meal)
                row += 2

                # ヘッダ出力
                if plate_list:
                    if [1 for cooking_direction, raw_plate in plate_list if raw_plate]:
                        # 原体情報が取得出来た場合
                        self.write_table_header(ws, row, plate_list)
                        row += 1
                    else:
                        # 原体情報が取得出来なかった場合
                        ws.cell(row, 2, '対象マスタ未登録')
                        logger.warning(f'原体マスタ未登録:{cooking_day}-{eating_day}-{meal}')
                        row += 2
                        continue
                else:
                    ws.cell(row, 2, '出力対象なし')
                    row += 2
                    continue


                unit_qs = UnitMaster.objects.filter(is_active=True).exclude(
                    unit_code__range=[80001, 80008]).order_by('unit_number')
                write_list = []
                for unit in unit_qs:
                    unit_start_row = row

                    # 対象喫食日・食事区分にユニットの注文があるか確認
                    unit_order_qs = Order.objects.filter(
                        eating_day=eating_day, meal_name__meal_name=meal, quantity__gt=0,
                        unit_name=unit,
                    ).select_related('menu_name')
                    if not unit_order_qs.exists():
                        logger.debug(f'原体注文無し：{unit.unit_number}.{unit.unit_name}-{eating_day}-{meal}')
                        continue

                    # 個食の出力はメインに含めるため、出力しない
                    if '個食' in unit.unit_name:
                        logger.debug(f'対象外(個食)：{unit.unit_number}-{eating_day}-{meal}')
                        continue

                    # 嚥下製造対象のみ、かつ献立種類が嚥下のみなら出力しない
                    if has_raw_enge_only:
                        basic_menu_list = [x for x in unit_order_qs if x.menu_name.menu_name == '常食']
                        if not basic_menu_list:
                            logger.info(f'嚥下製造対象-基本食なし：{unit.unit_number}-{eating_day}-{meal}')
                            basic_menu_list.clear()
                            del basic_menu_list
                            continue

                    if (unit.unit_number, unit.calc_name) in write_list:
                        is_write = True
                        row -= 1
                    else:
                        # 施設名の出力
                        self.write_unit_name(ws, row, unit)
                        write_list.append((unit.unit_number, unit.calc_name))
                        row += 1

                        # QRコード画像の表示
                        qr_image = self.get_qr_image(unit, meal, '04', eating_day.day)
                        ws.add_image(qr_image, ws.cell(row, 3).coordinate)
                        ws.row_dimensions[row].height = 18.75 * 4 + 5
                        del qr_image
                        is_write = False

                    # 袋数の出力
                    col = 4
                    for index, tpl_plates in enumerate(plate_list):
                        plate, raw_plate = tpl_plates
                        if raw_plate:
                            col_index = col + index * 2
                            ws.cell(row, col_index).font = self.bold_font_count
                            ws.cell(row, col_index).alignment = self.count_alignment
                            ws.cell(row, col_index + 1).font = self.bold_font_count
                            ws.cell(row, col_index + 1).alignment = self.count_alignment

                            # 直送の出力
                            dry_cold_type = unit.username.dry_cold_type
                            if dry_cold_type == '乾燥':
                                if raw_plate.is_direct_dry:
                                    logger.debug(f'乾燥直送：{unit.unit_number}-{eating_day}-{meal}')
                                    ws.cell(row, col_index, '直送')
                                else:
                                    raw_package_qs = PickingRawPlatePackage.objects.filter(
                                        unit_name__unit_number=unit.unit_number, unit_name__calc_name=unit.calc_name,
                                        eating_day=eating_day, meal_name=meal, package_master=raw_plate
                                    )
                                    if raw_package_qs.exists():
                                        fs = raw_package_qs.first()
                                        if fs.quantity and fs.package:
                                            if fs.dry_cold_type == '乾燥':
                                                ws.cell(row, col_index, math.ceil(fs.quantity / fs.package))
                                            else:
                                                ws.cell(row, col_index + 1, math.ceil(fs.quantity / fs.package))
                                        else:
                                            if not is_write:
                                                if fs.dry_cold_type == '乾燥':
                                                    ws.cell(row, col_index, 0)
                                                else:
                                                    ws.cell(row, col_index+1, 0)
                                    else:
                                        logger.info(f'乾燥袋情報登録なし：{unit.unit_number}.{unit.unit_name}-{eating_day}-{meal}')
                                        if not is_write:
                                            ws.cell(row, col_index, 0)
                            elif '冷凍' in dry_cold_type:
                                if raw_plate.is_direct_cold and (unit.username.dry_cold_type == '冷凍'):
                                    logger.info(f'冷凍直送：{unit.unit_number}-{eating_day}-{meal}')
                                    ws.cell(row, col_index + 1, '直送')
                                else:
                                    raw_package_qs = PickingRawPlatePackage.objects.filter(
                                        unit_name__unit_number=unit.unit_number, unit_name__calc_name=unit.calc_name,
                                        eating_day=eating_day, meal_name=meal, package_master=raw_plate
                                    )
                                    if raw_package_qs.exists():
                                        fs = raw_package_qs.first()
                                        if fs.quantity and fs.package:
                                            if fs.dry_cold_type == '乾燥':
                                                ws.cell(row, col_index, math.ceil(fs.quantity / fs.package))
                                            else:
                                                ws.cell(row, col_index + 1, math.ceil(fs.quantity / fs.package))
                                        else:
                                            if not is_write:
                                                if fs.dry_cold_type == '乾燥':
                                                    ws.cell(row, col_index, 0)
                                                else:
                                                    ws.cell(row, col_index+1, 0)
                                    else:
                                        if not is_write:
                                            logger.info(f'冷凍袋情報登録なし：{unit.unit_number}.{unit.unit_name}-{eating_day}-{meal}')
                                            ws.cell(row, col_index + 1, 0)

                    col_no = col + index * 2 + 1
                    self.write_unit_border(ws, row - 1, plate_list)
                    if max_col < col_no:
                        max_col = col_no

                    # 同一ユニットが別ページに分かれないようにする対応
                    page_row_diff = row - page_start_row
                    if page_row_diff > self.PAGE_MAX_ROW:
                        logger.debug(f'{unit}[Break-Page]:{page_row_diff}({row}-{page_start_row})')
                        break_list.append(Break(unit_start_row - 1))
                        logger.debug(f'{unit}[Break-Row]:{unit_start_row - 1}')
                        page_start_row = row + 1

                    # 対象施設の出力完了
                    row += 1

                # 喫食日・食事区分の区切り
                row += 1

        # 改ページの設定
        self.adjust_print_area(ws, row, break_list, max_col)

        # ファイルの保存
        self.save(ws, cooking_day, '原体')


class PickingDirectionOutputManagement:
    """
    ピッキング指示書出力管理クラス
    """
    def __init__(self, cooking_day, chillers):
        self.cooking_day = cooking_day
        self.meal_list = MealUtil.get_name_list_without_snak()
        self.chillers = chillers

    def _get_writer(self, picking_type: str):
        if picking_type == '011':
            return BasicPickingDirectionWriter(self.chillers, 1)
        elif picking_type == '012':
            return BasicPickingDirectionWriter(self.chillers, 2)
        elif picking_type == '013':
            return BasicPickingDirectionWriter(self.chillers, 3)
        elif picking_type == '014':
            return BasicPickingDirectionWriter(self.chillers, 4)
        elif picking_type == '02':
            return EngePickingDirectionWriter(self.chillers)
        elif picking_type == '03':
            return SoupPickingDirectionWriter(self.chillers)
        elif picking_type == '04':
            return RawPlatePickingDirectionWriter(self.chillers)

    def write_directions(self, picking_type: str):
        """
        対象製造日のピッキング指示書を出力
        """
        writer = self._get_writer(picking_type)
        if writer:
            writer.write(self.cooking_day, self.meal_list)
            logger.info(f'帳票出力完了(ピッキング指示書)-{self.cooking_day}製造-({picking_type})')

            return True

        return False


class EatingManagement:
    """
    喫食日情報を管理するクラス
    """
    @classmethod
    def get_dict_by_cooking_day(cls, cooking_day):
        """
        製造日に作成する喫食日・食事区分のdict(key:食事区分、value:喫食日リスト)を返す。
        """
        distinct_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day).distinct().values_list('eating_day', 'meal_name')
        eating_day_dict = {'朝食': [], '昼食': [], '夕食': []}
        for x in distinct_qs:
            eating_day = x[0]
            meal_name = x[1]
            eating_day_dict[meal_name].append(eating_day)

        return eating_day_dict

    @classmethod
    def get_meals_dict_by_cooking_day(cls, cooking_day):
        """
        製造日に作成する喫食日・食事区分のdict(key:喫食日、value:食事区分リスト)を返す。
        """
        distinct_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day).distinct().values_list('eating_day', 'meal_name')
        meal_dict = {}
        for x in distinct_qs:
            eating_day = x[0]
            meal_name = x[1]
            if eating_day in meal_dict:
                meal_dict[eating_day].append(meal_name)
            else:
                meal_dict[eating_day] = [meal_name,]

        return meal_dict

    @classmethod
    def get_meal_dict_by_cooking_day(cls, cooking_day):
        """
        製造日に作成する喫食日・食事区分のdict(key:喫食日、value:食事区分リスト)を返す。
        """
        distinct_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day).distinct().values_list('eating_day', 'meal_name')
        meal_dict = {}
        for x in distinct_qs:
            eating_day = x[0]
            meal_name = x[1]
            if eating_day in meal_dict:
                meal_dict[eating_day].append(meal_name)
            else:
                meal_dict[eating_day] = [meal_name,]

        return sorted([key for key in meal_dict.keys()])


class InnerPackageManagement:
    """
    中袋の袋数を管理するクラス
    """
    # 中袋1枚に収納可能な注文数
    ORDER_COUNT_PER_PACKAGE = 20

    # 中袋1枚に収納可能な注文数(汁・汁具)
    ORDER_COUNT_PER_SOUP_PACKAGE = 90

    def __init__(self, cooking_eating_dict):
        self.cooking_eating_dict = cooking_eating_dict

    def _is_miso_soup(self, name: str):
        if 'みそ汁' in name:
            return True
        elif '味噌汁' in name:
            return True
        elif 'みそしる' in name:
            return True
        else:
            return False

    def _is_raw_plate(self, plate):
        """
        原体送りの料理かどうかを判断する。
        """
        return '原体' in plate.plate_name

    def has_mix_rice(self, cooking_day, eating_day, meal):
        plates = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, eating_day=eating_day, meal_name=meal, is_mix_rice=True)
        return plates.exists()

    def has_soup(self, cooking_day, eating_day, meal):
        plates = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, eating_day=eating_day, meal_name=meal,
                                                      is_soup=True)

        return plates.exists()

    def get_mixrice_agg_counts(self):
        date_count_dict = {}

        # ユニットの元施設を取得
        unit_numbers = settings.MIX_RICE_AGGREGATE_UNITS[0]

        for meal_name, eating_day_list in self.cooking_eating_dict.items():
            if meal_name == '朝食':
                meal_index = 0
            elif meal_name == '昼食':
                meal_index = 1
            elif meal_name == '夕食':
                meal_index = 2
            else:
                raise ValueError('内部パラメータ異常')

            if not eating_day_list:
                continue

            for eating_day in eating_day_list:
                if eating_day in date_count_dict:
                    count_dict = date_count_dict[eating_day]
                else:
                    count_dict = {
                        '01': [0, 0, 0],  # 基本食
                        '02': [0, 0, 0],  # 嚥下
                        '03': [0, 0, 0],  # 汁・汁具
                        '04': [0, 0, 0]  # 原体
                    }
                    date_count_dict[eating_day] = count_dict

                # 基本食の中袋数の計算
                basic_count_qs = Order.objects.filter(
                    unit_name__unit_number__in=unit_numbers, quantity__gt=0,
                    eating_day=eating_day, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                ).values('quantity')
                basic_count = sum([x['quantity'] for x in basic_count_qs])
                count_dict['01'][meal_index] += math.ceil(basic_count / self.ORDER_COUNT_PER_PACKAGE)

                # 嚥下の中袋数の計算
                enge_count_qs = Order.objects.filter(
                    unit_name__unit_number__in=unit_numbers, quantity__gt=0,
                    eating_day=eating_day, meal_name__meal_name=meal_name,
                    menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                ).values('quantity')
                enge_count = sum([x['quantity'] for x in enge_count_qs])
                count_dict['02'][meal_index] += math.ceil(enge_count / self.ORDER_COUNT_PER_PACKAGE)

        return count_dict

    def convert_plate_name(self, source: str) -> str:
        """
        計量表出力時の同様の料理名へ変換する。
        """

        # 先頭の番号以降を対象
        converted = source[1:]

        # ※下記内容は、cooking_direction.pyの料理名変換処理に合わせること
        # 全角のｇを半角のgに揃える
        converted = converted.replace('ｇ', 'g')

        # 半角の%を全角の％に揃える
        converted = converted.replace('%', '％')

        # 半角の+を全角の＋に揃える
        converted = converted.replace('+', '＋')

        # 半角の()を半角スペースにする
        converted = converted.replace('(', ' ')
        converted = converted.replace(')', ' ')

        # 全角の（）を半角スペースにする
        converted = converted.replace('（', ' ')
        converted = converted.replace('）', ' ')

        converted = converted.replace('/', '÷')

        res = re.findall('(\D+)(\d|\s)+', converted)
        if res:
            return res[0][0].strip()
        else:
            raise ValueError('名称不正')

    def _convert_package_counts(self, dict):
        for inner_dict in dict.values():
            for key, value in inner_dict.items():
                if key == '01':
                    value[0] = (value[0], math.ceil(value[0] / self.ORDER_COUNT_PER_PACKAGE))
                    #value[1] = (value[1], math.ceil(value[1] / self.ORDER_COUNT_PER_PACKAGE))
                    #value[2] = (value[2], math.ceil(value[2] / self.ORDER_COUNT_PER_PACKAGE))
                elif key == '02':
                    value[0] = (value[0], math.ceil(value[0] / self.ORDER_COUNT_PER_PACKAGE))
                    #value[1] = (value[1], math.ceil(value[1] / self.ORDER_COUNT_PER_PACKAGE))
                    #value[2] = (value[2], math.ceil(value[2] / self.ORDER_COUNT_PER_PACKAGE))
                elif key == '03':
                    value[0] = (value[0], math.ceil(value[0] / self.ORDER_COUNT_PER_SOUP_PACKAGE))
                    #value[1] = (value[1], math.ceil(value[1] / self.ORDER_COUNT_PER_SOUP_PACKAGE))
                    #value[2] = (value[2], math.ceil(value[2] / self.ORDER_COUNT_PER_SOUP_PACKAGE))
                else:
                    value[0] = (value[0], 1 if value[0] else 0)
                    #value[1] = (value[1], 1 if value[1] else 0)
                    #value[2] = (value[2], 1 if value[2] else 0)

    def generate_eating_day(self, input_meal):
        for eating_day, meal_name_list in self.cooking_eating_dict.items():
            for meal_name in meal_name_list:
                if meal_name == '朝食':
                    if input_meal == "01":
                        yield eating_day, meal_name
                elif meal_name == '昼食':
                    if input_meal == "02":
                        yield eating_day, meal_name
                elif meal_name == '夕食':
                    if input_meal == "03":
                        yield eating_day, meal_name

    def has_plate_miso_soup(self, eating_day, meal_name):
        has_miso_soup = False
        plate_qs = CookingDirectionPlate.objects.filter(eating_day=eating_day, meal_name=meal_name,
                                                        is_soup=True).values('plate_name')
        for plate in plate_qs:
            if self._is_miso_soup(plate['plate_name']):
                has_miso_soup = True
                break

        return has_miso_soup

    def get_raw_plates(self, cooking_day, eating_day, meal_name):
        converted = []
        plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, eating_day=eating_day, meal_name=meal_name, is_soup=False)
        for plate in plate_qs:
            if self._is_raw_plate(plate):
                # 嚥下製造対象であっても、常食は普通に原体処理が必要なので、ここでは絞り込まない
                converted.append(self.convert_plate_name(plate.plate_name))

        result_dict = {}
        for raw_plate in RawPlatePackageMaster.objects.all():
            for plate_name in converted:
                if raw_plate.base_name in plate_name:
                    if raw_plate.id in result_dict:
                        pass
                    else:
                        result_dict[raw_plate.id] = raw_plate
                    break

        return result_dict.values()

    def _in_plate(self, plate_name):
        for raw_plate in RawPlatePackageMaster.objects.all():
            if raw_plate.base_name in plate_name:
                return raw_plate

        # 見つからなかった場合
        return None

    def get_package_counts(self, unit_number: int, short_name: str, cooking_day, input_meal):
        date_count_dict = {}

        # ユニットの元施設を取得
        first_unit = UnitMaster.objects.filter(unit_number=unit_number, short_name=short_name).first()
        first_user = first_unit.username

        for meal_name, eating_day_list in self.cooking_eating_dict.items():
            if meal_name == '朝食':
                meal_index = 0
                if input_meal != "01":
                    continue
            elif meal_name == '昼食':
                meal_index = 0
                if input_meal != "02":
                    continue
            elif meal_name == '夕食':
                meal_index = 0
                if input_meal != "03":
                    continue
            else:
                raise ValueError('内部パラメータ異常')

            if not eating_day_list:
                continue

            for eating_day in eating_day_list:
                if eating_day in date_count_dict:
                    count_dict = date_count_dict[eating_day]
                else:
                    count_dict = {
                        '01': [0],  # 基本食
                        '02': [0],  # 嚥下
                        '03': [0],  # 汁・汁具
                        '04': [0]  # 原体
                    }
                    date_count_dict[eating_day] = count_dict

                # 基本食の中袋数の計算
                basic_count_qs = Order.objects.filter(
                    unit_name__unit_number=unit_number, quantity__gt=0,
                    eating_day=eating_day, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                ).values('quantity')
                basic_count = sum([x['quantity'] for x in basic_count_qs])
                count_dict['01'][meal_index] += basic_count
                del basic_count_qs

                # 嚥下の中袋数の計算
                enge_count_qs = Order.objects.filter(
                    unit_name__unit_number=unit_number, quantity__gt=0,
                    eating_day=eating_day, meal_name__meal_name=meal_name,
                    menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                ).values('quantity')
                enge_count = sum([x['quantity'] for x in enge_count_qs])
                count_dict['02'][meal_index] += enge_count
                del enge_count_qs

                # 汁・味噌汁の中袋数の計算
                # 契約内容変更により、顧客別～の設定値と実際の注文の食事区分が異なる場合があるため、Orderから取得する
                first_meal = Order.objects.filter(
                    unit_name__username=first_user, meal_name__meal_name=meal_name, meal_name__soup=True
                ).first()
                if first_meal:
                    # 汁なし、具のみの場合は、出力対象外。汁具ありの場合のみ以下の判断を行う

                    # 料理に味噌汁があるかどうか
                    has_miso_soup = False
                    plate_qs = CookingDirectionPlate.objects.filter(eating_day=eating_day, meal_name=meal_name, is_soup=True).values('plate_name')
                    for plate in plate_qs:
                        if self._is_miso_soup(plate['plate_name']):
                            has_miso_soup = True
                            break
                    del plate_qs

                    if has_miso_soup:
                        if first_meal.meal_name.soup:
                            # 嚥下の味噌汁の汁は、汁・汁具の中袋に入れるため、基本食と嚥下の合計が必要
                            soup_count = basic_count + enge_count
                        else:
                            # 嚥下の汁(味噌汁、それ以外の汁・スープ)の汁具は、嚥下の中袋に入れるため、汁・汁具の中袋は基本食のみ
                            soup_count = basic_count
                    else:
                        # 嚥下の味噌汁以外の汁・汁具は嚥下の中袋に入れる
                        soup_count = basic_count
                    count_dict['03'][meal_index] += soup_count
                else:
                    filling_qs = Order.objects.filter(
                        unit_name__username=first_user, meal_name__meal_name=meal_name, quantity__gt=0,
                        meal_name__soup=False, meal_name__filling=True
                    ).values('quantity')
                    # 具のみの場合は、中袋の計算外(汁の注文数で計算を行う)だが0は実体に合わないので数を設定
                    if filling_qs.exists():
                        filling = filling_qs.first()
                        count_dict['03'][meal_index] += filling['quantity']

                # 原体の中袋数の計算
                has_raw_plate = False
                plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, meal_name=meal_name, is_soup=False)
                for plate in plate_qs:
                    if self._is_raw_plate(plate):
                        # 原体マスタの参照
                        s_name = self.convert_plate_name(plate.plate_name)
                        raw_plate = self._in_plate(s_name)
                        if raw_plate:
                            if first_unit.username.dry_cold_type == "乾燥":
                                is_direct = raw_plate.is_direct_dry
                            elif first_unit.username.dry_cold_type == "冷凍":
                                is_direct = raw_plate.is_direct_cold
                            elif first_unit.username.dry_cold_type == "冷蔵":
                                is_direct = raw_plate.is_direct_cold
                            else:
                                logger.warn(f'施設の冷凍乾燥区分異常：{first_unit.dry_cold_type}')
                                is_direct = False

                            # 原体送りの施設の区分が直送でないの場合、対象献立(基本 or 嚥下)のフラグを立てる
                            if not is_direct:
                                has_raw_plate = True
                                break
                        else:
                            logger.warn(f'原体マスタ未登録：{s_name}')
                raw_count = 0
                raw_count += basic_count if has_raw_plate else 0
                raw_count += enge_count if has_raw_plate else 0

                # 原体は常に1袋のみ
                if raw_count:
                    count_dict['04'][meal_index] += raw_count

        self._convert_package_counts(date_count_dict)
        return date_count_dict

    @classmethod
    def get_cooking_day_from_picking_day(cls, date):
        cooking_date = date
        """
        cooking_date = picking_date - relativedelta(days=settings.ADJUST_PICKING_DAY)

        # 日曜日の補正
        if cooking_date.weekday() == 6:
            cooking_date -= relativedelta(days=1)
        """

        return cooking_date

    def _in_plate(self, plate_name):
        for raw_plate in RawPlatePackageMaster.objects.all():
            if raw_plate.base_name in plate_name:
                return raw_plate

        # 見つからなかった場合
        return None

    def get_inner_package_info(self, unit_number: int, meal_name: str, type_name: str, cooking_day):

        # 対象ユニットの判定
        first_unit = UnitMaster.objects.filter(unit_number=unit_number).first()
        if first_unit:
            first_user = first_unit.username

            order_count = 0
            package_count = 0
            for eat_meal_name, eating_day_list in self.cooking_eating_dict.items():
                # コードに含まれない食事区分は除外する
                if eat_meal_name != meal_name:
                    continue

                # 注文数、中袋数の取得
                if type_name == '基本食':
                    if unit_number == 904:
                        # 集約対象の先頭ユニットの場合
                        plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day,
                                                                        meal_name=meal_name,
                                                                        is_mix_rice=True)
                        if plate_qs.exists():
                            basic_count_qs = Order.objects.filter(
                                unit_name__unit_number__in=settings.MIX_RICE_AGGREGATE_UNITS, quantity__gt=0,
                                eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                                menu_name__menu_name='常食'
                            )
                            basic_count = sum([x.quantity for x in basic_count_qs])
                        else:
                            basic_count = 0
                    else:
                        basic_count_qs = Order.objects.filter(
                            unit_name__unit_number=unit_number, quantity__gt=0,
                            eating_day__in=eating_day_list, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                        )
                        basic_count = sum([x.quantity for x in basic_count_qs])
                        order_count += basic_count
                    package_count += math.ceil(basic_count / self.ORDER_COUNT_PER_PACKAGE)

                elif type_name == '嚥下食':
                    enge_count_qs = Order.objects.filter(
                        unit_name__unit_number=unit_number, quantity__gt=0,
                        eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                        menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                    )
                    enge_count = sum([x.quantity for x in enge_count_qs])
                    order_count += enge_count
                    package_count += math.ceil(enge_count / self.ORDER_COUNT_PER_PACKAGE)
                elif type_name == '汁・汁具':
                    # ユニット単位の取得のため、汁ありとなしが混在はありえない前提(嚥下も同様)
                    basic_count_qs = Order.objects.filter(
                        unit_name__unit_number=unit_number, quantity__gt=0,
                        eating_day__in=eating_day_list, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                    )
                    basic_count = sum([x.quantity for x in basic_count_qs])

                    enge_count_qs = Order.objects.filter(
                        unit_name__unit_number=unit_number, quantity__gt=0,
                        eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                        menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                    )
                    enge_count = sum([x.quantity for x in enge_count_qs])

                    # 汁・味噌汁の中袋数の計算
                    # 契約内容変更により、顧客別～の設定値と実際の注文の食事区分が異なる場合があるため、Orderから取得する
                    # ※同一喫食日、食事区分(朝・昼・夕)で汁・汁具の有無が異なることは運用上ない前提
                    first_meal = Order.objects.filter(
                        unit_name__username=first_user, meal_name__meal_name=meal_name, meal_name__soup=True
                    ).first()
                    if first_meal:
                        # 汁なし、具のみの場合は、出力対象外。汁具の場合のみ以下の判断を行う

                        # 料理に味噌汁があるかどうか
                        has_miso_soup = False
                        plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, meal_name=meal_name,
                                                                        is_soup=True)
                        for plate in plate_qs:
                            if self._is_miso_soup(plate.plate_name):
                                has_miso_soup = True
                                break

                        soup_count = 0
                        if has_miso_soup:
                            if first_meal.meal_name.soup:
                                # 嚥下の味噌汁の汁は、汁・汁具の中袋に入れるため、基本食と嚥下の合計が必要
                                soup_count += basic_count + enge_count
                            else:
                                # 嚥下の汁(味噌汁、それ以外の汁・スープ)の汁具は、嚥下の中袋に入れるため、汁・汁具の中袋は基本食のみ
                                soup_count += basic_count
                        else:
                            # 嚥下の味噌汁以外の汁・汁具は嚥下の中袋に入れる
                            soup_count += basic_count
                        order_count += soup_count
                        package_count += math.ceil(soup_count / self.ORDER_COUNT_PER_SOUP_PACKAGE)
                    else:
                        filling_qs = Order.objects.filter(
                            unit_name__username=first_user, meal_name__meal_name=meal_name,
                            meal_name__soup=False, meal_name__filling=True
                        )
                        # 具のみの場合は、中袋の計算外だが0は実体に合わないので1を設定
                        if filling_qs.exists():
                            package_count += 1

                elif type_name == '原体':
                    basic_count_qs = Order.objects.filter(
                        unit_name__unit_number=unit_number, quantity__gt=0,
                        eating_day__in=eating_day_list, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                    )
                    basic_count = sum([x.quantity for x in basic_count_qs])

                    enge_count_qs = Order.objects.filter(
                        unit_name__unit_number=unit_number, quantity__gt=0,
                        eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                        menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                    )
                    enge_count = sum([x.quantity for x in enge_count_qs])

                    has_raw_plate = False
                    plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, meal_name=meal_name,
                                                                    is_soup=False)
                    for plate in plate_qs:
                        if self._is_raw_plate(plate):
                            # 原体マスタの参照
                            s_name = self.convert_plate_name(plate.plate_name)
                            raw_plate = self._in_plate(s_name)
                            if raw_plate:
                                if first_unit.username.dry_cold_type == "乾燥":
                                    is_direct = raw_plate.is_direct_dry
                                elif first_unit.username.dry_cold_type == "冷凍":
                                    is_direct = raw_plate.is_direct_cold
                                elif first_unit.username.dry_cold_type == "冷蔵":
                                    is_direct = raw_plate.is_direct_cold
                                else:
                                    logger.warn(f'施設の冷凍乾燥区分異常：{first_unit.dry_cold_type}')
                                    is_direct = False

                                # 原体送りの施設の区分が直送でないの場合、対象献立(基本 or 嚥下)のフラグを立てる
                                if not is_direct:
                                    has_raw_plate = True
                                    break
                            else:
                                logger.warn(f'原体マスタ未登録：{s_name}')
                    raw_count = 0
                    raw_count += basic_count if has_raw_plate else 0
                    raw_count += enge_count if has_raw_plate else 0
                    order_count += raw_count

                    # 原体は常に1袋
                    if raw_count:
                        package_count += 1
                else:
                    order_count = 0
                    package_count = 0

                # 原体の中袋数の計算

            return {
                'unit_number': unit_number,
                'unit_name': f'{unit_number}.{first_unit.calc_name}', 'meal_name': meal_name, 'type_name': type_name,
                'order_count': order_count, 'package_count': package_count
            }
        else:
            return {
                'unit_number': 0,
                'unit_name': 'any', 'meal_name': meal_name, 'type_name': type_name,
                'order_count': 0, 'package_count': 0
            }

    def is_mix_rice_agg_main_unit(self, unit_number) -> bool:
        """
        混ぜご飯集計用ユニットで使用する番号のユニットかどうかを取得する。
        """
        if unit_number == settings.MIX_RICE_AGGREGATE_UNITS[0][0]:
            return True
        else:
            return False

    def get_transfer_info(self, unit_number: int, meal_name: str, type_name: str, cooking_day):
        """
        対象ユーザーの配送用段ボールでの照合情報を取得する。
        """

        # 対象ユニットの判定
        first_unit = UnitMaster.objects.filter(unit_number=unit_number).first()
        if first_unit:
            first_user = first_unit.username

            order_count = 0
            package_count = 0
            for eat_meal_name, eating_day_list in self.cooking_eating_dict.items():
                if eat_meal_name != meal_name:
                    continue

                # 注文数、中袋数の取得
                # 基本食の中袋数
                basic_count_qs = Order.objects.filter(
                    unit_name__unit_number=unit_number, quantity__gt=0,
                    eating_day__in=eating_day_list, meal_name__meal_name=meal_name, menu_name__menu_name='常食'
                )
                basic_count = sum([x.quantity for x in basic_count_qs])
                order_count += basic_count
                package_count += math.ceil(basic_count / self.ORDER_COUNT_PER_PACKAGE)

                if unit_number == settings.MIX_RICE_AGGREGATE_UNITS[0][0]:
                    # 混ぜご飯集約対象の先頭ユニットの場合
                    plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day,
                                                                    meal_name=meal_name,
                                                                    is_mix_rice=True)
                    if plate_qs.exists():
                        # 混ぜご飯集約の中袋を追加
                        package_count += 1

                # 嚥下食の中袋数
                enge_count_qs = Order.objects.filter(
                    unit_name__unit_number=unit_number, quantity__gt=0,
                    eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                    menu_name__menu_name__in=['ソフト', 'ゼリー', 'ミキサー']
                )
                enge_count = sum([x.quantity for x in enge_count_qs])
                order_count += enge_count
                package_count += math.ceil(enge_count / self.ORDER_COUNT_PER_PACKAGE)

                # 汁・汁具の中袋数
                first_meal_order = Order.objects.filter(
                    unit_name__unit_number=unit_number,
                    eating_day__in=eating_day_list, meal_name__meal_name=meal_name,
                ).first()
                if first_meal_order:
                    first_meal = first_meal_order.meal_name
                    if first_meal.meal_name.soup:
                        # 汁なし、具のみの場合は、出力対象外。汁具の場合のみ以下の判断を行う

                        # 料理に味噌汁があるかどうか
                        has_miso_soup = False
                        plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, meal_name=meal_name,
                                                                        is_soup=True)
                        for plate in plate_qs:
                            if self._is_miso_soup(plate.plate_name):
                                has_miso_soup = True
                                break

                        soup_count = 0
                        if has_miso_soup:
                            if first_meal.meal_name.soup:
                                # 嚥下の味噌汁の汁は、汁・汁具の中袋に入れるため、基本食と嚥下の合計が必要
                                soup_count += basic_count + enge_count
                            else:
                                # 嚥下の汁(味噌汁、それ以外の汁・スープ)の汁具は、嚥下の中袋に入れるため、汁・汁具の中袋は基本食のみ
                                soup_count += basic_count
                        else:
                            # 嚥下の味噌汁以外の汁・汁具は嚥下の中袋に入れる
                            soup_count += basic_count
                        # 基本食・嚥下食に含まれるので、食数カウントしない
                        package_count += math.ceil(soup_count / self.ORDER_COUNT_PER_SOUP_PACKAGE)
                    elif first_meal.meal_name.filling:
                        # 汁具でカウントするが、具のみも中袋に格納が必要なため、袋数を入れる
                        package_count += 1

                # 原体袋数の計算
                has_raw_plate = False
                plate_qs = CookingDirectionPlate.objects.filter(cooking_day=cooking_day, meal_name=meal_name,
                                                                is_soup=False)
                for plate in plate_qs:
                    if self._is_raw_plate(plate):
                        # 原体マスタの参照
                        raw_plate = self._in_plate(plate.plate_name)
                        if raw_plate:
                            if first_unit.username.dry_cold_type == "乾燥":
                                is_direct = raw_plate.is_direct_dry
                            elif first_unit.username.dry_cold_type == "冷凍":
                                is_direct = raw_plate.is_direct_cold
                            elif first_unit.username.dry_cold_type == "冷蔵":
                                is_direct = raw_plate.is_direct_cold
                            else:
                                logger.warn(f'施設の冷凍乾燥区分異常：{first_unit.dry_cold_type}')
                                is_direct = False

                            # 原体送りの施設の区分が直送でないの場合、対象献立(基本 or 嚥下)のフラグを立てる
                            if not is_direct:
                                has_raw_plate = True
                                break
                        else:
                            logger.warn(f'原体マスタ未登録：{plate.plate_name}')

                raw_count = 0
                raw_count += basic_count if has_raw_plate else 0
                raw_count += enge_count if has_raw_plate else 0
                order_count += raw_count
                package_count += math.ceil(raw_count / self.ORDER_COUNT_PER_PACKAGE)

            return {
                'unit_number': unit_number,
                'unit_name': f'{unit_number}.{first_unit.calc_name}', 'meal_name': meal_name, 'type_name': '-',
                'order_count': order_count, 'package_count': package_count
            }
        else:
            return {
                'unit_number': 0,
                'unit_name': 'any', 'meal_name': meal_name, 'type_name': '-',
                'order_count': 0, 'package_count': 0
            }

    def get_package_info(self, qr_value: str, cooking_day):
        """
        ピッキング実施日、QRコード値を元に、照合対象の情報を出力する
        """

        # QRコード値の解析(中袋・段ボール)
        unit_number, meal_name, type_name = QrCodeUtil.perse_qr_value(qr_value)
        return self.get_inner_package_info(unit_number, meal_name, type_name, cooking_day)

    def get_required_units(self, cooking_day):
        """
        照合が必要な施設・食事区分・種類・フェーズの一覧を取得する。
        """
        results = []

        # 原体以外
        required_units = UnitPackage.objects.filter(cooking_day=cooking_day, count__gt=0).values(
            'unit_number', 'meal_name', 'menu_name'
        ).distinct()
        for key, group in groupby(required_units, key=lambda x: (x['unit_number'], x['meal_name'])):
            # 基本食・嚥下食の抽出
            # (ダンボールの照合では、施設全体で1行。常食か嚥下のどちらかで有無が確定する。)
            menu_list = [x['menu_name'] for x in group]
            is_hit = False
            if '常食' in menu_list:
                tpl = (key[0], key[1], '基本食', '中袋')
                if not (tpl in results):
                    results.append(tpl)

                tpl_trans = (key[0], key[1], '基本食', '段ボール')
                if not (tpl_trans in results):
                    results.append(tpl_trans)
                is_hit = True
            if ('ソフト' in menu_list) or ('ミキサー' in menu_list) or ('ゼリー' in menu_list):
                tpl = (key[0], key[1], '嚥下食', '中袋')
                if not (tpl in results):
                    results.append(tpl)

                tpl_trans = (key[0], key[1], '嚥下食', '段ボール')
                if not (tpl_trans in results):
                    results.append(tpl_trans)
                is_hit = True

            # 汁・汁具の抽出
            if is_hit:
                unit_number = key[0]
                unit = UnitMaster.objects.filter(unit_number=unit_number).first()
                if unit:
                    eating_qs = UnitPackage.objects.filter(cooking_day=cooking_day, meal_name=key[1], count__gt=0).values(
                        'eating_day'
                    ).first()
                    order = Order.objects.filter(unit_name=unit, eating_day=eating_qs['eating_day'],
                                                 meal_name__meal_name=key[1],
                                                 meal_name__filling=True).ordery_by('eating_day').first()
                    # qs = MealDisplay.objects.filter(username=unit.username, meal_name__filling=True)
                    if order.exists():
                        tpl = (unit_number, key[1], '汁・汁具', '中袋')
                        if not (tpl in results):
                            results.append(tpl)

                        tpl_trans = (unit_number, key[1], '汁・汁具', '段ボール')
                        if not (tpl_trans in results):
                            results.append(tpl_trans)
        # 原体
        required_units_for_raw_plate = PickingRawPlatePackage.objects.filter(cooking_day=cooking_day, quantity__gt=0).values(
            'unit_name__unit_number', 'meal_name'
        ).distinct()
        for ru_rp in required_units_for_raw_plate:
            tpl = (ru_rp['unit_name__unit_number'], ru_rp['meal_name'], '原体', '中袋')
            if not (tpl in results):
                results.append(tpl)

            tpl_trans = (ru_rp['unit_name__unit_number'], ru_rp['meal_name'], '原体', '段ボール')
            if not (tpl_trans in results):
                results.append(tpl_trans)

        return results

class PickingResultFileReader:
    def __init__(self, df, upload_file_name: str):
        self.df = df
        self.upload_file_name = upload_file_name

    def read_to_save(self):
        for index, data in self.df.iterrows():
            menu_file_no = data[0]
            menu_no = data[1]
            terminal = data[2]
            date = dt.datetime.strptime(f'{data[3]} {data[4]}', '%Y/%m/%d %H:%M:%S')

            # math.isnan()だと、正常時にエラーになるため注意
            if repr(data[5]) == 'nan':
                # QRコード値がない場合は表示に使えないので、除く
                continue
            if len(data[5]) <= 1:
                # QRコード値がない場合は表示に使えないので、除く
                continue
            value = str(data[5]).zfill(9)
            if 'nan' in value:
                # QRコード値がない場合は表示に使えないので、除く
                continue
            result = data[7]
            result_raw, is_create = PickingResultRaw.objects.get_or_create(
                menu_file_no=menu_file_no, menu_no=menu_no,
                terminal_no=terminal, picking_date=date, qr_value=value[:9], result=result,
                picking_phase='中袋' if menu_no == '1' else '段ボール')
            if is_create:
                result_raw.upload_file_name = self.upload_file_name
                result_raw.save()
