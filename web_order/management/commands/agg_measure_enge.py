import logging
import math
import os
import shutil
import datetime as dt
import platform
import openpyxl as excel

from django.conf import settings
from django_pandas.io import read_frame
from django.core.management.base import BaseCommand

from web_order.models import Order, OrderEveryday, UnitPackage, TmpPlateNamePackage
from .utils import AggEngePackageMixin, AggFixedOrderRule, ExcelOutputMixin
from web_order.cooking_direction_plates import PlateNameAnalizeUtil
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister, RawPlatePackageRegisterFactory

logger = logging.getLogger(__name__)


"""
    喫食日の食数を集計し、計量表を出力する処理

# 引数
    製造日（YYYY-MM-DD）
    喫食日（YYYY-MM-DD）
    食事区分（朝・昼・夕）
    料理名
    分量
    単位
    出汁の量

# 出力ファイル
    朝食の計量表
    昼食の計量表
    夕食の計量表
"""

class AggEngeMeasureWriter(AggEngePackageMixin, ExcelOutputMixin):

    def __init__(self):
        self.package_size = None

        # 販売固定商品注文数
        # (マスタ定義の存在するもの)
        # -保存用1人用
        self.pre_1pack_s = 0
        self.pre_1pack_z = 0
        self.pre_1pack_m = 0

        # -保存用50g
        self.pre_50g_j = 0
        self.pre_50g_s = 0
        self.pre_50g_z = 0
        self.pre_50g_m = 0

        self.in_index = 0
        self.in_cooking_day = None
        self.aggregation_day = None
        self.in_menu = None
        self.in_name = None
        self.in_is_soup_enge = False
        self.in_adjust = 0

        self.agg_day_short = 0

    def get_enge_50g_count(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.pre_50g_s
        elif menu_name == 'ゼリー':
            return self.pre_50g_z
        else:
            return self.pre_50g_m

    def create_excel_enge(self,
            dataframe, worksheet, meal, parts_name, quantity, unit, density, needle, preserve,
            is_show_enge, package):
        i = 10
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)  # 朝・昼・夕
        worksheet.cell(2, 2, parts_name)  # 料理名

        worksheet.cell(5, 4, needle)  # 針刺し用の食数
        worksheet.cell(7, 7, preserve)  # 保存用の食数
        worksheet.cell(6, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
        # 嚥下は基本、料理の内容量(g数)を出力しないため、コマンドのパラメータでする/しないを指定する
        if is_show_enge:
            worksheet.cell(2, 5, quantity)  # 料理の内容量

        menu = ''
        if 'ソフト' in worksheet.title:
            menu = 'ソフト'
        elif 'ゼリー' in worksheet.title:
            menu = 'ゼリー'
        elif 'ミキサー' in worksheet.title:
            menu = 'ミキサー'
        pack_50g = self.get_enge_50g_count(menu)
        worksheet.cell(6, 7, pack_50g)  # 保存用50gの食数

        package_count = 0
        package_count_1p = preserve
        size = self.get_filling_package_size(unit, quantity, package)
        self.package_size = size
        is_raw_save = False
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            #worksheet.cell(i, 4, row['注文数'])
            worksheet.cell(i, 7, self.get_filling_package_excel_function(unit, quantity, package, i))

            is_set_plate_package = False
            if PlateNameAnalizeUtil.is_raw_plate_name(parts_name):
                if PlateNameAnalizeUtil.is_raw_enge_plate_name(self.in_before_name, self.aggregation_day)[0]:
                    is_set_plate_package = True
                    is_raw_save = True
                    worksheet.cell(i, 4, row['注文数'])
                elif ('錦糸卵' in parts_name) and (not self.in_is_mixrice_part):
                    # 混ぜご飯は全て混ぜご飯のパウチに混ぜられるため、袋を作らない
                    is_raw_save = True
                    # 錦糸卵の場合は、乾燥だけ袋計算(直送かどうかで判断すべき?)
                    if row['乾燥冷凍区分'] == '乾燥':
                        is_set_plate_package = True
                        worksheet.cell(i, 4, row['注文数'])
            else:
                is_set_plate_package = True
                worksheet.cell(i, 4, row['注文数'])
                is_raw_save = False

            if is_set_plate_package:
                if row['注文数'] >= 2:
                    unit_package_count = math.ceil(row['注文数'] / size)
                    package_count += unit_package_count

                    # ピッキング指示書用のユニット毎袋数を保存
                    if row['注文数'] == 2:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], 1, 'ENGE_2', parts_name, menu, is_raw_to_enge=is_raw_save,)
                    else:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                   f'ENGE_{size}', parts_name, menu, is_raw_to_enge=is_raw_save,)
                if row['注文数'] == 1:
                    package_count_1p += 1

                    # ピッキング指示書用のユニット毎袋数を保存
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                               row['呼出番号'], row['ユニット名'], 1, 'ENGE_1', parts_name, menu, is_raw_to_enge=is_raw_save,)
                if is_show_enge:
                    worksheet.cell(i, 5, self.get_filling_quantity_function(quantity, row['注文数'], i))

            i = i + 1

        if self.in_adjust == 2:
            if index != -1:
                enge_index = self.in_index + 1
            else:
                enge_index = self.in_index
        else:
            enge_index = self.in_index

        if 'ソフト' in worksheet.title:
            P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, package_count,
                                                package_count_1p, 'ソフト',
                                                self.in_menu, pack_50g, is_raw_save)
        elif 'ゼリー' in worksheet.title:
            P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, package_count,
                                                package_count_1p, 'ゼリー',
                                                self.in_menu, pack_50g, is_raw_save)
        elif 'ミキサー' in worksheet.title:
            P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, package_count,
                                                package_count_1p, 'ミキサー',
                                                self.in_menu, pack_50g, is_raw_save)
        if self.in_adjust == 1:
            if 'ソフト' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index + 1, package_count,
                                                    package_count_1p, 'ソフト',
                                                    self.in_menu, pack_50g, is_raw_save)
            elif 'ゼリー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index + 1, package_count,
                                                    package_count_1p, 'ゼリー',
                                                    self.in_menu, pack_50g, is_raw_save)
            elif 'ミキサー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index + 1, package_count,
                                                    package_count_1p, 'ミキサー',
                                                    self.in_menu, pack_50g, is_raw_save)

    def get_enge_name(self):
        if self.in_is_soup_enge:
            base_name = self.in_name[0:self.in_name.find('具')]
            gu_name = self.in_name[0:self.in_name.find('＋')]
            soup_name = self.in_name[self.in_name.find('＋') + 1:]
            return (f'{gu_name}', f'{base_name}{soup_name}')
        else:
            return self.in_name

    def handle(self, manager, *args, **options):
        logger.info('嚥下用計量表解析開始')
        self.in_index = int(options['opt'][0])      # 呼び出し時の引数1つ目、調理表のインデックス(-1は対象不明)
        in_cook = options['opt'][1]         # 呼び出し時の引数2つ目、製造日（YYYY-MM-DD）
        in_date = options['opt'][2]         # 呼び出し時の引数3つ目、喫食日（YYYY-MM-DD）
        self.in_menu = options['opt'][3]         # 呼び出し時の引数4つ目、食事区分（朝・昼・夕）
        self.in_name = options['opt'][4]         # 呼び出し時の引数5つ目、料理名
        in_qty = options.get('quantity', 0)        # 呼び出し時の引数 --quantity 分量
        in_unit = options.get('unit', '')           # 呼び出し時の引数 --unit 単位
        in_dty = options.get('density', 0)         # 呼び出し時の引数 --density 出汁の量
        # in_fixed = options['fixed']         # 呼び出し時の引数 --fixed 出汁の液量が確定しているかどうか
        in_package_rule = options.get('package', '')       # 呼び出し時の引数 --package 袋数のルール(主菜扱い(main)/副菜扱い(sub))
        in_is_less = options.get('less', 0) == 1
        # in_is_same_thickness = options['thickness'] == 1    # 呼び出し時の引数 --thickness 常食と薄味で同じ濃さの液を使用するかどうか
        in_is_show_enge = options.get('showenge', 0) == 1          # 呼び出し時の引数 --showenge 嚥下のシートにg数出力を行うかどうか
        in_is_hide_enge = options.get('showenge', 0) == -1         # 呼び出し時の引数 --showenge -1の場合は、嚥下シート自体を表示しない
        # in_dilute = options['dilute'] == 1                  # 呼び出し時の引数 --dilute 食材を0.8倍提供するかどうか
        self.in_is_soup_enge = options.get('engesoup', 0) == 1          # 呼び出し時の引数 --engesoup 嚥下の液を別で出力するかどうか
        self.in_adjust = options.get('adjust', 0)                       # 呼び出し時の引数 --adjust 嚥下袋数保存時のindexの補正の状況(0:補正なし、1:補正の必要正発生、2:補正発生済み)
        self.in_before_name = options.get('before_name', '')
        self.in_is_mixrice_part = options.get('is_mix_rice_parts', 0) == 1          # 混ぜご飯の一部かどうか
        logger.info(f'[{self.in_name}]is_mixrice_part={self.in_is_mixrice_part}')

        # fixed_order = AggFixedOrderRule(self.in_name, float(in_qty), in_unit, in_is_less, float(in_dty))

        self.aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d').date()
        if platform.system() == 'Windows':
            self.agg_day_short = self.aggregation_day.strftime('%m/%d')
        else:
            self.agg_day_short = self.aggregation_day.strftime('%-m/%-d')
        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + in_cook + '_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        measure_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_enge.xlsx')  # 計量表のテンプレート

        # 製造日の取得
        self.in_cooking_day = dt.datetime.strptime(in_cook, '%Y-%m-%d').date()

        # 注文対象日・食事区分の設定
        manager.set_eating(self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # ピッキング指示書用袋数：原体の袋数の登録
        # ------------------------------------------------------------------------------
        package_register = RawPlatePackageRegisterFactory.create(self.in_before_name)
        if package_register.is_valid:
            package_register.register(manager.get_df_raw(), self.in_cooking_day, self.aggregation_day, self.in_menu)


        # ------------------------------------------------------------------------------
        # 針刺し用の食数
        # ------------------------------------------------------------------------------
        res_s, res_m, res_z = manager.get_needle_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        self.pre_1pack_s, self.pre_1pack_m, self.pre_1pack_z = manager.get_preserve_1p_enge_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(50g)
        # ------------------------------------------------------------------------------
        self.pre_50g_j, self.pre_50g_s, self.pre_50g_z, self.pre_50g_m = manager.get_preserve_50g_orders()

        # ------------------------------------------------------------------------------
        # qs（Queryset）df（Dataframe）
        # 朝食（b）昼食（l）夕食（d）
        # 常食（j）薄味（u）ソフト（s）ミキサー（m）ゼリー（z）
        # 具（g） 汁（s）
        # ------------------------------------------------------------------------------

        if self.in_menu == '朝食':

            book_b = excel.load_workbook(measure_template)

            # ソフト
            df_bs = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bs, book_b["ソフト"], '△ 朝', gu_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_bs, book_b["ソフト(液)"], '△ 朝', soup_name, in_dty, 'g', in_dty, res_s, self.pre_1pack_s, True, in_package_rule)
            else:
                self.create_excel_enge(df_bs, book_b["ソフト"], '△ 朝', self.in_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ソフト(液)"])

            # ミキサー食
            df_bm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bm, book_b["ミキサー"], '△ 朝', gu_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_bm, book_b["ミキサー(液)"], '△ 朝', soup_name, in_dty, 'g', in_dty, res_m, self.pre_1pack_m, True, in_package_rule)
            else:
                self.create_excel_enge(df_bm, book_b["ミキサー"], '△ 朝', self.in_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ミキサー(液)"])

            # ゼリー食
            df_bz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bz, book_b["ゼリー"], '△ 朝', gu_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_bz, book_b["ゼリー(液)"], '△ 朝', soup_name, in_dty, 'g', in_dty, res_z, self.pre_1pack_z, True, in_package_rule)
            else:
                self.create_excel_enge(df_bz, book_b["ゼリー"], '△ 朝', self.in_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ゼリー(液)"])

            breakfast_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_朝_' + self.in_name + '.xlsx')
            self.save_with_select(book_b, breakfast_output_file)

        if self.in_menu == '昼食':

            book_l = excel.load_workbook(measure_template)

            df_ls = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_ls, book_l["ソフト"], '○ 昼', gu_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_ls, book_l["ソフト(液)"], '○ 昼', soup_name, in_dty, 'g', in_dty, res_s, self.pre_1pack_s,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_ls, book_l["ソフト"], '○ 昼', self.in_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ソフト(液)"])

            df_lm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_lm, book_l["ミキサー"], '○ 昼', gu_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_lm, book_l["ミキサー(液)"], '○ 昼', soup_name, in_dty, 'g', in_dty, res_m, self.pre_1pack_m,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_lm, book_l["ミキサー"], '○ 昼', self.in_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ミキサー(液)"])

            df_lz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_lz, book_l["ゼリー"], '○ 昼', gu_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_lz, book_l["ゼリー(液)"], '○ 昼', soup_name, in_dty, 'g', in_dty, res_z, self.pre_1pack_z,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_lz, book_l["ゼリー"], '○ 昼', self.in_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ゼリー(液)"])

            lunch_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_昼_' + self.in_name + '.xlsx')
            self.save_with_select(book_l, lunch_output_file)

        if self.in_menu == '夕食':

            book_d = excel.load_workbook(measure_template)

            df_ds = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_ds, book_d["ソフト"], '□ 夕', gu_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_ds, book_d["ソフト(液)"], '□ 夕', soup_name, in_dty, 'g', in_dty, res_s, self.pre_1pack_s,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_ds, book_d["ソフト"], '□ 夕', self.in_name, in_qty, in_unit, in_dty, res_s, self.pre_1pack_s,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ソフト(液)"])

            df_dm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_dm, book_d["ミキサー"], '□ 夕', gu_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_dm, book_d["ミキサー(液)"], '□ 夕', soup_name, in_dty, 'g', in_dty, res_m, self.pre_1pack_m,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_dm, book_d["ミキサー"], '□ 夕', self.in_name, in_qty, in_unit, in_dty, res_m, self.pre_1pack_m,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ミキサー(液)"])

            df_dz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_dz, book_d["ゼリー"], '□ 夕', gu_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                self.create_excel_enge(df_dz, book_d["ゼリー(液)"], '□ 夕', soup_name, in_dty, 'g', in_dty, res_z, self.pre_1pack_z,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_dz, book_d["ゼリー"], '□ 夕', self.in_name, in_qty, in_unit, in_dty, res_z, self.pre_1pack_z,
                                  in_is_show_enge, in_package_rule)
                if in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ゼリー(液)"])

            dinner_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_夕_' + self.in_name + '.xlsx')
            self.save_with_select(book_d, dinner_output_file)

        if self.in_index == -1:
            TmpPlateNamePackage.objects.get_or_create(
                plate_name=self.in_before_name,
                cooking_day=self.in_cooking_day,
                size=self.package_size,
                menu_name='嚥下'
            )

        # shutil.make_archive(new_dir_path, 'zip', root_dir=new_dir_path)

