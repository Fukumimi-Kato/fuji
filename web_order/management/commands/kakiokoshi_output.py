import datetime as dt
import logging
import os
import re

from dateutil.relativedelta import relativedelta
from functools import cmp_to_key
from itertools import groupby
import openpyxl as excel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break

from django.conf import settings
from django.core.management.base import BaseCommand
from django.db.models import Sum
from django_pandas.io import read_frame

from web_order.models import CookingDirectionPlate, AllergenPlateRelations, Order, CommonAllergen, UncommonAllergen, UncommonAllergenHistory
from web_order.p7 import P7Util
from web_order.cooking_direction_plates import PlateNameAnalizeUtil


logger = logging.getLogger(__name__)


class Command(BaseCommand):
    DEFAULT_KOBACHI_GRAM = 40.0
    DEFAULT_BREAK_ROW = 38
    date_title_font = Font(name='游ゴシック', size=24)
    eating_day_font = Font(name='游ゴシック', size=36, bold=True)

    meal_mark_font = Font(name='ＭＳ Ｐゴシック', size=36, bold=True)
    meal_kana_font = Font(name='游ゴシック', size=12, bold=True)
    meal_label_font = Font(name='游ゴシック', size=48, bold=True)

    plate_p7_number_font = Font(name='游ゴシック', size=18, bold=True)
    plate_title_font = Font(name='游ゴシック', size=24, bold=True)
    plate_title_fill = PatternFill(patternType='solid', fgColor='E2EFDA')

    unit_font = Font(name='游ゴシック', size=18, bold=True)
    unit_fill = PatternFill(patternType='solid', fgColor='E2EFDA')

    allergen_font = Font(name='游ゴシック', size=18, bold=True)
    allergen_fill = PatternFill(patternType='solid', fgColor='E2EFDA')
    allergen_alter_font = Font(name='游ゴシック', size=24, bold=True, u='single')

    soft_fill = PatternFill(patternType='solid', fgColor='FFCCFF')
    jerry_fill = PatternFill(patternType='solid', fgColor='F8CBAD')
    mixer_fill = PatternFill(patternType='solid', fgColor='FFFF99')

    meal_kanji_font = Font(name='游ゴシック', size=48, bold=True)

    center_alignment = Alignment(vertical='center')
    center_center_alignment = Alignment(vertical='center', horizontal='center')
    top_alignment = Alignment(vertical='top', horizontal='center')
    bottom_alignment = Alignment(vertical='bottom', horizontal='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    detail_top_border = Border(top=Side(style='thin'))

    def add_arguments(self, parser):
        parser.add_argument('opt', nargs='*', type=str)

    def _write_lastupdated(self, ws, qs):
        last_updated = max([x.updated_at for x in qs])

        # UTCで出力されてしまうので補正
        last_updated += dt.timedelta(hours=9)
        ws.cell(2, 10, f'最終データ更新日時：{last_updated.strftime("%Y/%m/%d %H:%M:%S")}')

    def _get_weekday_for_excel(self, date):
        weekday = date.weekday()
        if weekday == 6:
            return '(日Sun.)'
        elif weekday == 0:
            return '(月Mon.)'
        elif weekday == 1:
            return '(火Tue.)'
        elif weekday == 2:
            return '(水Wed.)'
        elif weekday == 3:
            return '(木Thu.)'
        elif weekday == 4:
            return '(金Fri.)'
        else:
            return '(土Sat.)'

    def write_eating_day_and_meal(self, ws, row, eating_day, meal):
        ws.cell(row, 2).font = self.eating_day_font
        ws.cell(row, 2, f'{eating_day.strftime("%m/%d")}{self._get_weekday_for_excel(eating_day)}')

        ws.cell(row + 2, 2).font = self.meal_mark_font
        ws.cell(row + 4, 2).font = self.meal_kana_font
        ws.cell(row + 2, 2).alignment = self.center_center_alignment
        ws.cell(row + 4, 2).alignment = self.bottom_alignment
        ws.cell(row + 5, 2).font = self.meal_label_font
        ws.cell(row + 5, 2).alignment = self.top_alignment

        ws.merge_cells(f'B{row + 2}:B{row + 3}')

        # セル結合は別メソッドから行う
        if meal == '朝食':
            ws.cell(row + 2, 2, '△')
            ws.cell(row + 4, 2, 'あさ')
            ws.cell(row + 5, 2, '朝')
        elif meal == '昼食':
            ws.cell(row + 2, 2, '〇')
            ws.cell(row + 4, 2, 'ひる')
            ws.cell(row + 5, 2, '昼')
        elif meal == '夕食':
            ws.cell(row + 2, 2, '□')
            ws.cell(row + 4, 2, 'ゆう')
            ws.cell(row + 5, 2, '夕')

    def _is_required_dry_notice(self, plate):
        """
        乾燥品注意書き出力が必要かどうかを判定する。現在は、混ぜご飯に使われる錦糸卵のみ
        """
        if plate.is_mix_rice:
            if ('錦糸卵' in plate.plate_name) or ('きんしたまご' in plate.plate_name):
                return True
            else:
                return False
        else:
            False

    def _is_raw_plate(self, plate):
        """
        原体送りの料理かどうかを判断する。
        """
        return '原体' in plate.plate_name

    def write_plate_name(self, ws, row, plate, package, cooking_day, eating_day, is_adjusted, food_count):
        """
        書き起こし票に料理名を出力する。
        """
        # スタイルの設定
        ws.row_dimensions[row].height = 30
        ws.row_dimensions[row + 1].height = 56
        ws.cell(row, 3).alignment = self.left_alignment
        ws.cell(row, 3).font = self.plate_p7_number_font

        ws.cell(row, 3).fill = self.plate_title_fill
        ws.cell(row, 4).fill = self.plate_title_fill
        ws.cell(row, 5).fill = self.plate_title_fill
        ws.cell(row, 6).fill = self.plate_title_fill
        ws.cell(row, 7).fill = self.plate_title_fill
        ws.cell(row, 8).fill = self.plate_title_fill
        ws.cell(row, 9).fill = self.plate_title_fill
        ws.cell(row, 10).fill = self.plate_title_fill

        ws.cell(row + 1, 3).font = self.plate_title_font
        ws.cell(row + 1, 3).alignment = self.center_alignment

        ws.cell(row + 1, 3).fill = self.plate_title_fill
        ws.cell(row + 1, 4).fill = self.plate_title_fill
        ws.cell(row + 1, 5).fill = self.plate_title_fill
        ws.cell(row + 1, 6).fill = self.plate_title_fill
        ws.cell(row + 1, 7).fill = self.plate_title_fill
        ws.cell(row + 1, 8).fill = self.plate_title_fill
        ws.cell(row + 1, 9).fill = self.plate_title_fill
        ws.cell(row + 1, 10).fill = self.plate_title_fill

        # P7番号の出力
        if self._is_need_adjust_index(plate):
            basic_number_index = P7Util.get_number_index(food_count)
            enge_number_index = P7Util.get_number_index(food_count)
            enge_append_index = P7Util.get_number_index(food_count + 1)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, cooking_day)}{enge_number_index}'
            soft_append_number = f'{P7Util.get_number_prefix("ソフト", False, False, cooking_day)}{enge_append_index}'

            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, cooking_day)}{enge_number_index}'
            jelly_append_number = f'{P7Util.get_number_prefix("ゼリー", False, False, cooking_day)}{enge_append_index}'

            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, cooking_day)}{enge_number_index}'
            mixer_append_number = f'{P7Util.get_number_prefix("ミキサー", False, False, cooking_day)}{enge_append_index}'

            p7_numbers = "、".join([
                basic_number,
                soft_number, soft_append_number,
                jelly_number, jelly_append_number,
                mixer_number, mixer_append_number
            ])

            result = food_count + 2
        elif PlateNameAnalizeUtil.is_miso_soup(plate.plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate.plate_name):
            p7_numbers = P7Util.get_number_miso_soup()
            result = food_count
        else:
            basic_number_index = P7Util.get_number_index(food_count)
            if is_adjusted:
                enge_number_index = P7Util.get_number_index(food_count + 1)
            else:
                enge_number_index = P7Util.get_number_index(food_count)

            basic_number = f'{P7Util.get_number_prefix("常食", False, False, cooking_day)}{basic_number_index}'
            soft_number = f'{P7Util.get_number_prefix("ソフト", False, False, cooking_day)}{enge_number_index}'
            jelly_number = f'{P7Util.get_number_prefix("ゼリー", False, False, cooking_day)}{enge_number_index}'
            mixer_number = f'{P7Util.get_number_prefix("ミキサー", False, False, cooking_day)}{enge_number_index}'
            p7_numbers = "、".join([basic_number, soft_number, jelly_number, mixer_number])

            result = food_count + 1

        ws.cell(row, 3, p7_numbers)

        # 料理名の出力
        if self._is_raw_plate(plate):
            if PlateNameAnalizeUtil.is_raw_enge_plate_name(plate.plate_name, eating_day)[0]:
                if self._is_required_dry_notice(plate):
                    # 乾燥注意記載品で、嚥下製造対象はないため、packageは出力しない
                    ws.cell(row + 1, 3, f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)')
                else:
                    ws.cell(row + 1, 3, f'☆{plate.plate_name}　{package}')
            else:
                if self._is_required_dry_notice(plate):
                    ws.cell(row + 1, 3, f'☆{plate.plate_name}　(※乾燥(かんそう)は袋(ふくろ)あり)')
                else:
                    ws.cell(row + 1, 3, f'☆{plate.plate_name}')
        else:
            ws.cell(row + 1, 3, f'{plate.plate_name}　{package}')

        return result

    def _get_plate_gram(self, plate_name: str):
        """
        料理のg数を取得する。gでない単位の場合は小鉢のデフォルト値(中サイズ対象となるg数)を返却する。
        """
        # 全角半角変換(+は全角に、gは半角に揃える)
        replaced = plate_name.replace('+', '＋').replace('ｇ', 'g')
        plus_index = replaced.find('＋')
        if plus_index != -1:
            replaced = replaced[:plus_index]
        res = re.findall('(\d+|\d+\.\d+)g', replaced)
        if res:
            return float(res[0])

        return self.DEFAULT_KOBACHI_GRAM

    def _get_plate_total_gram(self, plate_name: str):
        """
        料理のg数を取得する。gでない単位の場合は小鉢のデフォルト値(中サイズ対象となるg数)を返却する。
        """
        # 全角半角変換(+は全角に、gは半角に揃える)
        replaced = plate_name.replace('+', '＋').replace('%', '％').replace('ｇ', 'g').replace('（', '(').replace('）', ')')

        enable_units = settings.MEASURE_ENABLE_UNITS + ['個', '丁']
        not_gram_unit_list = "|".join(enable_units)
        res = re.findall(f'({not_gram_unit_list})\((\d+|\d+\.\d+)g\)＋(\D*)(\d+|\d+\.\d+)％', replaced)
        if res:
            src = float(res[0][1])
            return src + src * float(res[0][3]) / 100

        res = re.findall('(\d+|\d+\.\d+)g＋(\D*)(\d+|\d+\.\d+)g', replaced)
        if res:
            return float(res[0][0]) + float(res[0][2])
        else:
            res = re.findall('(\d+|\d+\.\d+)g＋(\D*)(\d+|\d+\.\d+)％', replaced)
            if res:
                src = float(res[0][0])
                return src + src * float(res[0][2]) / 100
            else:
                plus_index = replaced.find('＋')
                if plus_index != -1:
                    replaced = replaced[:plus_index]
                res = re.findall('(\d+|\d+\.\d+)g', replaced)
                if res:
                    return float(res[0])

        return self.DEFAULT_KOBACHI_GRAM

    def _get_sub_plate_package(self, plate_name):
        """
        副菜の袋サイズ判断
        """
        total_gram = self._get_plate_total_gram(plate_name)
        if (total_gram < 20.0) or ('▼' in plate_name):
            return 'しせつごと'
        elif total_gram >= 50.0:
            return '大'
        elif total_gram >= 40.0:
            return '中'
        else:
            return '小'

    def _get_package(self, plate):
        """
        袋サイズを取得する
        """
        number = plate.plate_name[0]
        plate_name = plate.plate_name

        if number == '⑩':
            return '中'
        if number == '①':
            if PlateNameAnalizeUtil.is_sansyokudon_plate(plate_name):
                return '小'
            else:
                return '大'
        elif (number == '②') or (number == '③'):
            return self._get_sub_plate_package(plate_name)
        elif number == '④':
            if plate.is_mix_rice:
                return '中'
            else:
                gram = self._get_plate_gram(plate_name)
                if (gram < 20.0) or ('▼' in plate_name):
                    total_gram = self._get_plate_total_gram(plate_name)
                    if (total_gram < 20.0) or ('▼' in plate_name):
                        return 'しせつごと'
                    elif total_gram >= 50.0:
                        return '大'
                    elif total_gram >= 40.0:
                        return '中'
                    else:
                        return '小'
                else:
                    if gram >= 40:
                        return '中'
                    else:
                        return '小'
        elif number == '⑤':
            if not PlateNameAnalizeUtil.is_miso_soup(plate_name):
                if PlateNameAnalizeUtil.is_soup_liquid(plate_name):
                    # 味噌汁以外の汁・スープの液の料理は、小で出力する(年に1,2回程度中のケースがあるがそちらは運用対応)
                    return '小'

            return '中'
        else:
            # 本来この内容は出てはならない
            return '未定'

    def _is_requierd_unit(self, plate):
        """
        ユニット出力が必要かどうかを判定する
        """
        return plate.is_soup

    def _is_soup_filling(self, plate):
        return plate.is_soup and ('具' in plate.plate_name)

    def write_units(self, ws, row, plate):
        ws.cell(row, 3).border = self.detail_top_border
        ws.cell(row, 4).border = self.detail_top_border
        ws.cell(row, 5).border = self.detail_top_border
        ws.cell(row, 6).border = self.detail_top_border
        ws.cell(row, 7).border = self.detail_top_border
        ws.cell(row, 8).border = self.detail_top_border
        ws.cell(row, 9).border = self.detail_top_border
        ws.cell(row, 10).border = self.detail_top_border

        ws.cell(row, 3).font = self.unit_font
        ws.cell(row, 3).alignment = self.center_alignment
        ws.cell(row, 3, 'しせつ：')

        # 表示対象の施設の取得
        if self._is_soup_filling(plate):
            qs = Order.objects.filter(
                eating_day=plate.eating_day, meal_name__meal_name=plate.meal_name,
                meal_name__filling=True, quantity__gt=0, allergen__allergen_name='なし') \
                .select_related('unit_name')\
                .values('unit_name__unit_number', 'unit_name__short_name')\
                .exclude(unit_name__unit_code__range=[80001, 80008]).order_by('unit_name__unit_number')
        else:
            qs = Order.objects.filter(
                eating_day=plate.eating_day, meal_name__meal_name=plate.meal_name,
                meal_name__soup=True, quantity__gt=0, allergen__allergen_name='なし')\
                .select_related('unit_name')\
                .values('unit_name__unit_number', 'unit_name__short_name')\
                .exclude(unit_name__unit_code__range=[80001, 80008]).order_by('unit_name__unit_number')
        dataframe = read_frame(qs)
        dataframe = dataframe.drop_duplicates().reset_index()

        # 施設内容の出力
        col = 4
        for index, data in dataframe.iterrows():
            unit_row = row + int(index / 7)
            unit_col = col + (index % 7)
            ws.cell(unit_row, unit_col).font = self.unit_font
            ws.cell(unit_row, unit_col).alignment = self.center_alignment
            ws.cell(unit_row, unit_col, f'{data["unit_name__unit_number"]}.{data["unit_name__short_name"]}')

        end_row = row + int(index / 7) + 1
        for i in range(end_row - row):
            ws.cell(row + i, 3).fill = self.unit_fill
        return int(index / 7) + 1

    def get_kind_menu_name(self, code):
        """
        食種の献立種類を取得する
        """
        if '常' in code:
            return '常食'
        elif 'ソ' in code:
            return 'ソフト'
        elif 'ゼ' in code:
            return 'ゼリー'
        elif 'ミ' in code:
            return 'ミキサー'
        else:
            return ''

    def get_allergens_with_menu(self, code, cooking_day):
        if settings.KIZAWA_RAKUKON_CODE in code:
            if settings.KOSHOKU_UNIT_IDS:
                return ['個食'], '常食'
        if settings.FREEZE_RACKUKON_CODE in code:
            if settings.FREEZE_UNIT_IDS:
                return ['ﾌﾘｰｽﾞ'], '常食'

        # 献立種類名の取得
        menu_name = self.get_kind_menu_name(code)

        # 散発アレルギーから検索
        uncommon_qs = UncommonAllergenHistory.objects.filter(code=code, cooking_day=cooking_day, menu_name=menu_name)
        if uncommon_qs.exists():
            logger.info(f'menu={menu_name}/qs={uncommon_qs}')
            allergen = uncommon_qs.first().allergen
            logger.info(f'allergen={repr(allergen)}-{allergen.allergen_name}')
            return [allergen], menu_name
        else:
            common_qs = CommonAllergen.objects.filter(code=code, menu_name__menu_name=menu_name)
            if common_qs.exists():
                return [x.allergen for x in common_qs], menu_name
            else:
                logger.info(f'data is none.({code}-{menu_name})')
                return [], None

    def _get_enge_cell_color(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.soft_fill
        elif menu_name == 'ゼリー':
            return self.jerry_fill
        elif menu_name == 'ミキサー':
            return self.mixer_fill
        else:
            return None

    def compare_allergen_unit(self, unit_order1, unit_order2):
        seq_dict = {
            '常食': 0,
            'ソフト': 1,
            'ゼリー': 2,
            'ミキサー': 3
        }

        cmp_seq = seq_dict[unit_order1.menu_name.menu_name] - seq_dict[unit_order2.menu_name.menu_name]
        if cmp_seq:
            return cmp_seq
        else:
            cmp_allergen_id = unit_order1.allergen_id - unit_order2.allergen_id
            if cmp_allergen_id:
                return cmp_allergen_id
            return unit_order1.unit_name.unit_number - unit_order2.unit_name.unit_number

    def is_soe_allergen(self, plate_unit_list, prev_unit_list):
        if not prev_unit_list:
            return False
        if len(plate_unit_list) != len(prev_unit_list):
            return False

        for current in plate_unit_list:
            is_find = False
            for prev in prev_unit_list:
                if current == prev:
                    is_find = True
                    break

            if not is_find:
                return False

        return True

    def _get_display_unit_name(self, unit):
        display_unit_number = ''
        if unit.unit_number < 10:
            # 1桁の場合は、半角空白を入れて、2桁の施設名とも並ぶようにするう
            display_unit_number = f' {unit.unit_number}'
        else:
            display_unit_number = f'{unit.unit_number}'

        display_unit_name = unit.short_name
        if len(unit.short_name) < 5:
            diff = 5 - len(unit.short_name)
            # 5文字になるまで全角空白を入れる
            for i in range(diff):
                display_unit_name = display_unit_name + '　'

        return f'{display_unit_number}.{display_unit_name}'

    def _is_need_adjust_index(self, plate):
        logger.info(f'◆付料理判定：{plate.plate_name}')
        if '◆' in plate.plate_name:
            res = re.findall('具(\d+|\d+\.\d+)[g|ｇ]\s*\D液(\d+|\d+\.\d+)[g|ｇ]', plate.plate_name)
            logger.info(f'◆付料理判定結果：{res}')
            if res and (res[0][0] and res[0][1]):
                logger.info('◆付料理判定--HIT')
                return True
        if ('カレーライス' in plate.plate_name) or ('シチュー' in plate.plate_name):
            return True
        else:
            return False

    def write_allergen(self, ws, row, plate, cooking_day):
        # アレルギー情報の取得
        relations = AllergenPlateRelations.objects.filter(source=plate).exclude(plate=None).order_by('plate_id')

        col = 4
        index = 0
        prev_unit_list = None
        for key, group in groupby(relations, key=lambda x: x.plate):
            allergen_unit_list = []
            for allergen_plate in group:
                code = allergen_plate.code
                allergen_list, menu_name = self.get_allergens_with_menu(code, cooking_day)
                for allergen in allergen_list:
                    # 対象アレルギーを注文している施設を取得
                    if allergen == '個食':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.KOSHOKU_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    elif allergen == 'ﾌﾘｰｽﾞ':
                        order_qs = Order.objects.filter(
                            eating_day=plate.eating_day, allergen__allergen_name='なし', quantity__gt=0,
                            unit_name_id__in=settings.FREEZE_UNIT_IDS,
                            meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                        ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                            unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    else:
                        if key.is_soup:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name,
                                meal_name__filling=True
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                        else:
                            order_qs = Order.objects.filter(
                                eating_day=plate.eating_day, allergen=allergen, quantity__gt=0,
                                meal_name__meal_name=plate.meal_name, menu_name__menu_name=menu_name
                            ).exclude(unit_name__unit_code__range=[80001, 80008]).annotate(
                                unit_quantity=Sum('quantity')).order_by('menu_name__seq_order', 'unit_name__unit_number')
                    logger.info(f'{plate.eating_day}-{allergen}-{plate.meal_name}-{menu_name}:')
                    logger.info(f'{order_qs}:')
                    logger.info('')

                    allergen_unit_list += list(order_qs)

            # 施設情報のソート
            allergen_orders = sorted(allergen_unit_list, key=cmp_to_key(self.compare_allergen_unit))

            if self.is_soe_allergen(allergen_unit_list, prev_unit_list):
                logger.info(f'{ws.cell(row + index - 1, 10).value}、{key.plate_name}')
                logger.info(allergen_unit_list)
                logger.info(prev_unit_list)
                ws.cell(row + index - 1, 10).value = f'{ws.cell(row + index - 1, 10).value}、{key.plate_name}'
            else:
                prev_unit_list = []
                prev_unit_list += allergen_unit_list

                ws.cell(row + index, 3).border = self.detail_top_border
                ws.cell(row + index, 4).border = self.detail_top_border
                ws.cell(row + index, 5).border = self.detail_top_border
                ws.cell(row + index, 6).border = self.detail_top_border
                ws.cell(row + index, 7).border = self.detail_top_border
                ws.cell(row + index, 8).border = self.detail_top_border
                ws.cell(row + index, 9).border = self.detail_top_border
                ws.cell(row + index, 10).border = self.detail_top_border

                ws.cell(row + index, 3).font = self.allergen_font
                ws.cell(row + index, 3).alignment = self.center_alignment
                ws.cell(row + index, 3, 'アレルギー：')

                # 施設情報の出力
                allergen_index = 0
                for order in allergen_orders:
                    unit_row = row + index + (allergen_index // 2)
                    unit_col = col + (allergen_index % 2) * 3
                    if order.menu_name.menu_name == '常食':
                        kana_name = order.allergen.kana_name
                        if not kana_name:
                            if '個食' in order.unit_name.unit_name:
                                kana_name = 'こしょく'
                            elif 'ﾌﾘｰｽﾞ' in order.unit_name.unit_name:
                                kana_name = 'ﾌﾘｰｽﾞ'
                        # 出力内容(常食はセル色を変えない)
                        label = f'{self._get_display_unit_name(order.unit_name)}　{kana_name}　{order.unit_quantity}'
                    else:
                        # セルの色付け
                        fill_color = self._get_enge_cell_color(order.menu_name.menu_name)
                        if fill_color:
                            ws.cell(unit_row, unit_col).fill = fill_color
                            ws.cell(unit_row, unit_col + 1).fill = fill_color
                            ws.cell(unit_row, unit_col + 2).fill = fill_color
                            if allergen_index % 2:
                                ws.cell(unit_row, unit_col + 3).fill = fill_color

                        # 出力内容
                        label = f'{self._get_display_unit_name(order.unit_name)}　{order.allergen.kana_name}　{order.menu_name.menu_name} {order.unit_quantity}'
                    ws.cell(unit_row, unit_col).font = self.allergen_font
                    ws.cell(unit_row, unit_col).alignment = self.center_alignment
                    ws.cell(unit_row, unit_col, f'{label}')
                    allergen_index += 1
                index += (allergen_index // 2) + 1

                ws.cell(row + index, 10, f'{key.plate_name}')
                logger.info(f'{key}:{row}+{index}')
                ws.cell(row + index, 10).font = self.allergen_alter_font
                ws.cell(row + index, 10).alignment = self.right_alignment
                index += 1

        for i in range(index):
            ws.cell(row + i, 3).fill = self.allergen_fill

        return index

    def output_excel(self, ws, cooking_day):
        # 対象データの取得
        plates_qs = CookingDirectionPlate.objects.filter(
            cooking_day=cooking_day, is_basic_plate=True
        ).order_by('eating_day', 'seq_meal', 'index')

        if not plates_qs.exists():
            logger.error('plate data is none.')

        # データ最終更新日時の出力
        self._write_lastupdated(ws, plates_qs)

        # データの出力
        logger.debug(f'Breaks(1):{ws.row_breaks}')
        row = 3
        next_break = self.DEFAULT_BREAK_ROW
        food_count = 0
        is_first_eating = True
        for key, group in groupby(plates_qs, key=lambda x: (x.eating_day, x.meal_name)):
            # 改ページ位置調整
            if is_first_eating:
                is_first_eating = False
            else:
                logger.info(f'change_eating_day:page break:row={row},next_break={next_break}')
                ws.row_breaks.append(Break(row - 1))
                logger.debug(f'Breaks(2):{ws.row_breaks}')
                next_break = row - 1 + self.DEFAULT_BREAK_ROW

            start_row = row
            ws.row_dimensions[row].height = 57.75

            # 調理日の出力
            ws.cell(row, 7).font = self.date_title_font
            ws.cell(row, 7, f'調理：{cooking_day.strftime("%m/%d")}')

            # 消費期限の出力
            limit_day = cooking_day + relativedelta(days=6)
            ws.cell(row, 9).font = self.date_title_font
            ws.cell(row, 9, f'消費期限：{limit_day.strftime("%m/%d")}')

            eating_day, meal = key

            # 喫食日の出力
            self.write_eating_day_and_meal(ws, row, eating_day, meal)

            row += 2
            is_adjusted = False
            for plate in group:
                # 料理名の出力
                plate_start_row = row
                food_count = self.write_plate_name(ws, row, plate, self._get_package(plate), cooking_day, eating_day, is_adjusted, food_count)
                row += 2

                if self._is_need_adjust_index(plate):
                    is_adjusted = True

                # 対象施設の出力
                if self._is_requierd_unit(plate):
                    row += self.write_units(ws, row, plate)

                # アレルギーの出力
                row += self.write_allergen(ws, row, plate, cooking_day)

                # 次の料理のために空白行を入れる
                ws.row_dimensions[row].height = 30

                if row > next_break:
                    logger.info(f'page break(2):row={row},next_break={next_break}')

                    ws.row_breaks.append(Break(plate_start_row - 1))
                    logger.debug(f'Breaks(3):{ws.row_breaks}')

                    next_break += self.DEFAULT_BREAK_ROW
                    logger.info(f'next_break={next_break}')
                row += 1

            # 食事区分(漢字)のセル結合
            ws.cell(start_row + 5, 2).font = self.meal_kanji_font
            ws.cell(start_row + 5, 2).alignment = self.top_alignment
            if (start_row + 5) <= (row - 2):
                ws.merge_cells(f'B{start_row + 5}:B{row - 2}')

        logger.debug(f'Breaks(4):{ws.row_breaks}')

    def handle(self, *args, **options):
        in_cook = options['opt'][0]         # 呼び出し時の引数1つ目、製造日（YYYY-MM-DD）

        cooking_day = dt.datetime.strptime(in_cook, '%Y-%m-%d').date()

        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/kakiokoshi.xlsx')  # 書き起こし票のテンプレート
        workbook = excel.load_workbook(template_path)

        # 帳票出力
        self.output_excel(workbook['template'], cooking_day)

        output_dir = os.path.join(settings.OUTPUT_DIR, 'kakiokoshi')
        os.makedirs(output_dir, exist_ok=True)  # 上書きOK
        save_path = os.path.join(output_dir, f'書き起こし票_{str(cooking_day)}_製造.xlsx')
        workbook.save(save_path)
        workbook.close()
