import math
import os
import shutil
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP
import platform
import openpyxl as excel

from django.conf import settings
from django_pandas.io import read_frame
from django.core.management.base import BaseCommand

from web_order.models import Order, OrderEveryday, PlatePackageForPrint, UnitPackage, TmpPlateNamePackage
from .utils import AggEngePackageMixin, ExcelOutputMixin
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister, RawPlatePackageRegisterFactory


"""
    喫食日の食数を集計し、計量表を出力する処理

# 引数
    製造日（YYYY-MM-DD）
    喫食日（YYYY-MM-DD）
    食事区分（朝・昼・夕）
    料理名
    分量

# 出力ファイル
    朝食の味噌汁の計量表
    昼食の味噌汁の計量表
    夕食の味噌汁の計量表
"""

class MisoSoupMeasureWriter(AggEngePackageMixin, ExcelOutputMixin):

    def __init__(self):
        self.index = None
        self.agg_day_short = None
        self.basic_plate_enable = False
        self.in_cooking_day = None
        self.aggregation_day = None
        self.in_menu = None
        self.package_size = None
        self.enge_package_size = None
        self.in_adjust = None

        # 販売固定商品注文数
        # (マスタ定義の存在するもの)
        # -保存用1人用
        self.pre_1pack_j = 0
        self.pre_1pack_u = 0

        # -見本・写真用
        self.photo = 0

        # -保存用1人用
        self.pre_1pack_s = 0
        self.pre_1pack_z = 0
        self.pre_1pack_m = 0

        # -保存用50g
        self.pre_50g_j = 0
        self.pre_50g_s = 0
        self.pre_50g_z = 0
        self.pre_50g_m = 0

    def get_fixed_1pack_count(self, is_soup: bool = False):
        if is_soup:
            # 味噌汁の汁の場合は、写真用が不要だが、その代わりに1件必要
            return self.pre_1pack_j + 1
        else:
            return self.pre_1pack_j + self.photo

    def get_enge_50g_count(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.pre_50g_s
        elif menu_name == 'ゼリー':
            return self.pre_50g_z
        else:
            return self.pre_50g_m

    def create_excel_miso(self,
            dataframe, worksheet, meal, is_enge, parts_name, quantity, unit, needle, preserve,
            preserve_one_pack, photo, is_write_total_needle_preserve=False, is_enge_display=False):
        """
        具の出力を行う
        """
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)
        worksheet.cell(2, 2, parts_name)
        worksheet.cell(2, 6, unit)

        total = 0
        if is_enge:
            i = 10
            worksheet.cell(6, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
            worksheet.cell(5, 4, needle)        # 針刺し用の食数
            worksheet.cell(7, 7, preserve)      # 保存用の食数
            worksheet.cell(2, 9, quantity)      # 一人前数量
            if is_enge_display:
                worksheet.cell(2, 5, quantity)  # 一人前数量
                worksheet.cell(4, 5, quantity)  # 一人前数量
                worksheet.cell(5, 5, '=D5*E2')  # 針刺し用
                # 保存用(50g)は数量固定のため不要
                worksheet.cell(7, 5, quantity)  # 保存用(一人前)
        else:
            i = 14
            worksheet.cell(2, 5, quantity)      # 一人前数量
            worksheet.cell(7, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
            worksheet.cell(7, 7, self.pre_50g_j)        # 保存用(50)の袋数
            worksheet.cell(8, 7, preserve_one_pack)     # 保存用(一人前の袋)の袋数
            worksheet.cell(9, 7, photo)                 # 写真
            if self.basic_plate_enable:
                worksheet.cell(6, 7, 5)         # 保存用(10人袋)数量

        if is_write_total_needle_preserve:
            worksheet.cell(5, 2, 30)            # 針刺し用の袋の注文数(固定値)
            worksheet.cell(6, 2, 10)            # 保存用の袋の注文数(固定値)

        if is_enge:
            print_package_count = 0
            print_package_count_1p = preserve
        else:
            print_package_count = 2 + 5         # 保存用(10人袋)5、針刺し用(30人袋)2
            print_package_count_1p = self.get_fixed_1pack_count()

        # 袋サイズ
        if is_enge:
            package_size = self.get_miso_soup_package_size(unit, quantity)
            self.enge_package_size = package_size
            menu = ''
            if 'ソフト' in worksheet.title:
                menu = 'ソフト'
            elif 'ゼリー' in worksheet.title:
                menu = 'ゼリー'
            elif 'ミキサー' in worksheet.title:
                menu = 'ミキサー'
            enge_50g_count = self.get_enge_50g_count(menu)
            print_package_count_50g = enge_50g_count
            worksheet.cell(6, 7, enge_50g_count)      # 保存用(50)の袋数
            self.package_size = package_size
        else:
            self.package_size = 30
            print_package_count_50g = self.pre_50g_j

        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            if is_enge or (row['注文数'] > 1):
                worksheet.cell(i, 4, row['注文数'])
            else:
                worksheet.cell(i, 4, 0)
            total += int(row['注文数'])
            if row['注文数'] > 1:
                unit_package_count = math.ceil(row['注文数']/self.package_size)
                print_package_count += unit_package_count

                # ピッキング指示書用のユニット毎袋数を保存
                if is_enge:
                    # ピッキング指示書用のユニット毎袋数を保存
                    if row['注文数'] == 2:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index,
                            row['呼出番号'], row['ユニット名'], 1, 'ENGE_2', parts_name, menu, soup_type='filling')
                    else:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index,
                            row['呼出番号'], row['ユニット名'], unit_package_count, f'ENGE_{self.package_size}', parts_name, menu, soup_type='filling')
                else:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index,
                        row['呼出番号'], row['ユニット名'], unit_package_count, 'SOUP_UNIT', parts_name, soup_type='filling')
            if row['注文数'] == 1:
                print_package_count_1p += 1

                # ピッキング指示書用のユニット毎袋数を保存
                if is_enge:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index,
                        row['呼出番号'], row['ユニット名'], 1, 'ENGE_1', parts_name, menu, soup_type='filling')
                else:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index,
                        row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', parts_name, soup_type='filling')
            if is_enge and is_enge_display:
                worksheet.cell(i, 5, f'=D{i}*E2/G{i}')

            if is_enge:
                worksheet.cell(i, 7, self.get_miso_soup_package_function(unit, quantity, i))

            i = i + 1

        # 印刷用袋数の保存
        if is_enge:
            if self.in_adjust == 2:
                enge_index = self.index + 1
            else:
                enge_index = self.index

            if 'ソフト' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ソフト',
                                                    self.in_menu, print_package_count_50g)
            elif 'ゼリー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ゼリー',
                                                    self.in_menu, print_package_count_50g)
            elif 'ミキサー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ミキサー',
                                                    self.in_menu, print_package_count_50g)
        else:
            P7Util.save_package_count_for_print(
                self.in_cooking_day, self.aggregation_day, self.index, print_package_count, print_package_count_1p,
                '常食', self.in_menu, print_package_count_50g)

        return total

    def create_excel_miso_soup(self,
            dataframe, worksheet, meal, parts_name, quantity,
            is_write_total_needle_preserve=False):
        """
        味噌汁の汁の出力を行う。
        """
        i = 10
        worksheet.cell(1, 9, self.agg_day_short)
        worksheet.cell(1, 14, meal)
        worksheet.cell(2, 2, parts_name)

        total = 0
        worksheet.cell(2, 5, quantity)      # 一人前数量

        print_package_count = 0
        print_package_count_1p = 0
        print_package_count_50g = 0
        if is_write_total_needle_preserve:
            worksheet.cell(5, 2, 30)            # 針刺し用の袋の注文数(固定値)
            worksheet.cell(6, 2, 10)            # 保存用の袋の注文数(固定値)
            print_package_count += 2
            print_package_count_1p += self.get_fixed_1pack_count(True)  # 見本(味噌汁の汁では、写真用を使わない)
            print_package_count_50g = self.pre_50g_j
        if self.basic_plate_enable:
            worksheet.cell(6, 9, '=IF(B6=0,0,5)')         # 保存用(10人袋)袋数
            if is_write_total_needle_preserve:
                print_package_count += 5

        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            worksheet.cell(i, 4, row['注文数'])
            if row['注文数'] > 1:
                unit_package_count = math.ceil(row['注文数']/30)
                print_package_count += unit_package_count

                # ピッキング指示書用のユニット毎袋数を保存
                if self.index != -1:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index + 1,
                        row['呼出番号'], row['ユニット名'], unit_package_count, 'SOUP_UNIT', '味噌汁', soup_type='soup')
            if row['注文数'] == 1:
                print_package_count_1p += 1

                # ピッキング指示書用のユニット毎袋数を保存
                if self.index != -1:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.index + 1,
                        row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', '味噌汁', soup_type='soup')
            total += int(row['注文数'])

            i = i + 1

        self.save_soup_package_count_for_print(print_package_count, print_package_count_1p, print_package_count_50g)
        return total

    def read_excel_items_quantity(self, worksheet, quantity):
        """
        味噌汁の食材の1人分の数量を読み込む。常食と薄味は読み込むシートが異なるため注意。
        """
        worksheet.cell(4, 5, quantity)
        worksheet.cell(4, 6, 0)
        worksheet.cell(4, 7, 0)

        # 1人分の数量はテンプレートファイル上にしか定義していない
        katsukonmi = worksheet['E4'].value
        miso = worksheet['F4'].value
        water = worksheet['G4'].value

        # 結果を/区切りで返す
        return "{0}/{1}/{2}".format(katsukonmi, miso, water)

    def save_soup_package_count_for_print(self, count: int, count_1p: int, count_50g: int):
        if self.index == -1:
            return

        # 前のインデックスに汁の料理があったら、そちらで登録済みなので、ここでは登録しない
        plate_package_qs_lt = PlatePackageForPrint.objects.filter(
            cooking_day=self.in_cooking_day, eating_day=self.aggregation_day,
            is_basic_plate=True, meal_name=self.in_menu, index__lt=self.index, menu_name='常食')
        for plate_lt in plate_package_qs_lt:
            if '⑤' in plate_lt.plate_name:
                return

        # 汁の袋数を登録(汁の料理に「希釈」が含まれている前提)
        plate_package_qs = PlatePackageForPrint.objects.filter(
            cooking_day=self.in_cooking_day, eating_day=self.aggregation_day,
            is_basic_plate=True, meal_name=self.in_menu, index__gt=self.index, menu_name='常食')
        for plate_package in plate_package_qs:
            if '希釈' in plate_package.plate_name:
                plate_package.count = count
                plate_package.count_one_p = count_1p
                plate_package.count_one_50g = count_50g
                plate_package.save()
                break

    def handle(self, arg_index, arg_cook, arg_date, arg_menu, arg_name, arg_qty, arg_unit, arg_enge_display,
               arg_is_first, adjust, manager, before_name, liquid_quantity):
        self.index = arg_index      # 呼び出し時の引数1つ目、調理表のインデックス(-1は対象不明)
        in_cook = arg_cook         # 呼び出し時の引数2つ目、製造日（YYYY-MM-DD）
        in_date = arg_date         # 呼び出し時の引数3つ目、喫食日（YYYY-MM-DD）
        self.in_menu = arg_menu         # 呼び出し時の引数4つ目、食事区分（朝・昼・夕）
        in_name = arg_name         # 呼び出し時の引数5つ目、料理名
        in_qty = arg_qty   # 呼び出し時の引数 --quantity 分量
        in_unit = arg_unit           # 呼び出し時の引数 --unit 単位
        in_enge_display = arg_enge_display == 1   # 呼び出し時の引数 --is_enge_display 嚥下の内容量を出力するかどうか
        in_is_first_miso_soup = arg_is_first == 1   # 呼び出し時の引数 --is_first_miso_soup 初回の味噌汁出力かどうか
        self.in_adjust = adjust               # 呼び出し時の引数 --adjust 嚥下袋数保存時のindexの補正の状況(0:補正なし、1:補正の必要正発生、2:補正発生済み)

        self.aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d').date()
        if platform.system() == 'Windows':
            self.agg_day_short = self.aggregation_day.strftime('%m/%d')
        else:
            self.agg_day_short = self.aggregation_day.strftime('%-m/%-d')
        enable_day = dt.datetime.strptime(settings.BASIC_PLATE_ENABLE_DATE, '%Y-%m-%d').date()
        if self.aggregation_day >= enable_day:
            self.basic_plate_enable = True
        else:
            self.basic_plate_enable = False

        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + in_cook + '_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        if self.basic_plate_enable:
            miso_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_miso_v2.xlsx')  # 味噌汁用の計量表テンプレ
        else:
            miso_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_miso.xlsx')  # 味噌汁用の計量表テンプレ

        breakfast_output_miso = os.path.join(new_dir_path, str(self.aggregation_day) + '_朝汁_' + in_name + '.xlsx')
        lunch_output_miso = os.path.join(new_dir_path, str(self.aggregation_day) + '_昼汁_' + in_name + '.xlsx')
        dinner_output_miso = os.path.join(new_dir_path, str(self.aggregation_day) + '_夕汁_' + in_name + '.xlsx')

        # 製造日の取得
        self.in_cooking_day = dt.datetime.strptime(in_cook, '%Y-%m-%d').date()

        # 注文対象日・食事区分の設定
        manager.set_eating(self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # ピッキング指示書用袋数：原体の袋数の登録
        # ------------------------------------------------------------------------------
        package_register = RawPlatePackageRegisterFactory.create(before_name)
        if package_register.is_valid:
            package_register.register(manager.get_def_miso_raw_soup(), self.in_cooking_day, self.aggregation_day, self.in_menu)
            package_register.register(manager.get_def_miso_raw_filling(), self.in_cooking_day, self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # 針刺し用の食数
        # ------------------------------------------------------------------------------
        res_s, res_m, res_z = manager.get_needle_orders()

        # ------------------------------------------------------------------------------
        # 嚥下保存用(1人用)の食数
        # ------------------------------------------------------------------------------
        self.pre_1pack_s, self.pre_1pack_m, self.pre_1pack_z = manager.get_preserve_1p_enge_orders()

        # ------------------------------------------------------------------------------
        # 常食保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        self.pre_1pack_j, self.pre_1pack_u = manager.get_preserve_1p_orders()

        # ------------------------------------------------------------------------------
        # 写真用の食数
        # ------------------------------------------------------------------------------
        self.photo = manager.get_photo_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(50g)
        # ------------------------------------------------------------------------------
        self.pre_50g_j, self.pre_50g_s, self.pre_50g_z, self.pre_50g_m = manager.get_preserve_50g_orders()

        # ------------------------------------------------------------------------------
        # qs（Queryset）df（Dataframe）
        # 朝食（b）昼食（l）夕食（d）
        # 常食（j）薄味（u）ソフト（s）ミキサー（m）ゼリー（z）
        # 具（g） 汁（s）
        # ------------------------------------------------------------------------------
        book_bm = excel.load_workbook(miso_template)
        j_items_quantity = self.read_excel_items_quantity(book_bm["汁_常食"], liquid_quantity)
        u_items_quantity = self.read_excel_items_quantity(book_bm["汁_薄味"], liquid_quantity)
        j_total = 0
        if self.in_menu == '朝食':
            # 基本食(常食)
            df_bj_g = manager.get_def_miso_filling()
            self.create_excel_miso(df_bj_g, book_bm["具_常食"], '△ 朝', False, in_name, in_qty, in_unit, 0, 0, self.pre_1pack_j, self.photo)

            # 薄味
            book_bm.remove(book_bm["具_薄味"])

            # ソフト食
            df_bs_g = manager.get_def_miso_soft_filling()
            self.create_excel_miso(df_bs_g, book_bm["具_ソフト"], '△ 朝', True, in_name, in_qty, in_unit, res_s, self.pre_1pack_s, 0, 0, False, in_enge_display)

            # ミキサー食
            df_bm_g = manager.get_def_miso_mixer_filling()
            self.create_excel_miso(df_bm_g, book_bm["具_ミキサー"], '△ 朝', True, in_name, in_qty, in_unit, res_m, self.pre_1pack_m, 0, 0, False, in_enge_display)

            # ゼリー食
            df_bz_g = manager.get_def_miso_jelly_filling()
            self.create_excel_miso(df_bz_g, book_bm["具_ゼリー"], '△ 朝', True, in_name, in_qty, in_unit, res_z, self.pre_1pack_z, 0, 0, False, in_enge_display)

            # 汁(基本食)
            df_bj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_bj_s, book_bm["汁_常食"], '△ 朝', in_name, in_qty, in_is_first_miso_soup)

            # 汁(薄味)
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            self.save_with_select(book_bm, breakfast_output_miso)

        if self.in_menu == '昼食':
            # 基本食(常食)
            df_lj_g = manager.get_def_miso_filling()
            self.create_excel_miso(df_lj_g, book_bm["具_常食"], '○ 昼', False, in_name, in_qty, in_unit, 0, 0, self.pre_1pack_j, self.photo)

            # 薄味
            book_bm.remove(book_bm["具_薄味"])

            # ソフト食
            df_ls_g = manager.get_def_miso_soft_filling()
            self.create_excel_miso(df_ls_g, book_bm["具_ソフト"], '○ 昼', True, in_name, in_qty, in_unit, res_s, self.pre_1pack_s, 0, 0, False, in_enge_display)

            # ミキサー食
            df_lm_g = manager.get_def_miso_mixer_filling()
            self.create_excel_miso(df_lm_g, book_bm["具_ミキサー"], '○ 昼', True, in_name, in_qty, in_unit, res_m, self.pre_1pack_m, 0, 0, False, in_enge_display)

            # ゼリー食
            df_lz_g = manager.get_def_miso_jelly_filling()
            self.create_excel_miso(df_lz_g, book_bm["具_ゼリー"], '○ 昼', True, in_name, in_qty, in_unit, res_z, self.pre_1pack_z, 0, 0, False, in_enge_display)

            # 基本食(常食)-汁
            df_lj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_lj_s, book_bm["汁_常食"], '○ 昼',in_name, in_qty, in_is_first_miso_soup)

            # 薄味
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            self.save_with_select(book_bm, lunch_output_miso)

        if self.in_menu == '夕食':
            # 基本食(常食)
            df_dj_g = manager.get_def_miso_filling()
            self.create_excel_miso(df_dj_g, book_bm["具_常食"], '□ 夕', False, in_name, in_qty, in_unit, 0, 0, self.pre_1pack_j, self.photo)

            # 薄味
            book_bm.remove(book_bm["具_薄味"])

            # ソフト食
            df_ds_g = manager.get_def_miso_soft_filling()
            self.create_excel_miso(df_ds_g, book_bm["具_ソフト"], '□ 夕', True, in_name, in_qty, in_unit, res_s, self.pre_1pack_s, 0, 0, False, in_enge_display)

            # ミキサー食
            df_dm_g = manager.get_def_miso_mixer_filling()
            self.create_excel_miso(df_dm_g, book_bm["具_ミキサー"], '□ 夕', True, in_name, in_qty, in_unit, res_m, self.pre_1pack_m, 0, 0, False, in_enge_display)

            # ゼリー食
            df_dz_g = manager.get_def_miso_jelly_filling()
            self.create_excel_miso(df_dz_g, book_bm["具_ゼリー"], '□ 夕', True, in_name, in_qty, in_unit, res_z, self.pre_1pack_z, 0, 0, False, in_enge_display)

            # 基本食(常食)-汁
            df_dj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_dj_s, book_bm["汁_常食"], '□ 夕', in_name, in_qty, in_is_first_miso_soup)

            # 薄味
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            self.save_with_select(book_bm, dinner_output_miso)
        # ------------------------------------------------------------------------------

        if self.index == -1:
            if self.package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.package_size,
                    menu_name='常食'
                )
            if self.enge_package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.enge_package_size,
                    menu_name='嚥下'
                )

        return "{0}:{1}:{2}:{3}".format(j_items_quantity, u_items_quantity, j_total, u_total)
