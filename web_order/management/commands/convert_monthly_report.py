import os
import openpyxl as excel
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import platform

from django.conf import settings
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    def __init__(self):
        self.thin_side = Side(style='thin')
        self.header_top_border = Border(top=self.thin_side, left=self.thin_side, right=self.thin_side)
        self.header_bottom_border = Border(bottom=self.thin_side, left=self.thin_side, right=self.thin_side)
        self.cell_border = Border(top=self.thin_side, bottom=self.thin_side, left=self.thin_side, right=self.thin_side)

        # 条件入力シートのフォント
        self.condition_title_font = Font(name='ＭＳ Ｐゴシック', size=16, bold=True, u='single')
        self.condition_guide_font = Font(name='ＭＳ Ｐゴシック', size=11, bold=True, color='FF0000')
        self.condition_text_font = Font(name='ＭＳ Ｐゴシック', size=12)
        self.condition_readme_font = Font(name='ＭＳ Ｐゴシック', size=12, bold=True)
        self.condition_link_font = Font(name='ＭＳ Ｐゴシック', size=11, u='single', color='0000FF')
        self.medium_side = Side(style='medium')
        self.condition_cell_border = Border(top=self.medium_side, bottom=self.medium_side, left=self.medium_side, right=self.medium_side)

        self.condition_input_fill = PatternFill(patternType='solid', fgColor='FDE9D9')
        self.condition_readme_fill = PatternFill(patternType='solid', fgColor='EBF1DE')

    def add_arguments(self, parser):
        parser.add_argument('opt', nargs='+', type=str)

    def handle(self, *args, **options):
        in_file = options['opt'][0]

        # アップロードファイルパス
        report_file = os.path.join(settings.MEDIA_ROOT, 'upload', 'monthly', in_file)

        # 保存先
        output_path = os.path.join(settings.MEDIA_ROOT, 'output', 'monthly_report')
        os.makedirs(output_path, exist_ok=True)  # 上書きOK

        workbook = excel.load_workbook(report_file)
        ws_item = workbook.worksheets[0]
        ws_all = workbook.worksheets[1]

        # 検索条件シートのテンプレートからコピー
        condition_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/monthly_report_condition.xlsx')
        condition_workbook = excel.load_workbook(condition_template)
        condition_sheet = condition_workbook.worksheets[0]
        workbook._sheets.insert(0, condition_sheet)

        sheet_items = workbook.copy_worksheet(ws_item)
        sheet_items.title = '食材別_提供割合計算'
        sheet_items_all = workbook.copy_worksheet(ws_all)
        sheet_items_all.title = '食材別_全要素_提供割合計算'

        # 隠しセルに条件の計算結果を設定
        self.add_hide_column(workbook._sheets[1])

        # 食材別シートに最小値の設定を反映
        self.reflect_items(workbook._sheets[4], workbook._sheets[1].title)

        # 食座別_全要素シートに反映
        self.reflect_items_all(workbook._sheets[5], workbook._sheets[2].title, workbook._sheets[1].title)

        workbook.save(os.path.join(output_path, in_file))

        condition_workbook.close()
        workbook.close()

        book2 = excel.load_workbook(os.path.join(output_path, in_file))

        # 条件入力シートの見た目の設定
        self.set_condition_input_sheet(book2.worksheets[0])

        # 選択状態の変更
        for ws in book2.worksheets:
            ws.sheet_view.tabSelected = False
        book2.active = book2.worksheets[0]

        book2.save(os.path.join(output_path, in_file))
        book2.close()

    def set_condition_input_sheet(self, ws):
        ws.cell(2, 2).font = self.condition_title_font
        ws.cell(3, 2).font = self.condition_guide_font

        # 入力欄
        # -エネルギー
        ws.cell(5, 2).font = self.condition_text_font
        ws.cell(5, 4).border = self.condition_cell_border
        ws.cell(5, 4).fill = self.condition_input_fill
        # -炭水化物
        ws.cell(6, 2).font = self.condition_text_font
        ws.cell(6, 4).border = self.condition_cell_border
        ws.cell(6, 4).fill = self.condition_input_fill
        # -脂質
        ws.cell(7, 2).font = self.condition_text_font
        ws.cell(7, 4).border = self.condition_cell_border
        ws.cell(7, 4).fill = self.condition_input_fill
        # -炭水化物
        ws.cell(8, 2).font = self.condition_text_font
        ws.cell(8, 4).border = self.condition_cell_border
        ws.cell(8, 4).fill = self.condition_input_fill
        # -カリウム
        ws.cell(9, 2).font = self.condition_text_font
        ws.cell(9, 4).border = self.condition_cell_border
        ws.cell(9, 4).fill = self.condition_input_fill
        # -塩分
        ws.cell(10, 2).font = self.condition_text_font
        ws.cell(10, 4).border = self.condition_cell_border
        ws.cell(10, 4).fill = self.condition_input_fill

        # 説明欄
        for col in range(2, 14):
            ws.cell(13, col).fill = self.condition_readme_fill
        ws.cell(13, 2).font = self.condition_readme_font
        ws.cell(19, 2).hyperlink = ''
        ws.cell(19, 2).hyperlink.location = '食材別_提供割合計算!A1'
        ws.cell(19, 2).font = self.condition_link_font
        ws.cell(20, 2).hyperlink = ''
        ws.cell(20, 2).hyperlink.location = '食材別_全要素_提供割合計算!A1'
        ws.cell(20, 2).font = self.condition_link_font

    def add_hide_column(self, ws):
        # 参照用の列タイトル出力
        ws.cell(4, 20, 'エネルギー')
        ws.cell(4, 21, 'たんぱく質')
        ws.cell(4, 22, '脂質')
        ws.cell(4, 23, '炭水化物')
        ws.cell(4, 24, 'カリウム')
        ws.cell(4, 25, '食塩')
        ws.cell(4, 26, 'MIN')

        # 非表示設定
        ws.column_dimensions['T'].hidden = True
        ws.column_dimensions['U'].hidden = True
        ws.column_dimensions['V'].hidden = True
        ws.column_dimensions['W'].hidden = True
        ws.column_dimensions['X'].hidden = True
        ws.column_dimensions['Y'].hidden = True
        ws.column_dimensions['Z'].hidden = True

        # 計算列の出力
        i = 5
        for row in ws.iter_rows(min_row=i):
            try:
                if int(row[1].value) <= 31:
                    # エネルギー
                    ws.cell(i, 20, f'=IF(提供条件入力!$D$5="", 100, 提供条件入力!$D$5/$D{i}*100)')

                    # たんぱく質
                    ws.cell(i, 21, f'=IF(提供条件入力!$D$6="", 100, 提供条件入力!$D$6/$E{i}*100)')

                    # 脂質
                    ws.cell(i, 22, f'=IF(提供条件入力!$D$7="", 100, 提供条件入力!$D$7/$F{i}*100)')

                    # 炭水化物
                    ws.cell(i, 23, f'=IF(提供条件入力!$D$8="", 100, 提供条件入力!$D$8/$G{i}*100)')

                    # カリウム
                    ws.cell(i, 24, f'=IF(提供条件入力!$D$9="", 100, 提供条件入力!$D$9/$I{i}*100)')

                    # 食塩
                    ws.cell(i, 25, f'=IF(提供条件入力!$D$10="", 100, 提供条件入力!$D$10/$S{i}*100)')

                    # 最小値
                    ws.cell(i, 26, f'=MIN(T{i}:Y{i})')

                i += 1
            except Exception as e:
                # int変換エラーの場合、平均値の行に達したとみなす
                break

    def reflect_items(self, ws, org_sheet_title: str):
        # 提供割合列の追加
        ws.cell(3, 20, '提供割合')
        ws.cell(3, 20).border = self.header_top_border
        ws.cell(3, 20).alignment = Alignment(horizontal='center')
        ws.cell(4, 20, '%')
        ws.cell(4, 20).border = self.header_bottom_border
        ws.cell(4, 20).alignment = Alignment(horizontal='center')

        # 計算結果の反映
        column_start = 4
        column_end = 20
        i = 5
        for row in ws.iter_rows(min_row=i):
            try:
                if int(row[1].value) <= 31:
                    for col in range(column_start, column_end):
                        col_str = excel.utils.get_column_letter(col)
                        ws.cell(i, col, f"='{org_sheet_title}'!{col_str}{i}*$T{i}%")
                        ws.cell(i, col).number_format = '0.00'

                    # 提供割合
                    ws.cell(i, column_end, f"=IF('{org_sheet_title}'!Z{i} > 100, 100, '{org_sheet_title}'!Z{i})")
                    ws.cell(i, column_end).number_format = '0.00'
                    ws.cell(i, column_end).border = self.cell_border

                i += 1
            except Exception as e:
                # int変換エラーの場合、平均値の行に達したとみなす
                break

        # 平均値
        ws.cell(i, 20, "-")
        for col in range(column_start, column_end):
            col_str = excel.utils.get_column_letter(col)
            ws.cell(i, col, f"=AVERAGE({col_str}5:{col_str}{i-1})")
            ws.cell(i, col).number_format = '0.00'
        ws.cell(i, col + 1).border = self.cell_border

        ws.cell(i + 1, 20, "-")
        ws.cell(i + 1, 20).border = self.cell_border
        ws.cell(i + 2, 20, "-")
        ws.cell(i + 2, 20).border = self.cell_border
        ws.cell(i + 3, 20, "-")
        ws.cell(i + 3, 20).border = self.cell_border

    def reflect_items_all(self, ws, org_sheet_title: str, cond_sheet_title: str):
        # 提供割合列の追加
        ws.cell(3, 68, '提供割合')
        ws.cell(3, 68).border = self.header_top_border
        ws.cell(3, 68).alignment = Alignment(horizontal='center')
        ws.cell(4, 68, '%')
        ws.cell(4, 68).border = self.header_bottom_border
        ws.cell(4, 68).alignment = Alignment(horizontal='center')

        # 計算結果の反映
        i = 5
        column_start = 4
        column_end = 68
        for row in ws.iter_rows(min_row=i):
            try:
                if int(row[1].value) <= 31:
                    for col in range(column_start, column_end):
                        col_str = excel.utils.get_column_letter(col)
                        ws.cell(i, col, f"='{org_sheet_title}'!{col_str}{i}*$BP{i}%")
                        ws.cell(i, col).number_format = '0.00'

                    # 提供割合
                    ws.cell(i, column_end, f"=IF('{cond_sheet_title}'!Z{i} > 100, 100, '{cond_sheet_title}'!Z{i})")
                    ws.cell(i, column_end).number_format = '0.00'
                    ws.cell(i, column_end).border = self.cell_border

                i += 1
            except Exception as e:
                # int変換エラーの場合、平均値の行に達したとみなす
                break

        # 平均値
        ws.cell(i, 68, "-")
        for col in range(column_start, column_end):
            col_str = excel.utils.get_column_letter(col)
            ws.cell(i, col, f"=AVERAGE({col_str}5:{col_str}{i-1})")
            ws.cell(i, col).number_format = '0.00'
        ws.cell(i, col + 1).border = self.cell_border

        ws.cell(i + 1, 68, "-")
        ws.cell(i + 1, 68).border = self.cell_border
        ws.cell(i + 2, 68, "-")
        ws.cell(i + 2, 68).border = self.cell_border
        ws.cell(i + 3, 68, "-")
        ws.cell(i + 3, 68).border = self.cell_border
