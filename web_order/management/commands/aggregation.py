import os
import pandas as pd
import datetime as dt
from itertools import groupby
import logging
import openpyxl as excel
import shutil

from django_pandas.io import read_frame
from django.core.management.base import BaseCommand
from django.conf import settings

from web_order.models import Order, OrderEveryday, RakukonShortname, MenuMaster, AllergenMaster, CommonAllergen
from web_order.models import UncommonAllergen


class AllergenCount:
    def __init__(self, code, name, seq):
        self.code = code
        self.name = name
        self.seq = seq
        # 朝・昼・夕のカウント
        self.counts = [0, 0, 0]

class AllergenAnalyzer:
    def __init__(self):
        self.logger = logging.getLogger(__name__)

        self.commons = list(CommonAllergen.objects.all().order_by('seq_order'))
        self.uncommons = list(UncommonAllergen.objects.all().select_related('allergen', 'menu_name').order_by('seq_order'))
        self.quantity_list_common = []
        self.quantity_list_uncommon = []
        self.preserved_uncommons = [] # 更新予定の散発アレルギー
        self.chenged_uncommons = []

        self.quantity_list_common.append(AllergenCount("ﾌﾘｰｽﾞ", "フリーズ", 0))
        # らくらく献立の食種は、散発アレルギー->頻発アレルギーの順
        for uc in self.uncommons:
            self.quantity_list_uncommon.append(AllergenCount(uc.code, uc.name, uc.seq_order))
        for key, group in groupby(self.commons, key=lambda x: x.code):
            cm = group.__next__()
            self.quantity_list_common.append(AllergenCount(key, cm.name, cm.seq_order))

    def _get_uncommon_allergen(self, allergen, menu, meal, quantity):
        ucs = sorted([x for x in self.uncommons if x.menu_name.menu_name == menu], key=lambda uc: uc.last_use_date)
        for uncommon in ucs:
            if uncommon.name == allergen.get_rakukon_name():
                # 使用日時を更新する
                uncommon.save()
                return uncommon

        # 散発アレルギーの切り替え用に予約
        self.preserved_uncommons.append((allergen, menu, meal, quantity))
        return None

    def _get_common_allergen(self, name, menu, menu_group):
        if '個食' in name:
            commons = [x for x in self.commons if x.allergen.allergen_name == '個食']
            return commons[0]

        commons = [x for x in self.commons if x.menu_name.menu_name == menu]
        for common in commons:
            if common.allergen.allergen_name == name:
                self.logger.info(f'Allergen name:{name}(order:{common.seq_order})')
                return common

        # 見つからなかった場合、嚥下を常食に含める対応
        commons = [x for x in self.commons if x.menu_name.menu_name == menu_group]
        for common in commons:
            if common.allergen.allergen_name == name:
                self.logger.info(f'Allergen name:{name}(order:{common.seq_order})')
                return common

        # 通常ありえないはず
        self.logger.warning(f'Allergen name:{name}(menu:{menu}) common allergen not exist.')
        return None

    def add_common_for_kizawa(self, meal, quantity):
        kizawa_code = settings.KIZAWA_RAKUKON_CODE

        # 食種が設定で未定義の場合は、個別集計しない
        if not kizawa_code:
            return

        matches = [x for x in self.quantity_list_common if x.code == kizawa_code]
        if matches:
            counts = matches[0].counts
            if meal == "朝食":
                counts[0] += quantity
            elif meal == "昼食":
                counts[1] += quantity
            elif meal == "夕食":
                counts[2] += quantity

    def add_common(self, name, menu, menu_group, meal, quantity):
        # 頻発アレルギーは、らくらく献立アレルギー名の比較は行わない(CommonAllergenのcodeに正しい短縮名が入っていればよい)

        if name == "フリーズ":
            freeze_count = self.quantity_list_common[0].counts
            if meal == "朝食":
                freeze_count[0] += quantity
            elif meal == "昼食":
                freeze_count[1] += quantity
            elif meal == "夕食":
                freeze_count[2] += quantity
        else:
            common = self._get_common_allergen(name, menu, menu_group)
            if common:
                self.logger.info(f'Add Common:{common.name}(order:{common.seq_order})')
                matches = [x for x in self.quantity_list_common if x.code == common.code]
                if matches:
                    counts = matches[0].counts
                    if meal == "朝食":
                        counts[0] += quantity
                    elif meal == "昼食":
                        counts[1] += quantity
                    elif meal == "夕食":
                        counts[2] += quantity
            else:
                # 対応する頻発アレルギーが存在しなかったら、取りこぼしのないよう、散発アレルギーとして集計しておく
                self.add_uncommon(name, menu, meal, quantity)

    def _add_uncommon_order_count(self, code, meal, quantity):
        matches = [x for x in self.quantity_list_uncommon if x.code == code]
        if matches:
            counts = matches[0].counts
            if meal == "朝食":
                counts[0] += quantity
            elif meal == "昼食":
                counts[1] += quantity
            elif meal == "夕食":
                counts[2] += quantity

    def add_uncommon(self, name, menu, meal, quantity):
        # らくらく献立側に散発アレルギーを使いまわしてもらう(=食種情報を更新してもらう)ため、
        # らくらく献立のアレルギー名で比較を行う
        allergen_list = AllergenMaster.objects.filter(allergen_name=name)
        if allergen_list:
            allergen = allergen_list[0]
            uncommon = self._get_uncommon_allergen(allergen, menu, meal, quantity)
            if uncommon:
                self._add_uncommon_order_count(uncommon.code, meal, quantity)
        else:
            # 本来システムに存在しないアレルギー名が渡ってくることはないため、
            # 画面からの操作では、ここには到達しない想定
            pass

    def change_uncommon(self):
        for allergen, menu, meal, quantity in self.preserved_uncommons:
            ucs = sorted(self.uncommons, key=lambda x: x.last_use_date)
            for uc in ucs:
                # 献立種類(常食/薄味/ソフト/ゼリー/ミキサー)が等しいものを取得
                if menu == uc.menu_name.menu_name:
                    # 散発アレルギーの切り替え(データの更新)

                    # 更新前の情報を退避
                    self.chenged_uncommons.append((uc.code, uc.name))

                    uc.name = allergen.get_rakukon_name()
                    uc.menu_name = MenuMaster.objects.get(menu_name=menu)
                    uc.allergen = allergen
                    uc.save()

                    # 食数の加算
                    self._add_uncommon_order_count(uc.code, meal, quantity)
                    break

    def get_autoinput(self):
        prefix_list = []    # フリーズの食数情報
        result_list = []

        # フリーズの食数情報(現在は使っていないが、アレルギーのフリーズを使うようになる時のため残している)
        freeze = self.quantity_list_common[0]
        prefix_list += [freeze.counts[0], freeze.counts[1], 0, freeze.counts[2]]

        # 散発アレルギー
        for x in self.quantity_list_uncommon:
            result_list += [x.counts[0], x.counts[1], 0, x.counts[2]]
        for _ in range(9):
            result_list += [0, 0, 0, 0]

        # 頻発アレルギー
        index = 10
        for common in self.quantity_list_common[1:]:
            for _ in range(common.seq - index):
                result_list += [0, 0, 0, 0]
            result_list += [common.counts[0], common.counts[1], 0, common.counts[2]]
            index = common.seq + 1
        return prefix_list, result_list

    def _get_uncommon_list_item(self, short_name):
        for uc in self.uncommons:
            if short_name == uc.code:
                return uc
        return None

    def _get_chenged_uncommon(self, short_name):
        for tpl in self.chenged_uncommons:
            if short_name == tpl[0]:
                return tpl
        return None

    def write_uncommon_list(self, worksheet, eating_date):
        worksheet.cell(3, 3, dt.datetime.now().date())
        worksheet.cell(4, 3, eating_date)

        for row in worksheet.iter_rows(min_row=7):
            short_name = row[2].value
            uc = self._get_uncommon_list_item(short_name)
            if uc:
                changed = self._get_chenged_uncommon(short_name)
                row[5].value = uc.last_use_date.date()
                if changed:
                    row[1].value = '●'
                    row[3].value = uc.name
                    row[4].value = changed[1]
                else:
                    row[3].value = uc.name
                    row[4].value = '-'


# ログファイルの「↓」はpd.concatで「→」はpd.mergeでの結合を示す

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('date', nargs='+', type=str)

    def write_csv_sjis(self, df, path):
        with open(path, mode="w",
                  encoding="cp932", errors="backslashreplace") as f:
            # Windwows環境だと、余分な改行(空白行)が挿入される?
            df.to_csv(f, index=False)

    def exclude_fzreeze_order(self, df):
        """
        Dataframeの中から、フリーズの注文を除く。フリーズは、ユニット名に「フリーズ」を含むものを指すものとする。
        """
        exclude_list = []
        i = 0
        for index, row in df.iterrows():
            if 'フリーズ' in row.unit_name:
                exclude_list.append(i)
            i += 1
        excluded = df.drop(index=df.index[exclude_list])
        return excluded

    def handle(self, *args, **options):

        #######################################################
        # らくらく献立に食数を入力するため、食種別に集計する処理（新しい方式）
        #######################################################

        # 集計日時を喫食日で指定する
        # in_date = '2022-04-01'
        in_date = options['date'][0]  # 呼び出し時の引数1つ目

        aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d')
        aggregation_day = aggregation_day.date()  # 時刻部分を除外

        enable_day = dt.datetime.strptime(settings.BASIC_PLATE_ENABLE_DATE, '%Y-%m-%d').date()
        basic_plate_enable = aggregation_day >= enable_day

        rakukon_output_dir = os.path.join(settings.OUTPUT_DIR, settings.RAKUKON_DIR)

        new_dir_path = os.path.join(rakukon_output_dir, str(aggregation_day) + '_食数集計表（新方式）')
        os.makedirs(new_dir_path, exist_ok=True)

        # 食数自動入力用ファイル
        auto_input_file = os.path.join(rakukon_output_dir, 'autoinput', str(aggregation_day) + '.txt')

        # 散発アレルギー一覧
        allergen_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/uncommon_allergen.xlsx')
        uncommon_aggregation_file = os.path.join(rakukon_output_dir, '散発アレルギー一覧_' + str(aggregation_day) + '.xlsx')

        aggregation_file = os.path.join(new_dir_path, '_agg_log_' + str(aggregation_day) + '.csv')

        aggregation_log = pd.DataFrame(index=[],
                                       columns=[str(aggregation_day) + ' 喫食分集計開始 ' + str(dt.datetime.now()),
                                                'レコード数',
                                                '食数'])

        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-1_DataFrameを食数入力用に処理する内容
        #
        # 注文テーブルから集計日の「アレルギーなし」「個食」「フリーズ」を取得
        # アレルギーマスタのID（1:なし,2:個食,3:フリーズ...）なので1を指定
        # ------------------------------------------------------------------------------

        # ------------------------------------------------------------------------------
        # 当日の全データ
        # ------------------------------------------------------------------------------
        qs_all = Order.objects\
            .filter(eating_day=aggregation_day, quantity__gt=0)\
            .values('unit_name__unit_number', 'unit_name',
                    'meal_name__meal_name', 'meal_name__soup', 'meal_name__filling',
                    'menu_name', 'menu_name__group',
                    'allergen', 'quantity', 'eating_day')\
            .exclude(unit_name__unit_code__range=[80001, 80008])
        df_all = read_frame(qs_all)

        df_all.to_csv(new_dir_path + "/A-1-1_注文データ_アレルギー込.csv", index=False)

        shokusu = df_all['quantity'].sum()
        aggregation_log.loc[len(aggregation_log)] = ['A-1-1_注文データ_アレルギー込', len(df_all), shokusu]


        # ------------------------------------------------------------------------------
        # 食数固定製造データ
        # ------------------------------------------------------------------------------

        # サンシティあい 検食用は、通常注文にも、専用のユニットとして存在する。
        # 2重計上されてしまうので、ここでは取得しないようにする
        qs_everyday = OrderEveryday.objects.all()\
            .values('unit_name__unit_number', 'unit_name',
                    'meal_name__meal_name', 'meal_name__soup', 'meal_name__filling',
                    'menu_name', 'menu_name__group',
                    'allergen', 'quantity', 'eating_day')\
            .exclude(unit_name__unit_name='サンシティあい 検食用')

        df_everyday = read_frame(qs_everyday)

        shokusu = df_everyday['quantity'].sum()


        df_everyday = df_everyday.fillna({'unit_name__unit_number': 999, 'eating_day': aggregation_day})
        df_everyday = df_everyday.astype({'unit_name__unit_number': 'int64'})

        df_everyday.to_csv(new_dir_path + "/A-1-pre2_食数固定製造分(置き換え前).csv", index=False)

        # 薄味の固定分を置き換え
        if basic_plate_enable:
            # 薄味読込
            drop_index_list = []
            for index, row in df_everyday.iterrows():
                if (row['unit_name__unit_number'] == 999) and (row['menu_name'] == '薄味'):
                    if row['unit_name'] == '保存':
                        df_everyday.loc[index, 'menu_name'] = '常食'
                        df_everyday.loc[index, 'menu_name__group'] = '常食'
                    elif row['unit_name'] == '保存1人袋':
                        drop_index_list.append(index)

            df_everyday = df_everyday.drop(index=df_everyday.index[drop_index_list])

        df_everyday.to_csv(new_dir_path + "/A-1-2_食数固定製造分.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['A-1-2_食数固定製造分', len(df_everyday), shokusu]


        # 結合した際はindexをリセットしないとforループの結果が変になる
        df_all_everyday = pd.concat([df_all, df_everyday]).reset_index()

        shokusu = df_all_everyday['quantity'].sum()

        df_all_everyday.to_csv(new_dir_path + "/A-1-3_注文データ＋食数固定.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['A-1-3_注文データ＋食数固定↓', len(df_all_everyday), shokusu]


        for index, row in df_all_everyday.iterrows():

            if row['meal_name__soup'] is True:
                df_all_everyday.loc[index, 'quantity_soup'] = row['quantity']
            else:
                df_all_everyday.loc[index, 'quantity_soup'] = 0

            if row['meal_name__filling'] is True:
                df_all_everyday.loc[index, 'quantity_filling'] = row['quantity']
            else:
                df_all_everyday.loc[index, 'quantity_filling'] = 0

        df_all_everyday = df_all_everyday.astype({'quantity_soup': 'int64'})
        df_all_everyday = df_all_everyday.astype({'quantity_filling': 'int64'})

        df_all_everyday.to_csv(new_dir_path + "/A-1-4_注文データ＋食数固定_味噌汁食数列追加.csv", index=False)

        # ------------------------------------------------------------------------------
        # 食数内訳の出力(汁具は常食/薄味共通)
        # ファイル名先頭2文字はソートのための数字(1桁目：食種毎に1アップ、2桁目：朝昼夕の順にカウント)
        # フリーズの食数は除く。(フリーズ判定の仕様はexclude_fzreeze_order()参照)
        # フリーズは常食のみの想定
        # 個食のユニットでアレルギーの注文はありうる。個食とアレルギーはそれぞれでカウントする(どちらか一方ではない)
        # ------------------------------------------------------------------------------
        df_without_allergen = df_all_everyday[df_all_everyday.allergen == "なし"]
        df_kizawa = df_all_everyday[(df_all_everyday.unit_name == "木沢・個食1") | (df_all_everyday.unit_name == "木沢・個食2") ]
        df_all_everyday = df_all_everyday[~(df_all_everyday.unit_name == "木沢・個食1") & ~(df_all_everyday.unit_name == "木沢・個食2") ]

        # 常食-朝
        self.write_csv_sjis(self.exclude_fzreeze_order(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "朝食") &
            (df_all_everyday.menu_name__group == "常食")]), new_dir_path + "/01_常食_朝食_内訳.csv")

        # 常食-昼
        self.write_csv_sjis(self.exclude_fzreeze_order(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "昼食") &
            (df_all_everyday.menu_name__group == "常食")]), new_dir_path + "/02_常食_昼食_内訳.csv")

        # 常食-夕
        self.write_csv_sjis(self.exclude_fzreeze_order(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "夕食") &
            (df_all_everyday.menu_name__group == "常食")]), new_dir_path + "/03_常食_夕食_内訳.csv")

        # 汁具-朝
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "朝食") &
            (df_without_allergen.meal_name__filling)], new_dir_path + "/10_汁具_朝食_内訳.csv")

        # 汁具-昼
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "昼食") &
            (df_without_allergen.meal_name__filling)], new_dir_path + "/11_汁具_昼食_内訳.csv")

        # 汁具-夕
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "夕食") &
            (df_without_allergen.meal_name__filling)], new_dir_path + "/12_汁具_夕食_内訳.csv")

        # 常食-汁-朝
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "朝食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "常食")], new_dir_path + "/20_常食_汁_朝食_内訳.csv")

        # 常食-汁-昼
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "昼食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "常食")], new_dir_path + "/21_常食_汁_昼食_内訳.csv")

        # 常食-汁-夕
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "夕食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "常食")], new_dir_path + "/22_常食_汁_夕食_内訳.csv")

        # 薄味-朝
        self.write_csv_sjis(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "朝食") &
            (df_all_everyday.menu_name__group == "薄味")], new_dir_path + "/30_薄味_朝食_内訳.csv")

        # 薄味-昼
        self.write_csv_sjis(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "昼食") &
            (df_all_everyday.menu_name__group == "薄味")], new_dir_path + "/31_薄味_昼食_内訳.csv")

        # 薄味-夕
        self.write_csv_sjis(df_all_everyday[
            (df_all_everyday.meal_name__meal_name == "夕食") &
            (df_all_everyday.menu_name__group == "薄味")], new_dir_path + "/32_薄味_夕食_内訳.csv")

        # 薄味-汁-朝
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "朝食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "薄味")], new_dir_path + "/40_薄味_汁_朝食_内訳.csv")

        # 薄味-汁-昼
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "昼食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "薄味")], new_dir_path + "/41_薄味_汁_昼食_内訳.csv")

        # 薄味-汁-夕
        self.write_csv_sjis(df_without_allergen[
            (df_without_allergen.meal_name__meal_name == "夕食") &
            (df_without_allergen.meal_name__soup) &
            (df_without_allergen.menu_name__group == "薄味")], new_dir_path + "/42_薄味_汁_夕食_内訳.csv")

        # 木沢個食
        self.write_csv_sjis(df_kizawa, new_dir_path + "/50_木沢個食_内訳.csv")

        aggregation_log.loc[len(aggregation_log)] = ['A-1-4_注文データ＋食数固定↓_味噌汁食数列追加', len(df_all_everyday), '']
        df_without_allergen.to_csv(new_dir_path + "/A-1-5_注文データ＋食数固定_味噌汁食数列追加_アレルギーなし.csv", index=False)


        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-2_らくらく献立に入力する食数・全施設総数（朝・昼・夕・間食）
        #
        # 献立種類グループ（ソフト・ミキサー・ゼリーは常食グループ）と食事区分で全施設分を集計
        # ------------------------------------------------------------------------------

        # フリーズの集計(現在はユニットとして換算する仕様)
        f_b = f_l = f_k = f_d = 0
        for index, row in df_all_everyday.iterrows():
            name = row['unit_name']
            if name.find("フリーズ") != -1:
                if row['meal_name__meal_name'] == "朝食":
                    f_b += row['quantity']
                elif row['meal_name__meal_name'] == "昼食":
                    f_l += row['quantity']
                elif row['meal_name__meal_name'] == "間食":
                    f_k += row['quantity']
                elif row['meal_name__meal_name'] == "夕食":
                    f_d += row['quantity']
        freeze_input = [f_b, f_l, f_k, f_d]

        df_all_everyday = df_all_everyday.groupby(['menu_name__group', 'meal_name__meal_name']).sum().reset_index()
        df_without_allergen = df_without_allergen.groupby(['menu_name__group', 'meal_name__meal_name']).sum().reset_index()

        shokusu = df_all_everyday['quantity'].sum()
        shokusu_g = df_without_allergen['quantity_filling'].sum()
        shokusu_s = df_without_allergen['quantity_soup'].sum()

        df_all_everyday.to_csv(new_dir_path + "/A-2_注文データ_集計後.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['A-2_注文データ_集計後', len(df_all_everyday), shokusu]
        aggregation_log.loc[len(aggregation_log)] = ['味噌汁の具 総数', '', shokusu_g]
        aggregation_log.loc[len(aggregation_log)] = ['味噌汁の汁 総数', '', shokusu_s]


        '''
        らくらく献立の食事区分
        ---------------------------
        常食・朝食（j_b）
        常食・昼食（j_l）
        常食・間食（j_k）
        常食・夕食（j_d）
        ---------------------------
        薄味・朝食（u_b）
        薄味・昼食（u_l）
        薄味・間食（u_k）
        薄味・夕食（u_d）
        ---------------------------
        味噌汁の具は常食・薄味同じなので合算

        味噌汁の具・朝食（j_b_g + u_b_g）
        味噌汁の具・昼食（j_l_g + u_l_g）
        味噌汁の具・間食（存在しないので食数は常に0）
        味噌汁の具・夕食（j_d_g + u_d_g）
        ---------------------------
        味噌汁の汁・常食・朝食（j_b_s）
        味噌汁の汁・常食・昼食（j_l_s）
        味噌汁の汁・常食・間食（存在しないので食数は常に0）
        味噌汁の汁・常食・夕食（j_d_s）
        ---------------------------
        味噌汁の汁・薄味・朝食（u_b_s）
        味噌汁の汁・薄味・昼食（u_l_s）
        味噌汁の汁・薄味・間食（存在しないので食数は常に0）
        味噌汁の汁・薄味・夕食（u_d_s）
        ---------------------------
        常食（j）薄味（u） 朝食（b）昼食（l）間食（k）夕食（d） 具（g）汁（s）
        '''

        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-3_らくらく献立に入力時の並び順
        #
        # らくらく献立に自動入力する順番に合わせて、集計したものを「朝・昼・間・夕」に並び替える
        # ------------------------------------------------------------------------------

        j_b = j_l = j_k = j_d = 0
        j_b_g = j_l_g = j_d_g = 0
        j_b_s = j_l_s = j_d_s = 0

        u_b = u_l = u_k = u_d = 0
        u_b_g = u_l_g = u_d_g = 0
        u_b_s = u_l_s = u_d_s = 0

        for index, row in df_all_everyday.iterrows():
            if row['menu_name__group'] == "常食":
                if row['meal_name__meal_name'] == "朝食":
                    j_b = row['quantity']               # 常食・朝食
                elif row['meal_name__meal_name'] == "昼食":
                    j_l = row['quantity']               # 常食・昼食
                elif row['meal_name__meal_name'] == "間食":
                    j_k = row['quantity']               # 常食・間食
                elif row['meal_name__meal_name'] == "夕食":
                    j_d = row['quantity']               # 常食・夕食

            elif row['menu_name__group'] == "薄味":
                if row['meal_name__meal_name'] == "朝食":
                    u_b = row['quantity']               # 薄味・朝食
                elif row['meal_name__meal_name'] == "昼食":
                    u_l = row['quantity']               # 薄味・昼食
                elif row['meal_name__meal_name'] == "間食":
                    u_k = row['quantity']
                elif row['meal_name__meal_name'] == "夕食":
                    u_d = row['quantity']               # 薄味・夕食

        for index, row in df_all_everyday.iterrows():
            if row['menu_name__group'] == "常食":
                if row['meal_name__meal_name'] == "朝食":
                    j_b_g = row['quantity_filling']   # 常食・朝食・具
                    j_b_s = row['quantity_soup']      # 常食・朝食・汁
                elif row['meal_name__meal_name'] == "昼食":
                    j_l_g = row['quantity_filling']   # 常食・昼食・具
                    j_l_s = row['quantity_soup']      # 常食・昼食・汁
                elif row['meal_name__meal_name'] == "夕食":
                    j_d_g = row['quantity_filling']   # 常食・夕食・具
                    j_d_s = row['quantity_soup']      # 常食・夕食・汁

            elif row['menu_name__group'] == "薄味":
                if row['meal_name__meal_name'] == "朝食":
                    u_b_g = row['quantity_filling']   # 薄味・朝食・具
                    u_b_s = row['quantity_soup']      # 薄味・朝食・汁
                elif row['meal_name__meal_name'] == "昼食":
                    u_l_g = row['quantity_filling']   # 薄味・昼食・具
                    u_l_s = row['quantity_soup']      # 薄味・昼食・汁
                elif row['meal_name__meal_name'] == "夕食":
                    u_d_g = row['quantity_filling']   # 薄味・夕食・具
                    u_d_s = row['quantity_soup']      # 薄味・夕食・汁

        b_g = j_b_g + u_b_g
        l_g = j_l_g + u_l_g
        d_g = j_d_g + u_d_g

        # 常食の食数にはフリーズを含んだいるので、除いた数字を出力
        auto_input_j = [j_b - f_b, j_l - f_l, j_k - f_k, j_d - f_d]         # 常食の朝・昼・間・夕(フリーズ分を減算する)
        auto_input_u = [u_b, u_l, u_k, u_d]         # 薄味の朝・昼・間・夕
        auto_input_g = [b_g, l_g, 0, d_g]           # 味噌汁の具の朝・昼・間・夕
        auto_input_js = [j_b_s, j_l_s, 0, j_d_s]    # 常食の味噌汁の朝・昼・間・夕
        auto_input_us = [u_b_s, u_l_s, 0, u_d_s]    # 薄味の味噌汁の朝・昼・間・夕


        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-1_DataFrameを食数入力用に処理する内容
        #
        # 注文テーブルから集計日の「アレルギーあり」のレコードを取得
        # アレルギーマスタのID（1:なし,2:個食,3:フリーズ....）なので2以上を指定
        # ------------------------------------------------------------------------------

        qs_allergen = qs_all.filter(allergen_id__gte=2)
        df_allergen = read_frame(qs_allergen)
        df_allergen.to_csv(new_dir_path + "/A-4_注文_アレルギーあり.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['A-4_注文_アレルギーあり', len(df_allergen), '']


        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-5_らく献に入力する食数・全施設総数（アレルギー）
        #
        # 献立種類グループ（ソフト・ミキサー・ゼリーは常食グループ）とアレルギーで全施設分を集計。
        # 食事区分（朝昼夕）はひとまず個別に集計される
        # ------------------------------------------------------------------------------

        df_allergen = df_allergen.groupby(['menu_name__group',
                                           'menu_name',
                                           'allergen',
                                           'meal_name__meal_name']).sum().reset_index()
        df_allergen.to_csv(new_dir_path + "/A-5-1_集計_アレルギーあり.csv", index=False)


        # ------------------------------------------------------------------------------
        # 献立種類の並び順を結合するためにマスタデータを取り出す
        # ------------------------------------------------------------------------------
        qs_menu_seq = MenuMaster.objects.all().values('menu_name', 'seq_order')
        df_menu_seq = read_frame(qs_menu_seq)

        # 結合するために列名を変更
        df_menu_seq = df_menu_seq.rename(columns={'menu_name': 'menu_name__group'})
        df_menu_seq.to_csv(new_dir_path + "/A-5-2_献立種類順.csv", index=False)

        # 結合
        allergen_menu_seq = pd.merge(df_allergen, df_menu_seq, on='menu_name__group', how='left')


        # ------------------------------------------------------------------------------
        # アレルギーの並び順を結合するためにマスタデータを取り出す
        # ------------------------------------------------------------------------------
        qs_allergen_seq = AllergenMaster.objects.all().values('allergen_name', 'seq_order', 'is_common')
        df_allergen_seq = read_frame(qs_allergen_seq)

        # 結合するために列名を変更
        df_allergen_seq = df_allergen_seq.rename(columns={'allergen_name': 'allergen'})
        df_allergen_seq.to_csv(new_dir_path + "/A-5-3_アレルギー順.csv", index=False)

        # 結合
        allergen_seq = pd.merge(allergen_menu_seq, df_allergen_seq, on='allergen', how='left')

        allergen_seq = allergen_seq.sort_values(['seq_order_y', 'allergen', 'seq_order_x'])

        allergen_seq.to_csv(new_dir_path + "/A-5-4_集計_アレルギーあり_ソート.csv", index=False)


        # ------------------------------------------------------------------------------
        # 設計図： 食数入力から調理表へ
        # データの流れ図： A-6_らくらく献立に入力時の並び順（その日に対応するアレルギー分）
        #
        # 献立種類グループとアレルギー名が同じものを、らくらく献立の食種名「アレルギー1〜20」として朝昼夕の食数内訳を保持する
        # ------------------------------------------------------------------------------

        auto_input_a = []  # 散発アレルギー+頻発アレルギーの朝・昼・夕の食数を連結したもの

        analyzer = AllergenAnalyzer()
        for index, row in allergen_seq.iterrows():
            # if row['menu_name__group'] == menu_prev and row['allergen'] == allergen_prev:
            # ------------------------------------------------------------------
            # 献立種類とアレルギーの組み合わせが「同じ場合」は食数のみ反映させる
            # ------------------------------------------------------------------
            if row['is_common']:
                analyzer.add_common(row['allergen'], row['menu_name'], row['menu_name__group'], row['meal_name__meal_name'], row['quantity'])
            else:
                analyzer.add_uncommon(row['allergen'], row['menu_name'], row['meal_name__meal_name'], row['quantity'])
        analyzer.change_uncommon()

        # 木沢個食の集計(頻発アレルギーとして登録する)
        for index, row in df_kizawa.iterrows():
            analyzer.add_common_for_kizawa(row['meal_name__meal_name'], row['quantity'])

        prefix_auto_input_a, auto_input_a = analyzer.get_autoinput()

        df_allergen_input = pd.DataFrame(prefix_auto_input_a + auto_input_a)
        df_allergen_input.to_csv(new_dir_path + "/A-6_らく献_アレルギー.csv", header=False, index=False)

        wb = excel.load_workbook(allergen_template)
        analyzer.write_uncommon_list(wb['アレルギー一覧'], aggregation_day)
        wb.save(uncommon_aggregation_file)

        # ------------------------------------------------------------------
        # 食数自動入力用ファイルの出力
        # ------------------------------------------------------------------
        auto_input = []
        interval_array = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                          0, 0, 0, 0, 0, 0, 0, 0, ]

        if freeze_input:
            auto_input += freeze_input          # ﾌﾘｰｽﾞの朝・昼・間・夕
            auto_input += [0, 0, 0, 0, ]        # 食種1つ分
        else:
            auto_input += [0, 0, 0, 0, 0, 0, 0, 0, ]    # 食種2つ分

        auto_input += auto_input_j     # 常食の朝・昼・間・夕
        auto_input += auto_input_g     # 味噌汁の具の朝・昼・間・夕
        auto_input += auto_input_js    # 常食の味噌汁の朝・昼・間・夕

        auto_input += interval_array   # 帳票出力などで使っている区分（12種類ｘ朝・昼・間・夕の48行）

        auto_input += auto_input_u     # 薄味の朝・昼・間・夕
        auto_input += auto_input_us    # 薄味の味噌汁の朝・昼・間・夕
        auto_input += [0, 0, 0, 0, ]   # P7アレルギー献立作成用
        auto_input += auto_input_a     # アレルギー

        df_auto_input = pd.DataFrame(auto_input)
        # ローカルで実行する際に参照できるようmediaフォルダに書き出す

        df_auto_input.to_csv(auto_input_file, header=False, index=False)

        df_auto_input.to_csv(new_dir_path + "/A-3_らく献_自動入力.csv", header=False, index=False)

        # ------------------------------------------------------------------------------
        # ローカル端末での自動処理へ
        # ------------------------------------------------------------------------------

        # ログを出力
        aggregation_log.to_csv(aggregation_file, index=False, mode='a')
        shutil.make_archive(new_dir_path, 'zip', root_dir=new_dir_path)
