import logging

import math
import os
import shutil
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP
import platform
import openpyxl as excel

from django.conf import settings
from django_pandas.io import read_frame
from django.core.management.base import BaseCommand

from web_order.models import Order, OrderEveryday, PlatePackageForPrint, UnitPackage, TmpPlateNamePackage
from .utils import AggEngePackageMixin, AggFixedOrderRule, ExcelOutputMixin
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister, RawPlatePackageRegisterFactory

logger = logging.getLogger(__name__)

"""
    喫食日の食数を集計し、計量表を出力する処理

# 引数
    製造日（YYYY-MM-DD）
    喫食日（YYYY-MM-DD）
    食事区分（朝・昼・夕）
    料理名
    分量
    スープ名
    スープ分量

# 出力ファイル
    朝食の汁・スープの計量表
    昼食の汁・スープの計量表
    夕食の汁・スープの計量表
"""

class OtherSoupDevideMeasureWriter(AggEngePackageMixin, ExcelOutputMixin):

    def __init__(self):
        self.package_size = None
        self.enge_package_size = None

        # 販売固定商品注文数
        # (マスタ定義の存在するもの)
        # -保存用1人用
        self.pre_1pack_j = 0
        self.pre_1pack_u = 0

        # -見本・写真用
        self.photo = 0

        # -保存用1人用
        self.pre_1pack_s = 0
        self.pre_1pack_z = 0
        self.pre_1pack_m = 0

        # -保存用50g
        self.pre_50g_j = 0
        self.pre_50g_s = 0
        self.pre_50g_z = 0
        self.pre_50g_m = 0

        self.in_index = 0
        self.in_cooking_day = None
        self.aggregation_day = None
        self.in_menu = None
        self.is_reminder_output = False
        self.in_adjust = 0
        self.agg_day_short = None
        self.basic_plate_enable = False

    def get_fixed_1pack_count(self, is_soup: bool = False):
        if is_soup:
            # 味噌汁の汁の場合は、写真用が不要
            return self.pre_1pack_j + 1
        else:
            return self.pre_1pack_j + self.photo

    def get_enge_50g_count(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.pre_50g_s
        elif menu_name == 'ゼリー':
            return self.pre_50g_z
        else:
            return self.pre_50g_m

    def create_excel_filling(self,
            dataframe, worksheet, meal, is_enge, parts_name, quantity, unit, needle, preserve, preserve_1pack, photo, enge_list=[]):
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)
        worksheet.cell(2, 2, parts_name)

        total_count = 0                # 食数集計値(嚥下用)
        if not is_enge:
            worksheet.cell(2, 5, quantity)
        if is_enge:
            i = 10
            worksheet.cell(2, 6, unit)
            worksheet.cell(5, 4, needle)  # 針刺し用の食数
            total_count = needle  # 食数集計値

            # 50g用は後で出力
            worksheet.cell(7, 7, preserve)  # 保存用の食数
            # 嚥下は内容量を出力しない。fixed_packも常にNone
            total_count += preserve

            print_package_count_1p = preserve
            worksheet.cell(7, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)

            print_package_count = 0  # 針刺し用
        else:
            i = 13
            worksheet.cell(2, 6, unit)

            if self.basic_plate_enable:
                worksheet.cell(6, 7, 5)  # 保存用(10人袋)数量

            # 食数固定分の出力
            # 汁具の針刺し用、保存用は一定のため、テンプレートで対応
            worksheet.cell(7, 7, self.pre_50g_j)  # 保存用(50g)
            worksheet.cell(8, 7, preserve_1pack)  # 保存用(1人前袋)
            worksheet.cell(9, 7, photo)  # 写真用

            print_package_count_1p = preserve_1pack + photo
            worksheet.cell(7, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)

            print_package_count = 2 + 5  # 保存用

            # enge_listは、ソフト、ミキサー、ゼリーの順に並んでいる前提
            if enge_list:
                worksheet.cell(10, 4, enge_list[0])  # ソフト
                worksheet.cell(11, 4, enge_list[1])  # ミキサー
                worksheet.cell(12, 4, enge_list[2])  # ゼリー
        # 袋サイズ
        if is_enge:
            package_size = self.get_miso_soup_package_size(unit, quantity)
            menu = ''
            if 'ソフト' in worksheet.title:
                menu = 'ソフト'
            elif 'ゼリー' in worksheet.title:
                menu = 'ゼリー'
            elif 'ミキサー' in worksheet.title:
                menu = 'ミキサー'

            count_50g = self.get_enge_50g_count(menu)
            worksheet.cell(6, 7, count_50g)  # 保存用(50g)の食数
            print_package_count_50g = count_50g
            total_count += count_50g
        else:
            menu = '常食'
            package_size = 30
            print_package_count_50g = self.pre_50g_j
        self.package_size = package_size

        logger.debug(f'集計前{menu}:{print_package_count}-{print_package_count_1p}')
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            if is_enge or (row['注文数'] > 1):
                worksheet.cell(i, 4, row['注文数'])
                total_count += row['注文数']

                if row['注文数'] > 1:
                    unit_package_count = math.ceil(row['注文数'] / package_size)
                    print_package_count += unit_package_count
                    logger.debug(
                        f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')

                    if is_enge:
                        if row['注文数'] == 2:
                            PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                       self.in_index,
                                                                       row['呼出番号'], row['ユニット名'], 1, 'ENGE_2',
                                                                       parts_name, menu, soup_type='filling')
                        else:
                            PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                       self.in_index,
                                                                       row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                       f'ENGE_{package_size}',
                                                                       parts_name, menu, soup_type='filling')
                    else:
                        if unit == '個':
                            PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                       row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                       'SOUP_UNIT', parts_name, soup_type='filling')

                if row['注文数'] == 1:
                    print_package_count_1p += 1
                    logger.debug(
                        f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')
                    if is_enge:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], 1, 'ENGE_1', parts_name,
                                                                   menu, soup_type='filling')
                    else:
                        if unit == '個':
                            PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                       row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', parts_name,
                                                                       soup_type='filling')
            else:
                worksheet.cell(i, 4, 0)
            if is_enge:
                worksheet.cell(i, 7, self.get_miso_soup_package_function(unit, quantity, i))
            i = i + 1

        logger.debug(f'集計後{menu}:{print_package_count}-{print_package_count_1p}')
        if is_enge:
            if self.in_adjust == 2:
                if self.in_index != -1:
                    enge_index = self.in_index + 1
                else:
                    enge_index = self.in_index
            else:
                enge_index = self.in_index

            if 'ソフト' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ソフト',
                                                    self.in_menu, print_package_count_50g)
            elif 'ゼリー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ゼリー',
                                                    self.in_menu, print_package_count_50g)
            elif 'ミキサー' in worksheet.title:
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count,
                                                    print_package_count_1p, 'ミキサー',
                                                    self.in_menu, print_package_count_50g)
        else:
            P7Util.save_package_count_for_print(
                self.in_cooking_day, self.aggregation_day, self.in_index, print_package_count, print_package_count_1p, '常食',
                self.in_menu, print_package_count_50g)

        return total_count

    def create_excel_soup(self,
            dataframe, worksheet, meal, is_enge, soup_name, quantity, unit, needle, preserve, preserve_1pack, photo):
        worksheet.cell(1, 4, self.agg_day_short)
        worksheet.cell(1, 6, meal)
        worksheet.cell(3, 2, soup_name)
        worksheet.cell(3, 4, quantity)
        if unit == '個':
            worksheet.cell(3, 5, 'g')
        else:
            worksheet.cell(3, 5, unit)

        total = 0
        if is_enge:
            i = 12
            worksheet.cell(6, 2, needle)  # 針刺し用の食数
            # 50g用は後で設定
            worksheet.cell(9, 5, preserve)  # 保存用の食数

            worksheet.cell(2, 4, self.get_gram_package(quantity))  # 袋サイズ(〇人前の設定)

        else:
            # 汁の濃さを0.8倍にする
            if '汁_薄味' in worksheet.title:
                worksheet.cell(3, 4, f'={quantity}*0.8')

            # 食数固定分の出力
            needle_pack, needle_count = needle
            preserve_pack, preserve_count = preserve
            if self.is_reminder_output:
                # 端数出力の食数固定分出力
                # 1人用でない保存用、針刺し用はテンプレートファイル上で設定
                i = 13
                worksheet.cell(8, 8, self.pre_50g_j)  # 保存用(50g-注文数)
                worksheet.cell(8, 9, 0)  # 保存用(50g-単位袋10)
                worksheet.cell(8, 10, 0)  # 保存用(50g-単位袋5)
                worksheet.cell(8, 11, self.pre_50g_j)  # 保存用(50g-単位袋1)
                worksheet.cell(9, 8, preserve_1pack)  # 保存用(1人前袋-注文数)
                worksheet.cell(9, 9, 0)  # 保存用(1人前袋-単位袋10)
                worksheet.cell(9, 10, 0)  # 保存用(1人前袋-単位袋5)
                worksheet.cell(9, 11, preserve_1pack)  # 保存用(1人前袋-単位袋1)
                worksheet.cell(10, 8, photo)  # 保存用(袋数)
                worksheet.cell(10, 9, 0)  # 保存用(単位袋10)
                worksheet.cell(10, 10, 0)  # 保存用(単位袋5)
                worksheet.cell(10, 11, photo)  # 保存用(単位袋1)

                if self.basic_plate_enable:
                    worksheet.cell(7, 9, 5)  # 保存用(10人袋)袋数
                    worksheet.cell(7, 10, 12)  # 保存用(5人袋)袋数

            else:
                # 施設毎出力の食数固定分出力
                i = 11
                worksheet.cell(6, 2, needle_pack)  # 針刺し用(袋サイズ)
                worksheet.cell(6, 5, needle_count)  # 針刺し用(袋数)
                worksheet.cell(7, 2, preserve_pack)  # 保存用(袋サイズ)
                worksheet.cell(7, 5, preserve_count)  # 保存用(袋数)
                worksheet.cell(9, 5, self.pre_50g_j)  # 保存用(1人前袋-袋数)
                worksheet.cell(8, 5, preserve_1pack)  # 保存用(1人前袋-袋数)
                worksheet.cell(9, 5, photo)  # 写真用(袋数)
                if self.basic_plate_enable:
                    worksheet.cell(7, 5, '=IF(B7=0,0,5)')  # 保存用(袋数)

        if is_enge:
            print_package_count = 0
            print_package_count_1p = preserve
            menu = ''
            if 'ソフト' in worksheet.title:
                menu = 'ソフト'
            elif 'ゼリー' in worksheet.title:
                menu = 'ゼリー'
            elif 'ミキサー' in worksheet.title:
                menu = 'ミキサー'

            count_50g = self.get_enge_50g_count(menu)
            worksheet.cell(8, 5, count_50g)  # 保存用の食数
            print_package_count_50g = count_50g
        else:
            menu = '常食'
            print_package_count = 2 + 5
            print_package_count_1p = photo
            print_package_count_50g = self.pre_50g_j

        # 袋サイズの取得
        if is_enge:
            package_size = self.get_miso_soup_package_size(unit, quantity)
        else:
            package_size = 30

        logger.debug(f'集計前{menu}:{print_package_count}-{print_package_count_1p}')
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            if self.is_reminder_output and (not is_enge):
                # 端数はテンプレートファイルのサイズ指定・計算式から設定する

                worksheet.cell(i, 8, row['注文数'])
                worksheet.cell(i, 9, row['単位袋10'])
                print_package_count += row['単位袋10']
                if self.in_index > -1:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index + 1,
                                                               row['呼出番号'], row['ユニット名'], row['単位袋10'], 'SOUP_10',
                                                               soup_name, soup_type='soup')
                worksheet.cell(i, 10, row['10の端数袋・入数'])
                if row['10の端数袋・入数'] > 1:
                    print_package_count += 1
                    logger.debug(
                        f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')

                    if self.in_index > -1:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                   self.in_index + 1,
                                                                   row['呼出番号'], row['ユニット名'], 1, 'SOUP_FRACTION',
                                                                   soup_name, soup_type='soup')
                worksheet.cell(i, 11, row['10の1人用袋'])
                if row['10の1人用袋'] > 0:
                    print_package_count_1p += row['10の1人用袋']
                    logger.debug(
                        f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')

                    if self.in_index > -1:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                   self.in_index + 1,
                                                                   row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', soup_name,
                                                                   soup_type='soup')
                worksheet.cell(i, 12, row['単位袋7'])
                worksheet.cell(i, 13, row['7の端数袋・入数'])
                worksheet.cell(i, 14, row['7の1人用袋'])
                worksheet.cell(i, 15, row['単位袋5'])
                worksheet.cell(i, 16, row['5の端数袋・入数'])
                worksheet.cell(i, 17, row['5の1人用袋'])
            else:
                if is_enge or (row['注文数'] > 1):
                    worksheet.cell(i, 4, row['注文数'])
                    if row['注文数'] > 1:
                        unit_package_count = math.ceil(row['注文数'] / package_size)
                        print_package_count += unit_package_count
                        logger.debug(
                            f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')

                        if is_enge:
                            if row['注文数'] == 2:
                                if self.in_index > -1:
                                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                               self.in_index + 1,
                                                                               row['呼出番号'], row['ユニット名'], 1, 'ENGE_2',
                                                                               soup_name, menu, soup_type='soup')
                            else:
                                if self.in_index > -1:
                                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                               self.in_index + 1,
                                                                               row['呼出番号'], row['ユニット名'],
                                                                               unit_package_count,
                                                                               f'ENGE_{package_size}',
                                                                               soup_name, menu, soup_type='soup')
                        else:
                            if self.in_index > -1:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index + 1,
                                                                           row['呼出番号'], row['ユニット名'],
                                                                           unit_package_count, 'SOUP_UNIT', soup_name,
                                                                           soup_type='soup')

                    if row['注文数'] == 1:
                        print_package_count_1p += 1
                        logger.debug(
                            f'集計中{menu}-{row["呼出番号"]}.{row["ユニット名"]}:{print_package_count}-{print_package_count_1p}')

                        if is_enge:
                            if self.in_index > -1:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index + 1,
                                                                           row['呼出番号'], row['ユニット名'], 1, 'ENGE_1',
                                                                           soup_name, menu, soup_type='soup')
                        else:
                            if self.in_index > -1:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index + 1,
                                                                           row['呼出番号'], row['ユニット名'], 1, 'SOUP_1',
                                                                           soup_name, soup_type='soup')
                else:
                    worksheet.cell(i, 4, 0)
            total += int(row['注文数'])

            i = i + 1

        logger.debug(f'集計前{menu}:{print_package_count}-{print_package_count_1p}')
        self.save_soup_package_count_for_print(print_package_count, print_package_count_1p, print_package_count_50g, menu)

        return total

    def get_fixed_quantity(self, id: int):
        qs_fix = OrderEveryday.objects.filter(id=id)
        if qs_fix.exists():
            return qs_fix.first().quantity
        else:
            return 0

    def save_soup_package_count_for_print(self, count: int, count_1p: int, count_50g: int, menu):
        if self.in_index == -1:
            return

        if self.in_adjust == 2:
            if self.in_index != -1:
                enge_index = self.in_index + 1
            else:
                enge_index = self.in_index
        else:
            enge_index = self.in_index

        # スープは、具の次になる前提
        plate_package_qs = PlatePackageForPrint.objects.filter(
            cooking_day=self.in_cooking_day, eating_day=self.aggregation_day,
            is_basic_plate=True, meal_name=self.in_menu, index=enge_index + 1, menu_name=menu)
        if plate_package_qs.exists():
            plate_package = plate_package_qs.first()
            plate_package.count = count
            plate_package.count_one_p = count_1p
            plate_package.count_one_50g = count_50g
            plate_package.save()

    def handle(self, manager, *args, **options):

        self.in_index = int(options['opt'][0])      # 呼び出し時の引数1つ目、調理表のインデックス(-1は対象不明)
        in_cook = options['opt'][1]         # 呼び出し時の引数2つ目、製造日（YYYY-MM-DD）
        in_date = options['opt'][2]         # 呼び出し時の引数3つ目、喫食日（YYYY-MM-DD）
        self.in_menu = options['opt'][3]         # 呼び出し時の引数4つ目、食事区分（朝・昼・夕）
        in_name = options['opt'][4]         # 呼び出し時の引数5つ目、献立名称
        in_name_ko = options.get('name1', '')           # 呼び出し時の引数 --name1 名前(個数の食材)
        in_qty_ko = options.get('quantity1', 0)        # 呼び出し時の引数 --quantity1 分量(個数の食材)
        in_name_g = options.get('name2', '')            # 呼び出し時の引数 --name2 名前(g数の食材)
        in_qty_g = options.get('quantity2', 0)         # 呼び出し時の引数 --quantity2 分量(g数の食材)
        in_soup_name = options.get('soup_name', '')     # 呼び出し時の引数 --soup_name スープ名称
        in_short_name = options.get('short_name', '')     # 呼び出し時の引数 --soup_name スープ名称
        in_soup_qty = options.get('soup_quantity', 0)  # 呼び出し時の引数 --quantity スープ分量
        self.is_reminder_output = options.get('output', 0) == 'reminder'    # 呼び出し時の引数 --output
        self.in_adjust = options.get('adjust', 0)           # 呼び出し時の引数 --adjust 嚥下袋数保存時のindexの補正の状況(0:補正なし、1:補正の必要正発生、2:補正発生済み)
        self.before_name = options.get('before_name', '')

        soup_fixed_order = AggFixedOrderRule(in_name, float(in_soup_qty), 'g', False)

        self.aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d').date()
        if platform.system() == 'Windows':
            self.agg_day_short = self.aggregation_day.strftime('%m/%d')
        else:
            self.agg_day_short = self.aggregation_day.strftime('%-m/%-d')

        enable_day = dt.datetime.strptime(settings.BASIC_PLATE_ENABLE_DATE, '%Y-%m-%d').date()
        if self.aggregation_day >= enable_day:
            self.basic_plate_enable = True
        else:
            self.basic_plate_enable = False

        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + in_cook + '_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        if self.is_reminder_output:
            # 味噌汁以外の汁・スープ用の計量表テンプレ(端数出力)
            agg_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_soup_remainder.xlsx')
        else:
            # 味噌汁以外の汁・スープ用の計量表テンプレ(施設毎出力)
            agg_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_soup.xlsx')

        in_unit_ko = '個'
        in_unit_g = 'g'

        # 製造日の取得
        self.in_cooking_day = dt.datetime.strptime(in_cook, '%Y-%m-%d').date()

        # 注文対象日・食事区分の設定
        manager.set_eating(self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # 針刺し用の食数
        # ------------------------------------------------------------------------------
        res_s, res_m, res_z = manager.get_needle_orders()


        # ------------------------------------------------------------------------------
        # 嚥下保存用(1人用)の食数
        # ------------------------------------------------------------------------------
        self.pre_1pack_s, self.pre_1pack_m, self.pre_1pack_z = manager.get_preserve_1p_enge_orders()

        # ------------------------------------------------------------------------------
        # 常食保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        self.pre_1pack_j, self.pre_1pack_u = manager.get_preserve_1p_orders()

        # ------------------------------------------------------------------------------
        # 写真用の食数
        # ------------------------------------------------------------------------------
        self.photo = manager.get_photo_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(50g)
        # ------------------------------------------------------------------------------
        self.pre_50g_j, self.pre_50g_s, self.pre_50g_z, self.pre_50g_m = manager.get_preserve_50g_orders()

        # ------------------------------------------------------------------------------
        # qs（Queryset）df（Dataframe）
        # 朝食（b）昼食（l）夕食（d）
        # 常食（j）薄味（u）ソフト（s）ミキサー（m）ゼリー（z）
        # 具（g） 汁（s）
        # ------------------------------------------------------------------------------
        # 嚥下用
        # soup_enge_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_soup_enge.xlsx')  # 味噌汁以外のスープ用の具(嚥下)・汁の計量表テンプレ

        book_bm = excel.load_workbook(agg_template)
        needle_j_fo = soup_fixed_order.get_needle_j()
        needle_u_fo = soup_fixed_order.get_needle_u()
        preserve_j_fo = soup_fixed_order.get_preserve_j()
        preserve_u_fo = soup_fixed_order.get_preserve_u()

        # 嚥下の具を出力
        if self.in_menu == '朝食':
            # 出力ファイル名
            breakfast_output_soup_enge = os.path.join(new_dir_path,
                                                      f'{str(self.aggregation_day)}_朝汁_{in_short_name}_{in_name_ko + "・" + in_name_g}.xlsx')
            book_bm.remove(book_bm["具_常食"])
            book_bm.remove(book_bm["具_薄味"])

            # 嚥下具の出力
            # ソフト食
            df_bs_g = manager.get_def_miso_soft_filling()
            bs_total = self.create_excel_filling(df_bs_g, book_bm["具_ソフト"], '△ 朝', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_s, self.pre_1pack_s, 0, 0)

            # ミキサー
            df_bm_g = manager.get_def_miso_mixer_filling()
            bm_total = self.create_excel_filling(df_bm_g, book_bm["具_ミキサー"], '△ 朝', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_m, self.pre_1pack_m, 0, 0)

            # ゼリー
            df_bz_g = manager.get_def_miso_jelly_filling()
            bz_total = self.create_excel_filling(df_bz_g, book_bm["具_ゼリー"], '△ 朝', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_z, self.pre_1pack_z, 0, 0)

            # 常食汁の出力(スープは通常g)
            df_bj_s = manager.get_def_other_soup()
            self.create_excel_soup(
                df_bj_s, book_bm["汁_常食"], '△ 朝', False, in_soup_name, in_soup_qty, 'g', needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo)

            # 嚥下汁の出力
            # ソフト
            df_bs_s = manager.get_def_other_soup_enge('ソフト')
            self.create_excel_soup(
                df_bs_s, book_bm["汁_ソフト"], '△ 朝', True, in_soup_name, in_soup_qty, 'g', res_s, self.pre_1pack_s, self.pre_1pack_j, self.photo)

            # ミキサー
            df_bm_s = manager.get_def_other_soup_enge('ミキサー')
            self.create_excel_soup(
                df_bm_s, book_bm["汁_ミキサー"], '△ 朝', True, in_soup_name, in_soup_qty, 'g', res_m, self.pre_1pack_m, self.pre_1pack_j, self.photo)

            # ゼリー
            df_bz_s = manager.get_def_other_soup_enge('ゼリー')
            self.create_excel_soup(
                df_bz_s, book_bm["汁_ゼリー"], '△ 朝', True, in_soup_name, in_soup_qty, 'g', res_z, self.pre_1pack_z, self.pre_1pack_j, self.photo)

            # 薄味汁の出力(現在は出力しない=puerysetがnot exists)
            """
            qs_bu_s = qs.filter(meal_name__meal_name='朝食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_bu_s.exists():
                df_bu_s = self.make_df_soup(qs_bu_s)
                self.create_excel_soup(
                    df_bu_s, book_bm["汁_薄味"], '△ 朝', False, in_soup_name, in_soup_qty, 'g', needle_u_fo, preserve_u_fo, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["汁_薄味"])
            """
            book_bm.remove(book_bm["汁_薄味"])

            self.save_with_select(book_bm, breakfast_output_soup_enge)

        if self.in_menu == '昼食':
            # 出力ファイル名
            lunch_output_soup_enge = os.path.join(new_dir_path,
                                                  f'{str(self.aggregation_day)}_昼汁_{in_short_name}_{in_name_ko + "・" + in_name_g}.xlsx')

            # 嚥下具の出力
            book_bm.remove(book_bm["具_常食"])
            book_bm.remove(book_bm["具_薄味"])

            # 嚥下具の出力
            # ソフト
            df_ls_g = manager.get_def_miso_soft_filling()
            ls_total = self.create_excel_filling(df_ls_g, book_bm["具_ソフト"], '○ 昼', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_s, self.pre_1pack_s, 0, 0)

            # ミキサー
            df_lm_g = manager.get_def_miso_mixer_filling()
            lm_total = self.create_excel_filling(df_lm_g, book_bm["具_ミキサー"], '○ 昼', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_m, self.pre_1pack_m, 0, 0)

            # ゼリー
            df_lz_g = manager.get_def_miso_jelly_filling()
            lz_total = self.create_excel_filling(df_lz_g, book_bm["具_ゼリー"], '○ 昼', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_z, self.pre_1pack_z, 0, 0)

            # 常食汁の出力(スープは通常g)
            df_lj_s = manager.get_def_other_soup()
            self.create_excel_soup(
                df_lj_s, book_bm["汁_常食"], '○ 昼', False, in_soup_name, in_soup_qty, 'g', needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo)

            # 嚥下汁の出力
            # ソフト
            df_ls_s = manager.get_def_other_soup_enge('ソフト')
            self.create_excel_soup(
                df_ls_s, book_bm["汁_ソフト"], '○ 昼', True, in_soup_name, in_soup_qty, 'g', res_s, self.pre_1pack_s, self.pre_1pack_j, self.photo)

            # ミキサー
            df_lm_s = manager.get_def_other_soup_enge('ミキサー')
            self.create_excel_soup(
                df_lm_s, book_bm["汁_ミキサー"], '○ 昼', True, in_soup_name, in_soup_qty, 'g', res_m, self.pre_1pack_m, self.pre_1pack_j, self.photo)

            # ゼリー
            df_lz_s = manager.get_def_other_soup_enge('ゼリー')
            self.create_excel_soup(
                df_lz_s, book_bm["汁_ゼリー"], '○ 昼', True, in_soup_name, in_soup_qty, 'g', res_z, self.pre_1pack_z, self.pre_1pack_j, self.photo)

            # 薄味汁の出力(現在は出力しない=puerysetがnot exists)
            """
            qs_lu_s = qs.filter(meal_name__meal_name='昼食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_lu_s.exists():
                df_lu_s = self.make_df_soup(qs_lu_s)
                self.create_excel_soup(df_lu_s, book_bm["汁_薄味"], '○ 昼', False, in_soup_name, in_soup_qty, in_unit, needle_u_fo, preserve_u_fo, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["汁_薄味"])
            """
            book_bm.remove(book_bm["汁_薄味"])

            self.save_with_select(book_bm, lunch_output_soup_enge)

        if self.in_menu == '夕食':
            # 出力ファイル名
            dinner_output_miso_enge = os.path.join(new_dir_path,
                                                   f'{str(self.aggregation_day)}_夕汁_{in_short_name}_{in_name_ko + "・" + in_name_g}.xlsx')

            book_bm.remove(book_bm["具_常食"])
            book_bm.remove(book_bm["具_薄味"])

            # 嚥下具の出力
            # ソフト
            df_ds_g = manager.get_def_miso_soft_filling()
            ds_total = self.create_excel_filling(df_ds_g, book_bm["具_ソフト"], '□ 夕', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_s, self.pre_1pack_s, 0, 0)

            # ミキサー
            df_dm_g = manager.get_def_miso_mixer_filling()
            dm_total = self.create_excel_filling(df_dm_g, book_bm["具_ミキサー"], '□ 夕', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_m, self.pre_1pack_m, 0, 0)

            # ゼリー
            df_dz_g = manager.get_def_miso_jelly_filling()
            dz_total = self.create_excel_filling(df_dz_g, book_bm["具_ゼリー"], '□ 夕', True,
                                 in_name_ko + '・' + in_name_g, in_qty_g, in_unit_g, res_z, self.pre_1pack_z, 0, 0)

            # 常食汁の出力(スープは通常g)
            df_dj_s = manager.get_def_other_soup()
            self.create_excel_soup(df_dj_s, book_bm["汁_常食"], '□ 夕', False, in_soup_name, in_soup_qty, 'g', needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo)

            # 嚥下汁の出力
            # ソフト
            df_ds_s = manager.get_def_other_soup_enge('ソフト')
            self.create_excel_soup(
                df_ds_s, book_bm["汁_ソフト"], '□ 夕', True, in_soup_name, in_soup_qty, 'g', res_s, self.pre_1pack_s, self.pre_1pack_j, self.photo)

            # ミキサー
            df_dm_s = manager.get_def_other_soup_enge('ミキサー')
            self.create_excel_soup(
                df_dm_s, book_bm["汁_ミキサー"], '□ 夕', True, in_soup_name, in_soup_qty, 'g', res_m, self.pre_1pack_m, self.pre_1pack_j, self.photo)

            # ゼリー
            df_dz_s = manager.get_def_other_soup_enge('ゼリー')
            self.create_excel_soup(
                df_dz_s, book_bm["汁_ゼリー"], '□ 夕', True, in_soup_name, in_soup_qty, 'g', res_z, self.pre_1pack_z, self.pre_1pack_j, self.photo)

            # 薄味汁の出力(現在は出力しない=puerysetがnot exists)
            """
            qs_du_s = qs.filter(meal_name__meal_name='夕食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_du_s.exists():
                df_du_s = self.make_df_soup(qs_du_s)
                self.create_excel_soup(df_du_s, book_bm["汁_薄味"], '□ 夕', False, in_soup_name, in_soup_qty, in_unit, needle_u_fo, preserve_u_fo, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["汁_薄味"])
            """
            book_bm.remove(book_bm["汁_薄味"])

            self.save_with_select(book_bm, dinner_output_miso_enge)
        # ------------------------------------------------------------------------------
        soup_filling_template = os.path.join(settings.STATICFILES_DIRS[0],
                                             'excel/measure_miso_filling.xlsx')  # 味噌汁以外のスープ用の具(常食・薄味)の計量表テンプレ(みそ汁と同じものを使う)
        # 具(個数)用
        book_bm = excel.load_workbook(soup_filling_template)
        if self.in_menu == '朝食':
            # 基本食(常食)-具
            df_bj_g = manager.get_def_miso_filling()
            self.create_excel_filling(
                df_bj_g, book_bm["具_常食"], '△ 朝', False, in_name_ko, in_qty_ko, in_unit_ko, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [bs_total, bm_total, bz_total])

            # 薄味
            """
            qs_bu_g = qs.filter(meal_name__meal_name='朝食', menu_name=2, meal_name__filling=True)  # 薄味 具あり
            if qs_bu_g.exists():
                df_bu_g = self.make_df_filling(qs_bu_g)
                #create_excel_miso_from_filling_template(df_bu_g, book_bm["具_薄味"], '△ 朝', in_name_ko, in_qty_ko, in_unit_ko, pre_1pack_u_m, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            breakfast_output_soup_ko = os.path.join(new_dir_path,
                                                    f'{str(self.aggregation_day)}_朝汁_{in_short_name}_{in_name_ko}.xlsx')
            self.save_with_select(book_bm, breakfast_output_soup_ko)

        if self.in_menu == '昼食':
            # 基本食(常食)-具
            df_lj_g = manager.get_def_miso_filling()
            self.create_excel_filling(
                df_lj_g, book_bm["具_常食"], '○ 昼', False, in_name_ko, in_qty_ko, in_unit_ko, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [ls_total, lm_total, lz_total])

            # 薄味
            """
            qs_lu_g = qs.filter(meal_name__meal_name='昼食', menu_name=2, meal_name__filling=True)
            if qs_lu_g.exists():
                df_lu_g = self.make_df_filling(qs_lu_g)
                # create_excel_miso_from_filling_template(df_lu_g, book_bm["具_薄味"], '○ 昼', in_name_ko, in_qty_ko, in_unit_ko, pre_1pack_u_l, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            lunch_output_soup_ko = os.path.join(new_dir_path,
                                                f'{str(self.aggregation_day)}_昼汁_{in_short_name}_{in_name_ko}.xlsx')
            self.save_with_select(book_bm, lunch_output_soup_ko)

        if self.in_menu == '夕食':
            # 基本食(常食)-具
            df_dj_g = manager.get_def_miso_filling()
            self.create_excel_filling(
                df_dj_g, book_bm["具_常食"], '□ 夕', False, in_name_ko, in_qty_ko, in_unit_ko, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [ds_total, dm_total, dz_total])

            # 薄味
            """
            qs_du_g = qs.filter(meal_name__meal_name='夕食', menu_name=2, meal_name__filling=True)
            if qs_du_g.exists():
                df_du_g = self.make_df_filling(qs_du_g)
                # create_excel_miso_from_filling_template(df_du_g, book_bm["具_薄味"], '□ 夕', in_name_ko, in_qty_ko, in_unit_ko, pre_1pack_u_d, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            dinner_output_soup_ko = os.path.join(new_dir_path,
                                                 f'{str(self.aggregation_day)}_夕汁_{in_short_name}_{in_name_ko}.xlsx')
            self.save_with_select(book_bm, dinner_output_soup_ko)
        # ------------------------------------------------------------------------------
        # g用
        book_bm = excel.load_workbook(soup_filling_template)
        if self.in_menu == '朝食':
            self.create_excel_filling(
                df_bj_g, book_bm["具_常食"], '△ 朝', False, in_name_g, in_qty_g, in_unit_g, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [bs_total, bm_total, bz_total])

            # 薄味
            """
            if qs_bu_g.exists():
                #create_excel_miso_from_filling_template(df_bu_g, book_bm["具_薄味"], '△ 朝', in_name_g, in_qty_g, in_unit_g, pre_1pack_u_m, 0)
                pass
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            breakfast_output_miso_g = os.path.join(new_dir_path,
                                                   f'{str(self.aggregation_day)}_朝汁_{in_short_name}_{in_name_g}.xlsx')
            self.save_with_select(book_bm, breakfast_output_miso_g)

        if self.in_menu == '昼食':
            self.create_excel_filling(
                df_lj_g, book_bm["具_常食"], '○ 昼', False, in_name_g, in_qty_g, in_unit_g, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [ls_total, lm_total, lz_total])

            # 薄味
            """
            if qs_lu_g.exists():
                # create_excel_miso_from_filling_template(df_lu_g, book_bm["具_薄味"], '○ 昼', in_name_g, in_qty_g, in_unit_g, pre_1pack_u_l, 0)
                pass
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            lunch_output_miso_g = os.path.join(new_dir_path,
                                               f'{str(self.aggregation_day)}_昼汁_{in_short_name}_{in_name_g}.xlsx')
            self.save_with_select(book_bm, lunch_output_miso_g)

        if self.in_menu == '夕食':
            self.create_excel_filling(
                df_dj_g, book_bm["具_常食"], '□ 夕', False, in_name_g, in_qty_g, in_unit_g, needle_j_fo, preserve_j_fo, self.pre_1pack_j, self.photo, [ds_total, dm_total, dz_total])

            # 薄味
            """
            if qs_du_g.exists():
                #create_excel_miso_from_filling_template(df_du_g, book_bm["具_薄味"], '□ 夕', in_name_g, in_qty_g, in_unit_g, pre_1pack_u_d, 0)
                pass
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            dinner_output_miso_g = os.path.join(new_dir_path,
                                                f'{str(self.aggregation_day)}_夕汁_{in_short_name}_{in_name_g}.xlsx')
            self.save_with_select(book_bm, dinner_output_miso_g)
        # ------------------------------------------------------------------------------

        if self.in_index == -1:
            if self.package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=self.before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.package_size,
                    menu_name='常食'
                )
            if self.enge_package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=self.before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.enge_package_size,
                    menu_name='嚥下'
                )
