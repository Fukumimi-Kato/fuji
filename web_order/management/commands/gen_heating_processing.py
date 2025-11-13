import os
import datetime as dt
import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
import platform
import re
import shutil

from django.core.management.base import BaseCommand
from django.conf import settings

from .utils import ExcelOutputMixin, ExcelHellper

class AllergenPlateDetail:
    def __init__(self, name:str, menu_list:list):
        self.name = name
        self.menu_list = menu_list

    def has_menu(self, menu: str):
        return menu in self.menu_list

class CookPlate:
    """
    料理情報
    """
    def __init__(self, eating_day):
        self.eating_day = eating_day.replace('■', '')
        self.normal_plates = []
        self.allergen_plates = []

    def _is_raw_plate(self, raw_name: str):
        return '原体' in raw_name

    def _exists_allergen_plate(self, name: str):
        for x in self.allergen_plates:
            if x.name == name:
                return True
        return False

    def add_normal_plate(self, raw_name: str):
        # 原体を含む料理は、出力対象外とする
        if self._is_raw_plate(raw_name):
            return

        name = self._convert_plate_name(raw_name)
        if not (name in self.normal_plates):
            self.normal_plates.append(name)

    def add_allergen_plate(self, raw_name: str, menu_list: list):
        # 原体を含む料理は、出力対象外とする
        if self._is_raw_plate(raw_name):
            return

        name = self._convert_plate_name(raw_name)
        allergen_name = f'ア){name}'
        if not self._exists_allergen_plate(allergen_name):
            self.allergen_plates.append(AllergenPlateDetail(allergen_name, menu_list))

    def _convert_plate_name(self, raw_name: str):
        # 記号の削除
        converted = raw_name.replace('①', '')
        converted = converted.replace('②', '')
        converted = converted.replace('③', '')
        converted = converted.replace('④', '')
        converted = converted.replace('⑤', '')
        converted = raw_name.replace('⑩', '')

        converted = converted.replace('■', '')
        converted = converted.replace('▼', '')
        converted = converted.replace('△', '')
        converted = converted.replace('◎', '')

        # 数量の削除
        converted = re.sub('(\d+|\d+\.\d+)g', '', converted, 3)
        converted = re.sub('(\d+|\d+\.\d+)ｇ', '', converted, 3)
        converted = re.sub('(\d+|\d+\.\d+)個', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)切れ', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)本', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)尾', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)丁', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)枚', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)%', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)％', '', converted)
        converted = re.sub('(\d+|\d+\.\d+)cc', '', converted)

        converted = converted.replace('（', '(')
        converted = converted.replace('）', ')')
        converted = converted.replace('()', '')
        converted = converted.replace('(約)', '')
        converted = converted.replace('＋具入りの液', '')
        converted = re.sub('希釈(\d+|\d+\.\d+)', '希釈', converted)
        if converted[-1] == '+':
            converted = converted.replace('+', '')
        # 「③△煮物（かぼちゃ）2個+(赤キャップ4.5％＋水10.5％)で15％」の対応
        if converted[-1] == 'で':
            converted = converted.replace('で', '')

        return converted


class Command(BaseCommand, ExcelOutputMixin):
    def __init__(self):
        self.table_outer_side = Side(style='medium', color="000000")
        self.table_inner_side = Side(style='thin', color="000000")

    def add_arguments(self, parser):
        parser.add_argument('--date', nargs='+', type=str)
        parser.add_argument('--filename', nargs='+', type=str)

    def is_not_allergen(self, detail: str):
        without_count = detail[detail.find(' ') + 1:]
        if (without_count.find(settings.COOKING_DIRECTION_J_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_U_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_B_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_GU_J_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_GU_B_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_GU_U_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_SOUP_J_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_SOUP_B_CODE) == 0) or \
            (without_count.find(settings.COOKING_DIRECTION_SOUP_U_CODE) == 0):
            return True
        else:
            return False

    def get_menu_from_detail(self, detail: str):
        without_count = detail[detail.find(' ') + 1:]
        if without_count[0] == 'ソ':
            return 'ソフト'
        elif without_count[0] == 'ミ':
            return 'ミキサー'
        elif without_count[0] == 'ゼ':
            return 'ゼリー'
        else:
            return '通常'

    def convert_weekday(self, weekday: int):
        if weekday == 6:
            return '日'
        if weekday == 0:
            return '月'
        if weekday == 1:
            return '火'
        if weekday == 2:
            return '水'
        if weekday == 3:
            return '木'
        if weekday == 4:
            return '金'
        else:
            return '土'

    def is_miso_soup_filling(self, name: str):
        return ('具' in name) or ('個' in name) or ('・' in name)

    def is_miso_soup(self, name: str):
        if ('味噌汁' in name) or ('みそ汁' in name):
            # 汁具は対象外とする
            return False if self.is_miso_soup_filling(name) else True
        else:
            return False

    def get_with_enge_names_list(self, name: str):
        result = [name, ]
        if not self.is_miso_soup(name):
            for enge in ['ソフト', 'ミキサー', 'ゼリー']:
                enge_name = f'【{enge}】{name}'
                result.append(enge_name)

        return result

    def is_enge_plate_name(self, plate_name: str):
        return plate_name.find('【') == 0

    def write_file(self, workbook, output_plate_list, file_path: str, cooking_date: dt.date, is_enge: bool):
        base_heights = [40, 40, 60]
        for index, ws in enumerate(workbook.worksheets):

            # タイトルの出力(テンプレートのままでは、紙で印刷したときにどちらの記録簿か分かりづらいため)
            ws.cell(1, 1, f'加熱加工記録簿(嚥下食)' if is_enge else f'加熱加工記録簿(通常食)')

            # 調理日の出力
            if platform.system() == 'Windows':
                ws.cell(4, 1, f'調理日：{cooking_date.strftime("%Y年%m月%d日")}({self.convert_weekday(cooking_date.weekday())})')
            else:
                ws.cell(4, 1, f'調理日：{cooking_date.strftime("%Y年%-m月%-d日")}({self.convert_weekday(cooking_date.weekday())})')

            # 各料理名の出力
            row = 7
            for plate_name in output_plate_list:
                if is_enge:
                    if self.is_enge_plate_name(plate_name):
                        # 嚥下分の出力
                        count = settings.ENGE_WRITE_COUNT
                    else:
                        # 常食・薄味分の出力(スキップ)
                        continue
                else:
                    if self.is_enge_plate_name(plate_name):
                        # 嚥下分の出力(スキップ)
                        continue
                    else:
                        # 常食・薄味分の出力
                        count = settings.PLATE_WRITE_COUNT

                # 行の鷹さの設定
                base_height = base_heights[index]
                for _ in range(count):
                    ws.cell(row, 1, plate_name)
                    h = len(plate_name) / 7
                    if h > 1:
                        dh = 22 * h
                        if base_height > dh:
                            ws.row_dimensions[row].height = base_height
                        else:
                            ws.row_dimensions[row].height = dh
                    else:
                        ws.row_dimensions[row].height = base_height
                    row += 1

            # 余分な行の非表示
            while row <= settings.HEATING_PROCESSING_MAX_ROW + 1:
                ws.row_dimensions[row].hidden = True
                row += 1

            # テーブルの罫線の修復
            if index == 0:
                ExcelHellper.set_grid_border_without_top(ws, f'A7:A901', self.table_outer_side, self.table_inner_side)
                ExcelHellper.set_grid_border_without_top(ws, f'B7:E901', self.table_outer_side, self.table_inner_side)
            elif index == 1:
                ExcelHellper.set_grid_border_without_top(ws, f'A7:A901', self.table_outer_side, self.table_inner_side)
                ExcelHellper.set_grid_border_without_top(ws, f'B7:C901', self.table_outer_side, self.table_inner_side)
            elif index == 2:
                ExcelHellper.set_grid_border_without_top(ws, f'A7:A901', self.table_outer_side, self.table_inner_side)
                ExcelHellper.set_grid_border_without_top(ws, f'B7:E901', self.table_outer_side, self.table_inner_side)
                ExcelHellper.set_grid_border_without_top(ws, f'F7:H901', self.table_outer_side, self.table_inner_side)
                ExcelHellper.set_grid_border_without_top(ws, f'I7:I901', self.table_outer_side, self.table_inner_side)

            footer_text = '&P / &Nページ'
            ws.oddFooter.center.text = footer_text
            ws.page_margins.bottom = 1.0

        # 変換後ファイルの出力
        self.save_with_select(workbook, file_path)

    def handle(self, *args, **options):

        #######################################################
        # らくらく献立仕様外の形式に調理表を変換する
        # 食数内訳には常食+薄味でアレルギーを含めた全食数を記載し、別途その料理を使用するアレルギーを記述する
        # 食数(B列、合計)には、食数内訳の合計ではなく、常食+薄味でアレルギーを含めた全食数を記述する。
        #######################################################

        in_cook = options['date']          # 呼び出し時の引数:製造日（YYYY.MM.DD）
        in_file = options['filename']      # 呼び出し時の引数:調理表ファイル名
        cooking_date = dt.datetime.strptime(in_cook, '%Y.%m.%d')

        # 対象製造日の調理表取得
        cook_direc_file = os.path.join(settings.MEDIA_ROOT, 'upload', in_file)

        # 出力先の準備
        output_dir = os.path.join(settings.OUTPUT_DIR, 'heating_processing')
        os.makedirs(output_dir, exist_ok=True)  # 上書きOK

        # テンプレートファイル
        template_path = os.path.join(settings.STATICFILES_DIRS[0], 'excel/heating_processing.xlsx')  # 加熱加工記録簿のテンプレート

        # ------------------------------------------------------------------------------
        # 調理表の読み込み
        # ------------------------------------------------------------------------------
        workbook = excel.load_workbook(cook_direc_file)
        worksheet = workbook.worksheets[0]

        plate_list = []
        plate = None
        current_plate_name = None
        is_allergen = False
        row_index = 9
        menu_list = []
        for row in worksheet.iter_rows(min_row=row_index):
            # 食種(A列)に値
            if row[0].value:
                if plate:
                    plate_list.append(plate)
                plate = CookPlate(row[0].value)
            # 料理名(D列)に値
            if row[3].value:
                # 献立切り替え
                if current_plate_name:
                    # 処理中の料理の解析を終了する処理
                    if is_allergen:
                        plate.add_allergen_plate(current_plate_name, menu_list)
                    else:
                        plate.add_normal_plate(current_plate_name)

                current_plate_name = row[3].value
                is_allergen = True
                menu_list = []

                # 内訳の内容チェック(アレルギー対応代替食でないことが確定するかどうか)
                if row[2].value:
                    if self.is_not_allergen(row[2].value):
                        is_allergen = False
                    else:
                        menu = self.get_menu_from_detail(row[2].value)
                        menu_list.append(menu)
            elif row[2].value:
                # 内訳の内容チェック(アレルギー対応代替食でないことが確定するかどうか)
                if is_allergen:
                    if self.is_not_allergen(row[2].value):
                        is_allergen = False
                    else:
                        menu = self.get_menu_from_detail(row[2].value)
                        if not (menu in menu_list):
                            menu_list.append(menu)
            else:
                # 内訳にも値がなければ、何もしない
                pass

            row_index += 1

        # 最後に残ったplateを追加
        plate_list.append(plate)
        if is_allergen:
            plate.add_allergen_plate(current_plate_name, menu_list)
        else:
            plate.add_normal_plate(current_plate_name)

        output_plate_list = []
        first_index_dict = {}
        duplicated = []
        menu_list = []
        # 通常料理(アレルギー代替食以外)のリスト化
        for plate in plate_list:
            for plate_name in [x for x in plate.normal_plates]:
                # 重複対応
                if plate_name in output_plate_list:
                    name_with_day = f'{plate_name} {plate.eating_day}'
                    output_plate_list += self.get_with_enge_names_list(name_with_day)
                    if not (first_index_dict[plate_name] in duplicated):
                        duplicated.append(first_index_dict[plate_name])
                else:
                    index = len(output_plate_list)
                    names_list = self.get_with_enge_names_list(plate_name)
                    output_plate_list += names_list
                    first_index_dict[plate_name] = (index, plate.eating_day, len(names_list))

        # アレルギー代替食以外のリスト化
        for plate in plate_list:
            for allergen_plate in [x for x in plate.allergen_plates]:
                # 重複対応
                if allergen_plate.name in first_index_dict:
                    output_name = f'{allergen_plate.name} {plate.eating_day}'
                    if allergen_plate.has_menu('通常'):
                        output_plate_list.append(output_name)
                    if allergen_plate.has_menu('ソフト') and not (self.is_miso_soup(output_name)):
                        output_plate_list.append(f'【ソフト】{output_name}')
                    if allergen_plate.has_menu('ミキサー') and not (self.is_miso_soup(output_name)):
                        output_plate_list.append(f'【ミキサー】{output_name}')
                    if allergen_plate.has_menu('ゼリー') and not (self.is_miso_soup(output_name)):
                        output_plate_list.append(f'【ゼリー】{output_name}')

                    if not (first_index_dict[allergen_plate.name] in duplicated):
                        duplicated.append(first_index_dict[allergen_plate.name])
                else:
                    index = len(output_plate_list)
                    names_cnt = 0
                    if allergen_plate.has_menu('通常'):
                        output_plate_list.append(allergen_plate.name)
                        names_cnt += 1
                    if allergen_plate.has_menu('ソフト') and not (self.is_miso_soup(allergen_plate.name)):
                        output_plate_list.append(f'【ソフト】{allergen_plate.name}')
                        names_cnt += 1
                    if allergen_plate.has_menu('ミキサー') and not (self.is_miso_soup(allergen_plate.name)):
                        output_plate_list.append(f'【ミキサー】{allergen_plate.name}')
                        names_cnt += 1
                    if allergen_plate.has_menu('ゼリー') and not (self.is_miso_soup(allergen_plate.name)):
                        output_plate_list.append(f'【ゼリー】{allergen_plate.name}')
                        names_cnt += 1
                    first_index_dict[allergen_plate.name] = (index, plate.eating_day, names_cnt)

        # 初回の重複の名称変更
        for index, e_day, cnt in duplicated:
            for i in range(cnt):
                output_plate_list[index + i] = f'{output_plate_list[index + i]} {e_day}'
        # ------------------------------------------------------------------------------
        # 加熱加工記録簿の出力
        # ------------------------------------------------------------------------------
        # 常食・薄味の出力
        wb = excel.load_workbook(template_path)
        normal_file_path = os.path.join(output_dir, f'加熱加工記録簿_{in_cook}_通常食.xlsx')
        self.write_file(wb, output_plate_list, normal_file_path, cooking_date, is_enge=False)
        wb.close()

        # 嚥下の出力
        wb = excel.load_workbook(template_path)
        enge_file_path = os.path.join(output_dir, f'加熱加工記録簿_{in_cook}_嚥下食.xlsx')
        self.write_file(wb, output_plate_list, enge_file_path, cooking_date, is_enge=True)
        wb.close()
