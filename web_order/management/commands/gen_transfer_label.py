import logging
import os
import pandas as pd
import datetime as dt
import openpyxl as excel

from datetime import timedelta
from django.conf import settings
from django.core.management.base import BaseCommand
from django_pandas.io import read_frame

from web_order.models import Order, OrderEveryday, ProductMaster, MenuDisplay
from web_order.models import InvoiceException, SerialCount, EverydaySelling
from web_order.picking import QrCodeUtil


logger = logging.getLogger(__name__)


class Command(BaseCommand):


    def add_arguments(self, parser):  # 引数をとれるようにする
        parser.add_argument('date', nargs='+', type=str)

    def _is_fix_qr_sheet_target(self, eating_day, meal, unit_number):
        if (eating_day.day == 17) and (meal == '夕食') and (unit_number == 119):
            return True
        else:
            return False

    def handle(self, *args, **options):


        # 集計日時を喫食日で指定
        in_date = options['date'][0]  # 呼び出し時の引数1つ目

        aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d')
        aggregation_day = aggregation_day.date()  # 時刻部分を除外
        logger.info(f'配送リスト出力開始:{aggregation_day}')

        new_dir_path = 'tmp/transfer/' + str(aggregation_day)  # 一時ファイルを集計日別のフォルダに出力
        os.makedirs(new_dir_path, exist_ok=True)      # 上書きOK

        label_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/label.xlsx')  # 配送ラベルのテンプレート
        label_output_dir = os.path.join(settings.OUTPUT_DIR, settings.LABEL_XLS_DIR)
        label_output_file = os.path.join(label_output_dir, '配送ラベル_' + str(aggregation_day) + '_喫食分.xlsx')

        # ログ出力
        aggregation_log = pd.DataFrame(index=[],
                                       columns=[str(aggregation_day) + ' 喫食分 - 配送ラベル作成処理開始',
                                                dt.datetime.now()])


        # ------------------------------------------------------------------------------
        #
        # 請求データの作成(配送リストとしては不要な処理もあるが、デグレをおこさないため、ひとまず請求データ出力の処理を踏襲する)
        #
        # ------------------------------------------------------------------------------

        # ------------------------------------------------------------------------------
        # 注文データから食数が入っているものを抽出
        # ------------------------------------------------------------------------------
        qs_order = Order.objects.filter(eating_day=aggregation_day, quantity__gt=0)\
            .values('eating_day', 'unit_name__unit_code', 'unit_name__username', 'unit_name__unit_number', 'unit_name',
                    'meal_name__meal_name', 'meal_name__seq_order', 'menu_name', 'menu_name__seq_order',
                    'allergen', 'allergen__seq_order', 'quantity')\
            .exclude(unit_name__unit_code__range=[80001, 80008])

        df_order = read_frame(qs_order)
        df_order.to_csv(new_dir_path + "/E-1_注文データ.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-1_注文データ', len(df_order)]


        # ------------------------------------------------------------------------------
        # 食数固定製造データから検食分のみ抽出
        # ------------------------------------------------------------------------------

        qs_kenshoku = OrderEveryday.objects.filter(menu_name_id=6)\
            .values('eating_day', 'unit_name__unit_code', 'unit_name__username', 'unit_name__unit_number', 'unit_name',
                    'meal_name__meal_name', 'meal_name__seq_order', 'menu_name', 'menu_name__seq_order',
                    'allergen', 'allergen__seq_order', 'quantity')

        df_kenshoku = read_frame(qs_kenshoku)
        df_kenshoku = df_kenshoku.fillna(aggregation_day)  # 検食分は喫食日が入っていないので設定する

        df_kenshoku.to_csv(new_dir_path + "/E-2-1_検食分.csv", index=False)


        # ------------------------------------------------------------------------------
        # 検食分として抽出したレコードのユーザーIDをチェックして当日の注文が入っているか確認する
        # ------------------------------------------------------------------------------

        # コピー元の列名を使用して追加するDataFrameを初期化
        append_kenshoku = pd.DataFrame(index=[], columns=df_kenshoku.columns.values)

        # 喫食日に対象施設の注文が入っているかチェック、あれば結合用Dataframeに追加
        for index, data in df_kenshoku.iterrows():
            queryset = qs_order.filter(unit_name__username=data['unit_name__username'])
            if queryset.exists():
                append_kenshoku.loc[len(append_kenshoku)] = data


        append_kenshoku.to_csv(new_dir_path + "/E-2-2_検食_注文ある施設分.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-2-2_検食_注文ある施設分', len(append_kenshoku)]


        order_kenshoku = pd.concat([df_order, append_kenshoku]).reset_index(drop=True)  # 検食分を結合


        # ------------------------------------------------------------------------------
        # 検食分を含んだ全注文データの喫食日から売上日を算出する
        # ------------------------------------------------------------------------------

        order_kenshoku = order_kenshoku.sort_values(by=['unit_name__unit_code', 'meal_name__seq_order',
                                                    'menu_name__seq_order', 'allergen__seq_order'])

        # ソートし直す時用に連番付与
        start_num = 1
        stop_num = start_num + len(order_kenshoku.index)

        seq_order = pd.RangeIndex(start=start_num, stop=stop_num, step=1)
        order_kenshoku['seq_order'] = seq_order

        # 基本は喫食日の2日前、商品が施設に到着する日（事故・遅延は考慮しない）
        sold_day = aggregation_day - timedelta(days=2)

        order_kenshoku['売上日'] = sold_day  # 列の追加

        order_kenshoku.to_csv(new_dir_path + "/E-3_注文データ＋検食.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-3_注文データ＋検食↓', len(order_kenshoku)]


        # ------------------------------------------------------------------------------
        # 販売固定商品データから送料などのレコードを抽出
        # ------------------------------------------------------------------------------

        qs_selling = EverydaySelling.objects.all()\
            .values('unit_name__unit_code', 'unit_name__username', 'unit_name', 'quantity', 'price',
                    'product_code', 'product_name')

        df_selling = read_frame(qs_selling)

        df_selling['売上日'] = sold_day  # 列の追加

        df_selling.to_csv(new_dir_path + "/E-4-1_販売固定商品.csv", index=False)


        # コピー元の列名を使用して追加するDataFrameを初期化
        append_selling = pd.DataFrame(index=[], columns=df_selling.columns.values)

        # 作業日に対象施設の注文が入っているかチェック、あれば結合用Dataframeに追加
        for index, data in df_selling.iterrows():
            queryset = qs_order.filter(unit_name__username=data['unit_name__username'])
            if queryset.exists():
                append_selling.loc[len(append_selling)] = data

        append_selling.to_csv(new_dir_path + "/E-4-2_販売固定商品_注文ある施設分.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-4-2_販売固定商品_注文ある施設分', len(append_selling)]


        # ------------------------------------------------------------------------------
        # 請求内容を施設ごとに個別対応するため、例外内容を保持したテーブルを読み込む
        # ------------------------------------------------------------------------------
        qs_invoice_exception = InvoiceException.objects.all()\
            .values('unit_name__username', 'ng_saturday', 'ng_sunday', 'ng_holiday', 'reduced_rate')

        df_invoice_exception = read_frame(qs_invoice_exception)
        df_invoice_exception.to_csv(new_dir_path + "/E-5_請求書調整項目.csv", index=False)


        # ------------------------------------------------------------------------------
        # 売上日の調整
        # ------------------------------------------------------------------------------

        order_kenshoku_exception = pd.merge(order_kenshoku, df_invoice_exception, on='unit_name__username', how='left')
        append_selling_exception = pd.merge(append_selling, df_invoice_exception, on='unit_name__username', how='left')

        youbi = sold_day.weekday()  # 売上日の曜日(月曜日が0〜日曜日が6)

        # 日曜受取不可(6)なら調整日数を足す（マイナス数値を足すので前倒し、プラス数値を足すので後日にずらす）
        if youbi == 6:

            for index, data in order_kenshoku_exception.iterrows():
                if data['ng_sunday'] > 0 or data['ng_sunday'] < 0:
                    order_kenshoku_exception.loc[index, '売上日'] = sold_day + timedelta(days=data['ng_sunday'])

            for index, data in append_selling_exception.iterrows():
                if data['ng_sunday'] > 0 or data['ng_sunday'] < 0:
                    append_selling_exception.loc[index, '売上日'] = sold_day + timedelta(days=data['ng_sunday'])

        # 土曜受取不可(5)なら
        elif youbi == 5:

            for index, data in order_kenshoku_exception.iterrows():
                if data['ng_saturday'] > 0 or data['ng_saturday'] < 0:
                    order_kenshoku_exception.loc[index, '売上日'] = sold_day + timedelta(days=data['ng_saturday'])

            for index, data in append_selling_exception.iterrows():
                if data['ng_saturday'] > 0 or data['ng_saturday'] < 0:
                    append_selling_exception.loc[index, '売上日'] = sold_day + timedelta(days=data['ng_saturday'])

        # 業務委託の場合、喫食日が売上日になるため
        for index, data in order_kenshoku_exception.iterrows():
            if data['reduced_rate'] is True:  # 省略記法にすると正常に動作しないので is Trueまで記述
                order_kenshoku_exception.loc[index, '売上日'] = aggregation_day

        for index, data in append_selling_exception.iterrows():
            if data['reduced_rate'] is True:  # 省略記法にすると正常に動作しないので is Trueまで記述
                append_selling_exception.loc[index, '売上日'] = aggregation_day


        order_kenshoku_exception.to_csv(new_dir_path + "/E-6_注文データ＋検食＋売上日調整.csv", index=False)

        append_selling_exception.to_csv(new_dir_path + "/E-7_販売固定商品＋売上日調整.csv", index=False)


        # ------------------------------------------------------------------------------
        # 商品単価を結合するために商品マスタテーブルを読み込む
        # ------------------------------------------------------------------------------
        
        qs_price = MenuDisplay.objects.all()\
            .values('username__username', 'menu_name',
                    'price_breakfast', 'price_lunch', 'price_dinner', 'price_snack')
        df_price = read_frame(qs_price)

        # サービス付き高齢者向け住宅みどり・日曜朝のみ価格が100円に変更###############################

        if aggregation_day.weekday() == 6:  # 喫食日の曜日(月曜日が0〜日曜日が6)
            df_price.loc[df_price['username__username'] == '10014', 'price_breakfast'] = 100

        df_price.to_csv(new_dir_path + "/E-8_単価.csv", index=False)



        # ------------------------------------------------------------------------------
        # 商品単価を結合、該当する朝昼夕の単価のみ残す
        # ------------------------------------------------------------------------------

        # 結合するために列名を変更
        order = order_kenshoku_exception.rename(columns={'unit_name__username': 'username__username'})
        order_price = pd.merge(order, df_price, on=['username__username', 'menu_name'], how='left')

        for index, data in order_price.iterrows():
            # 商品コードを結合するために「肉」や「えび」などアレルギーありの名称をすべて「あり」に変更する
            if data['allergen'] == 'なし' or data['allergen'] == '個食' or data['allergen'] == 'フリーズ':
                pass
            else:
                order_price.loc[index, 'allergen'] = 'あり'  # locで元のDFにアクセスして値を代入、書き換える

            # 該当する食事区分の価格のみ残すため、username列をprice列として仮使用する（列名は後で変更）
            if data['meal_name__meal_name'] == '朝食':
                order_price.loc[index, 'username__username'] = data['price_breakfast']
            elif data['meal_name__meal_name'] == '昼食':
                order_price.loc[index, 'username__username'] = data['price_lunch']
            elif data['meal_name__meal_name'] == '夕食':
                order_price.loc[index, 'username__username'] = data['price_dinner']
            elif data['meal_name__meal_name'] == '間食':
                order_price.loc[index, 'username__username'] = data['price_snack']

        # 価格設定がされていないセルに0を埋める、floatになるのでintに変換
        order_price = order_price.fillna(0)
        order_price = order_price.astype({'username__username': 'int64'})

        order_price.to_csv(new_dir_path + "/E-9_注文データ＋検食＋単価.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-9_注文データ＋検食↓＋単価→', len(order_price)]


        # ------------------------------------------------------------------------------
        # 商品コードを結合するためにマスタデータを取り出す
        # ------------------------------------------------------------------------------
        qs_prod_code = ProductMaster.objects.all()\
            .values('product_code', 'product_name', 'meal_name__meal_name', 'menu_name', 'allergen')

        df_prod_code = read_frame(qs_prod_code)
        df_prod_code.to_csv(new_dir_path + "/E-10_商品コード.csv", index=False)


        # ------------------------------------------------------------------------------
        # 商品コードを結合
        # ------------------------------------------------------------------------------
        order_price_code = pd.merge(order_price, df_prod_code,
                                    on=['meal_name__meal_name', 'menu_name', 'allergen'], how='left')

        # 列名の変更
        order_price_code = order_price_code.rename(columns={'unit_name__unit_code': '得意先コード',
                                                            'username__username': '売上単価',
                                                            'unit_name': '得意先名称',
                                                            'quantity': '売上数量',
                                                            'product_code': '商品コード',
                                                            'product_name': '商品名'})

        # 列順の並べ替え（指定された列のみ残る）
        order_price_code = order_price_code.reindex(columns=['売上日', '得意先コード', '得意先名称',
                                                             '商品コード', '商品名', '売上数量', '売上単価',
                                                             'reduced_rate', 'seq_order'])

        order_price_code.to_csv(new_dir_path + "/E-11_注文データ＋検食＋単価＋コード.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-11_注文データ＋検食↓＋単価→＋コード→', len(order_price_code)]


        # ------------------------------------------------------------------------------
        # 固定販売データを結合するために列名を変更・並べ替え
        # ------------------------------------------------------------------------------

        # 列名の変更
        append_selling_exception = append_selling_exception.rename(columns={'unit_name__unit_code': '得意先コード',
                                                                            'unit_name': '得意先名称',
                                                                            'quantity': '売上数量',
                                                                            'price': '売上単価',
                                                                            'product_code': '商品コード',
                                                                            'product_name': '商品名'})
        # 列順の並べ替え（指定された列のみ残る）
        append_selling_exception = append_selling_exception.reindex(columns=['売上日', '得意先コード', '得意先名称',
                                                                             '商品コード', '商品名',
                                                                             '売上数量', '売上単価',
                                                                             'reduced_rate', 'seq_order'])

        append_selling_exception.to_csv(new_dir_path + "/E-12_販売固定商品＋売上日調整.csv", index=False)


        # ------------------------------------------------------------------------------
        # 固定販売データを結合
        # ------------------------------------------------------------------------------

        invoice_data = pd.concat([order_price_code, append_selling_exception]).reset_index(drop=True)

        invoice_data = invoice_data.sort_values(by=['得意先コード', 'seq_order'])

        invoice_data.to_csv(new_dir_path + "/E-13_請求データ結合完了.csv", index=False)
        aggregation_log.loc[len(aggregation_log)] = ['E-13_請求データ結合完了', len(invoice_data)]


        # ------------------------------------------------------------------------------
        # 行番号を付与
        # ------------------------------------------------------------------------------
        row_num_now = 1
        code_prev = ''



        # シリアル値管理テーブルから現在の最終伝票番号を取得
        qs_serial_num = SerialCount.objects.filter(serial_name='伝票NO')
        df_serial_num = read_frame(qs_serial_num)

        serial_now = df_serial_num.loc[0, 'serial_number']  # 現在の最終伝票番号

        for index, data in invoice_data.iterrows():
            code_now = data['得意先名称']
            # 得意先名称ごとに行番号を振り直す
            if code_now != code_prev:
                serial_now += 1
                row_num_now = 1
            invoice_data.loc[index, '伝票NO'] = serial_now
            invoice_data.loc[index, '行番号'] = row_num_now
            row_num_now += 1
            code_prev = invoice_data.loc[index, '得意先名称']

        invoice_data = invoice_data.astype({'伝票NO': 'int64', '行番号': 'int64'})


        qs_serial_num.update(serial_number=serial_now)  # 最終伝票番号をDB側で更新


        # ------------------------------------------------------------------------------
        # 列名の追加
        # ------------------------------------------------------------------------------

        invoice_data['レコード区分'] = 0
        invoice_data['データレコード区分'] = 'ur01'
        invoice_data['伝票区分'] = 0  # 0:掛売 1:現売
        invoice_data['取引区分'] = 1  # 1:売上
        invoice_data['取引性格'] = 1  # 1:売上
        invoice_data['取引区分名'] = '売上'
        invoice_data['数量入力区分'] = 1
        invoice_data['選択単価区分'] = 1
        invoice_data['原価単価'] = 0
        invoice_data['課税区分'] = 1  # 1:外税 2:内税 3:非課税 9:対象外
        invoice_data['消費税訂正額'] = 0

        invoice_data['消費税率コード'] = 5
        invoice_data['消費税率'] = 8

        # 業務委託の場合、喫食日が売上日になるため
        for index, data in invoice_data.iterrows():
            if data['商品名'] == '送料':
                invoice_data.loc[index, '消費税率コード'] = 4
                invoice_data.loc[index, '消費税率'] = 10
            if data['reduced_rate'] is True:    # 省略記法にすると正常に動作しないので is Trueまで記述
                invoice_data.loc[index, '消費税率コード'] = 4
                invoice_data.loc[index, '消費税率'] = 10

        # 列順の最終並べ替え（指定された列のみ残る）
        invoice_data = invoice_data.reindex(columns=['レコード区分', 'データレコード区分', '伝票区分', '伝票NO',
                                                     '売上日', '得意先コード', '得意先名称', '消費税率コード', '消費税率',
                                                     '行番号', '取引区分', '取引性格', '取引区分名',
                                                     '商品コード', '商品名', '数量入力区分', '選択単価区分',
                                                     '売上数量', '売上単価', '原価単価', '課税区分', '消費税訂正額'])

        invoice_data.to_csv(new_dir_path + "/E-14_請求データ完成.csv", index=False)  # encoding='cp932'
        aggregation_log.loc[len(aggregation_log)] = ['E-14_請求データ完成', len(invoice_data)]

        # ------------------------------------------------------------------------------
        #
        # 配送ラベルの作成
        # 要件： 朝昼夕の注文がある施設を先に出力し、1行で横に並ぶようにする
        #
        # ------------------------------------------------------------------------------

        meal_count = order_kenshoku  # 注文データに固定検食分を足したもの

        # ユニットごとに当日の食事区分が何種類あるか調べるため、まずユニットと食事区分でグループ化
        meal_count = meal_count.groupby(['unit_name__unit_number', 'meal_name__meal_name'])[['quantity']]\
            .count().reset_index()

        # meal_name__meal_nameが当日の食事区分の数（1〜4）になる
        # 間食があると4になるが、現在は間食の注文が無い状態
        meal_count = meal_count.groupby(['unit_name__unit_number'])[['meal_name__meal_name']]\
            .count().reset_index()

        # 結合するために列名を変更
        meal_count = meal_count.rename(columns={'unit_name__unit_number': 'unit_number',
                                                'meal_name__meal_name': 'meal_count'})

        meal_count.to_csv(new_dir_path + "/F-1_ユニット別食事区分数.csv", index=False)


        delivery_label = order_kenshoku  # 注文データに固定検食分を足したもの

        # 結合するために列名を変更
        delivery_label = delivery_label.rename(columns={'unit_name__unit_number': 'unit_number',
                                                        'meal_name__meal_name': 'meal_name'})

        delivery_label_meal = pd.merge(delivery_label, meal_count, on='unit_number', how='left')

        delivery_label_meal = delivery_label_meal.sort_values(['unit_number',
                                                               'meal_count',
                                                               'meal_name__seq_order',
                                                               'menu_name__seq_order',
                                                               'allergen__seq_order'],
                                                              ascending=[True, False, True, True, True])

        delivery_label_meal.to_csv(new_dir_path + "/F-2_朝昼夕あり施設優先ソート.csv", index=False)

        # 列順の並べ替え（指定された列のみ残る）
        delivery_label_meal = delivery_label_meal.reindex(columns=['eating_day', 'unit_number', 'unit_name',
                                                                   'meal_name', 'menu_name',
                                                                   'allergen', 'quantity'])

        # 食事区分を△○□付きの表記に変更
        for index, data in delivery_label_meal.iterrows():
            if data['meal_name'] == '朝食':
                delivery_label_meal.loc[index, 'meal_name'] = '△ 朝'  # locで元のDFにアクセスして値を代入、書き換える

            if data['meal_name'] == '昼食':
                delivery_label_meal.loc[index, 'meal_name'] = '○ 昼'

            if data['meal_name'] == '夕食':
                delivery_label_meal.loc[index, 'meal_name'] = '□ 夕'

            if data['allergen'] == 'なし':
                delivery_label_meal.loc[index, 'allergen'] = ''  # アレルギーなしの場合は「なし」の表記を外す

        delivery_label_meal.to_csv(new_dir_path + "/F-3_朝昼夕あり施設優先ソート記号変換.csv", index=False)


        # ------------------------------------------------------------------------------
        # Excelブックの作成
        # ------------------------------------------------------------------------------

        book = excel.load_workbook(label_template)

        unit_prev = meal_prev = ''
        i = 5

        for index, row in delivery_label_meal.iterrows():

            unit_now = row['unit_name']  # 現在行のユニット名
            meal_now = row['meal_name']  # 現在行の食事区分

            if unit_now != unit_prev or meal_now != meal_prev:  # ユニットと食事区分が変わったら

                if self._is_fix_qr_sheet_target(aggregation_day, row['meal_name'], row['unit_number']):
                    i = 6
                    sheet = book.copy_worksheet(book["119 □ 夕_原本"])
                    sheet.title = str(row['unit_number']) + ' ' + row['meal_name']

                    sheet.cell(4, 3, row['meal_name'])
                    sheet.cell(4, 4, row['eating_day'])
                else:
                    i = 6
                    sheet = book.copy_worksheet(book["原本"])
                    sheet.title = str(row['unit_number']) + ' ' + row['meal_name']

                    # 施設番号・名称出力
                    sheet.cell(1, 2, str(row['unit_number']))
                    sheet.cell(2, 2, row['unit_name'])

                    # ピッキング用QRコード画像対応
                    image_path = QrCodeUtil.get_transfer_imege_path_v2(row['unit_number'], row['meal_name'], aggregation_day.day)
                    logger.info(f"{row['unit_number']}-{row['meal_name']}-{aggregation_day.day}")
                    logger.info(image_path)
                    qr_image = excel.drawing.image.Image(image_path)
                    qr_image.width = 160
                    qr_image.height = 160
                    sheet.add_image(qr_image, sheet.cell(1, 5).coordinate)
                    #sheet.cell(3, 2, f"{row['unit_number']}-{row['meal_name']}-{aggregation_day.day}")

                    sheet.cell(4, 3, row['meal_name'])
                    sheet.cell(4, 4, row['eating_day'])

            sheet.cell(i, 3, row['menu_name'])
            sheet.cell(i, 4, row['allergen'])
            sheet.cell(i, 5, row['quantity'])

            unit_prev = unit_now
            meal_prev = meal_now

            i = i + 1

        book.remove(book["原本"])
        book.remove(book["119 □ 夕_原本"])
        book.save(label_output_file)

        # ログを出力
        aggregation_log.to_csv("tmp/gen_transfer_label_log.csv", index=False, mode='a')
        logger.info('配送リスト出力終了')
