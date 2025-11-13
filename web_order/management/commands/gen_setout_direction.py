import datetime as dt
import os

from dateutil.relativedelta import relativedelta

import openpyxl as excel

from openpyxl.drawing.image import Image
from openpyxl.worksheet.pagebreak import Break

from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand
from web_order.models import FoodPhoto, MonthlyMenu, EngeDirection, EngeFoodDirection, SetoutDuration

"""
    盛付指示書の作成

# 引数
    喫食日

# 献立DB
    Model: MonthlyMenu
    Model: FoodPhoto
    Model: EngeFoodDirection        
"""


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('date', nargs='+', type=str)

    def get_osechi_book(self):
        xls_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/setout_osechi.xlsx')  # 指示書のテンプレート
        book = excel.load_workbook(xls_template)
        return book

    def _set_direction_if_not_blank(self, sheet, row, column, direction):
        if direction == "■":
            pass
        else:
            sheet.cell(row, column, direction)

    def handle(self, *args, **options):
        in_date = options['date'][0]
        in_output_enge_option = options['date'][1] or 'off'

        is_for_new_year = '-01-01' in in_date

        setout_name = f'盛付指示書_{str(in_date)}'
        setout_output_dir = os.path.join(settings.OUTPUT_DIR, 'setout', setout_name)
        if os.path.isdir(setout_output_dir):
            is_create = False
        else:
            is_create = True
        os.makedirs(setout_output_dir, exist_ok=True)  # 上書きOK
        output_file = os.path.join(setout_output_dir, setout_name + '.xlsx')

        xls_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/setout.xlsx')  # 指示書のテンプレート

        # 画像が生成されておらずにエラーになる場合があるため、画像生成を実施
        call_command('generateimages')
        book = excel.load_workbook(xls_template)

        def create_excel(queryset, sheetname):

            worksheet = book[sheetname]

            worksheet.cell(2, 5, in_date)

            i = 4
            j = 6
            p = 5

            for q in queryset:
                qs2 = FoodPhoto.objects.filter(food_name=q, menu=q).first()

                if is_for_new_year and (sheetname == '昼食'):
                    worksheet.cell(i, 4, '別紙の『おせち用盛付指示書』にて')
                    break
                else:
                    worksheet.cell(i, 2, qs2.food_name)

                    self._set_direction_if_not_blank(worksheet, i, 4, qs2.direction)

                    self._set_direction_if_not_blank(worksheet, j, 4, qs2.direction2)

                    img_file = str(qs2.thumbnail)

                    if (img_file):

                        img_dir = os.path.join(settings.MEDIA_ROOT, str(qs2.basic))

                        # 画像を選択
                        img_to_excel = excel.drawing.image.Image(img_dir)
                        # 指定の位置に画像を添付
                        worksheet.add_image(img_to_excel, 'B' + str(p))

                i = i + 5
                j = j + 5
                p = p + 5

        def create_excel_for_osechi(queryset, sheetname):

            worksheet = osechi_book[sheetname]

            worksheet.cell(2, 5, in_date)

            i = 4
            j = 6
            p = 5
            food_count = 0
            page_count = 8
            max_row = 118

            for q in queryset:
                qs2 = FoodPhoto.objects.filter(food_name=q, menu=q).first()

                worksheet.cell(i, 2, qs2.food_name)

                self._set_direction_if_not_blank(worksheet, i, 4, qs2.direction)

                self._set_direction_if_not_blank(worksheet, j, 4, qs2.direction2)

                img_file = str(qs2.thumbnail)

                if (img_file):

                    img_dir = os.path.join(settings.MEDIA_ROOT, str(qs2.basic))

                    # 画像を選択
                    img_to_excel = excel.drawing.image.Image(img_dir)
                    # 指定の位置に画像を添付
                    worksheet.add_image(img_to_excel, 'B' + str(p))

                i = i + 5
                j = j + 5
                p = p + 5
                food_count += 1

            if food_count <= page_count:
                for row_index in range(max_row - i + 1):
                    worksheet.row_dimensions[i + row_index].hidden = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 1
                worksheet.row_breaks.brk = []
            elif food_count <= page_count * 2:
                for row_index in range(max_row - i + 1):
                    worksheet.row_dimensions[i + row_index].hidden = True
                new_brk = worksheet.row_breaks.brk[:1]
                new_brk.append(Break(123))
                worksheet.row_breaks.brk = new_brk

        qs_b = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='朝食', option=False).order_by('seq_order', 'id')
        qs_l = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='昼食', option=False).order_by('seq_order', 'id')
        qs_d = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='夕食', option=False).order_by('seq_order', 'id')
        qs_op_b = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='朝食', option=True).order_by('seq_order', 'id')
        qs_op_l = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='昼食', option=True).order_by('seq_order', 'id')
        qs_op_d = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='夕食', option=True).order_by('seq_order', 'id')

        create_excel(qs_b, '朝食')
        create_excel(qs_l, '昼食')
        create_excel(qs_d, '夕食')
        create_excel(qs_op_b, 'オプション・朝')
        create_excel(qs_op_l, 'オプション・昼')
        create_excel(qs_op_d, 'オプション・夕')

        # ------------------------------------------------------------------------------
        # 嚥下用の盛付指示書の取得
        # ------------------------------------------------------------------------------
        enge_ws = book['嚥下']
        enge_soup_ws = book['オプション・嚥下']

        def create_enge_excel(queryset, meal, worksheet):
            if meal == '朝食':
                start_row = 5
                max_index = 5
            elif meal == '昼食':
                start_row = 30
                max_index = 5
            else:
                start_row = 55
                max_index = 5

            r = start_row
            if is_for_new_year and (meal == '昼食'):
                worksheet.cell(r, 1, '昼　食')
                worksheet.cell(r, 3, '別紙の『おせち用盛付指示書』にて')

                # 次の料理の行まで
                r += 4

                # 夕食まで
                prev = r
                r += 5 * 4
                for row_index in range(r - prev + 1):
                    worksheet.row_dimensions[prev + row_index].hidden = True

            else:
                diff = 5
                for index, q in enumerate(queryset):
                    if index == max_index:
                        break

                    diff -= 1

                    excel_food_name = q.food_name
                    worksheet.cell(r, 2, excel_food_name)

                    direction = EngeFoodDirection.objects.get(menu=q)
                    # 説明欄1行目
                    self._set_direction_if_not_blank(worksheet, r, 3, direction.soft_direction)
                    self._set_direction_if_not_blank(worksheet, r, 4, direction.mixer_direction)
                    self._set_direction_if_not_blank(worksheet, r, 5, direction.jelly_direction)

                    # 説明欄2行目
                    self._set_direction_if_not_blank(worksheet, r + 1, 3, direction.soft_direction2)
                    self._set_direction_if_not_blank(worksheet, r + 1, 4, direction.mixer_direction2)
                    self._set_direction_if_not_blank(worksheet, r + 1, 5, direction.jelly_direction2)

                    # 説明欄3行目
                    self._set_direction_if_not_blank(worksheet, r + 2, 3, direction.soft_direction3)
                    self._set_direction_if_not_blank(worksheet, r + 2, 4, direction.mixer_direction3)
                    self._set_direction_if_not_blank(worksheet, r + 2, 5, direction.jelly_direction3)

                    # 説明欄4行目
                    self._set_direction_if_not_blank(worksheet, r + 3, 3, direction.soft_direction4)
                    self._set_direction_if_not_blank(worksheet, r + 3, 4, direction.mixer_direction4)
                    self._set_direction_if_not_blank(worksheet, r + 3, 5, direction.jelly_direction4)

                    # 説明欄5行目
                    self._set_direction_if_not_blank(worksheet, r + 4, 3, direction.soft_direction5)
                    self._set_direction_if_not_blank(worksheet, r + 4, 4, direction.mixer_direction5)
                    self._set_direction_if_not_blank(worksheet, r + 4, 5, direction.jelly_direction5)

                    r += 5

                for _ in range(diff):
                    worksheet.row_dimensions[r].hidden = True
                    worksheet.row_dimensions[r + 1].hidden = True
                    worksheet.row_dimensions[r + 2].hidden = True
                    worksheet.row_dimensions[r + 3].hidden = True
                    worksheet.row_dimensions[r + 4].hidden = True
                    r += 5

        # 喫食日
        enge_ws.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))
        enge_soup_ws.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))

        create_enge_excel(qs_b, '朝食', enge_ws)
        create_enge_excel(qs_l, '昼食', enge_ws)
        create_enge_excel(qs_d, '夕食', enge_ws)
        if in_output_enge_option == 'off':
            enge_soup_ws.sheet_state = 'hidden'
            visible_enge_option = 'hidden'
        else:
            create_enge_excel(qs_op_b, '朝食', enge_soup_ws)
            create_enge_excel(qs_op_l, '昼食', enge_soup_ws)
            create_enge_excel(qs_op_d, '夕食', enge_soup_ws)
            visible_enge_option = 'visible'

        book.save(output_file)

        # ------------------------------------------------------------------------------
        # 一部非表示にするパターンの保存
        # ------------------------------------------------------------------------------
        # 嚥下なし
        enge_ws.sheet_state = 'hidden'
        enge_soup_ws.sheet_state = 'hidden'
        no_enge_file = os.path.join(setout_output_dir, setout_name + '_noenge.xlsx')
        book.save(no_enge_file)

        # 一旦復元
        enge_ws.sheet_state = 'visible'
        enge_soup_ws.sheet_state = visible_enge_option

        # 汁オプションなし
        book['オプション・朝'].sheet_state = 'hidden'
        book['オプション・昼'].sheet_state = 'hidden'
        book['オプション・夕'].sheet_state = 'hidden'
        no_soup_file = os.path.join(setout_output_dir, setout_name + '_nosoup.xlsx')
        book.save(no_soup_file)

        # 両方なし
        enge_ws.sheet_state = 'hidden'
        enge_soup_ws.sheet_state = 'hidden'
        only_file = os.path.join(setout_output_dir, setout_name + '_only.xlsx')
        book.save(only_file)
        book.close()

        # ------------------------------------------------------------------------------
        # おせち用の盛付指示書の取得
        # ------------------------------------------------------------------------------
        def create_enge_exce_for_osechi(queryset, worksheet):
            start_row = 5
            max_item = 23

            r = start_row
            item_count = 0
            for index, q in enumerate(queryset):
                item_count += 1
                if item_count > max_item:
                    break
                excel_food_name = q.food_name
                worksheet.cell(r, 2, excel_food_name)

                direction = EngeFoodDirection.objects.get(menu=q)
                # 説明欄1行目
                self._set_direction_if_not_blank(worksheet, r, 3, direction.soft_direction1)
                self._set_direction_if_not_blank(worksheet, r, 4, direction.mixer_direction1)
                self._set_direction_if_not_blank(worksheet, r, 5, direction.jelly_direction1)

                # 説明欄2行目
                self._set_direction_if_not_blank(worksheet, r + 1, 3, direction.soft_direction2)
                self._set_direction_if_not_blank(worksheet, r + 1, 4, direction.mixer_direction2)
                self._set_direction_if_not_blank(worksheet, r + 1, 5, direction.jelly_direction2)

                # 説明欄3行目
                self._set_direction_if_not_blank(worksheet, r + 2, 3, direction.soft_direction3)
                self._set_direction_if_not_blank(worksheet, r + 2, 4, direction.mixer_direction3)
                self._set_direction_if_not_blank(worksheet, r + 2, 5, direction.jelly_direction3)

                # 説明欄4行目
                self._set_direction_if_not_blank(worksheet, r + 3, 3, direction.soft_direction4)
                self._set_direction_if_not_blank(worksheet, r + 3, 4, direction.mixer_direction4)
                self._set_direction_if_not_blank(worksheet, r + 3, 5, direction.jelly_direction4)

                # 説明欄5行目
                self._set_direction_if_not_blank(worksheet, r + 4, 3, direction.soft_direction5)
                self._set_direction_if_not_blank(worksheet, r + 4, 4, direction.mixer_direction5)
                self._set_direction_if_not_blank(worksheet, r + 4, 5, direction.jelly_direction5)

                r += 5

            if item_count <= max_item:
                for row_index in range(max_item - item_count):
                    worksheet.row_dimensions[r].hidden = True
                    worksheet.row_dimensions[r + 1].hidden = True
                    worksheet.row_dimensions[r + 2].hidden = True
                    worksheet.row_dimensions[r + 3].hidden = True
                    worksheet.row_dimensions[r + 4].hidden = True

                    r += 5

        if is_for_new_year:
            osechi_book = self.get_osechi_book()
            create_excel_for_osechi(qs_l, '昼食')

            osechi_enge_ws = osechi_book['嚥下']

            # 喫食日
            osechi_enge_ws.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))

            create_enge_exce_for_osechi(qs_l, osechi_enge_ws)

            setout_output_dir_osechi = os.path.join(settings.OUTPUT_DIR, 'setout', f'{setout_name}(おせち用)')
            os.makedirs(setout_output_dir_osechi, exist_ok=True)  # 上書きOK
            osechi_output_file = os.path.join(setout_output_dir_osechi, setout_name + '(おせち用).xlsx')

            osechi_book.save(osechi_output_file)

            # ------------------------------------------------------------------------------
            # 一部非表示にするパターンの保存
            # ------------------------------------------------------------------------------
            # 嚥下なし
            osechi_enge_ws.sheet_state = 'hidden'
            osechi_no_enge_file = os.path.join(setout_output_dir_osechi, setout_name + '(おせち用)_noenge.xlsx')
            osechi_book.save(osechi_no_enge_file)

        # ------------------------------------------------------------------------------
        # 朝食・昼食・夕食1枚シートの出力
        # ------------------------------------------------------------------------------
        setout_output_dir_1p = os.path.join(settings.OUTPUT_DIR, 'setout', f'{setout_name}(一枚表示)')
        os.makedirs(setout_output_dir_1p, exist_ok=True)  # 上書きOK
        output_file_1p = os.path.join(setout_output_dir_1p, f'{setout_name}(一枚表示).xlsx')

        xls_templat_1p = os.path.join(settings.STATICFILES_DIRS[0], 'excel/setout_1p.xlsx')  # 指示書のテンプレート

        book_1p = excel.load_workbook(xls_templat_1p)

        def create_excel_one_sheet(queryset_list, sheetname):

            worksheet = book_1p[sheetname]

            worksheet.cell(2, 5, in_date)

            i = 4
            j = 6
            p = 5

            # 朝食・昼食・夕食の最大料理数(テンプレートで有効にしている最大数を設定)
            max_plate = 5

            for index, set1 in enumerate(queryset_list):
                plate_index = 0
                if is_for_new_year and index == 1 and (sheetname == '基本食'):
                    # 元日で昼食の場合
                    worksheet.cell(i, 4, '別紙の『おせち用盛付指示書』にて')
                    i += 5
                    j += 5
                    p += 5
                    plate_index += 1
                else:
                    for q in set1:
                        qs2 = FoodPhoto.objects.filter(food_name=q, menu=q).first()

                        worksheet.cell(i, 2, qs2.food_name)

                        self._set_direction_if_not_blank(worksheet, i, 4, qs2.direction)

                        self._set_direction_if_not_blank(worksheet, j, 4, qs2.direction2)

                        img_file = str(qs2.thumbnail)

                        if (img_file):

                            img_dir = os.path.join(settings.MEDIA_ROOT, str(qs2.basic))

                            # 画像を選択
                            img_to_excel = excel.drawing.image.Image(img_dir)
                            # 指定の位置に画像を添付
                            worksheet.add_image(img_to_excel, 'B' + str(p))

                        i = i + 5
                        j = j + 5
                        p = p + 5
                        plate_index += 1

                # 不要な表示枠を非表示化
                hide_plate_count = max_plate - plate_index
                if hide_plate_count:
                    for _ in range(hide_plate_count):
                        worksheet.row_dimensions[i].hidden = True
                        worksheet.row_dimensions[i + 1].hidden = True
                        worksheet.row_dimensions[i + 2].hidden = True
                        worksheet.row_dimensions[i + 3].hidden = True
                        worksheet.row_dimensions[i + 4].hidden = True

                        i = i + 5
                        j = j + 5
                        p = p + 5

        create_excel_one_sheet([qs_b, qs_l, qs_d], '基本食')
        create_excel_one_sheet([qs_op_b, qs_op_l, qs_op_d], 'オプション')

        enge_ws_1p = book_1p['嚥下']
        enge_soup_ws_1p = book_1p['オプション・嚥下']

        # 喫食日
        enge_ws_1p.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))
        enge_soup_ws_1p.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))

        create_enge_excel(qs_b, '朝食', enge_ws_1p)
        create_enge_excel(qs_l, '昼食', enge_ws_1p)
        create_enge_excel(qs_d, '夕食', enge_ws_1p)
        if in_output_enge_option == 'off':
            enge_soup_ws_1p.sheet_state = 'hidden'
            visible_enge_option = 'hidden'
        else:
            create_enge_excel(qs_op_b, '朝食', enge_soup_ws_1p)
            create_enge_excel(qs_op_l, '昼食', enge_soup_ws_1p)
            create_enge_excel(qs_op_d, '夕食', enge_soup_ws_1p)
            visible_enge_option = 'visible'

        book_1p.save(output_file_1p)

        # ------------------------------------------------------------------------------
        # 一部非表示にするパターンの保存(一枚用)
        # ------------------------------------------------------------------------------
        # 嚥下なし
        enge_ws_1p.sheet_state = 'hidden'
        enge_soup_ws_1p.sheet_state = 'hidden'
        no_enge_file_1p = os.path.join(setout_output_dir_1p, f'{setout_name}(一枚表示)_noenge.xlsx')
        book_1p.save(no_enge_file_1p)

        # 一旦復元
        enge_ws_1p.sheet_state = 'visible'
        enge_soup_ws_1p.sheet_state = visible_enge_option

        # 汁オプションなし
        book_1p['オプション'].sheet_state = 'hidden'
        no_soup_file_1p = os.path.join(setout_output_dir_1p, f'{setout_name}(一枚表示)_nosoup.xlsx')
        book_1p.save(no_soup_file_1p)

        # 両方なし
        enge_ws_1p.sheet_state = 'hidden'
        enge_soup_ws_1p.sheet_state = 'hidden'
        only_file_1p = os.path.join(setout_output_dir_1p, f'{setout_name}(一枚表示)_only.xlsx')
        book_1p.save(only_file_1p)
        book_1p.close()

        # ------------------------------------------------------------------------------
        # 新規作成時、作成日時を保存
        # ------------------------------------------------------------------------------
        if is_create:
            obj = SetoutDuration(name=setout_name)
            # 登録日有効化のため、一旦保存
            obj.save()

            obj.last_enable = obj.create_at + relativedelta(days=14)
            obj.save()
