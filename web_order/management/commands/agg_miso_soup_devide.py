import math
import os
import shutil
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP
import platform
import openpyxl as excel

from django.conf import settings
from django_pandas.io import read_frame
from django.core.management.base import BaseCommand

from web_order.models import Order, OrderEveryday, PlatePackageForPrint, UnitPackage, TmpPlateNamePackage
from .utils import AggEngePackageMixin, ExcelOutputMixin
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister, RawPlatePackageRegisterFactory

"""
    喫食日の食数を集計し、計量表を出力する処理

# 引数
    製造日（YYYY-MM-DD）
    喫食日（YYYY-MM-DD）
    食事区分（朝・昼・夕）
    料理名
    分量

# 出力ファイル
    朝食の味噌汁の計量表
    昼食の味噌汁の計量表
    夕食の味噌汁の計量表
"""
class MisoSoupDevideMeasureWriter(AggEngePackageMixin, ExcelOutputMixin):

    def __init__(self):
        self.package_size = None
        self.enge_package_size = None
        self.in_index = None
        self.in_cook = None
        self.in_date = None
        self.in_menu = None
        self.in_name_ko = None
        self.in_qty_ko = None
        self.in_name_g = None
        self.in_qty_g = None
        self.in_is_first_miso_soup = None
        self.in_adjust = None
        self.in_unit_ko = '個'
        self.in_unit_g = 'g'
        self.agg_day_short = None
        self.aggregation_day = None
        self.basic_plate_enable = None

        # 販売固定商品注文数
        # (マスタ定義の存在するもの)
        # -保存用1人用
        self.pre_1pack_j = 0
        self.pre_1pack_u = 0

        # -見本・写真用
        self.photo = 0

        # -保存用1人用
        self.pre_1pack_s = 0
        self.pre_1pack_z = 0
        self.pre_1pack_m = 0

        # -保存用50g
        self.pre_50g_j = 0
        self.pre_50g_s = 0
        self.pre_50g_z = 0
        self.pre_50g_m = 0

    def get_fixed_1pack_count(self, is_soup: bool = False):
        if is_soup:
            # 味噌汁の汁の場合は、写真用が不要
            return self.pre_1pack_j + 1
        else:
            return self.pre_1pack_j + self.photo

    def get_enge_50g_count(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.pre_50g_s
        elif menu_name == 'ゼリー':
            return self.pre_50g_z
        else:
            return self.pre_50g_m

    def create_excel_miso_from_filling_template(self,
            dataframe, worksheet, meal, parts_name, quantity, unit, preserve, photo, enge_list=[]):
        i = 13
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)
        worksheet.cell(2, 2, parts_name)
        worksheet.cell(2, 5, quantity)
        worksheet.cell(2, 6, unit)
        worksheet.cell(7, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)

        # enge_listは、ソフト、ミキサー、ゼリーの順に並んでいる前提
        if enge_list:
            worksheet.cell(10, 4, enge_list[0])  # ソフト
            worksheet.cell(11, 4, enge_list[1])  # ミキサー
            worksheet.cell(12, 4, enge_list[2])  # ゼリー

        # 食数固定の項目を出力
        if self.basic_plate_enable:
            worksheet.cell(6, 7, 5)  # 保存用(10人袋)数量
        worksheet.cell(7, 7, self.pre_50g_j)
        worksheet.cell(8, 7, preserve)
        worksheet.cell(9, 7, photo)

        print_package_count = 7
        print_package_count_1p = self.get_fixed_1pack_count()
        print_package_count_50g = self.pre_50g_j
        self.package_size = 30
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            if row['注文数'] > 1:
                worksheet.cell(i, 4, row['注文数'])
                unit_package_count = math.ceil(row['注文数']/30)
                print_package_count += unit_package_count

                # ピッキング指示書用のユニット毎袋数を保存
                if unit == '個':
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                        row['呼出番号'], row['ユニット名'], unit_package_count, 'SOUP_UNIT', parts_name, soup_type='filling')
            else:
                worksheet.cell(i, 4, 0)
            if row['注文数'] == 1:
                print_package_count_1p += 1

                # ピッキング指示書用のユニット毎袋数を保存
                if unit == '個':
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                        row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', parts_name, soup_type='filling')

            i = i + 1

        # 袋数の出力は、分割で上書きされても同じ値になるため、特に処理は行わない
        P7Util.save_package_count_for_print(
            self.in_cooking_day, self.aggregation_day, self.in_index, print_package_count, print_package_count_1p,
            '常食', self.in_menu, print_package_count_50g)

    def create_excel_miso_from_enge_template(self,
            dataframe, worksheet, meal, name,
            needle, preserve, unit, quantity, is_write_total_needle_preserve=False):

        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)
        worksheet.cell(2, 2, name)

        worksheet.cell(5, 4, needle)        # 針刺し用の食数
        total_count = needle                # 食数集計値

        # 保存用は袋数で全体の数を表す
        worksheet.cell(7, 7, preserve)      # 保存用の食数(個)
        total_count += preserve

        if is_write_total_needle_preserve:
            worksheet.cell(6, 2, 30)            # 針刺し用の袋の注文数(固定値)

            worksheet.cell(7, 2, 10)            # 保存用の袋の注文数(固定値)

        # 袋サイズ
        is_enge = ('ソフト' in worksheet.title) or ('ゼリー' in worksheet.title) or ('ミキサー' in worksheet.title)
        if is_enge:
            package_size = self.get_miso_soup_package_size(unit, quantity)
            self.enge_package_size = package_size
            enge = ''
            if 'ソフト' in worksheet.title:
                enge = 'ソフト'
            elif 'ゼリー' in worksheet.title:
                enge = 'ゼリー'
            elif 'ミキサー' in worksheet.title:
                enge = 'ミキサー'
            enge_50g_count = self.get_enge_50g_count(enge)
            print_package_count_50g = enge_50g_count
            worksheet.cell(6, 7, enge_50g_count)  # 保存用(50)の袋数
        else:
            package_size = 30
            enge = 'なし'
            print_package_count_50g = self.pre_50g_j
        self.package_size = package_size
        total_count += print_package_count_50g

        i = 10
        print_package_count = 0             # 針刺し用(30人袋)と保存用(10人袋)の枚数
        print_package_count_1p = preserve
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            worksheet.cell(i, 4, row['注文数'])
            total_count += row['注文数']
            if row['注文数'] > 1:
                unit_package_count = math.ceil(row['注文数'] / package_size)
                print_package_count += unit_package_count


                # ピッキング指示書用のユニット毎袋数を保存
                if is_enge:
                    if row['注文数'] == 2:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                            row['呼出番号'], row['ユニット名'], 1, 'ENGE_2', name, enge, soup_type='filling')
                    else:
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                            row['呼出番号'], row['ユニット名'], unit_package_count, f'ENGE_{package_size}', name, enge, soup_type='filling')
                else:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                        row['呼出番号'], row['ユニット名'], unit_package_count, 'SOUP_UNIT', name, soup_type='filling')
            if row['注文数'] == 1:
                print_package_count_1p += 1
                if is_enge:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                        row['呼出番号'], row['ユニット名'], 1, 'ENGE_1', name, enge, soup_type='filling')
                else:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                        row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', name, soup_type='filling')

            worksheet.cell(i, 7, self.get_miso_soup_package_function(unit, quantity, i))

            i = i + 1

        if 'ソフト' in worksheet.title:
            menu = 'ソフト'
        elif 'ゼリー' in worksheet.title:
            menu = 'ゼリー'
        elif 'ミキサー' in worksheet.title:
            menu = 'ミキサー'

        if self.in_adjust == 2:
            if self.in_index != -1:
                enge_index = self.in_index + 1
            else:
                enge_index = self.index
        else:
            enge_index = self.in_index

        P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, print_package_count, print_package_count_1p, menu, self.in_menu, print_package_count_50g)

        # 汁・汁具でin_adjustが1になることは運用上ない。

        return total_count

    def create_excel_miso_soup(self,
            dataframe, worksheet, meal,
            is_write_total_needle_preserve=False):
        worksheet.cell(1, 9, self.agg_day_short)
        worksheet.cell(1, 14, meal)

        total = 0
        print_package_count = 0                 # サイズに関係なく、枚数を記録する
        print_package_count_1p = 0
        print_package_count_50g = 0
        if is_write_total_needle_preserve:
            worksheet.cell(5, 2, 30)            # 針刺し用の袋の注文数(固定値)
            worksheet.cell(6, 2, 10)            # 保存用の袋の注文数(固定値)
            print_package_count += 2            # 針刺し(30人の袋)
            print_package_count_1p += self.get_fixed_1pack_count(True)  # 見本(味噌汁の汁では、写真用を使わない)
            print_package_count_50g = self.pre_50g_j
        if self.basic_plate_enable:
            worksheet.cell(6, 9, '=IF(B6=0,0,5)')         # 保存用(10人袋)袋数
            if is_write_total_needle_preserve:
                print_package_count += 5                      # 保存用(10人の袋)

        i = 9
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            worksheet.cell(i, 4, row['注文数'])
            total += int(row['注文数'])
            if row['注文数'] > 1:
                unit_package_count = math.ceil(row['注文数'] / 30)
                print_package_count += unit_package_count

                if self.in_index != -1:
                    # アレルギーの計量表では、汁を出力しない
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index + 1,
                        row['呼出番号'], row['ユニット名'], unit_package_count, 'SOUP_UNIT', '味噌汁', soup_type='soup')
            if row['注文数'] == 1:
                print_package_count_1p += 1
                if self.in_index != -1:
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index + 1,
                        row['呼出番号'], row['ユニット名'], 1, 'SOUP_1', '味噌汁', soup_type='soup')

            i = i + 1

        self.save_soup_package_count_for_print(print_package_count, print_package_count_1p, print_package_count_50g)
        return total

    def read_excel_items_quantity(self, worksheet, quantity):
        """
        味噌汁の食材の1人分の数量を読み込む。常食と薄味は読み込むシートが異なるため注意。
        """
        worksheet.cell(4, 5, quantity)
        worksheet.cell(4, 6, 0)
        worksheet.cell(4, 7, 0)

        # 1人分の数量はテンプレートファイル上にしか定義していない
        katsukonmi = worksheet['E4'].value
        miso = worksheet['F4'].value
        water = worksheet['G4'].value

        # 結果を/区切りで返す
        return "{0}/{1}/{2}".format(katsukonmi, miso, water)

    def save_soup_package_count_for_print(self, count: int, count_1p: int, count_50g: int):
        if self.in_index == -1:
            return

        # 前のインデックスに汁の料理があったら、そちらで登録済みなので、ここでは登録しない
        plate_package_qs_lt = PlatePackageForPrint.objects.filter(
            cooking_day=self.in_cooking_day, eating_day=self.aggregation_day,
            is_basic_plate=True, meal_name=self.in_menu, index__lt=self.in_index, menu_name='常食')
        for plate_lt in plate_package_qs_lt:
            if '⑤' in plate_lt.plate_name:
                return

        # 汁の袋数を登録(汁の料理に「希釈」が含まれている前提)
        plate_package_qs = PlatePackageForPrint.objects.filter(
            cooking_day=self.in_cooking_day, eating_day=self.aggregation_day,
            is_basic_plate=True, meal_name=self.in_menu, index__gt=self.in_index, menu_name='常食')
        for plate_package in plate_package_qs:
            if '希釈' in plate_package.plate_name:
                plate_package.count = count
                plate_package.count_one_p = count_1p
                plate_package.count_one_50g = count_50g
                plate_package.save()
                break

    def handle(self, arg_index, arg_cook, arg_date, arg_menu, arg_name_ko, arg_qty_ko, arg_name_g, arg_qty_g,
               arg_is_first, adjust, manager, before_name, liquid_quantity):
        self.in_index = arg_index      # 呼び出し時の引数1つ目、インデックス(-1は対象不明)
        self.in_cook = arg_cook         # 呼び出し時の引数2つ目、製造日（YYYY-MM-DD）
        self.in_date = arg_date
        self.in_menu = arg_menu
        self.in_name_ko = arg_name_ko
        self.in_qty_ko = arg_qty_ko
        self.in_name_g = arg_name_g            # 呼び出し時の引数 --name2 名前(g数の食材)
        self.in_qty_g = arg_qty_g
        self.in_is_first_miso_soup = arg_is_first
        self.in_adjust = adjust

        self.aggregation_day = dt.datetime.strptime(self.in_date, '%Y-%m-%d').date()
        if platform.system() == 'Windows':
            self.agg_day_short = self.aggregation_day.strftime('%m/%d')
        else:
            self.agg_day_short = self.aggregation_day.strftime('%-m/%-d')
        enable_day = dt.datetime.strptime(settings.BASIC_PLATE_ENABLE_DATE, '%Y-%m-%d').date()
        if self.aggregation_day >= enable_day:
            self.basic_plate_enable = True
        else:
            self.basic_plate_enable = False

        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + self.in_cook + '_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        miso_filling_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_miso_filling.xlsx')  # 味噌汁用の具(常食・薄味)の計量表テンプレ

        # 製造日の取得
        self.in_cooking_day = dt.datetime.strptime(self.in_cook, '%Y-%m-%d').date()

        # 注文対象日・食事区分の設定
        manager.set_eating(self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # 針刺し用の食数
        # ------------------------------------------------------------------------------
        res_s, res_m, res_z = manager.get_needle_orders()


        # ------------------------------------------------------------------------------
        # 嚥下保存用(1人用)の食数
        # ------------------------------------------------------------------------------
        self.pre_1pack_s, self.pre_1pack_m, self.pre_1pack_z = manager.get_preserve_1p_enge_orders()

        # ------------------------------------------------------------------------------
        # 常食保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        self.pre_1pack_j, self.pre_1pack_u = manager.get_preserve_1p_orders()

        # ------------------------------------------------------------------------------
        # 写真用の食数
        # ------------------------------------------------------------------------------
        self.photo = manager.get_photo_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(50g)
        # ------------------------------------------------------------------------------
        self.pre_50g_j, self.pre_50g_s, self.pre_50g_z, self.pre_50g_m = manager.get_preserve_50g_orders()

        # ------------------------------------------------------------------------------
        # qs（Queryset）df（Dataframe）
        # 朝食（b）昼食（l）夕食（d）
        # 常食（j）薄味（u）ソフト（s）ミキサー（m）ゼリー（z）
        # 具（g） 汁（s）
        # ------------------------------------------------------------------------------

        # 嚥下用(今は仮で個だけを使用)
        if self.basic_plate_enable:
            miso_enge_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_miso_enge_v2.xlsx')  # 味噌汁用の具(嚥下)・汁の計量表テンプレ
        else:
            miso_enge_template = os.path.join(settings.STATICFILES_DIRS[0],
                                              'excel/measure_miso_enge.xlsx')  # 味噌汁用の具(嚥下)・汁の計量表テンプレ

        book_bm = excel.load_workbook(miso_enge_template)
        j_items_quantity = self.read_excel_items_quantity(book_bm["汁_常食"], liquid_quantity)
        u_items_quantity = self.read_excel_items_quantity(book_bm["汁_薄味"], liquid_quantity)
        if self.in_menu == '朝食':
            # ソフト食
            df_bs_g = manager.get_def_miso_soft_filling()
            bs_total = self.create_excel_miso_from_enge_template(
                df_bs_g, book_bm["具_ソフト"], '△ 朝',
                self.in_name_ko + '・' + self.in_name_g,
                res_s, self.pre_1pack_s, self.in_unit_ko, self.in_qty_ko)

            # ミキサー食
            df_bm_g = manager.get_def_miso_mixer_filling()
            bm_total = self.create_excel_miso_from_enge_template(
                df_bm_g, book_bm["具_ミキサー"], '△ 朝',
                self.in_name_ko + '・' + self.in_name_g,
                res_m, self.pre_1pack_m, self.in_unit_ko, self.in_qty_ko)

            # ゼリー食
            df_bz_g = manager.get_def_miso_jelly_filling()
            bz_total = self.create_excel_miso_from_enge_template(
                df_bz_g, book_bm["具_ゼリー"], '△ 朝',
                self.in_name_ko + '・' + self.in_name_g,
                res_z, self.pre_1pack_z, self.in_unit_ko, self.in_qty_ko)

            # 基本食(常食)-汁
            df_bj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_bj_s, book_bm["汁_常食"], '△ 朝', self.in_is_first_miso_soup)

            # 薄味
            """
            qs_bu_s = qs.filter(meal_name__meal_name='朝食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_bu_s.exists():
                df_bu_s = self.make_df_soup(qs_bu_s)
                u_total = self.create_excel_miso_soup(df_bu_s, book_bm["汁_薄味"], '△ 朝', self.in_is_first_miso_soup)
            else:
                book_bm.remove(book_bm["汁_薄味"])
                u_total = 0
            """
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            breakfast_output_miso_enge = os.path.join(new_dir_path,
                                                      str(self.aggregation_day) + '_朝汁_' + self.in_name_ko + '・' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, breakfast_output_miso_enge)

        if self.in_menu == '昼食':
            # ソフト食
            df_ls_g = manager.get_def_miso_soft_filling()
            ls_total = self.create_excel_miso_from_enge_template(
                df_ls_g, book_bm["具_ソフト"], '○ 昼',
                self.in_name_ko + '・' + self.in_name_g,
                res_s, self.pre_1pack_s, self.in_unit_ko, self.in_qty_ko)

            # ミキサー食
            df_lm_g = manager.get_def_miso_mixer_filling()
            lm_total = self.create_excel_miso_from_enge_template(
                df_lm_g, book_bm["具_ミキサー"], '○ 昼',
                self.in_name_ko + '・' + self.in_name_g,
                res_m, self.pre_1pack_m, self.in_unit_ko, self.in_qty_ko)

            # ゼリー食
            df_lz_g = manager.get_def_miso_jelly_filling()
            lz_total = self.create_excel_miso_from_enge_template(
                df_lz_g, book_bm["具_ゼリー"], '○ 昼',
                self.in_name_ko + '・' + self.in_name_g,
                res_z, self.pre_1pack_z, self.in_unit_ko, self.in_qty_ko)

            # 基本食(常食)-汁
            df_lj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_lj_s, book_bm["汁_常食"], '○ 昼', self.in_is_first_miso_soup)

            # 薄味
            """
            qs_lu_s = qs.filter(meal_name__meal_name='昼食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_lu_s.exists():
                df_lu_s = self.make_df_soup(qs_lu_s)
                u_total = self.create_excel_miso_soup(df_lu_s, book_bm["汁_薄味"], '○ 昼', self.in_is_first_miso_soup)
            else:
                book_bm.remove(book_bm["汁_薄味"])
                u_total = 0
            """
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            lunch_output_miso_enge = os.path.join(new_dir_path,
                                                  str(self.aggregation_day) + '_昼汁_' + self.in_name_ko + '・' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, lunch_output_miso_enge)

        if self.in_menu == '夕食':
            # ソフト食
            df_ds_g = manager.get_def_miso_soft_filling()
            ds_total = self.create_excel_miso_from_enge_template(
                df_ds_g, book_bm["具_ソフト"], '□ 夕',
                self.in_name_ko + '・' + self.in_name_g,
                res_s, self.pre_1pack_s, self.in_unit_ko, self.in_qty_ko)

            # ミキサー食
            df_dm_g = manager.get_def_miso_mixer_filling()
            dm_total = self.create_excel_miso_from_enge_template(
                df_dm_g, book_bm["具_ミキサー"], '□ 夕',
                self.in_name_ko + '・' + self.in_name_g,
                res_m, self.pre_1pack_m, self.in_unit_ko, self.in_qty_ko)

            # ゼリー
            df_dz_g = manager.get_def_miso_jelly_filling()
            dz_total = self.create_excel_miso_from_enge_template(
                df_dz_g, book_bm["具_ゼリー"], '□ 夕',
                self.in_name_ko + '・' + self.in_name_g,
                res_z, self.pre_1pack_z, self.in_unit_ko, self.in_qty_ko)

            # 基本食(常食)-汁
            df_dj_s = manager.get_def_miso_soup()
            j_total = self.create_excel_miso_soup(df_dj_s, book_bm["汁_常食"], '□ 夕', self.in_is_first_miso_soup)

            # 薄味
            """
            qs_du_s = qs.filter(meal_name__meal_name='夕食', menu_name__group='薄味', meal_name__soup=True)  # 薄味 汁あり
            if qs_du_s.exists():
                df_du_s = self.make_df_soup(qs_du_s)
                u_total = self.create_excel_miso_soup(df_du_s, book_bm["汁_薄味"], '□ 夕', self.in_is_first_miso_soup)
            else:
                book_bm.remove(book_bm["汁_薄味"])
                u_total = 0
            """
            book_bm.remove(book_bm["汁_薄味"])
            u_total = 0

            dinner_output_miso_enge = os.path.join(new_dir_path,
                                                   str(self.aggregation_day) + '_夕汁_' + self.in_name_ko + '・' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, dinner_output_miso_enge)
        # ------------------------------------------------------------------------------

        # 具(個数)用
        book_bm = excel.load_workbook(miso_filling_template)
        if self.in_menu == '朝食':
            # 基本食(常食)-具
            df_bj_g = manager.get_def_miso_filling()
            self.create_excel_miso_from_filling_template(
                df_bj_g, book_bm["具_常食"], '△ 朝', self.in_name_ko, self.in_qty_ko, self.in_unit_ko, self.pre_1pack_j, self.photo, [bs_total, bm_total, bz_total])

            # 薄味
            """
            qs_bu_g = qs.filter(meal_name__meal_name='朝食', menu_name=2, meal_name__filling=True)  # 薄味 具あり
            if qs_bu_g.exists():
                df_bu_g = self.make_df_filling(qs_bu_g)
                self.create_excel_miso_from_filling_template(df_bu_g, book_bm["具_薄味"], '△ 朝', self.in_name_ko, self.in_qty_ko, self.in_unit_ko, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
                u_total = 0
            """
            book_bm.remove(book_bm["具_薄味"])
            u_total = 0

            breakfast_output_miso_ko = os.path.join(new_dir_path,
                                                    str(self.aggregation_day) + '_朝汁_' + self.in_name_ko + '.xlsx')
            self.save_with_select(book_bm, breakfast_output_miso_ko)

        if self.in_menu == '昼食':
            # 基本食(常食)-具
            df_lj_g = manager.get_def_miso_filling()
            self.create_excel_miso_from_filling_template(
                df_lj_g, book_bm["具_常食"], '○ 昼', self.in_name_ko, self.in_qty_ko, self.in_unit_ko,self.pre_1pack_j, self.photo, [ls_total, lm_total, lz_total])

            # 薄味
            book_bm.remove(book_bm["具_薄味"])
            u_total = 0

            lunch_output_miso_ko = os.path.join(new_dir_path,
                                                str(self.aggregation_day) + '_昼汁_' + self.in_name_ko + '.xlsx')
            self.save_with_select(book_bm, lunch_output_miso_ko)

        if self.in_menu == '夕食':
            df_dj_g = manager.get_def_miso_filling()
            self.create_excel_miso_from_filling_template(
                df_dj_g, book_bm["具_常食"], '□ 夕', self.in_name_ko, self.in_qty_ko, self.in_unit_ko, self.pre_1pack_j, self.photo, [ds_total, dm_total, dz_total])

            # 薄味
            """
            qs_du_g = qs.filter(meal_name__meal_name='夕食', menu_name=2, meal_name__filling=True)
            if qs_du_g.exists():
                df_du_g = self.make_df_filling(qs_du_g)
                self.create_excel_miso_from_filling_template(df_du_g, book_bm["具_薄味"], '□ 夕', self.in_name_ko, self.in_qty_ko, self.in_unit_ko, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
                u_total = 0
            """
            book_bm.remove(book_bm["具_薄味"])
            u_total = 0

            dinner_output_miso_ko = os.path.join(new_dir_path,
                                                 str(self.aggregation_day) + '_夕汁_' + self.in_name_ko + '.xlsx')
            self.save_with_select(book_bm, dinner_output_miso_ko)
        # ------------------------------------------------------------------------------
        # g用
        book_bm = excel.load_workbook(miso_filling_template)
        if self.in_menu == '朝食':
            # 基本食(常食)-具
            self.create_excel_miso_from_filling_template(
                df_bj_g, book_bm["具_常食"], '△ 朝', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_j, self.photo, [bs_total, bm_total, bz_total])

            # 薄味
            """
            if qs_bu_g.exists():
                self.create_excel_miso_from_filling_template(df_bu_g, book_bm["具_薄味"], '△ 朝', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            breakfast_output_miso_g = os.path.join(new_dir_path,
                                                   str(self.aggregation_day) + '_朝汁_味噌汁 ' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, breakfast_output_miso_g)

        if self.in_menu == '昼食':
            # 基本食(常食)-具
            self.create_excel_miso_from_filling_template(
                df_lj_g, book_bm["具_常食"], '○ 昼', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_j, self.photo, [ls_total, lm_total, lz_total])

            # 薄味
            """
            if qs_lu_g.exists():
                self.create_excel_miso_from_filling_template(df_lu_g, book_bm["具_薄味"], '○ 昼', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            lunch_output_miso_g = os.path.join(new_dir_path,
                                               str(self.aggregation_day) + '_昼汁_味噌汁 ' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, lunch_output_miso_g)

        if self.in_menu == '夕食':
            # 基本食(常食)-具
            self.create_excel_miso_from_filling_template(
                df_dj_g, book_bm["具_常食"], '□ 夕', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_j, self.photo, [ds_total, dm_total, dz_total])

            # 薄味
            """
            if qs_du_g.exists():
                self.create_excel_miso_from_filling_template(df_du_g, book_bm["具_薄味"], '□ 夕', self.in_name_g, self.in_qty_g, self.in_unit_g, self.pre_1pack_u, 0)
            else:
                book_bm.remove(book_bm["具_薄味"])
            """
            book_bm.remove(book_bm["具_薄味"])

            dinner_output_miso_g = os.path.join(new_dir_path,
                                                str(self.aggregation_day) + '_夕汁_味噌汁 ' + self.in_name_g + '.xlsx')
            self.save_with_select(book_bm, dinner_output_miso_g)
        # ------------------------------------------------------------------------------

        if self.in_index == -1:
            if self.package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.package_size,
                    menu_name='常食'
                )
            if self.enge_package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.enge_package_size,
                    menu_name='嚥下'
                )

        return "{0}:{1}:{2}:{3}".format(j_items_quantity, u_items_quantity, j_total, u_total)

