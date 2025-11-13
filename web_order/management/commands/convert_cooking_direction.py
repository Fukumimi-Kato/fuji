import os
import pandas as pd
import datetime as dt
import openpyxl as excel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
import platform
import re
import shutil

from django_pandas.io import read_frame
from django.core.management.base import BaseCommand
from django.conf import settings
from django.db.models import Sum

from web_order.models import Order, OrderEveryday

# ログファイルの「↓」はpd.concatで「→」はpd.mergeでの結合を示す

class Command(BaseCommand):
    def __init__(self):
        self.thin_side = Side(style='thin')
        self.eating_left_border = Border(top=self.thin_side, bottom=self.thin_side, left=self.thin_side)
        self.eating_right_border = Border(top=self.thin_side, bottom=self.thin_side, right=self.thin_side)
        self.eating_other_border = Border(top=self.thin_side, bottom=self.thin_side)

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs='+', type=str)
        parser.add_argument('--type', default='normal', type=str)

    def get_count_from_detail(self, detail):
        sp = detail.split()
        return int(sp[0])

    def get_count_from_detail_by_joshoku(self, details):
        base_code = settings.COOKING_DIRECTION_J_CODE
        base_code2 = settings.COOKING_DIRECTION_B_CODE
        allergen_codes = ['常', '基', 'ソ', 'ミ', 'ゼ']
        allergen_codes_usuaji = ['薄', ]

        total = 0
        usuaji_quantity = 0
        for d in details:
            sp = d.split()
            if sp[1].find(base_code) == 0:
                total += int(sp[0])
            elif sp[1].find(base_code2) == 0:
                total += int(sp[0])
            else:
                filtered = [x for x in allergen_codes if sp[1].find(x) == 0]
                filtered_usuaji = [x for x in allergen_codes_usuaji if sp[1].find(x) == 0]
                if filtered:
                    pass
                elif filtered_usuaji:
                    # 薄味が常食と同じ分量の食材の場合
                    total += int(sp[0])
                    usuaji_quantity += int(sp[0])
                else:
                    # フリーズ分はここに入るため、totalに加算される
                    total += int(sp[0])

        return (total, usuaji_quantity)

    def get_count_from_detail_by_usuaji(self, details):
        base_code = settings.COOKING_DIRECTION_U_CODE
        allergen_codes = ['薄', ]

        total = 0
        for d in details:
            sp = d.split()
            if sp[1].find(base_code) == 0:
                total += int(sp[0])
            else:
                filtered = [x for x in allergen_codes if sp[1].find(x) == 0]
                if filtered:
                    pass
                else:
                    # 薄味でここにくることはないはず
                    total += int(sp[0])

        return (total, 0)

    def get_total_font(self, ws):
        default_font = ws['B10'].font
        return Font(name=default_font.name, color=default_font.color, size=12)

    def get_detail_font(self, ws):
        default_font = ws['C10'].font
        return Font(name=default_font.name, color=default_font.color, size=10)

    def re_calc(self, ws):
        count = None
        irow = 9
        for row in ws.iter_rows(min_row=irow):
            # 献立名(D列)に値
            if row[3].value:
                count = float(row[1].value)

            # 総量(I列)に値
            try:
                quantity_sum = float(row[8].value or 0.0)
            except:
                # 念のため変換エラーを対応
                quantity_sum = 0.0
            if quantity_sum:

                # 一人分使用量(数量)
                res_q = re.findall('(\d)/(\d)', row[6].value)
                if res_q:
                    quantity_one = float(res_q[0][0]) / float(res_q[0][1])
                else:
                    quantity_one = float(row[6].value)

                # 一人分使用量(単位)
                unit_one = row[7].value
                unit_one = unit_one.replace('ｇ', 'g')
                unit_one = unit_one.replace('ｋｇ', 'g')

                # 総量(単位)
                unit_sum = row[9].value
                unit_sum = unit_sum.replace('ｇ', 'g')
                unit_sum = unit_sum.replace('ｋｇ', 'g')

                calculated = quantity_one * count
                if (unit_one == 'g') and (unit_sum == 'kg'):
                    calculated = calculated / 1000

                row[8].value = calculated
                if (unit_sum == 'g') or (unit_sum == 'kg'):
                    if (calculated - int(calculated)) > 0:
                        row[8].number_format = '0.00'
                irow += 1

    def handle(self, *args, **options):

        #######################################################
        # らくらく献立仕様外の形式に調理表を変換する
        # 食数内訳には常食+薄味でアレルギーを含めた全食数を記載し、別途その料理を使用するアレルギーを記述する
        # 食数(B列、合計)には、食数内訳の合計ではなく、常食+薄味でアレルギーを含めた全食数を記述する。
        #######################################################

        in_file = options['filename'][0]  # 呼び出し時の引数1つ目
        in_type = options['type']          # 呼び出し時の引数 --type 調理表の種類('normal': 通常の調理表, 'filtered': 常食・薄味で分けた調理表)

        if in_type == 'normal':
            converted_output_dir = os.path.join(settings.OUTPUT_DIR, 'converted_cook')
        else:
            converted_output_dir = os.path.join(settings.OUTPUT_DIR, 'converted_cook_filter')

        os.makedirs(converted_output_dir, exist_ok=True)  # 上書きOK
        new_dir_path = os.path.join(converted_output_dir, in_file)

        # xlsだとopenpyxlで読み込めないため、変換が必要(<-運用で対処)
        cook_direc_file = os.path.join(settings.MEDIA_ROOT, 'upload', in_file)

        workbook = excel.load_workbook(cook_direc_file)
        worksheet = workbook.worksheets[0]

        # -------------------------------------------------------------
        # 食数の再計算
        # -------------------------------------------------------------
        eating_count = 0
        eating_time = None
        plate = None
        cell_of_sum = 0
        detail_list = []
        has_joshoku = False
        has_usuaji = False
        joshoku_count = 0
        usuaji_count = 0
        row_index = 9
        total_font = self.get_total_font(worksheet)
        detail_font = self.get_detail_font(worksheet)
        usiaji_allergen_in_joshoku_dict = {}
        for row in worksheet.iter_rows(min_row=row_index):
            # 食種(A列)に値
            if row[0].value:
                # 食種(A列)に記載の喫食日を、変換でA列を非表示にしても出力できるように、B列にコピー
                worksheet.unmerge_cells(f'A{row_index}:O{row_index}')
                worksheet.cell(row_index, 2, row[0].value)

                # 見た目の修正
                worksheet.row_dimensions[row_index].height = 20
                worksheet[f'B{row_index}'].border = self.eating_left_border
                worksheet[f'B{row_index}'].font = total_font
                worksheet[f'C{row_index}'].border = self.eating_other_border
                worksheet[f'D{row_index}'].border = self.eating_other_border
                worksheet[f'E{row_index}'].border = self.eating_other_border
                worksheet[f'F{row_index}'].border = self.eating_other_border
                worksheet[f'G{row_index}'].border = self.eating_other_border
                worksheet[f'H{row_index}'].border = self.eating_other_border
                worksheet[f'I{row_index}'].border = self.eating_other_border
                worksheet[f'J{row_index}'].border = self.eating_other_border
                worksheet[f'K{row_index}'].border = self.eating_right_border

                eating_time = row[0].value
            else:
                # 文字サイズを大きくする対応
                worksheet[f'B{row_index}'].font = total_font
                worksheet[f'D{row_index}'].alignment = Alignment(wrapText=True, vertical='center')
                worksheet[f'D{row_index}'].font = detail_font
                worksheet[f'F{row_index}'].alignment = Alignment(wrapText=True, vertical='center')
                worksheet[f'F{row_index}'].font = detail_font
                worksheet[f'G{row_index}'].font = detail_font
                worksheet[f'I{row_index}'].font = detail_font
                value = worksheet[f'D{row_index}'].value or worksheet[f'F{row_index}'].value
                if value:
                    name_len = int(len(str(value)))
                    if name_len:
                        div = name_len / 8
                        if div >= 1:
                            h = 20 + 8 * div
                            worksheet.row_dimensions[row_index].height = h

            # 食数(B列)に値
            if row[1].value:
                # 献立切り替え
                if eating_count:
                    # 処理中の献立の解析を終了する処理

                    # 常食の内訳の有無・食数を取得
                    joshoku = [x for x in detail_list if settings.COOKING_DIRECTION_J_CODE in x]
                    if joshoku:
                        has_joshoku = True
                        joshoku_count = self.get_count_from_detail(joshoku[0])

                    # 基本食の内訳の有無・食数を取得(常食に含める)
                    kihon = [x for x in detail_list if settings.COOKING_DIRECTION_B_CODE in x]
                    if kihon:
                        has_joshoku = True
                        joshoku_count += self.get_count_from_detail(kihon[0])

                    # 薄味の内訳の有無・食数を取得
                    usuaji = [x for x in detail_list if settings.COOKING_DIRECTION_U_CODE in x]
                    if usuaji:
                        has_usuaji = True
                        usuaji_count = self.get_count_from_detail(usuaji[0])

                    # フリーズの内訳の食数を取得
                    freeze = [x for x in detail_list if settings.COOKING_DIRECTION_F_CODE in x]
                    if freeze:
                        freeze_count = self.get_count_from_detail(freeze[0])
                    else:
                        freeze_count = 0

                    # 木沢個食の内訳の食数を取得
                    kizawa = [x for x in detail_list if settings.KIZAWA_RAKUKON_CODE in x]
                    if kizawa:
                        kizawa_count = self.get_count_from_detail(kizawa[0])
                    else:
                        kizawa_count = 0

                    detail_count = len(detail_list)

                    if (detail_count > 2) and has_joshoku and has_usuaji:
                        cell_of_sum.value = joshoku_count + usuaji_count + freeze_count + kizawa_count
                    elif detail_count >= 2:
                        # 常食・薄味の食数集計食種のどちらかがあるのに、その反対となるアレルギーが内訳に載ることはない前提
                        # 例)
                        # 「100 常・基本食(施設)」と「1 薄肉ｱﾚ・基本食(施設)」が内訳に記載、
                        # 「薄・基本食(施設)」がない、という状態はありえない想定。
                        if has_joshoku and (not has_usuaji):
                            cell_of_sum.value, usuaji_quatity = self.get_count_from_detail_by_joshoku(detail_list)
                            if usuaji_quatity:
                                usiaji_allergen_in_joshoku_dict[(eating_time, plate)] = usuaji_quatity

                        elif (not has_joshoku) and has_usuaji:
                            cell_of_sum.value, _ = self.get_count_from_detail_by_usuaji(detail_list)

                    detail_list = []
                    has_joshoku = False
                    has_usuaji = False
                    joshoku_count = 0
                    usuaji_count = 0

                eating_count = row[1].value
                cell_of_sum = row[1]
                plate = row[3].value
                if row[2].value:
                    detail_list.append(row[2].value)
            else:
                # 内訳に値あり
                if row[2].value:
                    detail_list.append(row[2].value)

            # 備考(コメント2)のクリア
            worksheet.cell(row_index, 11, '')

            row_index += 1

        # 最終料理の計算
        if detail_list:
            # 常食の内訳の有無・食数を取得
            joshoku = [x for x in detail_list if settings.COOKING_DIRECTION_J_CODE in x]
            if joshoku:
                has_joshoku = True
                joshoku_count = self.get_count_from_detail(joshoku[0])

            # 基本食の内訳の有無・食数を取得(常食に含める)
            kihon = [x for x in detail_list if settings.COOKING_DIRECTION_B_CODE in x]
            if kihon:
                has_joshoku = True
                joshoku_count += self.get_count_from_detail(kihon[0])

            # 薄味の内訳の有無・食数を取得
            usuaji = [x for x in detail_list if settings.COOKING_DIRECTION_U_CODE in x]
            if usuaji:
                has_usuaji = True
                usuaji_count = self.get_count_from_detail(usuaji[0])

            # フリーズの内訳の食数を取得
            freeze = [x for x in detail_list if settings.COOKING_DIRECTION_F_CODE in x]
            if freeze:
                freeze_count = self.get_count_from_detail(freeze[0])
            else:
                freeze_count = 0

            # 木沢個食の内訳の食数を取得
            kizawa = [x for x in detail_list if settings.KIZAWA_RAKUKON_CODE in x]
            if kizawa:
                kizawa_count = self.get_count_from_detail(kizawa[0])
            else:
                kizawa_count = 0

            detail_count = len(detail_list)

            if (detail_count > 2) and has_joshoku and has_usuaji:
                cell_of_sum.value = joshoku_count + usuaji_count + freeze_count + kizawa_count
            elif detail_count >= 2:
                # 常食・薄味の食数集計食種のどちらかがあるのに、その反対となるアレルギーが内訳に載ることはない前提
                # 例)
                # 「100 常・基本食(施設)」と「1 薄肉ｱﾚ・基本食(施設)」が内訳に記載、
                # 「薄・基本食(施設)」がない、という状態はありえない想定。
                if has_joshoku and (not has_usuaji):
                    cell_of_sum.value, usuaji_quatity = self.get_count_from_detail_by_joshoku(detail_list)
                    if usuaji_quatity:
                        usiaji_allergen_in_joshoku_dict[(eating_time, plate)] = usuaji_quatity

                elif (not has_joshoku) and has_usuaji:
                    cell_of_sum.value, _ = self.get_count_from_detail_by_usuaji(detail_list)

        # -------------------------------------------------------------
        # 常食のみの食事の中に薄味の食数があったものの食数を補正
        # -------------------------------------------------------------
        detail_list = []
        eating_count = 0
        eating_time = None
        plate = None
        has_joshoku = False
        has_usuaji = False
        if any(usiaji_allergen_in_joshoku_dict):
            row_index = 9
            for row in worksheet.iter_rows(min_row=row_index):
                # 食種(A列)に値
                if row[0].value:
                    eating_time = row[0].value

                # 食数(B列)に値
                if row[1].value:
                    # 献立切り替え
                    if eating_count:
                        # 処理中の献立の解析を終了する処理

                        # 常食の内訳の有無・食数を取得
                        joshoku = [x for x in detail_list if settings.COOKING_DIRECTION_J_CODE in x]
                        if joshoku:
                            has_joshoku = True

                        # 薄味の内訳の有無・食数を取得
                        usuaji = [x for x in detail_list if settings.COOKING_DIRECTION_U_CODE in x]
                        if usuaji:
                            has_usuaji = True

                        detail_count = len(detail_list)

                        if (detail_count > 2) and has_joshoku and has_usuaji:
                            # 常食,薄味の両方に記載がある場合は、調整しない
                            pass
                        elif detail_count >= 2:
                            # 薄味のみの内訳の場合
                            if (not has_joshoku) and has_usuaji:
                                if (eating_time, plate) in usiaji_allergen_in_joshoku_dict:
                                    eating_count -= usiaji_allergen_in_joshoku_dict[(eating_time, plate)]
                                    cell_of_sum.value = eating_count

                        detail_list = []
                        has_joshoku = False
                        has_usuaji = False

                    eating_count = row[1].value
                    cell_of_sum = row[1]
                    plate = row[3].value
                    if row[2].value:
                        detail_list.append(row[2].value)
                else:
                    # 内訳に値あり
                    if row[2].value:
                        detail_list.append(row[2].value)

                row_index += 1

        # -------------------------------------------------------------
        # 材料合計の再計算
        # -------------------------------------------------------------
        self.re_calc(worksheet)

        # -------------------------------------------------------------
        # その他レイアウトの設定
        # -------------------------------------------------------------
        # 印刷範囲の再設定
        worksheet.print_area = f'A1:K{row_index-1}'

        # 列の非表示
        worksheet.column_dimensions['A'].width = 0
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 27
        worksheet.column_dimensions['E'].width = 0
        worksheet.column_dimensions['H'].width = 3
        worksheet.column_dimensions['J'].width = 3
        worksheet.column_dimensions['G'].width = 7
        worksheet.column_dimensions['I'].width = 7

        # 余白
        worksheet.page_margins.left = 0.4
        worksheet.page_margins.right = 0.4

        # 列タイトルの変更
        worksheet.cell(7, 11, '備考')

        footer_text = '&P / &Nページ'
        for ws in workbook.worksheets:
            ws.oddFooter.center.text = footer_text
            ws.page_margins.bottom = 1.0

        # 変換後ファイルの出力
        workbook.save(new_dir_path)
