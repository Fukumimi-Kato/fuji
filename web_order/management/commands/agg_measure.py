import logging
import math
import os
import shutil
import datetime as dt
import platform
import openpyxl as excel

from django.conf import settings

from web_order.models import TmpPlateNamePackage
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister, RawPlatePackageRegisterFactory
from .utils import AggEngePackageMixin, AggFixedOrderRule, ExcelOutputMixin, AggFixedOrderRuleForBasic
from web_order.cooking_direction_plates import PlateNameAnalizeUtil

logger = logging.getLogger(__name__)

"""
    喫食日の食数を集計し、計量表を出力する処理

# 引数
    製造日（YYYY-MM-DD）
    喫食日（YYYY-MM-DD）
    食事区分（朝・昼・夕）
    料理名
    分量
    単位
    出汁の量

# 出力ファイル
    朝食の計量表
    昼食の計量表
    夕食の計量表
"""

class AggMeasureWriter(AggEngePackageMixin, ExcelOutputMixin):

    def __init__(self):
        self.package_size = None
        self.enge_package_size = None

        # 販売固定商品注文数
        # (マスタ定義の存在するもの)
        # -保存用1人用
        self.pre_1pack_j = 0
        self.pre_1pack_u = 0
        # -見本・写真用
        self.photo = 0

        # -保存用50g
        self.pre_50g_j = 0
        self.pre_50g_s = 0
        self.pre_50g_z = 0
        self.pre_50g_m = 0

        # マスタ定義の存在しないもの
        # -保存用1人用
        self.pre_1pack_s = 4
        self.pre_1pack_z = 2
        self.pre_1pack_m = 2

        self.in_index = 0
        self.in_cooking_day = None
        self.aggregation_day = None
        self.in_menu = None
        self.in_name = None
        self.in_qty = 0.0
        self.in_dty = None

        self.in_is_same_thickness = False
        self.in_dilute = False
        self.in_is_soup_enge =False

        self.in_inner_g = 0.0
        self.in_adjust = 0
        self.agg_day_short = None
        self.fixed_order = None

        self.in_is_hide_enge = False

    def get_fixed_1pack_count(self):
        return self.pre_1pack_j + self.photo

    def get_enge_50g_count(self, menu_name: str):
        if menu_name == 'ソフト':
            return self.pre_50g_s
        elif menu_name == 'ゼリー':
            return self.pre_50g_z
        else:
            return self.pre_50g_m

    def create_excel_unit(self,
            dataframe, worksheet, meal, parts_name, quantity, unit, density, is_usuaji, is_fixed):
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)  # 朝・昼・夕
        worksheet.cell(2, 2, parts_name)  # 料理名

        worksheet.cell(2, 5, quantity)  # 料理の内容量
        worksheet.cell(2, 6, unit)  # 料理の内容量の単位
        worksheet.cell(2, 7, density)  # 料理の出汁の量
        worksheet.cell(3, 7, density)  # 料理の出汁の量(他セル参照用)
        worksheet.cell(8, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
        if (unit != 'g') and (density > 0) and (not is_fixed):
            worksheet.cell(3, 7, '')  # 料理の出汁の量
        else:
            # 合計50gになるように対応
            if self.in_inner_g:
                inner = float(self.in_inner_g)
                qty_for_50 = inner + density / 100
                qty_for_50 = 50.0 / qty_for_50 * inner
                worksheet.cell(8, 5, qty_for_50)
                worksheet.cell(8, 8, 50.0 - qty_for_50)
                worksheet.cell(8, 17, 'g')
            else:
                rate_for_50 = 1.0 + density / 100
                qty_for_50 = 50.0 / rate_for_50
                worksheet.cell(8, 5, qty_for_50)
                worksheet.cell(8, 8, qty_for_50 * (density / 100))

        if is_usuaji:
            needle_pack, needle_count = self.fixed_order.get_needle_u()
            preserve_pack, preserve_count = self.fixed_order.get_preserve_u()

            # 常食と同じ濃さの液を使用する場合
            if self.in_is_same_thickness:
                worksheet.cell(2, 8, '=IF($G$2="","",G2)')  # 料理の出汁の量

            # 対象の料理を0.8倍提供する場合
            # (該当するのは、調味液となるため、液同と同時には発生しない)
            if self.in_dilute:
                worksheet.cell(2, 5, f'={quantity}*0.8')
        else:
            needle_pack, needle_count = self.fixed_order.get_needle_j()
            preserve_pack, preserve_count = self.fixed_order.get_preserve_j()

        # 針刺し用の出力
        worksheet.cell(6, 11, needle_count)  # 針刺し用 注文数

        # 保存用の出力
        worksheet.cell(7, 11, preserve_count)  # 保存用 注文数

        # 保存用(50g)の出力
        worksheet.cell(8, 11, self.pre_50g_j)  # 保存用(50g) 注文数
        worksheet.cell(9, 11, self.pre_1pack_j)  # 保存用(1人袋) 注文数
        if self.pre_1pack_j == 0:
            worksheet.cell(9, 16, self.pre_1pack_j)  # 保存用(1人袋) 合計
        worksheet.cell(10, 11, 1 if self.photo else 0)  # 写真用 注文数

        i = 14
        package_count = needle_count  # 針刺し用(30人の袋)
        package_count += preserve_count  # 保存用(10人の袋)
        package_count_1p = self.get_fixed_1pack_count()  # 写真用+保存用
        package_count_50g = self.pre_50g_j  # 保存用(50g)
        is_raw_save = False
        self.package_size = 30
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    logger.info(f"【施設単位】錦糸卵出力:{row['ユニット名']}-{row['乾燥冷凍区分']}")
                    is_raw_save = True
                    if row['注文数'] > 1:
                        worksheet.cell(i, 4, row['注文数'])
                        unit_package_count = math.ceil(row['注文数'] / 30)
                        package_count += unit_package_count

                        # ピッキング指示書用のユニット毎袋数を保存
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                   'BASIC_UNIT', parts_name, is_raw_to_enge=True)
                    else:
                        worksheet.cell(i, 4, 0)
                    if row['注文数'] == 1:
                        package_count_1p += 1

                        # ピッキング指示書用のユニット毎袋数を保存
                        PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], 1,
                                                                   'BASIC_1', parts_name, is_raw_to_enge=True)
                else:
                    # 施設用のテンプレートには計算式が入っていて、出力対象外でも袋数が出るため無効化する
                    worksheet.cell(i, 11, 0)
            else:
                if row['注文数'] > 1:
                    worksheet.cell(i, 4, row['注文数'])
                    qty = math.ceil(row['注文数'] / 30)
                    package_count += qty

                    # ピッキング指示書用のユニット毎袋数を保存
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                               self.in_index,
                                                               row['呼出番号'], row['ユニット名'], qty, 'BASIC_UNIT', parts_name)

                else:
                    worksheet.cell(i, 4, 0)
                if row['注文数'] == 1:
                    package_count_1p += 1

                    # ピッキング指示書用のユニット毎袋数を保存
                    PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                               row['呼出番号'], row['ユニット名'], 1, 'BASIC_1', parts_name)

            i = i + 1

        # CSV出力用の袋数を保存
        P7Util.save_package_count_for_print(
            self.in_cooking_day, self.aggregation_day, self.in_index, package_count, package_count_1p, '常食', self.in_menu,
            package_count_50g, is_raw_save)

    def create_excel(self,
            dataframe, worksheet, meal, parts_name, quantity, unit, density, is_usuaji,
            preserve_one_pack, photo, is_fixed):
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)  # 朝・昼・夕
        worksheet.cell(2, 2, parts_name)  # 料理名

        worksheet.cell(2, 5, quantity)  # 料理の内容量
        worksheet.cell(2, 6, unit)  # 料理の内容量の単位
        worksheet.cell(2, 7, density)  # 料理の出汁の量
        worksheet.cell(3, 7, density)  # 料理の出汁の量(他セル参照用)
        worksheet.cell(8, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
        if (unit != 'g') and (density > 0) and (not is_fixed):
            worksheet.cell(3, 7, '')  # 料理の出汁の量
        else:
            # 合計50gになるように対応
            if self.in_inner_g:
                inner = float(self.in_inner_g)
                qty_for_50 = inner + density / 100
                qty_for_50 = 50.0 / qty_for_50 * inner
                worksheet.cell(8, 5, qty_for_50)
                worksheet.cell(8, 8, 50.0 - qty_for_50)
                worksheet.cell(8, 17, 'g')
            else:
                rate_for_50 = 1.0 + density / 100
                qty_for_50 = 50.0 / rate_for_50
                worksheet.cell(8, 5, qty_for_50)
                worksheet.cell(8, 8, qty_for_50 * (density / 100))

        if is_usuaji:
            # 常食と同じ濃さの液を使用する場合
            if self.in_is_same_thickness:
                worksheet.cell(2, 8, '=IF($G$2="","",G2)')  # 料理の出汁の量

            # 対象の料理を0.8倍提供する場合
            # (該当するのは、調味液となるため、液同と同時には発生しない)
            if self.in_dilute:
                worksheet.cell(2, 5, f'={quantity}*0.8')

        package_count = 0

        # 10人用袋を使用するかどうかの判定(=煮魚でない場合)
        is_use_package_10p = not PlateNameAnalizeUtil.is_5p_package_plate(self.before_name)

        # 針刺し用の出力
        needle_10_count = self.fixed_order.get_needle_10()
        needle_5_count = self.fixed_order.get_needle_5()
        # worksheet.cell(6, 2, needle_pack)                     # 針刺し用 〇人前
        worksheet.cell(6, 14, '-')  # 針刺し用 注文数(プルダウンの選択により変わるため値を入れない)
        worksheet.cell(6, 15, needle_10_count)  # 針刺し用 単位袋10
        worksheet.cell(6, 16, needle_5_count)  # 針刺し用 単位袋5
        worksheet.cell(6, 17, 0)  # 針刺し用 単位袋1
        if is_use_package_10p:
            package_count += needle_10_count
        else:
            package_count += needle_5_count
            worksheet.cell(2, 15, 5)

        # 保存用の出力
        preserve_10_count = self.fixed_order.get_preserve_j_10()
        preserve_5_count = self.fixed_order.get_preserve_j_5()
        # worksheet.cell(7, 2, preserve_pack)                     # 保存用 〇人前
        worksheet.cell(7, 14, '-')  # 保存用 注文数(プルダウンの選択により変わるため値を入れない)
        worksheet.cell(7, 15, preserve_10_count)  # 保存用 単位袋10
        worksheet.cell(7, 16, self.fixed_order.get_preserve_j_5())  # 保存用 単位袋5
        worksheet.cell(7, 17, 0)  # 保存用 単位袋1
        if is_use_package_10p:
            package_count += preserve_10_count
        else:
            package_count += preserve_5_count

        # 保存用(50g)の出力
        worksheet.cell(8, 14, self.pre_50g_j)  # 保存用・50g 注文数
        worksheet.cell(8, 15, 0)  # 保存用・50g 単位袋10
        worksheet.cell(8, 16, 0)  # 保存用・50g 単位袋5
        worksheet.cell(8, 17, self.pre_50g_j)  # 保存用・50g 単位袋1
        worksheet.cell(8, 10, self.pre_50g_j)  # 保存用・50g 袋数

        # 保存用(一人前袋)の出力
        worksheet.cell(9, 14, preserve_one_pack)  # 写真用・1人用 注文数
        worksheet.cell(9, 15, 0)  # 保存用・1人用 単位袋10
        worksheet.cell(9, 16, 0)  # 保存用・1人用 単位袋5
        worksheet.cell(9, 17, preserve_one_pack)  # 保存用・1人用 単位袋1
        worksheet.cell(9, 10, preserve_one_pack)  # 保存用・1人用 袋数
        if preserve_one_pack == 0:
            worksheet.cell(9, 11, 0)  # 保存用・1人用 合計

        # 写真用の出力
        worksheet.cell(10, 14, photo)  # 写真用・1人用 注文数
        worksheet.cell(10, 15, 0)  # 写真用・1人用 単位袋10
        worksheet.cell(10, 16, 0)  # 写真用・1人用 単位袋5
        # 写真用・1人用 単位袋1
        worksheet.cell(10, 17, 1 if photo else 0)
        worksheet.cell(10, 10, 1 if photo else 0)  # 写真用・1人用 袋数

        package_count_1p = self.get_fixed_1pack_count()  # 保存用・1人用 + 写真用・2人用
        package_count_50g = self.pre_50g_j  # 保存用(50g)

        i = 14
        is_raw_save = False
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            # 端数はテンプレートファイルのセル上の袋サイズ指定・計算式から設定する

            worksheet.cell(i, 14, row['注文数'])
            worksheet.cell(i, 15, row['単位袋10'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    logger.info(f"【端数】錦糸卵出力-判定:{row['ユニット名']}-{row['乾燥冷凍区分']}")
                    is_raw_save = True
                    if is_use_package_10p:
                        package_count += row['単位袋10']
            elif is_use_package_10p:
                self.package_size = 10
                package_count += row['単位袋10']

                # ピッキング指示書用のユニット毎袋数を保存
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], row['単位袋10'], 'BASIC_10',
                                                           parts_name)
            else:
                self.package_size = 5

            worksheet.cell(i, 16, row['10の端数袋・入数'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    if is_use_package_10p and (row['10の端数袋・入数'] > 0):
                        package_count += 1
            elif is_use_package_10p and (row['10の端数袋・入数'] > 0):
                package_count += 1

                # ピッキング指示書用のユニット毎袋数を保存
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], 1, 'BASIC_FRACTION', parts_name)

            worksheet.cell(i, 17, row['10の1人用袋'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    if is_use_package_10p and (row['10の1人用袋'] > 0):
                        package_count += row['10の1人用袋']
            elif is_use_package_10p and (row['10の1人用袋'] > 0):
                package_count_1p += row['10の1人用袋']

                # ピッキング指示書用のユニット毎袋数を保存
                logger.info(f"1人袋登録:{row['ユニット名']}")
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], 1, 'BASIC_1', parts_name)

            worksheet.cell(i, 18, row['単位袋7'])
            worksheet.cell(i, 19, row['7の端数袋・入数'])
            worksheet.cell(i, 20, row['7の1人用袋'])
            worksheet.cell(i, 21, row['単位袋5'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    if not is_use_package_10p:
                        package_count += row['単位袋5']
            elif not is_use_package_10p:
                package_count += row['単位袋5']

                # ピッキング指示書用のユニット毎袋数を保存
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], row['単位袋5'], 'BASIC_5',
                                                           parts_name)

            worksheet.cell(i, 22, row['5の端数袋・入数'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    if (not is_use_package_10p) and (row['5の端数袋・入数'] > 0):
                        package_count += 1
            elif (not is_use_package_10p) and (row['5の端数袋・入数'] > 0):
                package_count += 1

                # ピッキング指示書用のユニット毎袋数を保存
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], 1, 'BASIC_FRACTION', parts_name)

            worksheet.cell(i, 23, row['5の1人用袋'])
            if '原体' in parts_name:
                if ('錦糸卵' in parts_name) and (row['乾燥冷凍区分'] == '乾燥'):
                    if (not is_use_package_10p) and (row['5の1人用袋'] > 0):
                        package_count += row['5の1人用袋']
            elif (not is_use_package_10p) and (row['5の1人用袋'] > 0):
                package_count_1p += row['5の1人用袋']

                # ピッキング指示書用のユニット毎袋数を保存
                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                           row['呼出番号'], row['ユニット名'], 1, 'BASIC_1', parts_name)

            i = i + 1

        P7Util.save_package_count_for_print(
            self.in_cooking_day, self.aggregation_day, self.in_index, package_count, package_count_1p, '常食', self.in_menu,
            package_count_50g, is_raw_save)

    def create_excel_enge(self,
            dataframe, worksheet, meal, parts_name, quantity, unit, density, needle, preserve,
            is_show_enge, package):
        i = 10
        worksheet.cell(1, 5, self.agg_day_short)
        worksheet.cell(1, 7, meal)  # 朝・昼・夕
        worksheet.cell(2, 2, parts_name)  # 料理名

        worksheet.cell(5, 4, needle)  # 針刺し用の食数
        worksheet.cell(7, 7, preserve)  # 保存用の食数
        worksheet.cell(6, 6, 'g')  # 保存用50g単位(料理の単位に関わらず、保存用50gはg表記)
        # 嚥下は基本、料理の内容量(g数)を出力しないため、コマンドのパラメータでする/しないを指定する
        if is_show_enge:
            worksheet.cell(2, 5, quantity)  # 料理の内容量

        menu = ''
        if 'ソフト' in worksheet.title:
            menu = 'ソフト'
        elif 'ゼリー' in worksheet.title:
            menu = 'ゼリー'
        elif 'ミキサー' in worksheet.title:
            menu = 'ミキサー'

        pre_50g_count = self.get_enge_50g_count(menu)
        worksheet.cell(6, 7, pre_50g_count)  # 保存用の食数

        package_count = 0
        package_count_1p = preserve
        package_count_50g = pre_50g_count  # 保存用(50g)
        is_raw_save = False
        is_soup_parts = (self.in_adjust == 1) and ('液' in worksheet.title)
        for index, row in dataframe.iterrows():
            worksheet.row_dimensions[i].hidden = False
            worksheet.cell(i, 2, row['呼出番号'])
            worksheet.cell(i, 3, row['ユニット名'])

            is_set_plate_package = False
            if self.in_inner_g:
                q = self.get_inner_combined_quantity(self.in_inner_g, self.in_dty)
                worksheet.cell(i, 7, self.get_filling_package_excel_function('g', q, package, i))
                size = self.get_filling_package_size('g', q, package)
                self.enge_package_size = size
                if PlateNameAnalizeUtil.is_raw_plate_name(parts_name):
                    if PlateNameAnalizeUtil.is_raw_enge_plate_name(self.before_name, self.aggregation_day)[0]:
                        is_set_plate_package = True
                        is_raw_save = True
                        worksheet.cell(i, 4, row['注文数'])
                    elif '錦糸卵' in parts_name:
                        is_raw_save = True
                        # 錦糸卵の場合は、乾燥だけ袋計算(直送かどうかで判断すべき?)
                        if row['乾燥冷凍区分'] == '乾燥':
                            is_set_plate_package = True
                            worksheet.cell(i, 4, row['注文数'])
                else:
                    is_set_plate_package = True
                    worksheet.cell(i, 4, row['注文数'])
                    is_raw_save = False

                if is_set_plate_package:
                    if row['注文数'] >= 2:
                        unit_package_count = math.ceil(row['注文数'] / size)
                        package_count += unit_package_count

                        # ピッキング指示書用のユニット毎袋数を保存
                        if row['注文数'] == 2:
                            if not self.in_is_hide_enge:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index,
                                                                           row['呼出番号'], row['ユニット名'], 1, 'ENGE_2',
                                                                           parts_name, menu, is_raw_to_enge=is_raw_save,
                                                                           is_soup_parts=is_soup_parts)
                        else:
                            if not self.in_is_hide_enge:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index,
                                                                           row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                           f'ENGE_{size}', parts_name, menu,
                                                                           is_raw_to_enge=is_raw_save,
                                                                           is_soup_parts=is_soup_parts)
                    if row['注文数'] == 1:
                        package_count_1p += 1
                        # ピッキング指示書用のユニット毎袋数を保存
                        if not self.in_is_hide_enge:
                            PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index,
                                                                   row['呼出番号'], row['ユニット名'], 1, 'ENGE_1', parts_name,
                                                                   menu, is_raw_to_enge=is_raw_save, is_soup_parts=is_soup_parts)
            else:
                q = self.get_combined_quantity(quantity, self.in_dty)
                worksheet.cell(i, 7, self.get_filling_package_excel_function(unit, q, package, i))
                size = self.get_filling_package_size(unit, q, package)
                self.enge_package_size = size
                if PlateNameAnalizeUtil.is_raw_plate_name(parts_name):
                    if PlateNameAnalizeUtil.is_raw_enge_plate_name(self.before_name, self.aggregation_day)[0]:
                        is_set_plate_package = True
                        is_raw_save = True
                        worksheet.cell(i, 4, row['注文数'])
                    elif '錦糸卵' in parts_name:
                        is_raw_save = True
                        # 錦糸卵の場合は、乾燥だけ袋計算(直送かどうかで判断すべき?)
                        if row['乾燥冷凍区分'] == '乾燥':
                            is_set_plate_package = True
                            worksheet.cell(i, 4, row['注文数'])
                else:
                    is_set_plate_package = True
                    is_raw_save = False
                    worksheet.cell(i, 4, row['注文数'])

                if is_set_plate_package:
                    if row['注文数'] >= 2:
                        unit_package_count = math.ceil(row['注文数'] / size)
                        package_count += unit_package_count

                        # ピッキング指示書用のユニット毎袋数を保存
                        if row['注文数'] == 2:
                            if not self.in_is_hide_enge:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index,
                                                                           row['呼出番号'], row['ユニット名'], 1, 'ENGE_2',
                                                                           parts_name, menu, is_raw_to_enge=is_raw_save,
                                                                           is_soup_parts=is_soup_parts)
                        else:
                            if not self.in_is_hide_enge:
                                PlatePackageRegister.register_unit_package(self.in_cooking_day, self.aggregation_day, self.in_menu,
                                                                           self.in_index,
                                                                           row['呼出番号'], row['ユニット名'], unit_package_count,
                                                                           f'ENGE_{size}', parts_name, menu,
                                                                           is_raw_to_enge=is_raw_save,
                                                                           is_soup_parts=is_soup_parts)

                    if row['注文数'] == 1:
                        package_count_1p += 1
                        # ピッキング指示書用のユニット毎袋数を保存
                        if not self.in_is_hide_enge:
                            PlatePackageRegister.register_unit_package(
                                self.in_cooking_day, self.aggregation_day, self.in_menu, self.in_index, row['呼出番号'],
                                row['ユニット名'], 1, 'ENGE_1', parts_name, menu, is_raw_to_enge=is_raw_save,
                                is_soup_parts=is_soup_parts)
            if is_show_enge:
                worksheet.cell(i, 5, self.get_filling_quantity_function(quantity, row['注文数'], i))

            i = i + 1

        menu = ''
        if 'ソフト' in worksheet.title:
            menu = 'ソフト'
        elif 'ゼリー' in worksheet.title:
            menu = 'ゼリー'
        elif 'ミキサー' in worksheet.title:
            menu = 'ミキサー'

        if self.in_adjust == 2:
            if self.in_index != -1:
                enge_index = self.in_index + 1
            else:
                enge_index = self.in_index
            P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, enge_index, package_count,
                                                package_count_1p, menu,
                                                self.in_menu, package_count_50g, is_raw_save)
        else:
            P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, self.in_index, package_count,
                                                package_count_1p, menu,
                                                self.in_menu, package_count_50g, is_raw_save)

            if self.in_adjust == 1:
                # 調整発生時は、ルー用の料理に袋数を保存する。
                P7Util.save_package_count_for_print(self.in_cooking_day, self.aggregation_day, self.in_index + 1, package_count,
                                                    package_count_1p, menu,
                                                    self.in_menu, package_count_50g, is_raw_save)

    def get_enge_name(self):
        if self.in_is_soup_enge:
            base_name = self.in_name[0:self.in_name.find('具')]
            gu_name = self.in_name[0:self.in_name.find('＋')]
            soup_name = self.in_name[self.in_name.find('＋') + 1:]
            return (f'{gu_name}', f'{base_name}{soup_name}')
        else:
            return self.in_name

    def get_soup_quantity(self):
        return round(float(self.in_qty) * float(self.in_dty) / 100, 2)

    def handle(self, manager, *args, **options):
        self.in_index = int(options['opt'][0])      # 呼び出し時の引数1つ目、調理表のインデックス(-1は対象不明)
        in_cook = options['opt'][1]         # 呼び出し時の引数2つ目、製造日（YYYY-MM-DD）
        in_date = options['opt'][2]         # 呼び出し時の引数3つ目、喫食日（YYYY-MM-DD）
        self.in_menu = options['opt'][3]         # 呼び出し時の引数4つ目、食事区分（朝・昼・夕）
        self.in_name = options['opt'][4]         # 呼び出し時の引数5つ目、料理名
        self.in_qty = options['quantity']        # 呼び出し時の引数 --quantity 分量
        in_unit = options['unit']           # 呼び出し時の引数 --unit 単位
        self.in_dty = options.get('density', 0)    # 呼び出し時の引数 --density 出汁の量
        in_fixed = options.get('fixed', 0)         # 呼び出し時の引数 --fixed 出汁の液量が確定しているかどうか
        in_package_rule = options['package']       # 呼び出し時の引数 --package 袋数のルール(主菜扱い(main)/副菜扱い(sub))
        in_is_less = options.get('less', 1)
        self.in_is_same_thickness = options.get('thickness', 0) == 1    # 呼び出し時の引数 --thickness 常食と薄味で同じ濃さの液を使用するかどうか
        in_is_show_enge = options.get('showenge', 0) == 1          # 呼び出し時の引数 --showenge 嚥下のシートにg数出力を行うかどうか
        self.in_is_hide_enge = options.get('showenge', 0) == -1         # 呼び出し時の引数 --showenge -1の場合は、嚥下シート自体を表示しない
        self.in_dilute = options.get('dilute', 0) == 1                  # 呼び出し時の引数 --dilute 食材を0.8倍提供するかどうか
        self.in_is_soup_enge = options.get('engesoup', 0) == 1          # 呼び出し時の引数 --engesoup 嚥下の液を別で出力するかどうか
        self.in_inner_g = options.get('innergram', 0)                   # 呼び出し時の引数 --innergram g以外の場合の内容量
        self.in_adjust = options.get('adjust', 0)                       # 呼び出し時の引数 --adjust 嚥下袋数保存時のindexの補正の状況(0:補正なし、1:補正の必要正発生、2:補正発生済み)
        self.total_qty = options.get('total', 0)                        # 呼び出し時の引数 --total 計量表が分割になった場合も、料理全体の袋サイズを計算できるように
        self.before_name = options.get('before_name', '')

        # adjustは、常食のインデックスは補正しない。

        self.aggregation_day = dt.datetime.strptime(in_date, '%Y-%m-%d').date()
        if platform.system() == 'Windows':
            self.agg_day_short = self.aggregation_day.strftime('%m/%d')
        else:
            self.agg_day_short = self.aggregation_day.strftime('%-m/%-d')

        self.in_inner_g = float(self.in_inner_g)
        enable_day = dt.datetime.strptime(settings.BASIC_PLATE_ENABLE_DATE, '%Y-%m-%d').date()
        if self.aggregation_day >= enable_day:
            qty = self.total_qty or self.in_qty
            logger.info(f'Rule({self.in_name}):{float(qty)}-{self.in_inner_g}-{float(self.in_dty)}')
            self.fixed_order = AggFixedOrderRuleForBasic(self.in_name, float(qty), in_unit, in_is_less, self.in_inner_g, float(self.in_dty))
        else:
            self.fixed_order = AggFixedOrderRule(self.in_name, float(self.in_qty), in_unit, in_is_less, self.in_inner_g, float(self.in_dty))

        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + in_cook + '_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        if self.fixed_order.is_use_unit_package:
            measure_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_unit.xlsx')  # 計量表のテンプレート
        else:
            measure_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure.xlsx')  # 計量表のテンプレート

        # 製造日の取得
        self.in_cooking_day = dt.datetime.strptime(in_cook, '%Y-%m-%d').date()

        # 注文対象日・食事区分の設定
        manager.set_eating(self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # ピッキング指示書用袋数：原体の袋数の登録
        # ------------------------------------------------------------------------------
        package_register = RawPlatePackageRegisterFactory.create(self.before_name)
        if package_register.is_valid:
            package_register.register(manager.get_df_raw(), self.in_cooking_day, self.aggregation_day, self.in_menu)

        # ------------------------------------------------------------------------------
        # 針刺し用の食数
        # ------------------------------------------------------------------------------
        res_s, res_m, res_z = manager.get_needle_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数
        # ------------------------------------------------------------------------------
        pre_s, pre_m, pre_z = manager.get_preserve_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        self.pre_1pack_j, self.pre_1pack_u = manager.get_preserve_1p_orders()

        # ------------------------------------------------------------------------------
        # 写真用の食数
        # ------------------------------------------------------------------------------
        self.photo = manager.get_photo_orders()

        # ------------------------------------------------------------------------------
        # 保存用の食数(50g)
        # ------------------------------------------------------------------------------
        self.pre_50g_j, self.pre_50g_s, self.pre_50g_z, self.pre_50g_m = manager.get_preserve_50g_orders()

        # ------------------------------------------------------------------------------
        # qs（Queryset）df（Dataframe）
        # 朝食（b）昼食（l）夕食（d）
        # 常食（j）薄味（u）ソフト（s）ミキサー（m）ゼリー（z）
        # 具（g） 汁（s）
        # ------------------------------------------------------------------------------

        if self.in_menu == '朝食':

            book_b = excel.load_workbook(measure_template)

            # 基本食(常食)
            df_bj = manager.get_df_basic()
            if self.fixed_order.is_use_unit_package:
                self.create_excel_unit(df_bj, book_b["常食"], '△ 朝', self.in_name, self.in_qty, in_unit, self.in_dty, False, in_fixed)
            else:
                self.create_excel(df_bj, book_b["常食"], '△ 朝', self.in_name, self.in_qty, in_unit, self.in_dty, False, self.pre_1pack_j, self.photo, in_fixed)

            # 薄味
            book_b.remove(book_b["薄味"])

            # ソフト
            df_bs = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bs, book_b["ソフト"], '△ 朝', gu_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_bs, book_b["ソフト(液)"], '△ 朝', soup_name, soup_qty, 'g', self.in_dty, res_s, pre_s, True, in_package_rule)
            else:
                self.create_excel_enge(df_bs, book_b["ソフト"], '△ 朝', self.in_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ソフト(液)"])

            # ミキサー食
            df_bm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bm, book_b["ミキサー"], '△ 朝', gu_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_bm, book_b["ミキサー(液)"], '△ 朝', soup_name, soup_qty, 'g', self.in_dty, res_m, pre_m, True, in_package_rule)
            else:
                self.create_excel_enge(df_bm, book_b["ミキサー"], '△ 朝', self.in_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ミキサー(液)"])

            # ゼリー食
            df_bz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_bz, book_b["ゼリー"], '△ 朝', gu_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_bz, book_b["ゼリー(液)"], '△ 朝', soup_name, soup_qty, 'g', self.in_dty, res_z, pre_z, True, in_package_rule)
            else:
                self.create_excel_enge(df_bz, book_b["ゼリー"], '△ 朝', self.in_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_b["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_b.remove(book_b["ゼリー(液)"])

            breakfast_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_朝_' + self.in_name + '.xlsx')
            self.save_with_select(book_b, breakfast_output_file)

        if self.in_menu == '昼食':

            book_l = excel.load_workbook(measure_template)

            # 基本食(常食)
            df_lj = manager.get_df_basic()
            if self.fixed_order.is_use_unit_package:
                self.create_excel_unit(df_lj, book_l["常食"], '○ 昼', self.in_name, self.in_qty, in_unit, self.in_dty, False, in_fixed)
            else:
                self.create_excel(df_lj, book_l["常食"], '○ 昼', self.in_name, self.in_qty, in_unit, self.in_dty, False, self.pre_1pack_j, self.photo, in_fixed)

            # 薄味
            book_l.remove(book_l["薄味"])

            # ソフト食
            df_ls = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_ls, book_l["ソフト"], '○ 昼', gu_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_ls, book_l["ソフト(液)"], '○ 昼', soup_name, soup_qty, 'g', self.in_dty, res_s, pre_s,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_ls, book_l["ソフト"], '○ 昼', self.in_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ソフト(液)"])

            # ミキサー食
            df_lm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()
                self.create_excel_enge(df_lm, book_l["ミキサー"], '○ 昼', gu_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_lm, book_l["ミキサー(液)"], '○ 昼', soup_name, soup_qty, 'g', self.in_dty, res_m, pre_m,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_lm, book_l["ミキサー"], '○ 昼', self.in_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ミキサー(液)"])

            # ゼリー食
            df_lz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_lz, book_l["ゼリー"], '○ 昼', gu_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_lz, book_l["ゼリー(液)"], '○ 昼', soup_name, soup_qty, 'g', self.in_dty, res_z, pre_z,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_lz, book_l["ゼリー"], '○ 昼', self.in_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_l["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_l.remove(book_l["ゼリー(液)"])

            lunch_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_昼_' + self.in_name + '.xlsx')
            self.save_with_select(book_l, lunch_output_file)

        if self.in_menu == '夕食':

            book_d = excel.load_workbook(measure_template)

            # 基本食(常食)
            df_dj = manager.get_df_basic()
            if self.fixed_order.is_use_unit_package:
                self.create_excel_unit(df_dj, book_d["常食"], '□ 夕', self.in_name, self.in_qty, in_unit, self.in_dty, False, in_fixed)
            else:
                self.create_excel(df_dj, book_d["常食"], '□ 夕', self.in_name, self.in_qty, in_unit, self.in_dty, False, self.pre_1pack_j, self.photo, in_fixed)

            # 薄味
            book_d.remove(book_d["薄味"])

            # ソフト食
            df_ds = manager.get_df_soft()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_ds, book_d["ソフト"], '□ 夕', gu_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_ds, book_d["ソフト(液)"], '□ 夕', soup_name, soup_qty, 'g', self.in_dty, res_s, pre_s,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_ds, book_d["ソフト"], '□ 夕', self.in_name, self.in_qty, in_unit, self.in_dty, res_s, pre_s,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ソフト"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ソフト(液)"])

            # ミキサー食
            df_dm = manager.get_df_mixer()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_dm, book_d["ミキサー"], '□ 夕', gu_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_dm, book_d["ミキサー(液)"], '□ 夕', soup_name, soup_qty, 'g', self.in_dty, res_m, pre_m,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_dm, book_d["ミキサー"], '□ 夕', self.in_name, self.in_qty, in_unit, self.in_dty, res_m, pre_m,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ミキサー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ミキサー(液)"])

            # ゼリー食
            df_dz = manager.get_df_jelly()
            if self.in_is_soup_enge:
                gu_name, soup_name = self.get_enge_name()

                self.create_excel_enge(df_dz, book_d["ゼリー"], '□ 夕', gu_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                # 液用の嚥下出力
                soup_qty = self.get_soup_quantity()
                self.create_excel_enge(df_dz, book_d["ゼリー(液)"], '□ 夕', soup_name, soup_qty, 'g', self.in_dty, res_z, pre_z,
                                  True, in_package_rule)
            else:
                self.create_excel_enge(df_dz, book_d["ゼリー"], '□ 夕', self.in_name, self.in_qty, in_unit, self.in_dty, res_z, pre_z,
                                  in_is_show_enge, in_package_rule)
                if self.in_is_hide_enge:
                    # 嚥下のシートを非表示にする
                    book_d["ゼリー"].sheet_state = 'hidden'

                # 液用のシートを削除する
                book_d.remove(book_d["ゼリー(液)"])

            dinner_output_file = os.path.join(new_dir_path, str(self.aggregation_day) + '_夕_' + self.in_name + '.xlsx')
            self.save_with_select(book_d, dinner_output_file)

        if (self.in_index == -1) and self.before_name:
            logger.info(f'tmp-name:{self.before_name}')
            if self.package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=self.before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.package_size,
                    menu_name='常食'
                )
            if self.enge_package_size:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=self.before_name,
                    cooking_day=self.in_cooking_day,
                    size=self.enge_package_size,
                    menu_name='嚥下'
                )

        # shutil.make_archive(new_dir_path, 'zip', root_dir=new_dir_path)

