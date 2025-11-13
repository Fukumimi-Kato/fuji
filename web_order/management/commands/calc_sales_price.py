import os
import datetime
import logging
import pandas as pd
from datetime import timedelta
import openpyxl as excel

from django.conf import settings
from django.core.management.base import BaseCommand
from django_pandas.io import read_frame

from web_order.date_management import SalesDayUtil
from web_order.models import Order, MenuDisplay, InvoiceException, MonthlySalesPrice, NewUnitPrice

class Command(BaseCommand):
    def __init__(self):
        self.logger = logging.getLogger(__name__)

        self.TRANSPORT_BREAKFAST_RATE = 1
        self.TRANSPORT_LUNCH_RATE = 2
        self.TRANSPORT_DINNER_RATE = 2

        self.OPTION_BOSE_BREAKFAST_RATE = 40
        self.OPTION_BOSE_LUNCH_RATE = 40
        self.OPTION_BOSE_DINNER_RATE = 40

        self.OPTION_BOSE_BREAKFAST_RATE = 40
        self.OPTION_BOSE_LUNCH_RATE = 40
        self.OPTION_BOSE_DINNER_RATE = 40

    def add_arguments(self, parser):
        parser.add_argument('target', nargs='+', type=str)
        parser.add_argument('transport', nargs='+', type=int)

    def get_to_date(self, from_date):
        to_date = from_date + timedelta(days=31)
        check = to_date - timedelta(days=1)
        days = 1
        while check.month != from_date.month:
            check -= timedelta(days=1)
            days += 1

        return to_date - timedelta(days=days)

    def get_price_by_user(self):
        """
        施設毎の単価(嚥下、嚥下以外)を取得
        """
        qs = MenuDisplay.objects.all().select_related('username', 'menu_name')
        enge_dict = {}
        other_dict = {}
        for menudisp in qs:
            if menudisp.menu_name.menu_name in ['ソフト', 'ミキサー', 'ゼリー']:
                if menudisp.username.facility_name in enge_dict:
                    pass
                else:
                    enge_dict[menudisp.username.facility_name] = \
                        [menudisp.price_breakfast, menudisp.price_lunch, menudisp.price_dinner]
            else:
                if menudisp.username.facility_name in other_dict:
                    pass
                else:
                    other_dict[menudisp.username.facility_name] = \
                        [menudisp.price_breakfast, menudisp.price_lunch, menudisp.price_dinner]
        return enge_dict, other_dict

    def add_price_column(self, label, df, index, row, price_dict):
        price = 0
        if row['unit_name__username__facility_name'] in price_dict:
            prices = price_dict[row['unit_name__username__facility_name']]
            if row['meal_name__meal_name'] == '朝食':
                price = prices[0]
            elif row['meal_name__meal_name'] == '昼食':
                price = prices[1]
            elif row['meal_name__meal_name'] == '夕食':
                price = prices[2]

        df.loc[index, label] = price
        return price

    def get_template_reference_params(self, wb):
        ws = wb['参照']

        dict = {}
        dict['bose_breakfast'] = ws['C6'].value
        dict['bose_lunch'] = ws['D6'].value
        dict['bose_dinner'] = ws['E6'].value

        dict['filling_breakfast'] = ws['C7'].value
        dict['filling_lunch'] = ws['D7'].value
        dict['filling_dinner'] = ws['E7'].value

        dict['rate_breakfast'] = ws['G6'].value
        dict['rate_lunch'] = ws['H6'].value
        dict['rate_dinner'] = ws['I6'].value

        return dict

    def handle(self, *args, **options):
        self.logger.info('売価計算表出力コマンド開始')

        in_target_month = options['target'][0]  # 呼び出し時の引数1つ目(Y-m)
        in_transport_price = options['transport'][0]  # 呼び出し時の引数2つ目
        full_date_str = f'{in_target_month}-01'
        from_date = datetime.datetime.strptime(full_date_str, '%Y-%m-%d').date()
        to_date = self.get_to_date(from_date)
        expand_from_date = from_date + timedelta(days=-7)
        expand_to_date = to_date + timedelta(days=7)

        new_dir_path = os.path.join(settings.OUTPUT_DIR, settings.SALES_PRICE_DIR)
        os.makedirs(new_dir_path, exist_ok=True)

        temp_dir_path = "tmp"
        os.makedirs(temp_dir_path, exist_ok=True)

        # 施設毎の食数取得
        # ------------------------------------------------------------------------------
        # 当月の全データ取得(+-一週間多く取得する)
        # ------------------------------------------------------------------------------
        qs_all = Order.objects\
            .filter(eating_day__range=[expand_from_date, expand_to_date], quantity__gt=0)\
            .values('unit_name', 'unit_name__username__facility_name', 'unit_name__username',
                    'meal_name__meal_name', 'meal_name__soup', 'meal_name__filling',
                    'menu_name', 'menu_name__group',
                    'allergen', 'quantity', 'eating_day') \
            .exclude(unit_name__unit_code__range=[80001, 80008]) \
            .exclude(unit_name__unit_number=3)
        df_all = read_frame(qs_all)
        df_all = df_all.sort_values([
            'unit_name__username__facility_name'])

        for index, row in df_all.iterrows():
            # 汁オプション内容
            if row['meal_name__soup'] is True:
                df_all.loc[index, 'soup_contract'] = '汁具'
            elif row['meal_name__filling'] is True:
                df_all.loc[index, 'soup_contract'] = '具のみ'
            else:
                df_all.loc[index, 'soup_contract'] = '汁なし'

            # 嚥下か否か
            if row['menu_name'] in ['ソフト', 'ゼリー', 'ミキサー']:
                df_all.loc[index, 'is_enge'] = 'True'
            else:
                df_all.loc[index, 'is_enge'] = 'False'

            # 売上日(標準)
            sales_date = SalesDayUtil.get_by_eating_day_by_settings(row['eating_day'])
            df_all.loc[index, 'sales_date'] = sales_date


        df_all.to_csv(temp_dir_path + "/S-1_注文データ_アレルギー込.csv", index=False)

        # ------------------------------------------------------------------------------
        # 請求内容を施設ごとに個別対応するため、例外内容を保持したテーブルを読み込む
        # ------------------------------------------------------------------------------
        qs_invoice_exception = InvoiceException.objects.all()\
            .values('unit_name__username', 'ng_saturday', 'ng_sunday', 'ng_holiday', 'reduced_rate', 'is_far')

        df_invoice_exception = read_frame(qs_invoice_exception)
        df_invoice_exception.to_csv(temp_dir_path + "/A-2_売上日程.csv", index=False)

        append_selling_exception = pd.merge(df_all, df_invoice_exception, on='unit_name__username', how='left')
        append_selling_exception.to_csv(temp_dir_path + "/a2_1_merge.csv", index=False)
        special_price_orders = 0
        special_price_user = ''
        for index, data in append_selling_exception.iterrows():
            date = data['sales_date']
            youbi = date.weekday()  # 売上日の曜日(月曜日が0〜日曜日が6)

            if youbi == 6:
                if (data['ng_sunday'] > 0) or (data['ng_sunday'] < 0):
                    append_selling_exception.loc[index, 'sales_date'] = date + timedelta(days=data['ng_sunday'])
            elif youbi == 5:
                if (data['ng_saturday'] > 0) or (data['ng_saturday'] < 0):
                    append_selling_exception.loc[index, 'sales_date'] = date + timedelta(days=data['ng_saturday'])
            elif youbi == 4:
                if data['is_far'] is True:
                    # 1日前(=木曜)に補正する
                    append_selling_exception.loc[index, 'sales_date'] = date - timedelta(days=1)

            if data['reduced_rate'] is True:  # 省略記法にすると正常に動作しないので is Trueまで記述
                append_selling_exception.loc[index, 'sales_date'] = data['eating_day']

        append_selling_exception.to_csv(temp_dir_path + "/a2_2_converted.csv", index=False)
        df_all.to_csv(temp_dir_path + "/A-3_売上日本設定.csv", index=False)
        del df_invoice_exception

        append_selling_exception = append_selling_exception.drop(columns=append_selling_exception.columns[[14, 15, 16, 17]])
        df_all_sales = append_selling_exception[append_selling_exception.sales_date >= from_date]
        df_all_sales = df_all_sales[df_all_sales.sales_date <= to_date]
        del append_selling_exception
        other_new_price_dict = {}
        enge_new_price_dict = {}
        for index, data in df_all_sales.iterrows():
            e_day = data['sales_date']
            e_youbi = e_day.weekday()

            price_qs = NewUnitPrice.objects.filter(
                username__username=data['unit_name__username'], eating_day__lte=e_day).order_by('-eating_day')
            if price_qs.exists():
                for new_price in price_qs:
                    is_enge = (new_price.menu_name == 'ソフト') or (new_price.menu_name == 'ゼリー') or (new_price.menu_name == 'ミキサー')
                    if is_enge:
                        if not data['unit_name__username__facility_name'] in enge_new_price_dict:
                            enge_new_price_dict[data['unit_name__username__facility_name']] = {
                                'quantity_dict': {},
                                'eating_day': new_price.eating_day,
                                'price_b': new_price.price_breakfast,
                                'price_l': new_price.price_lunch,
                                'price_d': new_price.price_dinner,
                            }
                    else:
                        if not data['unit_name__username__facility_name'] in other_new_price_dict:
                            other_new_price_dict[data['unit_name__username__facility_name']] = {
                                'quantity_b': 0,
                                'quantity_l': 0,
                                'quantity_d': 0,
                                'eating_day': new_price.eating_day,
                                'price_b': new_price.price_breakfast,
                                'price_l': new_price.price_lunch,
                                'price_d': new_price.price_dinner,
                            }

            if e_youbi == 6:
                # 喫食日が日曜朝のみ特別な価格になるため、情報を抽出
                if (data['unit_name__username'] == '10014') and (data['meal_name__meal_name'] == '朝食'):
                    special_price_orders += data['quantity']
                    special_price_user = data['unit_name__username__facility_name']

        # 単価変更対象の数量を判定するため、再度ループ
        # 嚥下はソフト・ゼリー・ミキサー必ず全て単価変更を登録する前提(ソフトだけでもミキサー・ゼリーが検出されるので注意!)
        for index, data in df_all_sales.iterrows():
            s_day = data['sales_date']
            is_enge = (data['menu_name'] == 'ソフト') or (data['menu_name'] == 'ゼリー') or (
                    data['menu_name'] == 'ミキサー')
            if is_enge:
                if data['unit_name__username__facility_name'] in enge_new_price_dict:
                    price_dict = enge_new_price_dict[data['unit_name__username__facility_name']]
                    if s_day >= price_dict['eating_day']:
                        qty_dict = price_dict['quantity_dict']
                        if data['menu_name'] in qty_dict:
                            if data['meal_name__meal_name'] == '朝食':
                                qty_dict[data['menu_name']][0] += data['quantity']
                            elif data['meal_name__meal_name'] == '昼食':
                                qty_dict[data['menu_name']][1] += data['quantity']
                            elif data['meal_name__meal_name'] == '夕食':
                                qty_dict[data['menu_name']][2] += data['quantity']
                        else:
                            if data['meal_name__meal_name'] == '朝食':
                                qty_dict[data['menu_name']] = [data['quantity'], 0, 0]
                            elif data['meal_name__meal_name'] == '昼食':
                                qty_dict[data['menu_name']] = [0, data['quantity'], 0]
                            elif data['meal_name__meal_name'] == '夕食':
                                qty_dict[data['menu_name']] = [0, 0, data['quantity']]
            else:
                if data['unit_name__username__facility_name'] in other_new_price_dict:
                    price_dict = other_new_price_dict[data['unit_name__username__facility_name']]
                    if s_day >= price_dict['eating_day']:
                        if data['meal_name__meal_name'] == '朝食':
                            price_dict['quantity_b'] += data['quantity']
                        elif data['meal_name__meal_name'] == '昼食':
                            price_dict['quantity_l'] += data['quantity']
                        elif data['meal_name__meal_name'] == '夕食':
                            price_dict['quantity_d'] += data['quantity']
        df_all_sales.to_csv(temp_dir_path + "/A-4_対象絞り込み.csv", index=False)
        self.logger.info(other_new_price_dict)

        # ------------------------------------------------------------------------------
        # 施設毎・食事区分毎に集計
        # ------------------------------------------------------------------------------
        df_month_by_user = df_all_sales.groupby([
            'unit_name__username__facility_name', 'meal_name__meal_name',
            'menu_name', 'soup_contract']).sum().reset_index()
        df_month_by_user = df_month_by_user.sort_values([
            'unit_name__username__facility_name', 'meal_name__meal_name', 'menu_name'])
        df_month_by_user = df_month_by_user.drop(columns=df_month_by_user.columns[[4, 5]])
        for index, row in df_month_by_user.iterrows():
            # 嚥下か否か、出力用献立種別(常食と薄味はまとめる)
            if row['menu_name'] in ['ソフト', 'ゼリー', 'ミキサー']:
                df_month_by_user.loc[index, 'menu_name_output'] = row['menu_name']
                df_month_by_user.loc[index, 'is_enge'] = True
            else:
                df_month_by_user.loc[index, 'menu_name_output'] = '嚥下以外'
                df_month_by_user.loc[index, 'is_enge'] = False
        # TODO:並べ方検討
        df_month_by_user.to_csv(temp_dir_path + "/S-2_注文データ_施設毎.csv", index=False)

        # 施設毎の単価
        # ------------------------------------------------------------------------------
        # 単価情報列の追加
        # ------------------------------------------------------------------------------
        enge_dict, other_dict = self.get_price_by_user()
        for index, row in df_month_by_user.iterrows():
            # 嚥下以外単価
            other_price = self.add_price_column('other_price', df_month_by_user, index, row, other_dict)

            # 嚥下単価
            enge_price = self.add_price_column('enge_price', df_month_by_user, index, row, enge_dict)

            # 単価差有無
            df_month_by_user.loc[index, 'is_diff_price'] = (other_price != enge_price)

        df_month_by_user.to_csv(temp_dir_path + "/S-3_注文データ_施設毎_単価列追加.csv", index=False)
        del enge_dict
        del other_dict

        # ------------------------------------------------------------------------------
        # 嚥下、嚥下以外で集計
        # ------------------------------------------------------------------------------
        df_agg = df_month_by_user.groupby([
            'unit_name__username__facility_name', 'meal_name__meal_name', 'menu_name_output',
            'soup_contract', 'is_enge', 'other_price', 'enge_price', 'is_diff_price']).sum().reset_index()
        df_agg.to_csv(temp_dir_path + "/S-4_注文データ_施設毎_食事区分毎_集計.csv", index=False)
        del df_month_by_user

        # 嚥下のみ、それ以外の表に分割し、結合
        df_agg_enge = (df_agg[df_agg.is_enge == True]).rename(columns={'quantity': 'enge_quantity'})
        df_agg_other = (df_agg[df_agg.is_enge == False]).rename(columns={'quantity': 'other_quantity'})
        df_merge = pd.merge(
            df_agg_enge, df_agg_other,
            on=['unit_name__username__facility_name', 'meal_name__meal_name', 'menu_name_output', 'soup_contract'],
            how='outer').reset_index()
        del df_agg
        del df_agg_enge
        del df_agg_other
        df_merge = df_merge.sort_values([
            'unit_name__username__facility_name', 'meal_name__meal_name', 'menu_name_output', 'soup_contract'])

        # 扱いやすいように、有効な単価、フラグを共通の項目に置き換え
        for index, row in df_merge.iterrows():
            if pd.isnull(row['other_price_x']):
                df_merge.loc[index, 'is_enge'] = row['is_enge_y']
                df_merge.loc[index, 'enge_price'] = row['enge_price_y']
                df_merge.loc[index, 'other_price'] = row['other_price_y']
                df_merge.loc[index, 'is_diff_price'] = row['is_diff_price_y']
            else:
                df_merge.loc[index, 'is_enge'] = row['is_enge_x']
                df_merge.loc[index, 'enge_price'] = row['enge_price_x']
                df_merge.loc[index, 'other_price'] = row['other_price_x']
                df_merge.loc[index, 'is_diff_price'] = row['is_diff_price_x']

        # 不要列の削除
        df_merge = df_merge.drop(columns=df_merge.columns[[5, 6, 7, 8, 10, 11, 12, 13]])

        # 売上の追加
        for index, row in df_merge.iterrows():
            self.logger.info(f"売上金額の調整:{row['unit_name__username__facility_name']}")
            price = 0 if pd.isnull(row['enge_price']) else row['enge_price']
            quantity = 0 if pd.isnull(row['enge_quantity']) else row['enge_quantity']
            total_price = price * quantity
            if quantity and (row['unit_name__username__facility_name'] in enge_new_price_dict):
                price_dict = enge_new_price_dict[row['unit_name__username__facility_name']]
                qty_dict = price_dict['quantity_dict']
                if row['menu_name_output'] in qty_dict:
                    qty_array = qty_dict[row['menu_name_output']]
                    if row['meal_name__meal_name'] == '朝食':
                        total_price += (price_dict['price_b'] - price) * qty_array[0]
                        df_merge.loc[index, 'enge_price'] = price_dict['price_b']
                    if row['meal_name__meal_name'] == '昼食':
                        total_price += (price_dict['price_l'] - price) * qty_array[1]
                        df_merge.loc[index, 'enge_price'] = price_dict['price_l']
                    if row['meal_name__meal_name'] == '夕食':
                        total_price += (price_dict['price_d'] - price) * qty_array[2]
                        df_merge.loc[index, 'enge_price'] = price_dict['price_d']
            df_merge.loc[index, 'enge_sales'] = total_price

            price = 0 if pd.isnull(row['other_price']) else row['other_price']
            quantity = 0 if pd.isnull(row['other_quantity']) else row['other_quantity']
            total_price = price * quantity
            if (row['unit_name__username__facility_name'] == special_price_user) and (row['meal_name__meal_name'] == '朝食'):
                # 正規の価格からの差分を計算
                total_price -= special_price_orders * 50
            elif quantity and (row['unit_name__username__facility_name'] in other_new_price_dict):
                price_dict = other_new_price_dict[row['unit_name__username__facility_name']]
                if row['meal_name__meal_name'] == '朝食':
                    diff_price = (price_dict['price_b'] - price) * price_dict['quantity_b']
                    self.logger.info(f"朝食:price({price})-{price_dict['price_b']}*{price_dict['quantity_b']}")
                    total_price += diff_price
                    df_merge.loc[index, 'other_price'] = price_dict['price_b']
                if row['meal_name__meal_name'] == '昼食':
                    diff_price = (price_dict['price_l'] - price) * price_dict['quantity_l']
                    self.logger.info(f"昼食:price({price})-{price_dict['price_l']}*{price_dict['quantity_l']}")
                    total_price += diff_price
                    df_merge.loc[index, 'other_price'] = price_dict['price_l']
                if row['meal_name__meal_name'] == '夕食':
                    diff_price = (price_dict['price_d'] - price) * price_dict['quantity_d']
                    self.logger.info(f"夕食:price({price})-{price_dict['price_d']}*{price_dict['quantity_d']}")
                    total_price += diff_price
                    df_merge.loc[index, 'other_price'] = price_dict['price_d']
            df_merge.loc[index, 'other_sales'] = total_price

        df_merge.to_csv(temp_dir_path + "/S-5_注文データ_施設毎_食事区分毎_マージ.csv", index=False)

        # フォーマットファイルOpen
        template_filepath = os.path.join(settings.STATICFILES_DIRS[0], 'excel/sales.xlsx')
        workbook = excel.load_workbook(template_filepath)

        # 施設毎・食事区分毎情報出力
        ws_unit = workbook['施設毎詳細']
        i = 4
        for index, row in df_merge.iterrows():
            ws_unit.cell(i, 2, row['unit_name__username__facility_name'])
            ws_unit.cell(i, 3, row['meal_name__meal_name'])
            ws_unit.cell(i, 4, row['menu_name_output'])
            ws_unit.cell(i, 5, row['soup_contract'])
            ws_unit.cell(i, 6, row['other_price'])
            ws_unit.cell(i, 7, row['enge_price'])
            ws_unit.cell(i, 8, '別価格' if row['is_diff_price'] else '同価格')
            ws_unit.cell(i, 9, 0 if pd.isnull(row['other_quantity']) else row['other_quantity'])
            ws_unit.cell(i, 10, 0 if pd.isnull(row['enge_quantity']) else row['enge_quantity'])
            ws_unit.cell(i, 11, f'=I{i}+J{i}')
            ws_unit.cell(i, 12, row['other_sales'])
            ws_unit.cell(i, 13, row['enge_sales'])
            ws_unit.cell(i, 14, f'=L{i}+M{i}')
            i += 1

        # ------------------------------------------------------------------------------
        # 参照情報読込
        # ------------------------------------------------------------------------------
        temp_ref_params = self.get_template_reference_params(workbook)

        # ------------------------------------------------------------------------------
        # 食事区分毎情報出力
        # ------------------------------------------------------------------------------
        ws_meal = workbook['食事区分毎詳細']

        # 全体
        df_total_b = df_merge[df_merge.meal_name__meal_name == '朝食']
        df_total_l = df_merge[df_merge.meal_name__meal_name == '昼食']
        df_total_d = df_merge[df_merge.meal_name__meal_name == '夕食']
        ws_meal.cell(5, 3, df_total_b['other_quantity'].sum() + df_total_b['enge_quantity'].sum())
        ws_meal.cell(5, 4, df_total_l['other_quantity'].sum() + df_total_l['enge_quantity'].sum())
        ws_meal.cell(5, 5, df_total_d['other_quantity'].sum() + df_total_d['enge_quantity'].sum())
        # -売上合計(オプション分も含めた合計)の計算
        all_total_sales_b = df_total_b['other_sales'].sum() + df_total_b['enge_sales'].sum()
        all_total_sales_l = df_total_l['other_sales'].sum() + df_total_l['enge_sales'].sum()
        all_total_sales_d = df_total_d['other_sales'].sum() + df_total_d['enge_sales'].sum()

        ws_meal.cell(5, 11, all_total_sales_b)
        ws_meal.cell(5, 12, all_total_sales_l)
        ws_meal.cell(5, 13, all_total_sales_d)

        # 具のみ
        df_soup_filling = df_merge[df_merge.soup_contract == '具のみ']
        df_soup_filling_b = df_soup_filling[df_soup_filling.meal_name__meal_name == '朝食']
        df_soup_filling_l = df_soup_filling[df_soup_filling.meal_name__meal_name == '昼食']
        df_soup_filling_d = df_soup_filling[df_soup_filling.meal_name__meal_name == '夕食']
        df_soup_filling_b_count = df_soup_filling_b['other_quantity'].sum() + df_soup_filling_b['enge_quantity'].sum()
        df_soup_filling_l_count = df_soup_filling_l['other_quantity'].sum() + df_soup_filling_l['enge_quantity'].sum()
        df_soup_filling_d_count = df_soup_filling_d['other_quantity'].sum() + df_soup_filling_d['enge_quantity'].sum()
        ws_meal.cell(6, 3, df_soup_filling_b_count)
        ws_meal.cell(6, 4, df_soup_filling_l_count)
        ws_meal.cell(6, 5, df_soup_filling_d_count)
        # 売上はフォーマットファイルで計算

        # 汁具
        df_soup_bose = df_merge[df_merge.soup_contract == '汁具']
        df_soup_bose_b = df_soup_bose[df_soup_bose.meal_name__meal_name == '朝食']
        df_soup_bose_l = df_soup_bose[df_soup_bose.meal_name__meal_name == '昼食']
        df_soup_bose_d = df_soup_bose[df_soup_bose.meal_name__meal_name == '夕食']
        df_soup_bose_b_count = df_soup_bose_b['other_quantity'].sum() + df_soup_bose_b['enge_quantity'].sum()
        df_soup_bose_l_count = df_soup_bose_l['other_quantity'].sum() + df_soup_bose_l['enge_quantity'].sum()
        df_soup_bose_d_count = df_soup_bose_d['other_quantity'].sum() + df_soup_bose_d['enge_quantity'].sum()
        ws_meal.cell(7, 3, df_soup_bose_b_count)
        ws_meal.cell(7, 4, df_soup_bose_l_count)
        ws_meal.cell(7, 5, df_soup_bose_d_count)
        # 売上はフォーマットファイルで計算

        # 具のみ(嚥下以外)
        basic_filling_sum_count_b = df_soup_filling_b['other_quantity'].sum()
        basic_filling_sum_count_l = df_soup_filling_l['other_quantity'].sum()
        basic_filling_sum_count_d = df_soup_filling_d['other_quantity'].sum()
        ws_meal.cell(8, 3, basic_filling_sum_count_b)
        ws_meal.cell(8, 4, basic_filling_sum_count_l)
        ws_meal.cell(8, 5, basic_filling_sum_count_d)
        # 売上はフォーマットファイルで計算して出力。

        # 具のみ(嚥下)
        enge_filling_sum_count_b = df_soup_filling_b['enge_quantity'].sum()
        enge_filling_sum_count_l = df_soup_filling_l['enge_quantity'].sum()
        enge_filling_sum_count_d = df_soup_filling_d['enge_quantity'].sum()
        ws_meal.cell(9, 3, df_soup_filling_b['enge_quantity'].sum())
        ws_meal.cell(9, 4, df_soup_filling_l['enge_quantity'].sum())
        ws_meal.cell(9, 5, df_soup_filling_d['enge_quantity'].sum())

        # 売上計算-具のみ(ソフト)
        df_soft_soup_filling_b = df_soup_filling_b[df_soup_filling_b.menu_name_output == 'ソフト']
        df_soft_soup_filling_l = df_soup_filling_l[df_soup_filling_l.menu_name_output == 'ソフト']
        df_soft_soup_filling_d = df_soup_filling_d[df_soup_filling_d.menu_name_output == 'ソフト']
        option_filling_sales_soft_b = df_soft_soup_filling_b['enge_quantity'].sum() * temp_ref_params['filling_breakfast']
        option_filling_sales_soft_l = df_soft_soup_filling_l['enge_quantity'].sum() * temp_ref_params['filling_lunch']
        option_filling_sales_soft_d = df_soft_soup_filling_d['enge_quantity'].sum() * temp_ref_params['filling_dinner']

        # 売上計算-具のみ(ミキサー)
        df_mixer_soup_filling_b = df_soup_filling_b[df_soup_filling_b.menu_name_output == 'ミキサー']
        df_mixer_soup_filling_l = df_soup_filling_l[df_soup_filling_l.menu_name_output == 'ミキサー']
        df_mixer_soup_filling_d = df_soup_filling_d[df_soup_filling_d.menu_name_output == 'ミキサー']
        option_filling_sales_mixer_b = df_mixer_soup_filling_b['enge_quantity'].sum() * temp_ref_params['filling_breakfast']
        option_filling_sales_mixer_l = df_mixer_soup_filling_l['enge_quantity'].sum() * temp_ref_params['filling_lunch']
        option_filling_sales_mixer_d = df_mixer_soup_filling_d['enge_quantity'].sum() * temp_ref_params['filling_dinner']

        # 売上計算-具のみ(ゼリー)
        df_jerry_soup_filling_b = df_soup_filling_b[df_soup_filling_b.menu_name_output == 'ゼリー']
        df_jerry_soup_filling_l = df_soup_filling_l[df_soup_filling_l.menu_name_output == 'ゼリー']
        df_jerry_soup_filling_d = df_soup_filling_d[df_soup_filling_d.menu_name_output == 'ゼリー']
        option_filling_sales_jerry_b = df_jerry_soup_filling_b['enge_quantity'].sum() * temp_ref_params['filling_breakfast']
        option_filling_sales_jerry_l = df_jerry_soup_filling_l['enge_quantity'].sum() * temp_ref_params['filling_lunch']
        option_filling_sales_jerry_d = df_jerry_soup_filling_d['enge_quantity'].sum() * temp_ref_params['filling_dinner']

        """
        # 検算用
        option_filling_sales_all_b = option_filling_sales_soft_b + option_filling_sales_mixer_b + option_filling_sales_jerry_b
        option_filling_sales_all_l = option_filling_sales_soft_l + option_filling_sales_mixer_l + option_filling_sales_jerry_l
        option_filling_sales_all_d = option_filling_sales_soft_d + option_filling_sales_mixer_d + option_filling_sales_jerry_d
        """

        # 汁具(嚥下以外)
        basic_bose_sum_count_b = df_soup_bose_b['other_quantity'].sum()
        basic_bose_sum_count_l = df_soup_bose_l['other_quantity'].sum()
        basic_bose_sum_count_d = df_soup_bose_d['other_quantity'].sum()
        ws_meal.cell(10, 3, basic_bose_sum_count_b)
        ws_meal.cell(10, 4, basic_bose_sum_count_l)
        ws_meal.cell(10, 5, basic_bose_sum_count_d)

        # 汁具(嚥下)
        enge_bose_sum_count_b = df_soup_bose_b['enge_quantity'].sum()
        enge_bose_sum_count_l = df_soup_bose_l['enge_quantity'].sum()
        enge_bose_sum_count_d = df_soup_bose_d['enge_quantity'].sum()
        ws_meal.cell(11, 3, enge_bose_sum_count_b)
        ws_meal.cell(11, 4, enge_bose_sum_count_l)
        ws_meal.cell(11, 5, enge_bose_sum_count_d)

        # 売上計算-汁具(ソフト)
        df_soft_soup_bose_b = df_soup_bose_b[df_soup_bose_b.menu_name_output == 'ソフト']
        df_soft_soup_bose_l = df_soup_bose_l[df_soup_bose_l.menu_name_output == 'ソフト']
        df_soft_soup_bose_d = df_soup_bose_d[df_soup_bose_d.menu_name_output == 'ソフト']
        option_bose_sales_soft_b = df_soft_soup_bose_b['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        option_bose_sales_soft_l = df_soft_soup_bose_l['enge_quantity'].sum() * temp_ref_params['bose_lunch']
        option_bose_sales_soft_d = df_soft_soup_bose_d['enge_quantity'].sum() * temp_ref_params['bose_dinner']

        # 売上計算-汁具(ミキサー)
        df_mixer_soup_bose_b = df_soup_bose_b[df_soup_bose_b.menu_name_output == 'ミキサー']
        df_mixer_soup_bose_l = df_soup_bose_l[df_soup_bose_l.menu_name_output == 'ミキサー']
        df_mixer_soup_bose_d = df_soup_bose_d[df_soup_bose_d.menu_name_output == 'ミキサー']
        option_bose_sales_mixer_b = df_mixer_soup_bose_b['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        option_bose_sales_mixer_l = df_mixer_soup_bose_l['enge_quantity'].sum() * temp_ref_params['bose_lunch']
        option_bose_sales_mixer_d = df_mixer_soup_bose_d['enge_quantity'].sum() * temp_ref_params['bose_dinner']

        # 売上計算-汁具(ゼリー)
        df_jerry_soup_bose_b = df_soup_bose_b[df_soup_bose_b.menu_name_output == 'ゼリー']
        df_jerry_soup_bose_l = df_soup_bose_l[df_soup_bose_l.menu_name_output == 'ゼリー']
        df_jerry_soup_bose_d = df_soup_bose_d[df_soup_bose_d.menu_name_output == 'ゼリー']
        option_bose_sales_jerry_b = df_jerry_soup_bose_b['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        option_bose_sales_jerry_l = df_jerry_soup_bose_l['enge_quantity'].sum() * temp_ref_params['bose_lunch']
        option_bose_sales_jerry_d = df_jerry_soup_bose_d['enge_quantity'].sum() * temp_ref_params['bose_dinner']

        """
        # 検算用
        option_bose_sales_all_b = option_bose_sales_soft_b + option_bose_sales_mixer_b + option_bose_sales_jerry_b
        option_bose_sales_all_l = option_bose_sales_soft_l + option_bose_sales_mixer_l + option_bose_sales_jerry_l
        option_bose_sales_all_d = option_bose_sales_soft_d + option_bose_sales_mixer_d + option_bose_sales_jerry_d
        """

        # ソフト(嚥下のみ)
        df_soft = df_merge[df_merge.menu_name_output == 'ソフト']
        df_soft_b = df_soft[df_soft.meal_name__meal_name == '朝食']
        df_soft_l = df_soft[df_soft.meal_name__meal_name == '昼食']
        df_soft_d = df_soft[df_soft.meal_name__meal_name == '夕食']
        # 食数
        ws_meal.cell(12, 3, df_soft_b['enge_quantity'].sum())
        ws_meal.cell(12, 4, df_soft_l['enge_quantity'].sum())
        ws_meal.cell(12, 5, df_soft_d['enge_quantity'].sum())
        # 売上
        soft_sales_sum_b = df_soft_b['enge_sales'].sum()
        soft_sales_sum_l = df_soft_l['enge_sales'].sum()
        soft_sales_sum_d = df_soft_d['enge_sales'].sum()
        # -汁の売上を除く
        ws_meal.cell(12, 7, soft_sales_sum_b - (option_filling_sales_soft_b + option_bose_sales_soft_b))
        ws_meal.cell(12, 8, soft_sales_sum_l - (option_filling_sales_soft_l + option_bose_sales_soft_l))
        ws_meal.cell(12, 9, soft_sales_sum_d - (option_filling_sales_soft_d + option_bose_sales_soft_d))
        # -汁の売上を含む
        ws_meal.cell(12, 11, soft_sales_sum_b)
        ws_meal.cell(12, 12, soft_sales_sum_l)
        ws_meal.cell(12, 13, soft_sales_sum_d)

        # ミキサー(嚥下のみ)
        df_mixer = df_merge[df_merge.menu_name_output == 'ミキサー']
        df_mixer_b = df_mixer[df_mixer.meal_name__meal_name == '朝食']
        df_mixer_l = df_mixer[df_mixer.meal_name__meal_name == '昼食']
        df_mixer_d = df_mixer[df_mixer.meal_name__meal_name == '夕食']
        # 食数
        ws_meal.cell(13, 3, df_mixer_b['enge_quantity'].sum())
        ws_meal.cell(13, 4, df_mixer_l['enge_quantity'].sum())
        ws_meal.cell(13, 5, df_mixer_d['enge_quantity'].sum())
        # 売上
        mixer_sales_sum_b = df_mixer_b['enge_sales'].sum()
        mixer_sales_sum_l = df_mixer_l['enge_sales'].sum()
        mixer_sales_sum_d = df_mixer_d['enge_sales'].sum()
        # -汁の売上を除く
        ws_meal.cell(13, 7, mixer_sales_sum_b - (option_filling_sales_mixer_b + option_bose_sales_mixer_b))
        ws_meal.cell(13, 8, mixer_sales_sum_l - (option_filling_sales_mixer_l + option_bose_sales_mixer_l))
        ws_meal.cell(13, 9, mixer_sales_sum_d - (option_filling_sales_mixer_d + option_bose_sales_mixer_d))
        # -汁の売上を含む
        ws_meal.cell(13, 11, mixer_sales_sum_b)
        ws_meal.cell(13, 12, mixer_sales_sum_l)
        ws_meal.cell(13, 13, mixer_sales_sum_d)

        # ゼリー(嚥下のみ)
        df_jerry = df_merge[df_merge.menu_name_output == 'ゼリー']
        df_jerry_b = df_jerry[df_jerry.meal_name__meal_name == '朝食']
        df_jerry_l = df_jerry[df_jerry.meal_name__meal_name == '昼食']
        df_jerry_d = df_jerry[df_jerry.meal_name__meal_name == '夕食']
        # 食数
        ws_meal.cell(14, 3, df_jerry_b['enge_quantity'].sum())
        ws_meal.cell(14, 4, df_jerry_l['enge_quantity'].sum())
        ws_meal.cell(14, 5, df_jerry_d['enge_quantity'].sum())
        # 売上
        jerry_sales_sum_b = df_jerry_b['enge_sales'].sum()
        jerry_sales_sum_l = df_jerry_l['enge_sales'].sum()
        jerry_sales_sum_d = df_jerry_d['enge_sales'].sum()
        # -汁の売上を除く
        ws_meal.cell(14, 7, jerry_sales_sum_b - (option_filling_sales_jerry_b + option_bose_sales_jerry_b))
        ws_meal.cell(14, 8, jerry_sales_sum_l - (option_filling_sales_jerry_l + option_bose_sales_jerry_l))
        ws_meal.cell(14, 9, jerry_sales_sum_d - (option_filling_sales_jerry_d + option_bose_sales_jerry_d))
        # -汁の売上を含む
        ws_meal.cell(14, 11, jerry_sales_sum_b)
        ws_meal.cell(14, 12, jerry_sales_sum_l)
        ws_meal.cell(14, 13, jerry_sales_sum_d)

        # 常食・薄味・個食・フリーズのみ
        # 売価が別の施設の嚥下(売価)
        df_diff_price = df_merge[df_merge.is_diff_price == True]
        df_diff_price_b = df_diff_price[df_diff_price.meal_name__meal_name == '朝食']
        df_diff_price_l = df_diff_price[df_diff_price.meal_name__meal_name == '昼食']
        df_diff_price_d = df_diff_price[df_diff_price.meal_name__meal_name == '夕食']
        # -具のみ
        df_diff_price_b_filling_option = df_diff_price_b[df_diff_price_b.soup_contract == '具のみ']
        df_diff_price_l_filling_option = df_diff_price_l[df_diff_price_l.soup_contract == '具のみ']
        df_diff_price_d_filling_option = df_diff_price_d[df_diff_price_d.soup_contract == '具のみ']
        # -汁具
        df_diff_price_b_bose_option = df_diff_price_b[df_diff_price_b.soup_contract == '汁具']
        df_diff_price_l_bose_option = df_diff_price_l[df_diff_price_l.soup_contract == '汁具']
        df_diff_price_d_bose_option = df_diff_price_d[df_diff_price_d.soup_contract == '汁具']
        # -オプション分の売上計算
        diff_filling_option_sales_b = \
            df_diff_price_b_filling_option['enge_quantity'].sum() * temp_ref_params['filling_breakfast'] \
            + df_diff_price_b_bose_option['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        diff_filling_option_sales_l = \
            df_diff_price_l_filling_option['enge_quantity'].sum() * temp_ref_params['filling_lunch'] \
            + df_diff_price_l_bose_option['enge_quantity'].sum() * temp_ref_params['bose_lunch']
        diff_filling_option_sales_d = \
            df_diff_price_d_filling_option['enge_quantity'].sum() * temp_ref_params['filling_dinner'] \
            + df_diff_price_d_bose_option['enge_quantity'].sum() * temp_ref_params['bose_dinner']
        # 食数
        ws_meal.cell(17, 3, df_diff_price_b['enge_quantity'].sum())
        ws_meal.cell(17, 4, df_diff_price_l['enge_quantity'].sum())
        ws_meal.cell(17, 5, df_diff_price_d['enge_quantity'].sum())
        # 売上
        diff_enge_sales_b = df_diff_price_b['enge_sales'].sum()
        diff_enge_sales_l = df_diff_price_l['enge_sales'].sum()
        diff_enge_sales_d = df_diff_price_d['enge_sales'].sum()
        # -汁の売上を除く
        ws_meal.cell(17, 7, diff_enge_sales_b - diff_filling_option_sales_b)
        ws_meal.cell(17, 8, diff_enge_sales_l - diff_filling_option_sales_l)
        ws_meal.cell(17, 9, diff_enge_sales_d - diff_filling_option_sales_d)
        # -汁の売上を含む
        ws_meal.cell(17, 11, diff_enge_sales_b)
        ws_meal.cell(17, 12, diff_enge_sales_l)
        ws_meal.cell(17, 13, diff_enge_sales_d)

        # 売価が同じ施設の嚥下(売価)
        df_same_price = df_merge[df_merge.is_diff_price == False]
        df_same_price_b = df_same_price[df_same_price.meal_name__meal_name == '朝食']
        df_same_price_l = df_same_price[df_same_price.meal_name__meal_name == '昼食']
        df_same_price_d = df_same_price[df_same_price.meal_name__meal_name == '夕食']
        # -具のみ
        df_same_price_b_filling_option = df_same_price_b[df_same_price_b.soup_contract == '具のみ']
        df_same_price_l_filling_option = df_same_price_l[df_same_price_l.soup_contract == '具のみ']
        df_same_price_d_filling_option = df_same_price_d[df_same_price_d.soup_contract == '具のみ']
        # -汁具
        df_same_price_b_bose_option = df_same_price_b[df_same_price_b.soup_contract == '汁具']
        df_same_price_l_bose_option = df_same_price_l[df_same_price_l.soup_contract == '汁具']
        df_same_price_d_bose_option = df_same_price_d[df_same_price_d.soup_contract == '汁具']
        # -オプション分の売上計算
        same_bose_option_sales_b = \
            df_same_price_b_filling_option['enge_quantity'].sum() * temp_ref_params['filling_breakfast'] \
            + df_same_price_b_bose_option['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        same_bose_option_sales_l = \
            df_same_price_l_filling_option['enge_quantity'].sum() * temp_ref_params['filling_breakfast'] \
            + df_same_price_l_bose_option['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        same_bose_option_sales_d = \
            df_same_price_d_filling_option['enge_quantity'].sum() * temp_ref_params['filling_breakfast'] \
            + df_same_price_d_bose_option['enge_quantity'].sum() * temp_ref_params['bose_breakfast']
        # 食数
        ws_meal.cell(18, 3, df_same_price_b['enge_quantity'].sum())
        ws_meal.cell(18, 4, df_same_price_l['enge_quantity'].sum())
        ws_meal.cell(18, 5, df_same_price_d['enge_quantity'].sum())
        # 売上
        same_enge_sales_b = df_same_price_b['enge_sales'].sum()
        same_enge_sales_l = df_same_price_l['enge_sales'].sum()
        same_enge_sales_d = df_same_price_d['enge_sales'].sum()
        # -汁の売上を除く
        ws_meal.cell(18, 7, same_enge_sales_b - same_bose_option_sales_b)
        ws_meal.cell(18, 8, same_enge_sales_l - same_bose_option_sales_l)
        ws_meal.cell(18, 9, same_enge_sales_d - same_bose_option_sales_d)
        # -汁の売上を含む
        ws_meal.cell(18, 11, same_enge_sales_b)
        ws_meal.cell(18, 12, same_enge_sales_l)
        ws_meal.cell(18, 13, same_enge_sales_d)

        # ------------------------------------------------------------------------------
        # サマリ情報出力
        # ------------------------------------------------------------------------------
        ws_summery = workbook['サマリ']

        # タイトルに出力年月を記載
        summery_title = ws_summery.cell(1, 2).value
        ws_summery.cell(1, 2, f'{from_date.year}年{from_date.month}月{summery_title}')

        # 配送料請求額
        ws_summery['B19'].value = in_transport_price

        # ファイル保存
        output_path = os.path.join(new_dir_path, in_target_month + '_売価計算表.xlsx')
        workbook.save(output_path)
        workbook.close()

        # ------------------------------------------------------------------------------
        # サマリ情報出力
        # ------------------------------------------------------------------------------
        sales_price, is_create = MonthlySalesPrice.objects.get_or_create(
            year=from_date.year, month=from_date.month
        )

        sales_price.transport_price = in_transport_price

        # 嚥下以外の合計情報取得
        # -基本食数
        basic_count_total_b = df_total_b['other_quantity'].sum()
        basic_count_total_l = df_total_l['other_quantity'].sum()
        basic_count_total_d = df_total_d['other_quantity'].sum()

        # -オプション売上合計(基本と嚥下の合計)
        total_opion_sales_b = df_soup_filling_b_count * temp_ref_params['filling_breakfast'] + df_soup_bose_b_count * temp_ref_params['bose_breakfast']
        total_opion_sales_l = df_soup_filling_l_count * temp_ref_params['filling_lunch'] + df_soup_bose_l_count * temp_ref_params['bose_lunch']
        total_opion_sales_d = df_soup_filling_d_count * temp_ref_params['filling_dinner'] + df_soup_bose_d_count * temp_ref_params['bose_dinner']

        # -基本オプション売上合計
        basic_opion_sales_b = basic_filling_sum_count_b * temp_ref_params['filling_breakfast'] + basic_bose_sum_count_b * temp_ref_params['bose_breakfast']
        basic_opion_sales_l = basic_filling_sum_count_l * temp_ref_params['filling_lunch'] + basic_bose_sum_count_l * temp_ref_params['bose_lunch']
        basic_opion_sales_d = basic_filling_sum_count_d * temp_ref_params['filling_dinner'] + basic_bose_sum_count_d * temp_ref_params['bose_dinner']

        # -汁を含まない基本売上の計算(全体売上-基本オプション合計-嚥下合計)
        basic_sales_total_b = all_total_sales_b - basic_opion_sales_b - diff_enge_sales_b - same_enge_sales_b
        basic_sales_total_l = all_total_sales_l - basic_opion_sales_l - diff_enge_sales_l - same_enge_sales_l
        basic_sales_total_d = all_total_sales_d - basic_opion_sales_d - diff_enge_sales_d - same_enge_sales_d

        # -嚥下食数
        enge_count_total_b = (df_soft_b['enge_quantity'].sum() + df_jerry_b['enge_quantity'].sum() + df_mixer_b['enge_quantity'].sum())
        enge_count_total_l = (df_soft_l['enge_quantity'].sum() + df_jerry_l['enge_quantity'].sum() + df_mixer_l['enge_quantity'].sum())
        enge_count_total_d = (df_soft_d['enge_quantity'].sum() + df_jerry_d['enge_quantity'].sum() + df_mixer_d['enge_quantity'].sum())

        # -嚥下売上
        enge_sales_total_b = (diff_enge_sales_b + same_enge_sales_b)
        enge_sales_total_l = (diff_enge_sales_l + same_enge_sales_l)
        enge_sales_total_d = (diff_enge_sales_d + same_enge_sales_d)

        sales_price.basic_breakfast_count = basic_count_total_b
        sales_price.basic_lunch_count = basic_count_total_l
        sales_price.basic_dinner_count = basic_count_total_d

        sales_price.basic_breakfast_sales = basic_sales_total_b
        sales_price.basic_lunch_sales = basic_sales_total_l
        sales_price.basic_dinner_sales = basic_sales_total_d

        sales_price.enge_breakfast_count = enge_count_total_b
        sales_price.enge_lunch_count = enge_count_total_l
        sales_price.enge_dinner_count = enge_count_total_d

        sales_price.enge_breakfast_soup_sales = \
            (option_filling_sales_soft_b + option_bose_sales_soft_b) + \
            (option_filling_sales_mixer_b + option_bose_sales_mixer_b) + \
            (option_filling_sales_jerry_b + option_bose_sales_jerry_b)
        sales_price.enge_lunch_soup_sales = \
            (option_filling_sales_soft_l + option_bose_sales_soft_l) + \
            (option_filling_sales_mixer_l + option_bose_sales_mixer_l) + \
            (option_filling_sales_jerry_l + option_bose_sales_jerry_l)
        sales_price.enge_dinner_soup_sales = \
            (option_filling_sales_soft_d + option_bose_sales_soft_d) + \
            (option_filling_sales_mixer_d + option_bose_sales_mixer_d) + \
            (option_filling_sales_jerry_d + option_bose_sales_jerry_d)

        sales_price.enge_breakfast_sales = enge_sales_total_b - sales_price.enge_breakfast_soup_sales
        sales_price.enge_lunch_sales = enge_sales_total_l - sales_price.enge_lunch_soup_sales
        sales_price.enge_dinner_sales = enge_sales_total_d - sales_price.enge_dinner_soup_sales

        sales_price.basic_breakfast_soup_sales = total_opion_sales_b - sales_price.enge_breakfast_soup_sales
        sales_price.basic_lunch_soup_sales = total_opion_sales_l - sales_price.enge_lunch_soup_sales
        sales_price.basic_dinner_soup_sales = total_opion_sales_d - sales_price.enge_dinner_soup_sales

        sales_price.transport_breakfast_rate = temp_ref_params['rate_breakfast']
        sales_price.transport_lunch_rate = temp_ref_params['rate_lunch']
        sales_price.transport_dinner_rate = temp_ref_params['rate_dinner']

        sales_price.save()
        self.logger.info('売価計算表出力コマンド終了')
