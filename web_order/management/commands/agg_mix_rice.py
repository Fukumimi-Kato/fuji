import logging
import math
import os
import openpyxl as excel
import pandas as pd
import re

from django.conf import settings
from django.db.models import Sum
from django_pandas.io import read_frame

from web_order.models import Order, OrderRice, MixRicePackageMaster, TmpPlateNamePackage, OrderEveryday
from .utils import ExcelOutputMixin
from web_order.p7 import P7Util
from web_order.picking import PlatePackageRegister
from web_order.cooking_direction_plates import PlateNameAnalizeUtil


logger = logging.getLogger(__name__)


class MixRiceMeasureWriter(ExcelOutputMixin):
    """
    混ぜご飯計量表出力
    """

    def __init__(self, mix_rice):
        self.agg_mix_rice = mix_rice

        self.LIST_HIDDEN_PLATE_QUANTITY_ROW = 2
        self.LIST_HEADER_PLATE_ROW = 3

        self.LIST_ORDER_ROW_MIN = 30
        self.LIST_ORDER_ROW_MAX = 90
        # 食数固定製造分の行インデックス(テンプレートに合わせること)
        self.LIST_NEEDLE_PACKAGE_ROW = self.LIST_ORDER_ROW_MAX + 2
        self.LIST_STOCK_PACKAGE_ROW = self.LIST_ORDER_ROW_MAX + 3
        self.LIST_1_PACK_QUANTITY_ROW = self.LIST_ORDER_ROW_MAX + 4
        self.LIST_ENGE_START_QUANTITY_ROW = self.LIST_ORDER_ROW_MAX + 5
        self.LIST_TOTAL_QUANTITY_ROW = self.LIST_ORDER_ROW_MAX + 8

        # ソフト食固定製造分(針刺し用10人前*1、保存用1人前*4)
        self.SOFT_FIX_QUANTITY = 10 + 4
        # ゼリー食固定製造分(保存用1人前*2)
        self.JELLY_FIX_QUANTITY = 2
        # ミキサー食固定製造分(保存用1人前*2)
        self.MIXER_FIX_QUANTITY = 2

        # 袋サイズのデフォルト値(g/袋)
        self.DEFAULT_QUANTITY_PER_PACKAGE = 1000

        # セル背景色
        # -乾燥
        self.FILL_DRY_ITEM = excel.styles.PatternFill(patternType='solid', fgColor='FFEEFF')
        # -出汁・液
        self.FILL_SOUP_ITEM = excel.styles.PatternFill(patternType='solid', fgColor='DDEBF7')
        # -調味料(酢など)
        self.FILL_SEASONING_ITEM = excel.styles.PatternFill(patternType='solid', fgColor='E2EFDA')
        # -通常
        self.FILL_DEFAULT_ITEM = excel.styles.PatternFill(patternType='solid', fgColor='FFF2CC')

        # 計量表シートの最大行数(テンプレートと合わせること)
        self.MEASURE_MAX_UNIT_ROW = 90

    def write(self):
        if self.agg_mix_rice.is_disable:
            # 出力対象が無効化中なら出力しない

            # アレルギー袋登録用の処理は実施する。
            if self.agg_mix_rice.index == -1:
                TmpPlateNamePackage.objects.get_or_create(
                    plate_name=self.agg_mix_rice.before_name,
                    cooking_day=self.agg_mix_rice.cooking_day,
                    size=10,
                    menu_name='常食'
                )
            return

        # 入出力ファイルパス
        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, f'計量表_{self.agg_mix_rice.cooking_day}_製造')
        os.makedirs(new_dir_path, exist_ok=True)  # 上書きOK

        measure_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_rice.xlsx')  # 計量表のテンプレート

        output_file = os.path.join(new_dir_path, f'混ぜご飯計量表_{self.agg_mix_rice.eating_day}_{self.agg_mix_rice.meal}_{self.agg_mix_rice.name}.xlsx')

        # ------------------------------------------------------------------------------
        # 当日の混ぜご飯合数
        # ------------------------------------------------------------------------------
        # 通常、同じユニットが同じ日にレコードを作成することはないため、集計はしていない
        qs = OrderRice.objects \
            .filter(eating_day=self.agg_mix_rice.eating_day, quantity__gt=0) \
            .values('unit_name__unit_number', 'unit_name__calc_name', 'unit_name__username__username',
                    'quantity', 'eating_day', 'unit_name__username__dry_cold_type') \
            .exclude(unit_name__unit_code__range=[80001, 80008]) \
            .order_by('unit_name__unit_number')

        book_b = excel.load_workbook(measure_template)
        df = self.make_dataframe(qs)

        temp_dir_path = "tmp"
        df.to_csv(temp_dir_path + "/T-1_orderrice.csv", index=False)

        # ------------------------------------------------------------------------------
        # 当日の注文数(常食(=基本食)のみ抽出)
        # ------------------------------------------------------------------------------
        order_qs = Order.objects \
            .filter(eating_day=self.agg_mix_rice.eating_day, meal_name__meal_name=self.agg_mix_rice.meal, quantity__gt=0, menu_name=1) \
            .values('unit_name__unit_number', 'unit_name__calc_name', 'unit_name__username__username', 'unit_name__username__facility_name',
                    'meal_name__meal_name', 'quantity', 'eating_day') \
            .exclude(unit_name__unit_code__range=[80001, 80008]) \
            .order_by('unit_name__unit_number')

        df_qs = self.make_dataframe_order(order_qs)
        df_qs.to_csv(temp_dir_path + "/T-2_order.csv", index=False)

        # 1人用袋数の計算
        pack_1p_default = self._get_1pack_default(self.agg_mix_rice.meal) # 保存用1人用、写真用1人用*2
        pack_1p = pack_1p_default
        raw_1p = 0
        for index, row in df_qs.iterrows():
            if row['order_quantity'] == 1:
                pack_1p += 1
                raw_1p += 1

        # 結合
        joined_df = pd.merge(df, df_qs, on=['unit_name__unit_number', 'unit_name__calc_name'], how='inner')
        joined_df.to_csv(temp_dir_path + "/T-3_joined.csv", index=False)

        # ユニットの統合表示
        df_list = self.make_dataframes_for_aggregate(joined_df)

        dict_list = []
        df_dict = {}
        for df1 in df_list:
            total = df1['quantity'].sum()
            df_dict['base'] = df1[0:1]['unit_name__unit_number']
            df_dict['others'] = []
            for index, row in df1[1:].iterrows():
                df_dict['others'].append(row['unit_name__unit_number'])
            df_dict['total'] = total

            dict_list.append(df_dict)

        del_index_list = []
        others_list = []
        for x in dict_list:
            others_list += x['others']

        for index, row in joined_df.iterrows():
            number = row['unit_name__unit_number']

            if number in others_list:
                del_index_list.append(index)
            else:
                for x in dict_list:
                    if number in x['base'].head().values:
                        joined_df.loc[index, 'unit_name__calc_name'] = row['unit_name__username__facility_name']
                        joined_df.loc[index, 'quantity'] = x['total']
                        break

        # 統合された行の削除
        joined_df = joined_df.drop(joined_df.index[del_index_list])
        joined_df.to_csv(temp_dir_path + "/T-4_joined2.csv", index=False)

        joined_df = self.groupby_dataframe(joined_df)
        joined_df.to_csv(temp_dir_path + "/T-5_joined_comp.csv", index=False)

        # 計量表シート出力
        self.write_measure_sheet(book_b["計量表"], joined_df)

        # ------------------------------------------------------------------------------
        # 当日の注文数(嚥下のみ抽出)
        # ------------------------------------------------------------------------------
        agg_soft = self._get_enge_order_count(5) + self.SOFT_FIX_QUANTITY
        agg_jelly = self._get_enge_order_count(3) + self.JELLY_FIX_QUANTITY
        agg_mixer = self._get_enge_order_count(4) + self.MIXER_FIX_QUANTITY

        # 合数シート出力
        order_dict = self.write_list_sheet(book_b, joined_df, agg_soft, agg_jelly, agg_mixer)

        self.save_with_select(book_b, output_file)
        book_b.close()

        try:
            # 袋数出力のため、読込
            book_b2 = excel.load_workbook(output_file)
            ws2 = book_b2["合数一覧"]

            pack_50g_count = self._get_50g_pack_count(self.agg_mix_rice.meal)

            # 袋数の合計
            base_col = 3
            total_package = 0
            for row in ws2.iter_rows(min_row=4, max_row=90):
                if self.agg_mix_rice.mix_rice.is_mix_package:
                    # 同封の場合
                    if (not row[1].value) or (row[1].value == ''):
                        continue
                    if (not row[base_col].value) or (row[base_col].value == ''):
                        continue
                    if (not row[base_col + 3].value) or (row[base_col + 3].value == ''):
                        continue
                    gou_value = float(row[2].value)
                    amount1 = float(ws2.cell(2, base_col + 1).value)
                    amount2 = float(ws2.cell(2, base_col + 4).value)
                    package_size = float(ws2.cell(2, base_col + 5).value)
                    package_count = math.ceil(gou_value*(amount1 + amount2)/package_size)
                    total_package += package_count
                else:
                    if (not row[1].value) or (row[1].value == ''):
                        continue
                    if (not row[base_col].value) or (row[base_col].value == ''):
                        continue
                    gou_value = float(row[2].value)
                    amount = float(ws2.cell(2, base_col + 1).value)
                    package_size = float(ws2.cell(2, base_col + 2).value)
                    package_count = math.ceil(gou_value*amount/package_size)
                    total_package += package_count
                number, name = self._parse_unit_name(row[1].value)
                if self.agg_mix_rice.unit == 'g':
                    PlatePackageRegister.register_unit_package(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                               self.agg_mix_rice.meal, self.agg_mix_rice.index,
                        number, name, package_count, 'BASIC_UNIT', self.agg_mix_rice.name, mix_rice_type='main')

            # メイン(1人用の登録)
            pack_col = 6
            if self.agg_mix_rice.mix_rice.is_mix_package:
                pack_col += 3
            pack_t = total_package
            pack_n = int(ws2.cell(92, pack_col).value)
            pack_s = int(ws2.cell(93, pack_col).value)
            P7Util.save_package_count_for_print(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day, self.agg_mix_rice.index,
                                                pack_t + pack_n + pack_s, pack_1p,
                                                '常食', self.agg_mix_rice.meal, pack_50g_count)
            PlatePackageRegister.register_unit_package(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                       self.agg_mix_rice.meal, self.agg_mix_rice.index,
                                                       number, name, raw_1p, 'BASIC_1', self.agg_mix_rice.name,
                                                       mix_rice_type='main')
            # 1人用袋数の保存
            ws2.cell(self.LIST_1_PACK_QUANTITY_ROW, pack_col, pack_1p + pack_50g_count)

            for index, parts in enumerate(self.agg_mix_rice.parts):
                # 混ぜご飯パーツの登録
                base_col += 3
                total_package = 0
                order_index_col = base_col + 1
                if order_index_col in order_dict:
                    total_package += order_dict[order_index_col]
                else:
                    unit_index = -1
                    for row in ws2.iter_rows(min_row=4, max_row=90):
                        unit_index += 1 # ここの初回を0にするため、初期値を-1にした
                        if (not row[1].value) or (row[1].value == ''):
                            continue
                        if (not row[base_col].value) or (row[base_col].value == ''):
                            continue
                        gou_value = float(row[2].value)
                        amount = float(ws2.cell(2, base_col + 1).value)
                        package_size = float(ws2.cell(2, base_col + 2).value)
                        package_count = math.ceil(gou_value * amount / package_size)
                        total_package += package_count

                        # 袋数保存
                        number, name = self._parse_unit_name(ws2.cell(self.LIST_HEADER_PLATE_ROW + 1 + unit_index, 2).value)
                        if parts.unit == 'g':
                            PlatePackageRegister.register_unit_package(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                               self.agg_mix_rice.meal, self.agg_mix_rice.index + index + 1,
                                number, name, package_count, 'BASIC_UNIT', parts.name, mix_rice_type='parts', is_raw_to_enge=True)

                pack_col += 3
                pack_t = total_package
                pack_n = int(ws2.cell(92, pack_col).value)
                pack_s = int(ws2.cell(93, pack_col).value)
                save_flg = False
                if PlateNameAnalizeUtil.is_raw_plate_name(parts.name):
                    if self._is_dry_parts(parts.name):
                        save_flg = True
                        parts_pack_1p = pack_1p_default # 針刺し・保存用のみの数にする
                    else:
                        parts_pack_1p = pack_1p
                else:
                    parts_pack_1p = pack_1p
                P7Util.save_package_count_for_print(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                    self.agg_mix_rice.index + index + 1,
                                                    pack_t + pack_n + pack_s, parts_pack_1p,
                                                    '常食', self.agg_mix_rice.meal, pack_50g_count, save_flg)
                # 錦糸卵を登録できるように、is_raw_to_engeにTrueを渡す
                PlatePackageRegister.register_unit_package(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                           self.agg_mix_rice.meal, self.agg_mix_rice.index + index + 1,
                                                           number, name, raw_1p, 'BASIC_1', parts.name,
                                                           mix_rice_type='parts', is_raw_to_enge=True)

                # 1人用袋数の保存
                ws2.cell(self.LIST_1_PACK_QUANTITY_ROW, pack_col, parts_pack_1p + pack_50g_count)

            self.save_with_select(book_b2, output_file)
            book_b2.close()

            # 合数ログとの比較用に、計算後の内容を読み取る
            book_b3 = excel.load_workbook(output_file)
            pd_ws = book_b3['材料配分']
            pd_ws.column_dimensions["Y"].hidden = False
            pd_ws.sheet_view.selection[0].activeCell = "Y5"
            pd_ws.sheet_view.selection[0].sqref = "Y5"
            pd_dir = os.path.join(settings.OUTPUT_DIR, 'gosu_check')
            checkfile_path = os.path.join(pd_dir, f'混ぜご飯計量表_{self.agg_mix_rice.eating_day}_{self.agg_mix_rice.meal}_{self.agg_mix_rice.name}.xlsx')
            book_b3.save(checkfile_path)
            book_b3.close()

        except Exception as er:
            pass

        if self.agg_mix_rice.index == -1:
            TmpPlateNamePackage.objects.get_or_create(
                plate_name=self.agg_mix_rice.before_name,
                cooking_day=self.agg_mix_rice.cooking_day,
                size=10,
                menu_name='常食'
            )

    def get_fixed_quantity(self, id: int):
        qs_fix = OrderEveryday.objects.filter(id=id)
        if qs_fix.exists():
            return qs_fix.first().quantity
        else:
            return 0

    def _get_1pack_default(self, meal: str):
        """
        食数固定注文情報を元に1人用袋のデフォルト数を取得する。
        """
        # ------------------------------------------------------------------------------
        # 常食保存用の食数(1人用)
        # ------------------------------------------------------------------------------
        if meal == '朝食':
            pre_1pack_j = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_1PACK_ID_J[0])  # 保存用・朝・常食
        elif meal == '昼食':
            pre_1pack_j = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_1PACK_ID_J[1])  # 保存用・昼・常食
        elif meal == '夕食':
            pre_1pack_j = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_1PACK_ID_J[2])  # 保存用・夕・常食
        else:
            pre_1pack_j = 0

        # ------------------------------------------------------------------------------
        # 写真用の食数
        # ------------------------------------------------------------------------------
        if meal == '朝食':
            photo = self.get_fixed_quantity(settings.ORDER_EVERYDAY_FOR_PHOTO_ID_J[0])  # 写真用・朝・常食
        elif meal == '昼食':
            photo = self.get_fixed_quantity(settings.ORDER_EVERYDAY_FOR_PHOTO_ID_J[1])  # 写真用・昼・常食
        elif meal == '夕食':
            photo = self.get_fixed_quantity(settings.ORDER_EVERYDAY_FOR_PHOTO_ID_J[2])  # 写真用・夕・常食
        else:
            photo = 0

        return pre_1pack_j + photo

    def _parse_unit_name(self, unit_name):
        index = unit_name.find('.')
        if index == -1:
            return None
        else:
            number = unit_name[:index]
            name = unit_name[index + 1:]
            return number, name

    def _get_50g_pack_count(self, meal: str):
        """
        食数固定注文情報から元に保存用50g袋の注文数を取得する。
        """
        if meal == '朝食':
            count = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_50G_ID_J[0])  # 保存用50g・朝・常食
        elif meal == '昼食':
            count = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_50G_ID_J[1])  # 保存用50g・昼・常食
        elif meal == '夕食':
            count = self.get_fixed_quantity(settings.ORDER_EVERYDAY_PRESERVE_50G_ID_J[0])  # 保存用50g・昼・常食
        else:
            count = 0

        return count

    def _get_enge_order_count(self, menu_name_id):
        agg = Order.objects \
            .filter(eating_day=self.agg_mix_rice.eating_day, meal_name__meal_name=self.agg_mix_rice.meal, quantity__gt=0, menu_name=menu_name_id) \
            .exclude(unit_name__unit_code__range=[80001, 80008]) \
            .aggregate(total=Sum("quantity"))
        return agg["total"]

    def _is_soup_parts(self, parts):
        if '出汁' in parts.name:
            return True
        elif '栗ご飯の液' in parts.name:
            return True
        elif 'チキンライスの素' in parts.name:
            return True
        else:
            return False

    def make_dataframe(self, qs):
        dataframe = read_frame(qs)
        dataframe = dataframe.groupby(['unit_name__unit_number',
                                       'unit_name__calc_name',
                                       'unit_name__username__dry_cold_type',
                                       'eating_day']).sum().reset_index()
        return dataframe

    def make_dataframes_for_aggregate(self, df):
        df_list = []
        df_list2 = []
        for agg_stting in settings.MIX_RICE_AGGREGATE_UNITS:
            s_list = []
            for s1 in df.unit_name__unit_number:
                if s1 in agg_stting:
                    s_list.append(s1)
            for x in s_list:
                df_list.append(
                    df[(df.unit_name__unit_number == x)]
                )
            if df_list:
                hoge_df = df_list[0]
                for df1 in df_list[1:]:
                    hoge_df = hoge_df.append(df1)
                df_list2.append(hoge_df)
        return df_list2

    def make_dataframe_order(self, qs):
        dataframe = read_frame(qs)
        dataframe = dataframe.groupby(['unit_name__unit_number',
                                       'unit_name__calc_name', 'unit_name__username__facility_name',
                                       'meal_name__meal_name']).sum().reset_index()
        dataframe = dataframe.rename(columns={'quantity': 'order_quantity'})
        return dataframe

    def groupby_dataframe(self, df):
        dataframe = df.groupby(['unit_name__unit_number',
                                       'unit_name__calc_name',
                                       'unit_name__username__dry_cold_type',
                                       'eating_day',
                                       'unit_name__username__username', 'quantity',
                                       'meal_name__meal_name'
                                ]).sum().reset_index()
        return dataframe

    def fill_cell_bg_color(self, ws, columun, is_dry, is_soup, is_seasoning):
        """
        一覧上のセル背景色を変更する。対象列の料理名～合計までの全行が対象
        """
        cell_from = ws.cell(self.LIST_HEADER_PLATE_ROW, columun).coordinate
        cell_to = ws.cell(self.LIST_TOTAL_QUANTITY_ROW, columun).coordinate
        if is_dry:
            fill = self.FILL_DRY_ITEM
            is_enge_clear = True
        elif is_soup:
            fill = self.FILL_SOUP_ITEM
            is_enge_clear = False
        elif is_seasoning:
            fill = self.FILL_SEASONING_ITEM
            is_enge_clear = True
        else:
            fill = self.FILL_DEFAULT_ITEM
            is_enge_clear = False

        for cell in ws[f'{cell_from}:{cell_to}']:
            # cellはtuppleで取得
            cell[0].fill = fill

            if is_enge_clear and cell[0].row in [95, 96, 97]:
                cell[0].value = ''

    def write_list_sheet(self, wb, df, soft_count: int, jelly_count: int, mixer_count: int):
        ws = wb['合数一覧']

        # 嚥下数量を出力
        soft_cell = ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, 3)
        jelly_cell = ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, 3)
        mixer_cell = ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, 3)
        soft_cell.value = soft_count
        jelly_cell.value = jelly_count
        mixer_cell.value = mixer_count

        # タイトル
        ws.cell(1, 2, f'混ぜご飯合数一覧:{self._get_title()}')

        # 具材を出力
        col = 4
        dry_parts_col = []
        order_parts_col = []
        assembly_col = []

        if self.agg_mix_rice.mix_rice.is_mix_package:
            # 具と汁が一体になっているもの
            if self.agg_mix_rice.base_soup:
                # 料理名に具と汁の指定があるもの(同封の場合に%指定はない想定)
                # 今のところ、個数指定はない

                # 具
                filling_name = self._get_display_rice_name(self.agg_mix_rice.get_mix_rice_name())

                # -合数
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_quantity_1_gou())

                # -針刺し用
                filling_quantity = float(self.agg_mix_rice.quantity)
                soup_quantity = self._get_soup_quantity()
                filling_package_size = self._get_package_size(filling_name, self.agg_mix_rice.unit)
                filling_soup_rate = filling_package_size / (filling_quantity + soup_quantity)
                # ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, filling_quantity * filling_soup_rate)

                # -保存用(10人前の袋)
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={filling_quantity}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={self.agg_mix_rice.quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={self.agg_mix_rice.quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={self.agg_mix_rice.quantity}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, filling_package_size)
                # -料理名(具)
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, filling_name)
                self.fill_cell_bg_color(ws, col, False, False, False)
                # -具の袋数を非表示(col + 1は元々非表示)
                col_str = ws.cell(self.LIST_HEADER_PLATE_ROW, col + 2).coordinate[:-1]
                ws.column_dimensions[col_str].hidden = True

                col += 3

                # 汁
                soup_name = f'{self.agg_mix_rice.mix_rice.name}の液'
                # -合数
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_soup_quantity_1_gou())
                # -針刺し用(袋サイズは仮)
                # ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, soup_quantity * filling_soup_rate)
                # -保存用(10人前の袋)
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={soup_quantity}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={soup_quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={soup_quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={soup_quantity}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, filling_package_size)
                # -料理名(汁)
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, soup_name)
                self.fill_cell_bg_color(ws, col, False, True, False)

                assembly_col.append(col + 1)

                col += 3
            else:
                # 汁(液)がpartで設定されている

                # 具のみ出力
                filling_name = self._get_display_rice_name(self.agg_mix_rice.get_mix_rice_name())
                # -合数
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_quantity_1_gou())
                # -針刺し用(袋サイズは仮)
                filling_quantity = float(self.agg_mix_rice.quantity)
                soup_quantity = self._get_soup_quantity()
                filling_package_size = self._get_package_size(filling_name, self.agg_mix_rice.unit)
                filling_soup_rate = filling_package_size / (filling_quantity + soup_quantity)
                # ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, filling_quantity * filling_soup_rate)
                # -保存用(10人前の袋)
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={filling_quantity}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={self.agg_mix_rice.quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col,
                        f'={self.agg_mix_rice.quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col,
                        f'={self.agg_mix_rice.quantity}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, filling_package_size)
                # -料理名(具)
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, filling_name)
                self.fill_cell_bg_color(ws, col, False, False, False)
                # -具の袋数を非表示(col + 1は元々非表示)
                col_str = ws.cell(self.LIST_HEADER_PLATE_ROW, col + 2).coordinate[:-1]
                ws.column_dimensions[col_str].hidden = True

                # 単位が個の場合は非表示にする
                if self.agg_mix_rice.unit != 'g':
                    ws.column_dimensions['D'].hidden = True
                    ws.column_dimensions['E'].hidden = True
                    ws.column_dimensions['F'].hidden = True

                col += 3
        else:
            if self.agg_mix_rice.base_soup or self.agg_mix_rice.percentage:
                # 具
                display_name = self.agg_mix_rice.name
                # -合数
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_quantity_1_gou())
                # -針刺し用(袋サイズは仮)
                package_size = self._get_package_size(display_name, self.agg_mix_rice.unit)
                # ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, package_size)
                # -保存用(10人前の袋)
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={self.agg_mix_rice.quantity}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={self.agg_mix_rice.quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={self.agg_mix_rice.quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={self.agg_mix_rice.quantity}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, package_size)
                # -料理名
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}')
                self.fill_cell_bg_color(ws, col, False, False, False)

                """
                col += 3

                # 汁(液・出汁)
                display_name = f'{self.agg_mix_rice.get_mix_rice_name()}出汁'
                # -合数
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_soup_quantity_1_gou())
                # -針刺し用(空になるように対応)
                package_size = self._get_package_size(display_name, 'g')
                ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, '')
                ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col + 2, '-')
                # -保存用(10人前の袋)
                soup_quantity = self._get_soup_quantity()
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={soup_quantity}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={soup_quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={soup_quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={soup_quantity}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, package_size)
                # -料理名
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}')
                self.fill_cell_bg_color(ws, col, False, False, False)
                """
            else:
                # 具と汁が一体になっているもの(同封はしない。パーツ料理に汁が存在する)
                display_name = self.agg_mix_rice.name
                # -合数
                q = self._get_with_soup_quantity_1_gou()
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, q)
                # -針刺し用(袋サイズは仮)
                quantity_for_order = self._get_with_soup_quantity()
                package_size = self._get_package_size(display_name, self.agg_mix_rice.unit)
                #ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, '')
                # -保存用(10人前の袋)
                ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={quantity_for_order}*10')
                # 嚥下
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={quantity_for_order}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={quantity_for_order}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={quantity_for_order}*{mixer_cell.coordinate}')
                # -袋サイズ
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, package_size)
                # -料理名
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}')
                self.fill_cell_bg_color(ws, col, False, False, False)

            # 単位が個の場合は非表示にする
            if self.agg_mix_rice.unit != 'g':
                ws.column_dimensions['D'].hidden = True
                ws.column_dimensions['E'].hidden = True
                ws.column_dimensions['F'].hidden = True

            col += 3

        parts_list = self.agg_mix_rice.parts

        # パーツの出力
        for parts in parts_list:
            display_name = self._get_display_parts_name(parts.name)
            # -合数
            ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_parts_quantity_1_gou(parts))
            # -針刺し用(テンプレートにて対応)
            parts_package_size = self._get_package_size(display_name, parts.unit)
            #ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col, '')
            #ws.cell(self.LIST_NEEDLE_PACKAGE_ROW, col + 2, '-')
            # -保存用(10人前の袋)
            parts_quantity = self._get_parts_quantity_order(parts)
            ws.cell(self.LIST_STOCK_PACKAGE_ROW, col, f'={parts_quantity}*10')
            # 嚥下
            if not self._is_soup_parts(parts):
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW, col, f'={parts_quantity}*{soft_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 1, col, f'={parts_quantity}*{jelly_cell.coordinate}')
                ws.cell(self.LIST_ENGE_START_QUANTITY_ROW + 2, col, f'={parts_quantity}*{mixer_cell.coordinate}')
            # -袋サイズ
            ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col + 1, parts_package_size)
            # -料理名
            if self._is_dry_parts(parts.name):
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}(乾燥)')
                self.fill_cell_bg_color(ws, col, True, False, False)
                dry_parts_col.append(col)
            elif self._is_order_pattern_parts(parts.name):
                ws.cell(self.LIST_HIDDEN_PLATE_QUANTITY_ROW, col, self._get_parts_quantity_order(parts))
                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}')
                self.fill_cell_bg_color(ws, col, False, False, False)
                order_parts_col.append(col)
            else:
                # セル背景色の設定
                if self._is_seasoning_parts(parts.name):
                    self.fill_cell_bg_color(ws, col, False, False, True)
                elif self._is_extra_soup_parts(parts.name):
                    self.fill_cell_bg_color(ws, col, False, True, False)
                else:
                    self.fill_cell_bg_color(ws, col, False, False, False)

                ws.cell(self.LIST_HEADER_PLATE_ROW, col, f'{display_name}')

            # 単位が個の場合は非表示にする
            if parts.unit != 'g':
                for i in range(3):
                    hidden_col = ws.cell(2, col + i).coordinate[:-1]
                    ws.column_dimensions[hidden_col].hidden = True

            col += 3

        max_col = col

        # 各施設の合数を出力
        i = 4
        order_parts_dict = {}
        for index, row in df.iterrows():
            ws.row_dimensions[i].hidden = False
            ws.cell(i, 2, f'{row["unit_name__unit_number"]}.{row["unit_name__calc_name"]}')
            ws.cell(i, 3, float(row["quantity"]))

            # 乾燥・冷凍区分が冷凍のユニットは、乾燥専用品で空欄を表示する
            if row["unit_name__username__dry_cold_type"] == '乾燥':
                pass
            else:
                for c in dry_parts_col:
                    ws.cell(i, c, '')

            # 汁と具を同封するものは、汁と具の合計で袋数を算出する
            # (テンプレートは汁の数量で計算するようになっているため、ここで置き換え)
            for assembly in assembly_col:
                soup_cel = ws.cell(i, assembly - 1).coordinate
                filling_cel = ws.cell(i, assembly - 4).coordinate
                rating_cel = ws.cell(2, assembly).coordinate
                ws.cell(i, assembly, f'=SUM({soup_cel}+{filling_cel})/{rating_cel}')

            # 注文数出力
            for order_col in order_parts_col:
                rating_cel = ws.cell(2, order_col + 1).coordinate
                ws.cell(i, order_col, f'={row["order_quantity"]}*{ws.cell(2, order_col).coordinate}')
                ws.cell(i, order_col + 1, f'={ws.cell(i, order_col).coordinate}/{rating_cel}')

                # 袋数計算
                rating_value = ws.cell(2, order_col + 1).value
                base_value = ws.cell(2, order_col).value
                order_qty = row["order_quantity"] * base_value
                package_count = math.ceil(order_qty/rating_value)
                if order_col in order_parts_dict:
                    order_parts_dict[order_col] += package_count
                else:
                    order_parts_dict[order_col] = package_count

                # ピッキング指示書用袋数保存
                PlatePackageRegister.register_unit_package(self.agg_mix_rice.cooking_day, self.agg_mix_rice.eating_day,
                                                           self.agg_mix_rice.meal, self.agg_mix_rice.index + index + 1,
                    row["unit_name__unit_number"], row["unit_name__calc_name"], package_count, 'BASIC_UNIT',
                    ws.cell(3, order_col).value)

            i += 1

        if i < self.LIST_ORDER_ROW_MIN:
            i = self.LIST_ORDER_ROW_MIN
        while i < self.LIST_ORDER_ROW_MAX:
            ws.row_dimensions[i].hidden = True
            i += 1

        # 材料比率を出力
        self.write_items_rate_list(wb['材料配分'], f'混ぜご飯材料配分:{self._get_title()}', ws)

        # 余分な列を非表示
        limit_col = ord('U') - ord('A') + 1
        for c in range(limit_col - max_col + 1):
            index = chr(ord('A') + max_col + c - 1)
            ws.column_dimensions[index].hidden = True

        # 注文数換算の袋数を返却
        return order_parts_dict

    def _is_oil_item(self, item_name):
        """
        油の具材かどうかを判定する。～の素の液の分量から差し引く油の分量を取得するために使用する。
        ->要件がなくなったため、全てFalseとする
        """
        #if '大豆白絞油' in item_name:
            #return True
        #else:
        return False

    def _write_parts_item_rate(self, ws, parts, row, base_column, column, index):
        """
        材料配分シート下段の表に、パーツを具材単位で記述する
        """
        write_line = 0
        diff = 7
        if self.agg_mix_rice.mix_rice.is_mix_package:
            diff -= 2
        calc_base_column = 11
        if self.agg_mix_rice.mix_rice.is_mix_package:
            calc_base_column += 2
        item_calc_qty_min_cell = ws.cell(10, (calc_base_column + 1) + index * diff).coordinate
        item_calc_qty_max_cell = ws.cell(91, (calc_base_column + 1) + index * diff).coordinate
        item_calc_pack_min_cell = ws.cell(10, calc_base_column + index * diff).coordinate
        item_calc_pack_max_cell = ws.cell(91, calc_base_column + index * diff).coordinate
        item_calc_cnt_min_cell = ws.cell(10, (calc_base_column + 5) + index * diff).coordinate
        item_calc_cnt_max_cell = ws.cell(91, (calc_base_column + 5) + index * diff).coordinate

        # -1袋分量最大数
        ws.cell(row, base_column + 12, f'=MAX(計量表!{item_calc_qty_min_cell}:{item_calc_qty_max_cell})')
        item_max_cell = ws.cell(row, base_column + 12).coordinate
        # -1袋分量最大施設の合数
        func = f'=INDEX(計量表!{item_calc_pack_min_cell}:{item_calc_pack_max_cell}, MATCH(材料配分!{item_max_cell}, 計量表!{item_calc_qty_min_cell}:{item_calc_qty_max_cell}, 0))'
        ws.cell(row, base_column + 13,
                f'=INDEX(計量表!{item_calc_pack_min_cell}:{item_calc_pack_max_cell}, MATCH(材料配分!{item_max_cell}, 計量表!{item_calc_qty_min_cell}:{item_calc_qty_max_cell}, 0))')
        item_max_gosu_cell = ws.cell(row, base_column + 13).coordinate

        # -1袋分量最大施設の袋数
        ws.cell(row, base_column + 14,
                f'=INDEX(計量表!{item_calc_cnt_min_cell}:{item_calc_cnt_max_cell}, MATCH(材料配分!{item_max_cell}, 計量表!{item_calc_qty_min_cell}:{item_calc_qty_max_cell}, 0))')
        item_max_pack_cell = ws.cell(row, base_column + 14).coordinate
        # -針刺し用1袋の食数
        ws.cell(row, base_column + 15, f'={item_max_gosu_cell}*3/{item_max_pack_cell}')
        item_needle_cell = ws.cell(row, base_column + 15).coordinate
        total_cell = ws.cell(self.LIST_ORDER_ROW_MAX + 2, 9 + 3 * index).coordinate

        item_1p_cell = ws.cell(self.LIST_1_PACK_QUANTITY_ROW, column + 1).coordinate

        for i_index, values in enumerate(parts.items):
            item_name, quantity = values
            ws.cell(row + i_index, base_column, item_name)

            # 具の分量の設定
            ws.cell(row + i_index, base_column + 1, quantity)
            item_per_cell = ws.cell(row + i_index, base_column + 1).coordinate

            # 針刺し用
            ws.cell(row + i_index, base_column + 2, f'=IF({item_per_cell}, {item_per_cell}*{item_needle_cell}*合数一覧!{total_cell}, 0)')
            #item_needle_cell = ws.cell(row + i_index, base_column + 2).coordinate

            # -1人用袋
            ws.cell(row + i_index, base_column + 5, f'={item_per_cell}*合数一覧!{item_1p_cell}')

            # 補正無し合計
            item_saved_cell = ws.cell(row + i_index, base_column + 4).coordinate
            tmp_item_1p_cell = ws.cell(row + i_index, base_column + 5).coordinate
            item_soft_cell = ws.cell(row + i_index, base_column + 6).coordinate
            item_jelly_cell = ws.cell(row + i_index, base_column + 7).coordinate
            item_mixer_cell = ws.cell(row + i_index, base_column + 8).coordinate
            """
            # 補正ありに設定されていたため、コメントアウト。補正なしは、エクセル上の計算式を使用する。
            # なお、補正ありも具材補正倍率を使用したエクセル上の計算式を使用する形式でOKとなった(2025.06.18))
            sum_function = f'SUM({item_needle_cell},{item_saved_cell},{tmp_item_1p_cell},{item_soft_cell},{item_jelly_cell},{item_mixer_cell})'
            ws.cell(row + i_index, base_column + 10,
                        f'=IF({item_per_cell}, {item_per_cell}*合数一覧!{item_needle_cell}*3+{sum_function}, 0)')
            logger.info(f'set_function(row={row + i_index}, column={base_column+10}')
            """

            ws.row_dimensions[row + i_index].hidden = False
            write_line += 1

        return write_line

    def write_items_rate_list(self, ws_out, title, ws_gosu):
        # タイトルの出力
        ws_out['W1'].value = title

        row = 7
        column = 23

        # 本体の材料の出力
        oil_item_quantity = 0.0
        oil_name = None
        ws_out.cell(row - 1, column, self.agg_mix_rice.get_mix_rice_name())
        if self.agg_mix_rice.mix_rice.is_write_rate:
            total_quantity = sum([x[1] for x in self.agg_mix_rice.items])
            for item_name, quantity in self.agg_mix_rice.items:
                ws_out.cell(row, column, item_name)
                ws_out.cell(row, column + 1, quantity)
                plate_1p_cell = ws_out.cell(self.LIST_1_PACK_QUANTITY_ROW, 9).coordinate
                ws_out.cell(row, column + 5, f"=X{row}*合数一覧!{plate_1p_cell}")
                if self._is_oil_item(item_name):
                    oil_item_quantity = quantity
                    oil_name = item_name
                row += 1

        # パーツの出力(比率なし)
        parts_row = 25
        item_index = 0
        if self.agg_mix_rice.mix_rice.is_mix_package:
            # 具(本体)と汁の数量を出す
            filling_name = f'{self.agg_mix_rice.mix_rice.name}の具'
            soup_name = f'{self.agg_mix_rice.mix_rice.name}の液'

            filling_total_cell = ws_out.cell(self.LIST_ORDER_ROW_MAX + 1, 3).coordinate
            soup_total_cell = ws_out.cell(self.LIST_ORDER_ROW_MAX + 1, 3).coordinate
            ws_out.cell(parts_row, column, filling_name)
            # ws_out.cell(parts_row, column + 9, f'=合数一覧!{filling_total_cell}')
            ws_out.cell(parts_row + 1, column, soup_name)
            # ws_out.cell(parts_row + 1, column + 9, f'=合数一覧!{soup_total_cell}')

            # 具の分量の設定
            filling_quantity_cell = ws_out.cell(2, 4).coordinate
            # -具の1人当たりの分量
            ws_out.cell(parts_row, column + 1, f'=合数一覧!{filling_quantity_cell}/3')
            filling_per_cell = ws_out.cell(parts_row, column + 1).coordinate
            # -1袋分量最大数
            ws_out.cell(parts_row, column + 12, f'=MAX(計量表!C10:C91)')
            filling_max_cell = ws_out.cell(parts_row, column + 12).coordinate
            # -1袋分量最大施設の合数
            ws_out.cell(parts_row, column + 13, f'=INDEX(計量表!B10:B91, MATCH(材料配分!{filling_max_cell}, 計量表!C10:C91, 0))')
            filling_max_gosu_cell = ws_out.cell(parts_row, column + 13).coordinate
            # -1袋分量最大施設の袋数
            ws_out.cell(parts_row, column + 14, f'=INDEX(計量表!I10:I91, MATCH(材料配分!{filling_max_cell}, 計量表!C10:C91, 0))')
            filling_max_pack_cell = ws_out.cell(parts_row, column + 14).coordinate
            # -針刺し用1袋の食数
            ws_out.cell(parts_row, column + 15, f'={filling_max_gosu_cell}*3/{filling_max_pack_cell}')
            filling_needle_cell = ws_out.cell(parts_row, column + 15).coordinate
            filling_needle_pack_cell = ws_out.cell(self.LIST_ORDER_ROW_MAX + 2, 9).coordinate
            # -針刺し分量
            ws_out.cell(parts_row, column + 2, f'=IF({filling_per_cell}, {filling_per_cell}*{filling_needle_cell}*合数一覧!{filling_needle_pack_cell}, 0)')
            # -1人用袋
            filling_1p_cell = ws_out.cell(self.LIST_1_PACK_QUANTITY_ROW, 9).coordinate
            ws_out.cell(parts_row, column + 5, f'={filling_per_cell}*合数一覧!{filling_1p_cell}')

            # 汁の分量の設定
            soup_quantity = float(ws_gosu.cell(2, 7).value)/3 - oil_item_quantity
            # -汁の1人当たりの分量
            ws_out.cell(parts_row + 1, column + 1, soup_quantity)
            soup_per_cell = ws_out.cell(parts_row + 1, column + 1).coordinate
            # -1袋分量最大数
            ws_out.cell(parts_row + 1, column + 12, f'=MAX(計量表!L10:L91)')
            soup_max_cell = ws_out.cell(parts_row + 1, column + 12).coordinate
            # -1袋分量最大施設の合数
            ws_out.cell(parts_row + 1, column + 13, f'=INDEX(計量表!K10:K91, MATCH(材料配分!{soup_max_cell}, 計量表!L10:L91, 0))')
            soup_max_gosu_cell = ws_out.cell(parts_row + 1, column + 13).coordinate
            # -1袋分量最大施設の袋数
            ws_out.cell(parts_row + 1, column + 14, f'=INDEX(計量表!P10:P91, MATCH(材料配分!{soup_max_cell}, 計量表!L10:l91, 0))')
            soup_max_pack_cell = ws_out.cell(parts_row + 1, column + 14).coordinate
            # -針刺し用1袋の食数
            ws_out.cell(parts_row + 1, column + 15, f'={soup_max_gosu_cell}*3/{soup_max_pack_cell}')
            soup_needle_cell = ws_out.cell(parts_row + 1, column + 15).coordinate
            soup_needle_pack_cell = ws_out.cell(self.LIST_ORDER_ROW_MAX + 2, 9).coordinate
            # -針刺し分量
            ws_out.cell(parts_row + 1, column + 2, f'=IF({soup_per_cell}, {soup_per_cell}*{soup_needle_cell}*合数一覧!{soup_needle_pack_cell}, 0)')
            # -1人用袋
            soup_1p_cell = ws_out.cell(self.LIST_1_PACK_QUANTITY_ROW, 9).coordinate
            ws_out.cell(parts_row + 1, column + 5, f'={soup_per_cell}*合数一覧!{soup_1p_cell}')

            if oil_item_quantity > 0:
                ws_out.cell(parts_row + 2, column, oil_name)

                # 油の分量の設定
                # -汁の1人当たりの分量
                ws_out.cell(parts_row + 2, column + 1, oil_item_quantity)
                oil_per_cell = ws_out.cell(parts_row + 2, column + 1).coordinate
                """
                # -1袋分量最大数
                ws_out.cell(parts_row + 2, column + 12, f'=MAX(計量表!S10:S91)')
                oil_max_cell = ws_out.cell(parts_row + 2, column + 12).coordinate
                # -1袋分量最大施設の合数
                ws_out.cell(parts_row + 2, column + 13,
                            f'=INDEX(計量表!R10:R91, MATCH(材料配分!{oil_max_cell}, 計量表!S10:S91, 0))')
                oil_max_gosu_cell = ws_out.cell(parts_row + 2, column + 13).coordinate
                # -1袋分量最大施設の袋数
                ws_out.cell(parts_row + 2, column + 14,
                            f'=INDEX(計量表!W10:W91, MATCH(材料配分!{oil_max_cell}, 計量表!S10:S91, 0))')
                oil_max_pack_cell = ws_out.cell(parts_row + 2, column + 14).coordinate
                # -針刺し用1袋の食数
                ws_out.cell(parts_row + 2, column + 15, f'={oil_max_gosu_cell}*3/{oil_max_pack_cell}')
                oil_needle_cell = ws_out.cell(parts_row + 2, column + 14).coordinate
                """
                # -1人分分量(表示セル)
                ws_out.cell(parts_row + 2, column + 2,
                            f'=IF({oil_per_cell}, {oil_per_cell}*{soup_needle_cell}*合数一覧!{soup_needle_pack_cell}, 0)')
                oil_needle_cell = ws_out.cell(parts_row + 2, column + 2).coordinate
                ws_out.row_dimensions[parts_row + 2].hidden = False

                # -1人用袋
                soup_1p_cell = ws_out.cell(self.LIST_1_PACK_QUANTITY_ROW, 9).coordinate
                ws_out.cell(parts_row + 2, column + 5, f'={oil_per_cell}*合数一覧!{soup_1p_cell}')

                oil_saved_cell = ws_out.cell(parts_row + 2, column + 4).coordinate
                oil_1p_cell = ws_out.cell(parts_row + 2, column + 5).coordinate
                oil_soft_cell = ws_out.cell(parts_row + 2, column + 6).coordinate
                oil_jelly_cell = ws_out.cell(parts_row + 2, column + 7).coordinate
                oil_mixer_cell = ws_out.cell(parts_row + 2, column + 8).coordinate
                sum_function = f'SUM({oil_needle_cell},{oil_saved_cell},{oil_1p_cell},{oil_soft_cell},{oil_jelly_cell},{oil_mixer_cell})'
                ws_out.cell(parts_row + 2, column + 9,
                            f'=IF({oil_per_cell}, {oil_per_cell}*合数一覧!{soup_total_cell}*3+{sum_function}, 0)')

                parts_row += 3
                # 合数合計列
                parts_col = 11
                item_index += 1
            else:
                parts_row += 2
                parts_col = 11
                item_index += 1
        else:
            parts_col = 8

        for parts in self.agg_mix_rice.parts:
            write_line = self._write_parts_item_rate(ws_out, parts, parts_row, column, parts_col, item_index)
            parts_row += write_line
            parts_col += 3
            item_index += 1

    def _get_package_size(self, name: str, unit: str) -> int:
        qs = MixRicePackageMaster.objects.all()

        for x in qs:
            if x.parts_name in name:
                return x.package_size

        if unit == 'g':
            return self.DEFAULT_QUANTITY_PER_PACKAGE
        else:
            # 個・枚などのデフォルト
            return 10

    def _is_dry_parts(self, name: str):
        if "錦糸卵" in name:
            return True
        else:
            return False

    def _is_seasoning_parts(self, name: str):
        if "酢" in name:
            return True
        elif re.findall('ゆず(\d|\d.\d)g', name):
            return True
        else:
            return False

    def _is_extra_soup_parts(self, name: str):
        """
        名称そのものは汁・液・出汁ではないが、出汁扱いにするもの
        """
        if "チキンライスの素" in name:
            return True
        else:
            return False

    def _is_order_pattern_parts(self, name: str):
        if "栗ご飯の栗" in name:
            return True
        else:
            return False

    def _get_quantity(self):
        return float(self.agg_mix_rice.quantity)

    def _is_quantity_for_gou(self):
        if ('１合' in self.agg_mix_rice.name) or ('1合' in self.agg_mix_rice.name) or ('一合' in self.agg_mix_rice.name):
            if self.agg_mix_rice.name.find('合') < self.agg_mix_rice.name.find('g'):
                return True
            else:
                return False
        else:
            return False

    def _is_quantity_for_parts_gou(self, parts):
        if ('１合' in parts.name) or ('1合' in parts.name) or ('一合' in parts.name):
            if parts.name.find('合') < parts.name.find('g'):
                return True
            else:
                return False
        else:
            return False

    def _get_quantity_1_gou(self):
        if self._is_quantity_for_gou():
            return self._get_quantity()
        else:
            return self._get_quantity() * settings.MIX_RICE_RATE_GOU

    def _get_with_soup_quantity(self):
        if self.agg_mix_rice.percentage:
            soup_q = float(self.agg_mix_rice.quantity) * float(self.agg_mix_rice.percentage) / 100.0
        elif self.agg_mix_rice.base_soup:
            soup_q = float(self.agg_mix_rice.base_soup)
        else:
            #if self.agg_mix_rice.mix_rice.has_base:
            #    soup_q = float(self.agg_mix_rice.quantity) * 9.0 / 100.0
            #else:
            soup_q = 0.0
        return float(self.agg_mix_rice.quantity) + soup_q

    def _get_with_soup_quantity_1_gou(self):
        quantity = self._get_with_soup_quantity()
        if self._is_quantity_for_gou():
            return quantity
        else:
            return quantity * settings.MIX_RICE_RATE_GOU

    def _get_soup_quantity_measure_sheet(self):
        # %指定の出汁を別献立としては出力しない。
        if self.agg_mix_rice.base_soup:
            return float(self.agg_mix_rice.base_soup)
        else:
            if self.agg_mix_rice.mix_rice.is_mix_package:
                return float(self.agg_mix_rice.parts[0].quantity)
            return 0.0

    def _get_soup_quantity(self):
        if self.agg_mix_rice.percentage:
            return float(self.agg_mix_rice.quantity) * float(self.agg_mix_rice.percentage) / 100.0
        elif self.agg_mix_rice.base_soup:
            return float(self.agg_mix_rice.base_soup)
        else:
            if self.agg_mix_rice.mix_rice.is_mix_package:
                return float(self.agg_mix_rice.parts[0].quantity)
            return 0.0

    def _get_soup_quantity_1_gou(self):
        if self._is_quantity_for_gou():
            return self._get_soup_quantity()
        else:
            return self._get_soup_quantity() * settings.MIX_RICE_RATE_GOU

    def _get_parts_quantity_1_gou(self, parts):
        if self._is_quantity_for_parts_gou(parts):
            return float(parts.quantity)
        else:
            return float(parts.quantity) * settings.MIX_RICE_RATE_GOU

    def _get_parts_quantity_order(self, parts):
        return float(parts.quantity)

    def _get_display_rice_name(self, name: str):
        return f'{name}の具'

    def _get_display_parts_name(self, name: str):
        return name

    def _get_title(self):
        sp = self.agg_mix_rice.eating_day.split("-")
        return f'{sp[0]}年{sp[1]}月{sp[2]}日({self.agg_mix_rice.meal})'

    def write_measure_sheet(self, ws, df):
        # 出汁の非表示(液の%指定が無い場合は、「出汁」(%計算列のあるもの）は出力しない)
        if self.agg_mix_rice.percentage:
            ws.cell(1, 5, float(self.agg_mix_rice.percentage) / 100.0)
        else:
            ws.column_dimensions['E'].hidden = True
            ws.column_dimensions['F'].hidden = True
            ws.cell(1, 5, 0)

            # 具(メイン)料理の合計欄、袋数を非表示にする
            ws.column_dimensions['G'].hidden = True
            ws.column_dimensions['H'].hidden = True

        # タイトル
        sheet_header_title = f'混ぜご飯計量表:{self._get_title()}'
        ws.cell(1, 1, sheet_header_title)
        PAGE_MAX = 1
        page_count = 0

        # 混ぜご飯本体の出力
        ws.cell(2, 1, self._get_display_rice_name(self.agg_mix_rice.get_mix_rice_name()))
        ws.cell(2, 3, self.agg_mix_rice.quantity)
        ws.cell(2, 4, self.agg_mix_rice.unit)
        if self._is_quantity_for_gou():
            ws.cell(4, 1, '1合')
            ws.cell(5, 1, '10合')
            ws.cell(4, 7, '=(C4+E4)')
            ws.cell(5, 7, '=(C5+E5)')
        page_count += 1

        # 単位がgでない場合は、非表示にする
        if self.agg_mix_rice.unit != 'g':
            ws.column_dimensions['A'].hidden = True
            ws.column_dimensions['B'].hidden = True
            ws.column_dimensions['C'].hidden = True
            ws.column_dimensions['D'].hidden = True
            ws.column_dimensions['E'].hidden = True
            ws.column_dimensions['F'].hidden = True
            ws.column_dimensions['G'].hidden = True
            ws.column_dimensions['H'].hidden = True
            ws.column_dimensions['I'].hidden = True
            ws.cell(1, 10, sheet_header_title)
            page_count -= 1

        # 混ぜご飯の各パーツを出力
        col = 10

        # 施設毎の情報を出力する起点となる列(列番号と注文数出力かどうかのフラグのタプル)
        col_list = []

        # 汁・液
        soup_quantity = self._get_soup_quantity_measure_sheet()

        if soup_quantity:
            if self.agg_mix_rice.mix_rice.is_mix_package:
                ws.cell(2, col, f'{self.agg_mix_rice.mix_rice.name}の液')
            else:
                ws.cell(2, col, f'{self.agg_mix_rice.mix_rice.name}出汁')
            ws.cell(2, col + 2, f'{soup_quantity}')

            col += 7

        if self.agg_mix_rice.mix_rice.is_mix_package:
            ws.cell(1, 10, sheet_header_title)
            col_list.append((col, False))

            # 具(メイン)料理の袋数の内容を揃える
            for i in range(self.MEASURE_MAX_UNIT_ROW - 4):
                ws.cell(4 + i, 9, f'=p{4 + i}')

            # 汁の料理の合計欄表示
            ws.column_dimensions['N'].hidden = False
            ws.column_dimensions['O'].hidden = False

            # 合計欄の計算式設定
            # -1人、10人
            ws.cell(4, 14, '=(C4+L4)')
            ws.cell(5, 14, '=(C5+L5)')

            # -合計
            mix_row = 6
            for cell in ws[f'N6:N{self.MEASURE_MAX_UNIT_ROW}']:
                cell[0].value = f'=C{mix_row}+L{mix_row}'

                # 本体側の袋数の参照を修正
                ws.cell(mix_row, 9).value = f'=P{mix_row}'
                mix_row += 1

            page_count += 1

        # その他
        dry_parts_col = []
        for parts in self.agg_mix_rice.parts:
            if page_count == PAGE_MAX:
                page_count = 0
                ws.cell(1, col, sheet_header_title)

            ws.cell(2, col, self._get_display_parts_name(parts.name))
            ws.cell(2, col + 2, parts.quantity)
            ws.cell(2, col + 3, parts.unit)

            if self._is_quantity_for_parts_gou(parts):
                ws.cell(4, col, '1合')
                ws.cell(5, col, '10合')

            if self._is_order_pattern_parts(parts.name):
                # 注文数出力の場合
                ws.cell(3, col + 1, '注文数')
                col_list.append((col, True))
            else:
                col_list.append((col, False))
            if self._is_dry_parts(parts.name):
                dry_parts_col.append(col)

            # 単位がgでない場合は非表示
            if parts.unit != 'g':
                for i in range(7):
                    hidden_col = ws.cell(2, col + i).coordinate[:-1]
                    ws.column_dimensions[hidden_col].hidden = True

            col += 7
            page_count += 1

        # 各施設の合数を出力
        i = 10
        for index, row in df.iterrows():
            ws.row_dimensions[i].hidden = False
            ws.cell(i, 1, f'{row["unit_name__unit_number"]}.{row["unit_name__calc_name"]}')
            ws.cell(i, 2, f'{row["quantity"]}')

            for c, is_order_pattrn in col_list:
                ws.cell(i, c, f'{row["unit_name__unit_number"]}.{row["unit_name__calc_name"]}')
                if is_order_pattrn:
                    # 注文数出力の場合
                    ws.cell(i, c + 1, f'{row["order_quantity"]}')
                else:
                    ws.cell(i, c + 1, f'{row["quantity"]}')

            # 乾燥・冷凍区分が冷凍のユニットは、乾燥専用品で空欄を表示する
            if row["unit_name__username__dry_cold_type"] == '乾燥':
                pass
            else:
                for dpc in dry_parts_col:
                    ws.cell(i, dpc + 1, '')
            i += 1

        # 余分な列を非表示
        max_col_index = 'AS' # テンプレートに表示のある列の次を指定する
        current_col = ws.cell(1, col).coordinate[:-1]
        hidden_cnt = 0
        while current_col != max_col_index:
            ws.column_dimensions[current_col].hidden = True
            col += 1
            current_col = ws.cell(1, col).coordinate[:-1]
            hidden_cnt += 1
        hidden_page_cnt = int(hidden_cnt / 7)
        if hidden_page_cnt > 0:
            new_brk = ws.col_breaks.brk[:hidden_page_cnt * -1]
            ws.col_breaks.brk = new_brk
