import datetime
from itertools import groupby
import logging
import os
import re
import shutil
import traceback
from decimal import Decimal

import numpy as np
import openpyxl as excel
import pandas as pd

from django.conf import settings
from django.core.management import call_command
from django.core.management.base import BaseCommand

from web_order.models import AggMeasureSoupMaster, AggMeasureMixRiceMaster, PlatePackageForPrint, UnitPackage, TmpPlateNamePackage
from .agg_measure_analyzed import AggMeasureTarget, AggMeasurePlate, AggMeasurePlateWithDensity, AggMeasurePlateKoGram
from .agg_measure_analyzed import AggMeasurePlateKoGramDensity, AggMeasurePlateGramGram, AggMeasurePlateWithAnotherUnit
from .agg_measure_analyzed import AggMeasureMisoDevide, AggMeasureSoupDevide, AggMeasureMiso, AggMeasureSoupFilling
from .agg_measure_analyzed import AggMeasureSoupLiquid, AggMeasureMisoNone, AggMeasureSoupNone, AggMeasureLiquidSeasoning
from .agg_measure_analyzed import AggMeasureMixRice, AggMeasureMixRiceParts, AggMeasureOrdersManager, AggMeasurePlateKoGramPercent
from .agg_mix_rice import MixRiceMeasureWriter
from web_order.cooking_direction_plates import CookingDirectionPlatesManager, PlateNameAnalizeUtil


logger = logging.getLogger(__name__)


class MisoSoupAggregation:
    """
    味噌汁の合計量を管理するクラス
    """
    def __init__(self):
        self.j_total = 0
        self.u_total = 0

        # 針刺し用固定数量(30人前2袋)
        self.j_needles = 20 * 3
        self.u_needles = 20 * 3

        # 保存用固定数量(常食:10人前3袋、薄味:10人前2袋)
        self.j_preserves = 10 * 3
        self.u_preserves = 10 * 2

        self.j_items = None
        self.u_items = None

        # 追加済みの喫食日の朝/昼/晩(同一の喫食日を追加しないようにする)
        self.added_eating_day_menus = []

        self.item_quantity_list = None

    def __str__(self):
        base = {
            '常食数量': self.j_total,
            '薄味数量': self.u_total,
            '常食針刺': self.j_needles,
            '薄味針刺': self.u_needles,
            '常食保存用': self.j_preserves,
            '薄味保存用': self.u_preserves,
            '常食構成量': self.j_items,
            '薄味構成量': self.u_items
        }
        return str(base)

    def add_soup_quantity(self, quantity_message: str, eating_day: str, menu_name: str):
        """
        味噌汁の計量表の集計結果を登録する
        """
        eating_day_menu = eating_day + menu_name
        if not (eating_day_menu in self.added_eating_day_menus):
            self.added_eating_day_menus.append(eating_day_menu)

            # quantity_messageのフォーマット「常食の食材構成(※1):薄味の食材構成(※1):常食の数量:薄味の数量」
            # ※1:食材の各食材の1人分の数量を/区切りで設定
            quantities = quantity_message.split(':')

            # 食材の構成量は計量表が変わっても同じなので、一回だけ設定
            # 別途調理表の内容を設定することになったため、コメントアウト
            """
            if self.j_items is None:
                self.j_items = [Decimal(x) for x in quantities[0].split('/')]
            if self.u_items is None:
                self.u_items = [Decimal(x) for x in quantities[1].split('/')]
            """

            # 数量の加算
            logger.info(f'味噌汁加算({eating_day}):{quantities}')
            self.j_total += int(quantities[2])
            self.u_total += int(quantities[3])

    def get_j_needle_gram_quantity_function(self) -> str:
        """
        針刺し用のg数量(常食用)の取得
        """
        items = [str(value) for name, value in self.item_quantity_list]

        return "=({0})*{1}".format('+'.join(items), self.j_needles)

    def get_u_needle_gram_quantity_function(self) -> str:
        """
        針刺し用のg数量(薄味用)の取得
        """
        items = [str(x) for x in self.u_items]

        return "=({0})*{1}".format('+'.join(items), self.u_needles)

    def get_j_item_gram_quantity_function(self, value) -> str:
        """
        食材のg数量(常食用)の取得
        """
        # 薄味=0を薄味統合後と判断
        if self.u_total == 0:
            self.j_preserves = 10 * 5

        # 本来の値の1.1倍で出力することになった
        return "=({0})*({1}+{2}+{3})*1.1".format(value, self.j_total, self.j_needles, self.j_preserves)

    def get_u_item_gram_quantity_function(self, index: int) -> str:
        """
        食材のg数量(薄味用)の取得
        """
        # 本来の値の1.1倍で出力することになった
        return "=({0})*({1}+{2}+{3})*1.1".format(self.u_items[index], self.u_total, self.u_needles, self.u_preserves)


class PrintPlatePackageManager:
    """
    計量表出力中の献立袋数情報を管理するクラス
    """
    def __init__(self, package_list):
        self.raw_list = package_list
        self.regex_pattern = re.compile('■(\d+)/(\d+)(\D+)')
        self.adjust_sause_mix = 0
        self.adjust_timing = None

    def _parse_eating_meal(self, eating_meal: str, cooking_date):
        """
        調理表の喫食日・食事区分(A列)を解析し、喫食日と食事区分を取得する。
        喫食日・食事区分例)■3/1朝食
        """
        res_parse = self.regex_pattern.findall(eating_meal)
        if res_parse:
            date = datetime.datetime.strptime(cooking_date, '%Y-%m-%d').date()
            year = date.year
            month = int(res_parse[0][0])
            day = int(res_parse[0][1])
            if date.month == 12 and date.day >= 20 and month == 1:
                year += 1
            meal = res_parse[0][2]
            return datetime.date(year, month, day), meal
        else:
            raise ValueError("喫食日・食事区分変換:喫食食事区分不正")

    def _is_filling_and_sause_mix(self, name):
        if '◆' in name:
            res = re.findall('具(\d+|\d+\.\d+)[g|ｇ]\s*\D液(\d+|\d+\.\d+)[g|ｇ]', name)
            if res and (res[0][0] and res[0][1]):
                return True

        if 'カレーライス' in name:
            return True
        if 'シチュー' in name:
            return True

        return False

    def _save_print_for_package(self, name, eating_day, cooking_day, meal, is_basic_plate, index):
        # 常食
        package = PlatePackageForPrint(
            plate_name=name,
            eating_day=eating_day,
            cooking_day=cooking_day,
            meal_name=meal,
            menu_name='常食',
            is_basic_plate=is_basic_plate,
            index=index
        )
        package.save()

        if self.adjust_timing:
            ad_eating_day, ad_meal = self.adjust_timing
            if (ad_eating_day != eating_day) or (ad_meal != meal):
                self.adjust_sause_mix = 0
                self.adjust_timing = None
        enge_index = index + self.adjust_sause_mix

        # ソフト
        package = PlatePackageForPrint(
            plate_name=name,
            eating_day=eating_day,
            cooking_day=cooking_day,
            meal_name=meal,
            menu_name='ソフト',
            is_basic_plate=is_basic_plate,
            index=enge_index
        )
        package.save()

        # ゼリー
        package = PlatePackageForPrint(
            plate_name=name,
            eating_day=eating_day,
            cooking_day=cooking_day,
            meal_name=meal,
            menu_name='ゼリー',
            is_basic_plate=is_basic_plate,
            index=enge_index
        )
        package.save()

        # ミキサー
        package = PlatePackageForPrint(
            plate_name=name,
            eating_day=eating_day,
            cooking_day=cooking_day,
            meal_name=meal,
            menu_name='ミキサー',
            is_basic_plate=is_basic_plate,
            index=enge_index
        )
        package.save()

        if self._is_filling_and_sause_mix(name):
            # ソフト
            package = PlatePackageForPrint(
                plate_name=name + 'のルー',
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ソフト',
                is_basic_plate=is_basic_plate,
                index=enge_index + 1
            )
            package.save()

            # ゼリー
            package = PlatePackageForPrint(
                plate_name=name + 'のルー',
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ゼリー',
                is_basic_plate=is_basic_plate,
                index=enge_index + 1
            )
            package.save()

            # ミキサー
            package = PlatePackageForPrint(
                plate_name=name + 'のルー',
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ミキサー',
                is_basic_plate=is_basic_plate,
                index=enge_index + 1
            )
            package.save()

            self.adjust_timing = (eating_day, meal)
            self.adjust_sause_mix += 1

    def get_kind_menu_name(self, code):
        """
        食種の献立種類を取得する
        """
        if '常' in code:
            return '常食'
        elif 'ソ' in code:
            return 'ソフト'
        elif 'ゼ' in code:
            return 'ゼリー'
        elif 'ミ' in code:
            return 'ミキサー'
        elif '木沢個' in code:
            return '常食'
        else:
            return ''

    def _save_print_for_package_allergen(self, name, eating_day, cooking_day, meal, is_basic_plate, index_dict, eating_type_list):
        """
        通常食(アレルギーでない料理)用の登録処理
        """
        converted_list = [self.get_kind_menu_name(x) for x in eating_type_list]

        # 常食
        if '常食' in converted_list:
            logger.info(f'PlatePackageForPrint常食アレルギー登録:{eating_day}-{meal}-{name}-{index_dict["常食"]}')
            package = PlatePackageForPrint(
                plate_name=name,
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='常食',
                is_basic_plate=is_basic_plate,
                index=index_dict['常食']
            )
            package.save()

            index_dict['常食'] += 1

        # ソフト
        if 'ソフト' in converted_list:
            package = PlatePackageForPrint(
                plate_name=name,
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ソフト',
                is_basic_plate=is_basic_plate,
                index=index_dict['ソフト']
            )
            package.save()
            index_dict['ソフト'] += 1

            if self._is_filling_and_sause_mix(name):
                # ソフト
                package = PlatePackageForPrint(
                    plate_name=name + 'のルー',
                    eating_day=eating_day,
                    cooking_day=cooking_day,
                    meal_name=meal,
                    menu_name='ソフト',
                    is_basic_plate=is_basic_plate,
                    index=index_dict['ソフト']
                )
                package.save()
                index_dict['ソフト'] += 1

        # ゼリー
        if 'ゼリー' in converted_list:
            package = PlatePackageForPrint(
                plate_name=name,
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ゼリー',
                is_basic_plate=is_basic_plate,
                index=index_dict['ゼリー']
            )
            package.save()
            index_dict['ゼリー'] += 1

            if self._is_filling_and_sause_mix(name):
                # ゼリー
                package = PlatePackageForPrint(
                    plate_name=name + 'のルー',
                    eating_day=eating_day,
                    cooking_day=cooking_day,
                    meal_name=meal,
                    menu_name='ゼリー',
                    is_basic_plate=is_basic_plate,
                    index=index_dict['ゼリー']
                )
                package.save()
                index_dict['ゼリー'] += 1

        # ミキサー
        if 'ミキサー' in converted_list:
            package = PlatePackageForPrint(
                plate_name=name,
                eating_day=eating_day,
                cooking_day=cooking_day,
                meal_name=meal,
                menu_name='ミキサー',
                is_basic_plate=is_basic_plate,
                index=index_dict['ミキサー']
            )
            package.save()
            index_dict['ミキサー'] += 1

            if self._is_filling_and_sause_mix(name):
                # ミキサー
                package = PlatePackageForPrint(
                    plate_name=name + 'のルー',
                    eating_day=eating_day,
                    cooking_day=cooking_day,
                    meal_name=meal,
                    menu_name='ミキサー',
                    is_basic_plate=is_basic_plate,
                    index=index_dict['ミキサー']
                )
                package.save()
                index_dict['ミキサー'] += 1

    def save_new(self, cooking_date):
        """
        新規データとして、調理表から解析した内容で袋数情報を登録する。
        """
        # 対象製造日の全データを削除
        PlatePackageForPrint.objects.filter(cooking_day=cooking_date).delete()

        prev_key = None
        basic_plate_index = 0   # 喫食日・食事区分の単位で採番
        allergen_plate_index_dict = {'常食': 0, 'ソフト': 0, 'ミキサー': 0, 'ゼリー': 0}    # 調理日全体で1つのインデックス
        for index, package_dict in enumerate(self.raw_list):
            if (not package_dict['plate']) or package_dict['plate'] == '':
                continue

            if prev_key != package_dict['eating_meal']:
                prev_key = package_dict['eating_meal']
                basic_plate_index = 0
                allergen_plate_index_dict = {'常食': 0, 'ソフト': 0, 'ミキサー': 0, 'ゼリー': 0}
            eating_day, meal = self._parse_eating_meal(package_dict['eating_meal'], cooking_date)
            meal_name = meal.strip()
            if package_dict['is_basic_plate']:
                # 基本食の場合
                self._save_print_for_package(
                    name=package_dict['plate'],
                    eating_day=eating_day,
                    cooking_day=cooking_date,
                    meal=meal_name,
                    is_basic_plate=package_dict['is_basic_plate'],
                    index=basic_plate_index
                )

                basic_plate_index += 1
            else:
                # アレルギー食の場合
                self._save_print_for_package_allergen(
                    name=package_dict['plate'],
                    eating_day=eating_day,
                    cooking_day=cooking_date,
                    meal=meal_name,
                    is_basic_plate=package_dict['is_basic_plate'],
                    index_dict=allergen_plate_index_dict,
                    eating_type_list=package_dict['eating_type_list']
                )


class AggMeasureTargetAnalyzer:
    """
    計量対象を判定するクラス
    """
    def __init__(self, cooking_day):
        self.logger = logging.getLogger(__name__)

        self.cooking_day = cooking_day
        self.analyzed_dict = {}
        self.is_first_miso_soup = True
        self.soup_group_list = [x for x in AggMeasureSoupMaster.objects.filter(soup_group='スープ')]
        self.other_liquid_group_list = [x for x in AggMeasureSoupMaster.objects.filter(soup_group='汁')]

        # 混ぜご飯判定制御用
        self.finding_mix_rice = None
        self.mix_rice_list = [x for x in AggMeasureMixRiceMaster.objects.all()]

        # 汁・スープの具の単位が個のみの場合の判定
        self.regex_soup_ko = re.compile('(\D*)(\d+|\d+\.\d+)個')

        # '③サラダ（枝豆・豆腐）1/2丁'
        # ※/は本処理内で「÷」に置き換えられる
        self.regex_cho = re.compile('(\d+|\d+÷\d+)丁')

        self.regex_density_g = re.compile('(\+|＋)\s*(液同|液|汁|汁【しる】|)(\d+|\d+.\d+)g')
        self.regex_density_p = re.compile('(\+|＋)\s*(液同|液|)(\d+|\d+.\d+)％')

        self.regex_units = {}
        for unit in settings.MEASURE_ENABLE_UNITS:
            self.regex_units[unit] = [
                re.compile(f'(\d+){unit}'),
                re.compile(f'{unit}\s*(\d+|\d+.\d+)g\s*(\+|＋)')
            ]

        # 〇個 ●g + ▲%
        # のg数を取得する正規表現
        self.regex_inner_quantity_ko = re.compile('個\s*(\d*|\d*.\d*)g\s*(\+|＋)')
        self.regex_inner_quantity_cho = re.compile('丁\s*(\d*|\d*.\d*)g\s*(\+|＋)')

    def _is_miso_soup(self, name: str):
        if 'みそ汁' in name:
            return True
        elif '味噌汁' in name:
            return True
        elif 'みそしる' in name:
            return True
        else:
            return False

    def _get_soup(self, name: str):
        if '具' in name:
            # 「具」が入っていれば汁具とみなす
            return None

        if ('スープ' in name) or ('汁' in name):
            return AggMeasureSoupMaster(name=name, search_word='')
        else:
            # 「スープ」「汁」がつかないが、スープ・汁扱いになるもの(現在はお吸い物のみ)
            for shiru in self.other_liquid_group_list:
                if shiru.search_word in name:
                    # 名前の形式をスープと合わせる
                    shiru.name = name
                    return shiru

        return None

    def _get_percentage(self, name: str):
        res_unit_percentage = re.findall('(\d*|\d*\.\d*)％', name)
        if res_unit_percentage:
            # 2個以上の想定例：
            # ①かぼちゃの煮物2個+(赤キャップ4.5％＋水10.5％)で15％
            if len(res_unit_percentage) >= 3:
                return float(res_unit_percentage[-1])
            else:
                return float(res_unit_percentage[0])
        else:
            return 0

    def add_cook(self, index, name, eating_day, meal, items, before_name):
        eating_day_meal = eating_day + meal
        eating_day_meal_analyzed = self.analyzed_dict.get(eating_day_meal, None)
        analyzed = self.analyze(index, name, eating_day, meal, items)
        analyzed.before_name = before_name
        if eating_day_meal_analyzed:
            eating_day_meal_analyzed.append(analyzed)
        else:
            self.analyzed_dict[eating_day_meal] = [analyzed]

        if type(analyzed) in [AggMeasureMisoDevide, AggMeasureMiso]:
            self.is_first_miso_soup = False

    def is_same_thickness(self, name):
        # 「液同」は濃さの指定以外に使われない前提
        return '液同' in name

    def generate_analyzed(self):
        # 汁具と汁の関連の設定
        for key in self.analyzed_dict.keys():
            # 汁の情報を抽出
            items = self.analyzed_dict[key]
            for index, item in enumerate(items):
                if type(item) is AggMeasureSoupFilling:
                    # 以降の要素から、スープを検出
                    search_targets = items[index:]
                    for target in search_targets:
                        if type(target) is AggMeasureSoupLiquid:
                            target.add_filling(item)
                            item.liquid = target
                            break
                if type(item) is AggMeasureSoupDevide:
                    # 以降の要素から、スープを検出
                    search_targets = items[index:]
                    for target in search_targets:
                        if type(target) is AggMeasureSoupLiquid:
                            target.add_filling(item)
                            item.liquid = target
                            break
                elif type(item) is AggMeasureMixRiceParts:
                    # 以前の要素から、混ぜご飯本体を検出
                    search_targets = items[:index]
                    for target in reversed(search_targets):
                        if type(target) is AggMeasureMixRice:
                            target.add_parts(item)
                            break
                elif type(item) is AggMeasureMixRice:
                    # 以前の要素から、別の混ぜご飯本体を検出
                    search_targets = items[:index]
                    for target in reversed(search_targets):
                        if type(target) is AggMeasureMixRice:
                            if target.mix_rice.name == item.mix_rice.name:
                                # 計量表出力しないように無効化
                                item.is_disable = True
                                break


        for values in self.analyzed_dict.values():
            for value in values:
                yield value

    def get_inner_density(self, percentage, inner_value, quantity):
        percentage = float(percentage) / 100.0
        src = float(inner_value) * percentage
        density = src / float(quantity) * 100.0
        return density

    def analyze(self, index, name, eating_day, meal, items):
        # -------------------------------------------------------------
        # 味噌汁、スープなどの計量表
        # -------------------------------------------------------------
        res = re.findall('⑤', name)  # 先頭に⑤があるとき

        if res:
            self.finding_mix_rice = None

            # 先頭の丸数字を除外
            numberless_name = name[1:]

            # ---------------------------------------------------------
            # ⑤味噌汁（里芋2個・さつま揚げ5g）
            # ---------------------------------------------------------
            res_ko_gram = re.findall('(\D*)(\d+)個(\D*)(\d+|\d+\.\d+)g', numberless_name)

            if res_ko_gram:
                if self._is_miso_soup(numberless_name):
                    return AggMeasureMisoDevide(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        res_ko_gram[0][0],
                        res_ko_gram[0][1],
                        '個',
                        res_ko_gram[0][2],
                        res_ko_gram[0][3],
                        'g',
                        self.is_first_miso_soup
                    )
                else:
                    return AggMeasureSoupDevide(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        res_ko_gram[0][0],
                        res_ko_gram[0][1],
                        '個',
                        res_ko_gram[0][2],
                        res_ko_gram[0][3],
                        'g'
                    )

            # ---------------------------------------------------------
            # ⑤味噌汁具（玉葱・しめじ）16g
            # ⑤味噌汁具（あげ5g）
            # ⑤味噌汁具（わかめ0.9g）
            # ⑤みそ汁（菜の花・人参）17g
            # ⑤みそ汁（人参5g）
            # ⑤コンソメ（玉葱・人参）16g
            # ⑤コーンスープ17ｇ 水150g
            # ⑤スープ希釈（コンソメ）29.45ｇ（150ｇ水入れる）
            # ⑤ポタージュスープ　16g　水150g
            # ---------------------------------------------------------
            #　g指定のある献立
            res_gram = re.findall('(\D*)(\d+|\d+\.\d+)g', numberless_name)

            # 個指定のある献立
            res_soup_ko = self.regex_soup_ko.findall(numberless_name)

            if res_gram:
                # 味噌汁
                if self._is_miso_soup(numberless_name):
                    return AggMeasureMiso(index, self.cooking_day, eating_day, meal, numberless_name, res_gram[0][1], 'g', self.is_first_miso_soup)
                # それ以外
                else:
                    soup = self._get_soup(numberless_name)
                    if soup:
                        return AggMeasureSoupLiquid(
                            index, self.cooking_day, eating_day, meal, numberless_name, res_gram[0][1], 'g', soup)
                    else:
                        return AggMeasureSoupFilling(
                            index, self.cooking_day, eating_day, meal, numberless_name, res_gram[0][1], 'g')
            # ---------------------------------------------------------
            # ⑤スープ具（ギョーザ）1個
            # ---------------------------------------------------------
            elif res_soup_ko:
                # 味噌汁
                if self._is_miso_soup(numberless_name):
                    return AggMeasureMiso(index, self.cooking_day, eating_day, meal, numberless_name, res_soup_ko[0][1], '個', self.is_first_miso_soup)
                # それ以外
                else:
                    soup = self._get_soup(numberless_name)
                    if soup:
                        return AggMeasureSoupLiquid(
                            index, self.cooking_day, eating_day, meal, numberless_name, res_soup_ko[0][1], '個', soup)
                    else:
                        return AggMeasureSoupFilling(
                            index, self.cooking_day, eating_day, meal, numberless_name, res_soup_ko[0][1], '個')
            else:
                # ⑤スープの具（コーン）
                # ⑤味噌汁30cc 希釈140
                # ⑤みそ汁30cc 希釈130
                # 味噌汁
                if self._is_miso_soup(numberless_name):
                    return AggMeasureMisoNone(index, self.cooking_day, eating_day, meal, numberless_name, None)
                # それ以外
                else:
                    return AggMeasureSoupNone(index, self.cooking_day, eating_day, meal, numberless_name, None)

        # 先頭の丸数字を除外
        number = name[0]
        numberless_name = name[1:] if number in ['⑩', '①', '②', '③', '④', '⑤'] else name

        # -------------------------------------------------------------
        # 共通パラメータ(出汁の量(g)、%)
        # -------------------------------------------------------------
        res_has_soup_stock_g = self.regex_density_g.findall(numberless_name)
        res_has_soup_stock_p = self.regex_density_p.findall(numberless_name)

        # -------------------------------------------------------------
        # 混ぜご飯判定(本体)
        # -------------------------------------------------------------
        if number == '①':
            for mix_rice_master in self.mix_rice_list:
                if mix_rice_master.search_word in numberless_name:
                    mix_rice_unit = 'g'
                    quantity = 0
                    for key, value in self.regex_units.items():
                        res_unit = value[0].findall(numberless_name)
                        if res_unit:
                            mix_rice_unit = key
                            quantity = res_unit[0]
                            break

                    if mix_rice_unit == 'g':
                        res_mr_ko = re.findall('(\d+)個', numberless_name)
                        if res_mr_ko:
                            mix_rice_unit = '個'
                            quantity = res_mr_ko[0]

                    if mix_rice_unit == 'g':
                        res_mr = re.findall('(\d+|\d+\.\d+)g', numberless_name)
                        if res_mr:
                            quantity = res_mr[0]
                        else:
                            self.logger.info(f'混ぜご飯計量_数量不明({numberless_name})')
                            quantity = 0

                    base_soup = res_has_soup_stock_g[0][2] if res_has_soup_stock_g else None
                    percentage = res_has_soup_stock_p[0][2] if res_has_soup_stock_p else None
                    if self.finding_mix_rice:
                        pass
                        """
                        if mix_rice_master.name == self.finding_mix_rice:
                            # 既に同一名称の混ぜご飯検出中なら、具の一部と判断
                            return AggMeasureMixRiceParts(index, self.cooking_day, eating_day, meal, numberless_name, quantity, mix_rice_unit, number, items)
                        else:
                            # 別名称の混ぜご飯が検出
                            self.finding_mix_rice = mix_rice_master.name
                            return AggMeasureMixRice(
                                index, self.cooking_day, eating_day, meal, numberless_name, quantity, mix_rice_unit, number,
                                mix_rice_master, base_soup, percentage, items)
                        """
                    else:
                        self.finding_mix_rice = mix_rice_master.name
                        return AggMeasureMixRice(
                            index, self.cooking_day, eating_day, meal, numberless_name, quantity, mix_rice_unit, number,
                            mix_rice_master, base_soup, percentage, items)

        # -------------------------------------------------------------
        # 混ぜご飯判定(具・酢など)
        # -------------------------------------------------------------
        if self.finding_mix_rice and (number == '④'):
            # 混ぜご飯検出中の④は無条件で混ぜご飯の一部と判断
            res_mr_parts = re.findall('(\d+|\d+\.\d+)g', numberless_name)
            if res_mr_parts:
                return AggMeasureMixRiceParts(index, self.cooking_day, eating_day, meal, numberless_name, res_mr_parts[0], 'g', number, items)
            else:
                res_mr_ko_parts = re.findall('(\d+|\d+\.\d+)個', numberless_name)
                if res_mr_ko_parts:
                    return AggMeasureMixRiceParts(index, self.cooking_day, eating_day, meal, numberless_name, res_mr_ko_parts[0],
                                                  '個', number, items)
                else:
                    self.logger.info(f'混ぜご飯計量_数量不明({numberless_name})')
                    return AggMeasureMixRiceParts(index, self.cooking_day, eating_day, meal, numberless_name, 0, 'g', number, items)

        self.finding_mix_rice = None

        # -------------------------------------------------------------
        # ④■ポン酢7g
        # -------------------------------------------------------------
        res_small = re.findall('■\D*(\d+|\d+\.\d+)g', numberless_name)

        if res_small:
            # 具なし固定
            return AggMeasureLiquidSeasoning(index, self.cooking_day, eating_day, meal, numberless_name[1:], res_small[0], 'g', number, False)

        # -------------------------------------------------------------
        # ④タルタルソース7g
        # -------------------------------------------------------------
        res_liquid = re.findall('\D*(\d+|\d+\.\d+)g', numberless_name)
        res_has_gu = re.findall('\[(.*)\]', numberless_name)
        sub_liquid = re.sub('\[.*\]', '', numberless_name)
        sub_liquid = re.sub('(\d*|\d*\.\d*)g', '', sub_liquid)
        res_end_liquid = re.findall('ソース$', sub_liquid)

        if res_liquid and res_end_liquid:
            has_gu = True if res_has_gu else False
            return AggMeasureLiquidSeasoning(index, self.cooking_day, eating_day, meal, numberless_name, res_liquid[0], 'g', number, has_gu)

        # -------------------------------------------------------------
        # ①鮭の塩焼き60ｇ1尾
        # ①鰆照り焼き1切れ
        # ④いんげん2本
        # -------------------------------------------------------------
        for key, value in self.regex_units.items():
            res_unit = value[0].findall(numberless_name)
            if res_unit:
                if res_has_soup_stock_g:
                    density = float(res_has_soup_stock_g[0][2]) / int(res_unit[0]) * 100
                    return AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        res_unit[0],
                        key, number,
                        density, res_has_soup_stock_g[0][1] == '液同', True)
                elif res_has_soup_stock_p:
                    inner = value[1].findall(numberless_name)
                    if inner:
                        # 出汁の数に対する割合を計算
                        density = self.get_inner_density(res_has_soup_stock_p[0][2], inner[0][0], res_unit[0])
                        return AggMeasurePlateWithDensity(
                            index, self.cooking_day,
                            eating_day,
                            meal,
                            numberless_name,
                            res_unit[0],
                            key, number,
                            density, res_has_soup_stock_p[0][1] == '液同', True, inner=inner[0][0])
                    else:
                        return AggMeasurePlateWithDensity(
                            index, self.cooking_day,
                            eating_day,
                            meal,
                            numberless_name,
                            res_unit[0],
                            key, number,
                            self._get_percentage(numberless_name), res_has_soup_stock_p[0][1] == '液同')
                else:
                    return AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        res_unit[0],
                        key, number,
                        self._get_percentage(numberless_name), self.is_same_thickness(numberless_name))

        # -------------------------------------------------------------
        # ③サラダ（枝豆・豆腐）1/2丁
        # -------------------------------------------------------------
        res_unit_cho = self.regex_cho.findall(numberless_name)
        if res_unit_cho:
            cho_values = res_unit_cho[0].split('÷')
            if len(cho_values) == 2:
                numerator = float(cho_values[0])
                denominator = float(cho_values[1])
                cho_value_float = numerator / denominator
            else:
                cho_value_float = float(cho_values[0])

            if res_has_soup_stock_g:
                density = float(res_has_soup_stock_g[0][2]) / cho_value_float * 100
                return AggMeasurePlateWithDensity(
                    index, self.cooking_day,
                    eating_day,
                    meal,
                    numberless_name,
                    cho_value_float,
                    '丁', number,
                    density, res_has_soup_stock_g[0][1] == '液同', True)
            elif res_has_soup_stock_p:
                inner = self.regex_inner_quantity_cho.findall(numberless_name)
                if inner:
                    # 出汁の本数に対する割合を計算
                    density = self.get_inner_density(res_has_soup_stock_p[0][2], inner[0][0], cho_value_float)
                    agg_measure = AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        cho_value_float,
                        '丁', number,
                        density, res_has_soup_stock_p[0][1] == '液同', True)
                    agg_measure.inner_gram = inner[0][0]
                    return agg_measure
                else:
                    return AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        cho_value_float,
                        '丁', number,
                        self._get_percentage(numberless_name), res_has_soup_stock_p[0][1] == '液同')
            else:
                return AggMeasurePlateWithDensity(
                    index, self.cooking_day,
                    eating_day,
                    meal,
                    numberless_name,
                    cho_value_float,
                    '丁', number,
                    self._get_percentage(numberless_name), self.is_same_thickness(numberless_name))

        # -------------------------------------------------------------
        # ①トマトソース煮込みハンバーグ1個 + 液20g
        # -------------------------------------------------------------
        res_ko_gram_fill = re.findall('(\d+)個\D*液(\d+|\d+\.\d+)g', numberless_name)

        # ％がある場合は除く(次の判定で処理する)
        if res_ko_gram_fill and (self._get_percentage(numberless_name) == 0):
            # 液量が1個あたりの%になるように計算
            count, gram = res_ko_gram_fill[0]
            density = float(gram) / int(count) * 100
            return AggMeasurePlateWithDensity(
                index, self.cooking_day,
                eating_day,
                meal,
                numberless_name,
                count,
                '個', number,
                density, self.is_same_thickness(numberless_name), True)

        # -------------------------------------------------------------
        # ②煮物（肉団子3個＋小松菜36g）+液60g
        # -------------------------------------------------------------
        res_unit_ko_g_den = re.findall('(\D+)\s(\D+)(\d+)個＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)g', numberless_name)

        if res_unit_ko_g_den:
            names = [res_unit_ko_g_den[0][0], res_unit_ko_g_den[0][1], res_unit_ko_g_den[0][3]]

            return AggMeasurePlateKoGramDensity(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                res_unit_ko_g_den[0][2],
                number,
                res_unit_ko_g_den[0][5], numberless_name,
                res_unit_ko_g_den[0][4], self.is_same_thickness(numberless_name)
                )

        # -------------------------------------------------------------
        # ②煮物（肉団子3個(10g)＋小松菜36g）+10％
        # -------------------------------------------------------------
        res_unit_ko_g_den_p = re.findall('(\D+)\s(\D+)(\d+)個\s(\d+|\d+\.\d+)g\s＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)％', numberless_name)

        if res_unit_ko_g_den_p:
            names = [res_unit_ko_g_den_p[0][0], res_unit_ko_g_den_p[0][1], res_unit_ko_g_den_p[0][4]]
            quantity_g1 = float(res_unit_ko_g_den_p[0][3])
            quantity_g2 = float(res_unit_ko_g_den_p[0][5])
            density = round((quantity_g1 + quantity_g2) / 100 * float(res_unit_ko_g_den_p[0][6]), 2)

            return AggMeasurePlateKoGramDensity(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                res_unit_ko_g_den_p[0][2],
                number,
                density, numberless_name,
                res_unit_ko_g_den_p[0][5], self.is_same_thickness(numberless_name), res_unit_ko_g_den_p[0][3]
                )

        # -------------------------------------------------------------
        # ②れんこん煮物（れんこん1個+いんげん4ｇ）+9%
        # -------------------------------------------------------------
        res_unit_ko_g_p = re.findall('(\D+)\s(\D+)(\d+)個\s*＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)％', numberless_name)

        if res_unit_ko_g_p:
            names = [res_unit_ko_g_p[0][0], res_unit_ko_g_p[0][1], res_unit_ko_g_p[0][3]]
            quantity_ko = int(res_unit_ko_g_p[0][2])
            quantity_g = float(res_unit_ko_g_p[0][4])
            density = float(res_unit_ko_g_p[0][5])

            return AggMeasurePlateKoGramPercent(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                quantity_ko,
                number,
                density,
                numberless_name,
                quantity_g, self.is_same_thickness(numberless_name)
                )

        # -------------------------------------------------------------
        # ②煮物（肉団子3個(10g)＋小松菜36g）+10g
        # -------------------------------------------------------------
        res_unit_ko_g_den_g = re.findall('(\D+)\s(\D+)(\d+)個\s(\d+|\d+\.\d+)g\s＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)g', numberless_name)

        if res_unit_ko_g_den_g:
            names = [res_unit_ko_g_den_g[0][0], res_unit_ko_g_den_g[0][1], res_unit_ko_g_den_g[0][4]]

            return AggMeasurePlateKoGramDensity(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                res_unit_ko_g_den_g[0][2],
                number,
                res_unit_ko_g_den_g[0][6], numberless_name,
                res_unit_ko_g_den_g[0][5], self.is_same_thickness(numberless_name), res_unit_ko_g_den_g[0][3]
                )

        # -------------------------------------------------------------
        # ①食【た】べるスープの具【ぐ】（団子4個+具60ｇ）
        # -------------------------------------------------------------
        res_unit_ko_g_den_z = re.findall('(\D+)\s(\D+)(\d+)個＋(\D+)(\d+|\d+\.\d+)g', numberless_name)

        if res_unit_ko_g_den_z:
            names = [res_unit_ko_g_den_z[0][0], res_unit_ko_g_den_z[0][1], res_unit_ko_g_den_z[0][3]]

            return AggMeasurePlateKoGramDensity(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                res_unit_ko_g_den_z[0][2],
                number,
                0, numberless_name,
                res_unit_ko_g_den_z[0][4], self.is_same_thickness(numberless_name)
                )

        # -------------------------------------------------------------
        # ②麻婆豆腐（豆腐90g+ミンチ22g）+25％
        # -------------------------------------------------------------
        res_unit_g_g_per = re.findall('(\D+)\s(\D+)(\d+|\d+\.\d+)g＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)％', numberless_name)

        if res_unit_g_g_per:
            names = [res_unit_g_g_per[0][0], res_unit_g_g_per[0][1], res_unit_g_g_per[0][3]]
            quantity_g1 = float(res_unit_g_g_per[0][2])
            quantity_g2 = float(res_unit_g_g_per[0][4])
            density = round((quantity_g1 + quantity_g2) / 100 * float(res_unit_g_g_per[0][5]), 2)

            return AggMeasurePlateGramGram(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                quantity_g1,
                number,
                density, numberless_name,
                quantity_g2, self.is_same_thickness(numberless_name)
                )

        # -------------------------------------------------------------
        # ②麻婆豆腐（豆腐90g+ミンチ22g）+液25g
        # -------------------------------------------------------------
        res_unit_g_g_g = re.findall('(\D+)\s(\D+)(\d+|\d+\.\d+)g＋(\D+)(\d+|\d+\.\d+)g\s＋\D*(\d+|\d+\.\d+)g', numberless_name)

        if res_unit_g_g_g:
            names = [res_unit_g_g_g[0][0], res_unit_g_g_g[0][1], res_unit_g_g_g[0][3]]
            quantity_g1 = float(res_unit_g_g_g[0][2])
            quantity_g2 = float(res_unit_g_g_g[0][4])
            density = round(float(res_unit_g_g_g[0][5]), 2)

            return AggMeasurePlateGramGram(
                index, self.cooking_day,
                eating_day,
                meal,
                names,
                quantity_g1,
                number,
                density, numberless_name,
                quantity_g2, self.is_same_thickness(numberless_name)
                )

        # -------------------------------------------------------------
        # ①白身フライ60g1個
        # -------------------------------------------------------------
        res_unit_ko = re.findall('(\d+)個', numberless_name)

        if res_unit_ko:
            if res_has_soup_stock_g:
                density = float(res_has_soup_stock_g[0][2]) / int(res_unit_ko[0]) * 100
                return AggMeasurePlateWithDensity(
                    index, self.cooking_day,
                    eating_day,
                    meal,
                    numberless_name,
                    res_unit_ko[0],
                    '個', number,
                    density, res_has_soup_stock_g[0][1] == '液同', True)
            elif res_has_soup_stock_p:
                inner = self.regex_inner_quantity_ko.findall(numberless_name)
                if inner:
                    # 出汁の個数に対する割合を計算
                    density = self.get_inner_density(res_has_soup_stock_p[0][2], inner[0][0], res_unit_ko[0])
                    agg_measure = AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        res_unit_ko[0],
                        '個', number,
                        density, res_has_soup_stock_p[0][1] == '液同', True)
                    agg_measure.inner_gram = inner[0][0]
                    return agg_measure
                else:
                    return AggMeasurePlateWithDensity(
                        index, self.cooking_day,
                        eating_day,
                        meal,
                        numberless_name,
                        res_unit_ko[0],
                        '個', number,
                        self._get_percentage(numberless_name), res_has_soup_stock_p[0][1] == '液同')
            else:
                return AggMeasurePlateWithDensity(
                    index, self.cooking_day,
                    eating_day,
                    meal,
                    numberless_name,
                    res_unit_ko[0],
                    '個', number,
                    self._get_percentage(numberless_name), self.is_same_thickness(numberless_name))

        # -------------------------------------------------------------
        # ①すきやき76g+48g
        # ②ほうれん草山葵和え42.75g+10％
        # -------------------------------------------------------------
        res_unit_gram = re.findall('(\d+|\d+\.\d+)g', numberless_name)

        if len(res_unit_gram) > 0:
            if len(res_unit_gram) > 1:
                quantity2 = res_unit_gram[1]
                unit2 = 'g'
            else:
                quantity2 = None
                unit2 = None
            return AggMeasurePlateWithAnotherUnit(
                index, self.cooking_day,
                eating_day,
                meal,
                numberless_name,
                res_unit_gram[0],
                'g', number,
                self._get_percentage(numberless_name),
                quantity2, unit2, self.is_same_thickness(numberless_name)
            )

        return AggMeasureTarget(index, self.cooking_day, eating_day, meal, numberless_name, 0, None)


"""
    調理表から料理名、1食あたりの内容量、出汁の量を抜き出す処理

# 引数
    filename: 調理表_YYYY.MM.DD_施設給食.xls

# 出力ファイル
    料理名のリスト
    
# 概要
    Webで受注食数確定
    ↓
    らくらく献立に入力
    ↓
    5/1調理日指定で調理表出力
    ↓
    システムに調理表登録
    ↓
    調理表から料理名、1食あたりの内容量、出汁の量を抜き出し、
    料理名ごとに計量表を作成するプログラムを呼び出す（このプログラム）
    ↓
    5/5喫食分で食数集計→計量表作成
    5/6喫食分で食数集計→計量表作成
    
    調理表を登録することで、計量表まで自動的に作成ができるということ。

"""

class Command(BaseCommand):

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.plate_items = {}
        self.miso_items = None
        self.miso_liquid_quantity = None

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs='+', type=str)

    def is_miso_soup(self, name: str):
        if 'みそ汁' in name:
            return True
        elif '味噌汁' in name:
            return True
        elif 'みそしる' in name:
            return True
        else:
            return False

    def backup_parts_order_list(self, df):
        sort_order_list = []    # 調理表の出力順を再現するためのリスト
        prev_parts = ''
        for index, row in df.iterrows():
            eating_day = row['eating_day']
            meal_name = row['meal_name__meal_name']
            parts = row['parts_name']
            if parts != prev_parts:
                sort_order_list.append((eating_day, meal_name, parts))
                prev_parts = parts

        return sort_order_list

    def restore_parts_order(self, df, order_list):
        sorted = []
        # 調理表出現位置の算出(同一喫食日・同一食事区分(朝・昼・夕)で同一料理が存在した場合、先に登場する料理のインデックスを採用する)
        # 味噌汁の汁(液)に針刺し用等を出力する対象を特定するためのソート機能なので、汁の先頭が分かれば後は問題ない仕様
        for index, row in df.iterrows():
            parts_obj = (row['eating_day'], row['meal_name__meal_name'], row['parts_name'])
            for index2, sort_row in enumerate(order_list):
                if parts_obj == sort_row:
                    sorted.append(index2)
                    break

        # 復元したソート順で並べなおし(元々喫食日、朝・昼・夕の単位ではかたまっている)
        df['sorted'] = sorted
        return df.sort_values(['sorted', 'parts_name'])

    def store_miso_items(self):
        for no, plate_dict in self.plate_items.items():
            plate_name = plate_dict['name']
            if PlateNameAnalizeUtil.is_miso_soup(plate_name) and PlateNameAnalizeUtil.is_soup_liquid(plate_name):
                if not self.miso_liquid_quantity:
                    res = re.findall('(\d+|\d+.\d+)cc', plate_name)
                    if res:
                        self.miso_liquid_quantity = float(res[0])

                self.miso_items = plate_dict['items']
                return

    def read_plate_items(self, df):
        item_list = None
        plate = None
        for index, row in df.iterrows():
            #　献立名の検出
            if row['Unnamed: 3'] != '':
                # 検出中の情報を保存
                if plate:
                    self.plate_items[index] = {"name": plate, "items": item_list}

                item_list = []
                if row['Unnamed: 5'] != '':
                    if row['Unnamed: 6'] == '':
                        item_list.append(
                            (row['Unnamed: 5'],
                             0.0)
                        )
                    else:
                        res_q = re.findall('(\d)/(\d)', row['Unnamed: 6'])
                        if res_q:
                            item_list.append(
                                (row['Unnamed: 5'],
                                 float(res_q[0][0]) / float(res_q[0][1]))
                            )
                        else:
                            item_list.append(
                                (row['Unnamed: 5'],
                                 float(row['Unnamed: 6']))
                            )
                plate = row['Unnamed: 3']
            elif row['Unnamed: 5'] != '':
                # 登録済み食材と同名をチェック
                is_updated = False
                for i, item_value in enumerate(item_list):
                    if item_value[0] == row['Unnamed: 5']:
                        if row['Unnamed: 6'] == '':
                            item_list[i] = (item_value[0], item_value[1])
                        else:
                            res_q = re.findall('(\d)/(\d)',row['Unnamed: 6'])
                            if res_q:
                                item_list[i] = (item_value[0], item_value[1] + float(row['Unnamed: 6']) + float(res_q[0][0]) / float(res_q[0][1]))
                            else:
                                item_list[i] = (item_value[0], item_value[1] + float(row['Unnamed: 6']))
                        is_updated = True
                        break

                if not is_updated:
                    if row['Unnamed: 6'] == '':
                        item_list.append(
                            (row['Unnamed: 5'],
                             0.0)
                        )
                    else:
                        res_q = re.findall('(\d)/(\d)', row['Unnamed: 6'])
                        if res_q:
                            item_list.append(
                                (row['Unnamed: 5'],
                                 float(res_q[0][0]) / float(res_q[0][1]))
                            )
                        else:
                            item_list.append(
                                (row['Unnamed: 5'],
                                 float(row['Unnamed: 6']))
                            )

        if plate and item_list:
            index += 1
            self.plate_items[index] = {"name": plate, "items": item_list}

    def read_plate_for_package(self, df):
        """
        P7対応CSV出力のため、献立と食種の情報を読み込む
        """
        # 基本食と判定する食種のリスト
        BASIC_PLATE_TYPES = [
            settings.COOKING_DIRECTION_J_CODE,
            settings.COOKING_DIRECTION_SOUP_J_CODE,
            settings.COOKING_DIRECTION_GU_J_CODE,
            settings.COOKING_DIRECTION_B_CODE,
            settings.COOKING_DIRECTION_SOUP_B_CODE,
            settings.COOKING_DIRECTION_GU_B_CODE,
        ]

        result_list = []
        # 喫食日と食事区分の複合
        eating_meal = None
        plate = None
        eating_type_list = []
        for index, row in df.iterrows():
            if row['Unnamed: 0'] != '':
                # 喫食日+食事区分の検出
                if eating_meal:
                    # eating_meal切り替え
                    is_basic_plate = None
                    is_find = False
                    for eating_type in eating_type_list:
                        for basic_type in BASIC_PLATE_TYPES:
                            if basic_type in eating_type:
                                is_basic_plate = True
                                is_find = True
                                break
                        if is_find:
                            break

                    if not is_find:
                        is_basic_plate = False

                    dict = {
                        'eating_meal': eating_meal,
                        'plate': plate,
                        'is_basic_plate': is_basic_plate,
                        'eating_type_list': eating_type_list
                    }
                    result_list.append(dict)

                    # 献立名の検出で拾ってしまわないよう対応
                    plate = None
                eating_meal = row['Unnamed: 0']
                # 他の列は空白、かならず次行に献立、食事区分の切り替えがあるためこれ以上処理しなくても大丈夫

            if row['Unnamed: 3'] != '':
                # 献立名の検出
                if plate:
                    is_basic_plate = None
                    is_find = False
                    for eating_type in eating_type_list:
                        for basic_type in BASIC_PLATE_TYPES:
                            if basic_type in eating_type:
                                is_basic_plate = True
                                is_find = True
                                break
                        if is_find:
                            break

                    if not is_find:
                        is_basic_plate = False

                    dict = {
                        'eating_meal': eating_meal,
                        'plate': plate,
                        'is_basic_plate': is_basic_plate,
                        'eating_type_list': eating_type_list
                    }
                    result_list.append(dict)
                plate = row['Unnamed: 3']
                eating_type_list = []
            if row['Unnamed: 2'] != '':
                # 注文数＋食種の検出
                plate_type = row['Unnamed: 2']
                eating_type_list.append(plate_type)

        # 残った情報の登録
        is_basic_plate = None
        is_find = False
        for eating_type in eating_type_list:
            for basic_type in BASIC_PLATE_TYPES:
                if basic_type in eating_type:
                    is_basic_plate = True
                    is_find = True
                    break
            if is_find:
                break

        if not is_find:
            is_basic_plate = False

        dict = {
            'eating_meal': eating_meal,
            'plate': plate,
            'is_basic_plate': is_basic_plate,
            'eating_type_list': eating_type_list
        }
        result_list.append(dict)

        # 献立名「ごはん」の料理は除外して出力
        excluded = [x for x in result_list if x['plate'] != "ごはん"]
        self.logger.debug(excluded)
        return excluded

    def copy_allergen_base(self, plate, prev_plate):
        current_dict = plate['allergen_base']
        prev_dict = prev_plate['allergen_base']
        for key, value in prev_dict.items():
            current_dict[key] = value

    def is_allergen_soe_plate(self, plate, prev_plate):
        """
        アレルギー代替で、添えになる料理かどうかを判断する
        """
        if not prev_plate:
            return False

        # 連続した料理のみ対象
        if (plate['anarize_index'] - prev_plate['anarize_index']) != 1:
            return False

        # ①(主菜)に④(添え)が付く場合に限定する
        if plate['plate'][0] != '④':
            return False
        if prev_plate['plate'][0] != '①':
            return False

        # 全ての食種内容が一致する場合のみ代替の添えと判断
        current_list = plate['eating_type_list']
        prev_list = prev_plate['eating_type_list']
        if len(current_list) != len(prev_list):
            return False

        for current, prev in zip(current_list, prev_list):
            if current != prev:
                return False

        return True

    def allergen_anarize(self, plate_dict_list):
        eating_meal_list = []
        # 同一喫食・食事区分でグルーピング
        for key, group in groupby(plate_dict_list, key=lambda x: x['eating_meal']):
            eating_meal_list.append(list(group))

        # find_enable_kidnd_regex = re.compile('\d+ (\D+)')
        for plate_list in eating_meal_list:
            # フラグの設定
            index = 0
            for plate in plate_list:
                plate['is_soup'] = plate['plate'][0] == '⑤'
                if plate['is_basic_plate']:
                    plate['is_allergen'] = False
                else:
                    plate['is_allergen'] = True

                # 連続した料理かどうかを判定するため、本判定内のインデックスを付与する
                plate['anarize_index'] = index
                index += 1

            # アレルギー元の判定
            basic_list = [x for x in plate_list if x['is_basic_plate']]
            prev_plate = None
            allergen_plate_list = [x for x in plate_list if not x['is_basic_plate']]
            already_dict = {}
            unieque_allergen_kind_list = []
            for index, allergen_plate in enumerate(allergen_plate_list):
                allergen_base_dict = {}
                allergen_plate['allergen_base'] = allergen_base_dict

                # 1:Nになるアレルギー代替食の判定
                if self.is_allergen_soe_plate(allergen_plate, prev_plate):
                    self.copy_allergen_base(allergen_plate, prev_plate)
                else:
                    # アレルギー食種毎に、変更元を検索
                    for kind in allergen_plate['eating_type_list']:
                        if not kind in unieque_allergen_kind_list:
                            unieque_allergen_kind_list.append(kind)
                        allergen_base_list = []
                        for basic_plate in [x for x in basic_list if x['is_soup'] == allergen_plate['is_soup']]:
                            is_hit = False
                            for base_kind in basic_plate['eating_type_list']:
                                kind_value = kind[kind.find(' ') + 1:]
                                base_kind_value = base_kind[base_kind.find(' ') + 1:]
                                if kind_value == base_kind_value:
                                    is_hit = True
                                    break
                            if is_hit:
                                # 同じ食種が見つかった場合は、別の基本料理のため、スルー
                                is_hit = False
                            else:
                                # 対象アレルギー元料理(対象食種の食数を持たない料理)が見つかった
                                allergen_base_list.append(basic_plate['plate'])

                        # 料理名先頭の番号が同じ料理を優先してアレルギー元として採用する
                        plate_number = allergen_plate['plate'][0]
                        ar_tmp_list = [x for x in allergen_base_list if plate_number in x]
                        if ar_tmp_list:
                            # 同一番号の料理が存在する場合
                            if kind in already_dict:
                                already_list = already_dict[kind]
                                if len(already_list) < len(allergen_base_list):
                                    # 代替元候補のうち、まだ代替先が見つかっていないものがあったらそちらを優先
                                    not_already_source_list = [x for x in allergen_base_list if not x in already_list]

                                    already_dict[kind] += [not_already_source_list[0]]
                                    allergen_base_dict[kind] = not_already_source_list[0]
                                else:
                                    already_dict[kind] += [ar_tmp_list[0]]
                                    allergen_base_dict[kind] = ar_tmp_list[0]
                            else:
                                already_dict[kind] = [ar_tmp_list[0]]
                                allergen_base_dict[kind] = ar_tmp_list[0]
                        else:
                            if allergen_base_list:
                                # 小鉢(②、③)の元が④になることは通常ありえない
                                if plate_number != '④':
                                    ar_tmp_list2 = [x for x in allergen_base_list if not('④' in x)]
                                    if ar_tmp_list2:
                                        allergen_base_dict[kind] = ar_tmp_list2[0]

                                        # 代替元候補に①を含んだ複数があり、今回①が選ばれた場合
                                        if (ar_tmp_list2[0][0] == '①') and (len(ar_tmp_list2) > 1):
                                            # 後の①アレルギー代替食に、同じ食種があるかどうか?
                                            apl = allergen_plate_list[index+1:]
                                            is_change = False
                                            for arp in apl:
                                                if arp['plate'][0] != '①':
                                                    continue
                                                for apl_kind in arp['eating_type_list']:
                                                    if apl_kind == kind:
                                                        # 通常①の料理が代替で、②や③が同時に対応づくことはないため、
                                                        # 別の候補を採用する
                                                        allergen_base_dict[kind] = ar_tmp_list2[1]
                                                        if kind in already_dict:
                                                            already_dict[kind] += ar_tmp_list2[1]
                                                        else:
                                                            already_dict[kind] = [ar_tmp_list2[1]]
                                                        is_change = True
                                                        break
                                                if is_change:
                                                    break

                                            if not is_change:
                                                # 前のアレルギー代替食を確認
                                                apl = allergen_plate_list[:index]
                                                for arp2 in apl:
                                                    if arp2['plate'][0] == '①':
                                                        for apl_kind2 in arp2['eating_type_list']:
                                                            if apl_kind2 == kind:
                                                                # 通常①の料理が代替で、②や③が同時に対応づくことはないため、
                                                                # 別の候補を採用する
                                                                allergen_base_dict[kind] = ar_tmp_list2[1]
                                                                if kind in already_dict:
                                                                    already_dict[kind] += ar_tmp_list2[1]
                                                                else:
                                                                    already_dict[kind] = [ar_tmp_list2[1]]
                                                                is_change = True
                                                                break
                                                    if is_change:
                                                        break

                                            # 他の候補がなければ、①の代替として②、③が選ばれたと判断
                                    else:
                                        self.logger.error(f'アレルギー紐づけ先なし:{allergen_plate["plate"]}-{kind}')
                                else:
                                    allergen_base_dict[kind] = allergen_base_list[0]
                                    if kind in already_dict:
                                        already_dict[kind] += allergen_base_list[0]
                                    else:
                                        already_dict[kind] = [allergen_base_list[0]]
                            else:
                                self.logger.error(f'アレルギー紐づけ先なし:{allergen_plate["plate"]}-{kind}')
                    prev_plate = allergen_plate

        return eating_meal_list

    def generate_eating_meal_list(self, eating_meal_list):
        for plate_list in eating_meal_list:
            # plate_list:喫食日・食事区分単位
            for plate in plate_list:
                yield plate


    def handle(self, *args, **options):

        # 呼び出し時の引数1つ目「調理表_YYYY.MM.DD_施設給食.xls」
        in_file = options['filename'][0]

        # ファイル名から日時をYYYY-MM-DD形式で抽出
        cooking_day = re.sub('.*(\d{4})\.(\d{2})\.(\d{2}).*', '\\1-\\2-\\3', in_file)

        cook_direc_file = os.path.join(settings.MEDIA_ROOT, 'upload', in_file)
        cook_direc = pd.read_excel(cook_direc_file)

        # ------------------------------------------------------------------------------
        # 『2022年5月11日(水) 調理』のように月と日は0埋めされていない形式
        cooking_year = cook_direc.iloc[3, 5][0:4]
        cooking_month = cook_direc.iloc[3, 5][5:7]  # 「1月」〜「9月」、「10」「11」「12」のどれか

        # ７行目まで削除
        cook_direc = cook_direc.drop(index=cook_direc.index[[0, 1, 2, 3, 4, 5, 6]])
        cook_direc = cook_direc.replace(np.nan, '', regex=True)  # NaNを空文字列に変更しておく

        # 材料読込
        self.read_plate_items(cook_direc)
        self.store_miso_items()

        # 袋出力用情報読込
        plates_for_package = self.read_plate_for_package(cook_direc)
        logger.info(f'plates_for_package={plates_for_package}')
        analyzed_plates = self.allergen_anarize(plates_for_package)
        package_manager = PrintPlatePackageManager(plates_for_package)
        package_manager.save_new(cooking_day)

        # B列削除、E列以降L以外削除
        cook_direc = cook_direc.drop(columns=cook_direc.columns[[1, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]])

        # cook_direc.to_csv("tmp/C-1.csv", index=False)

        # ------------------------------------------------------------------------------
        # cook_direc = cook_direc.replace(np.nan, '', regex=True)  # NaNを空文字列に変更しておく
        c_direc = cook_direc.copy()
        del cook_direc
        del cook_direc_file

        date_now = menu_now = ""

        for index, row in c_direc.iterrows():

            # A列の「喫食日」と「食事区分」が入っていない行に、1行上の値をコピーする
            if row['Unnamed: 0'] == '':
                c_direc.loc[index, 'Unnamed: 0'] = date_now

            date_now = row['Unnamed: 0']

            # C列の「献立のパーツ」が入っていない行に、1行上の値をコピーする
            if row['Unnamed: 3'] == '':
                c_direc.loc[index, 'Unnamed: 3'] = menu_now

            menu_now = row['Unnamed: 3']

        # c_direc.to_csv("tmp/C-2.csv", index=False)

        # ------------------------------------------------------------------------------
        for index, row in c_direc.iterrows():

            # B列が空の場合は不要な行なので削除
            if row['Unnamed: 2'] == '':
                c_direc.drop(index=[index], axis=0, inplace=True)
            # C列のごはんは製造に関係ないので削除
            elif row['Unnamed: 3'] == 'ごはん':
                c_direc.drop(index=[index], axis=0, inplace=True)

        # 不要な表記を削除、変更
        c_direc['Unnamed: 0'] = c_direc['Unnamed: 0'].str.strip('■')
        c_direc['Unnamed: 0'] = c_direc['Unnamed: 0'].str.replace('間食昼', '間食', regex=True)
        c_direc['Unnamed: 2'] = c_direc['Unnamed: 2'].str.replace('･基本食\(施設\)', '', regex=True)

        # A列は「喫食日」と「食事区分」

        # A列の食事区分を削除し、新規に列名を「eating_day」として喫食日のみの列を追加
        c_direc['eating_day'] = c_direc['Unnamed: 0'].str.replace('\s.+$', '', regex=True)

        # 0埋めして MM/DD を -MM-DD に（YYYYは後で追加）
        c_direc['eating_day'] = c_direc['eating_day'].str.replace(r'^(\d{1})/(\d{1})$', r'-0\1-0\2', regex=True)
        c_direc['eating_day'] = c_direc['eating_day'].str.replace(r'^(\d{1})/(\d{2})$', r'-0\1-\2', regex=True)
        c_direc['eating_day'] = c_direc['eating_day'].str.replace(r'^(\d{2})/(\d{1})$', r'-\1-0\2', regex=True)
        c_direc['eating_day'] = c_direc['eating_day'].str.replace(r'^(\d{2})/(\d{2})$', r'-\1-\2', regex=True)

        # A列の喫食日を削除し、新規に列名を「meal_name」として食事区分のみの列を追加
        c_direc['meal_name__meal_name'] = c_direc['Unnamed: 0'].str.replace('^\d+/\d+\s', '', regex=True)

        # B列の注文数を削除し、新規に列名を「short_name」として短縮コードの列を追加
        c_direc['short_name'] = c_direc['Unnamed: 2'].str.replace('^\d+\s', '', regex=True)

        # C列はそのまま、新規に列名を「parts_name」として追加
        c_direc['parts_name'] = c_direc['Unnamed: 3']

        # A〜C列を削除
        c_direc = c_direc.drop(columns=c_direc.columns[[0, 1, 2]])

        # 「eating_day」にYYYYの追加
        for index, row in c_direc.iterrows():

            if cooking_month == '12' and row['eating_day'][1:3] == '01':
                # 製造日が12月で、喫食日に翌年の1月のものがあればYYYYは翌年にする
                next_year = int(cooking_year) + 1
                row['eating_day'] = str(next_year) + row['eating_day']
            else:
                row['eating_day'] = cooking_year + row['eating_day']

        # c_direc.to_csv("tmp/C-3.csv", index=False)

        # 調理表の登場順を再現するためのリストの取得
        sort_order_list = self.backup_parts_order_list(c_direc)

        # ------------------------------------------------------------------------------
        c_direc = c_direc.sort_values(
            by=['eating_day', 'meal_name__meal_name', 'short_name', 'parts_name'])

        for index, row in c_direc.iterrows():

            if row['short_name'][0] == '2':
                c_direc.drop(index=[index], axis=0, inplace=True)

        # c_direc.to_csv("tmp/C-4.csv", index=False)

        # ------------------------------------------------------------------------------
        c_direc = c_direc.groupby(['eating_day', 'meal_name__meal_name', 'parts_name']).count().reset_index()

        # 調理表の登場順の再現
        c_direc = self.restore_parts_order(c_direc, sort_order_list)
        sort_order_list.clear()

        # group byすると集計したshort_nameの列が追加されてしまうのでreindexする
        c_direc = c_direc.reindex(columns=['eating_day', 'meal_name__meal_name', 'parts_name'])

        # c_direc.to_csv("tmp/C-5_工程5.csv", index=False)

        # ------------------------------------------------------------------------------
        # 変換前の名称を記憶
        c_direc['before_name'] = c_direc['parts_name']

        # 全角のｇを半角のgに揃える
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('ｇ', 'g', regex=True)

        # 半角の%を全角の％に揃える
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('%', '％', regex=True)

        # 半角の+を全角の＋に揃える
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('+', '＋', regex=True)

        # 半角の()を半角スペースにする
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('(', ' ', regex=True)
        c_direc['parts_name'] = c_direc['parts_name'].str.replace(')', ' ', regex=True)

        # 全角の（）を半角スペースにする
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('（', ' ', regex=True)
        c_direc['parts_name'] = c_direc['parts_name'].str.replace('）', ' ', regex=True)

        c_direc['parts_name'] = c_direc['parts_name'].str.replace('/', '÷', regex=True)

        # 既存のフォルダをクリア
        measure_output_dir = os.path.join(settings.OUTPUT_DIR, 'measure')
        new_dir_path = os.path.join(measure_output_dir, '計量表_' + cooking_day + '_製造')
        if os.path.exists(new_dir_path):
            shutil.rmtree(new_dir_path)

        # 調理表の内容を解析する
        analyzer = AggMeasureTargetAnalyzer(cooking_day)
        for index, df_param in enumerate(c_direc.iterrows()):
            df_index, df_row = df_param
            is_find = False
            for items_dict in self.plate_items.values():
                row_name = df_row['before_name']
                dict_name = items_dict['name']
                package_qs = PlatePackageForPrint.objects.filter(
                    plate_name=df_row['before_name'], eating_day=df_row['eating_day'], meal_name=df_row['meal_name__meal_name']).values('id', 'index', 'is_basic_plate')
                if package_qs.exists():
                    package = package_qs.first()
                    if row_name == dict_name:
                        if package['is_basic_plate']:
                            analyzer.add_cook(package['index'], df_row['parts_name'], df_row['eating_day'], df_row['meal_name__meal_name'], items_dict['items'], row_name)
                        else:
                            analyzer.add_cook(-1, df_row['parts_name'], df_row['eating_day'],
                                              df_row['meal_name__meal_name'], items_dict['items'], row_name)
                        is_find = True
                        break

            # 万が一見つからない場合でも計量表は出力
            if not is_find:
                self.logger.warning(f'add to analyzer: not found :{df_row["before_name"]}')
                analyzer.add_cook(-1, df_row['parts_name'], df_row['eating_day'], df_row['meal_name__meal_name'], [], df_row['before_name'])

        # メモリ節約のため、要素削除
        self.plate_items.clear()
        del c_direc

        # ピッキング指示書用袋数をクリア
        UnitPackage.objects.filter(cooking_day=cooking_day).delete()
        TmpPlateNamePackage.objects.filter(cooking_day=cooking_day).delete()

        aggregation = MisoSoupAggregation()
        error_list = []
        plate_generator = self.generate_eating_meal_list(analyzed_plates)
        enge_adjust_status = 0
        enge_adjust_timing = None
        order_manager = AggMeasureOrdersManager()
        store_first_miso_flag = False
        for index, analyzed in enumerate(analyzer.generate_analyzed()):
            self.logger.info(f'measure_start:{analyzed.name}')

            # 前回ループでstatusが1になっていた場合は、2へ移行
            if enge_adjust_status == 1:
                enge_adjust_status = 2
            if enge_adjust_timing:
                adjust_eating_day, adjust_meal = enge_adjust_timing
                if (adjust_eating_day != analyzed.eating_day) or (adjust_meal != analyzed.meal):
                    enge_adjust_status = 0

            # 本来zippedにすべきだが、既存への影響を考慮し中で処理とした
            try:
                plate_dict_for_kakiokoshi = next(plate_generator)
            except Exception:
                self.logger.error(f'read eating_meal_list error:{index}')
                plate_dict_for_kakiokoshi = None

            if plate_dict_for_kakiokoshi:
                if CookingDirectionPlatesManager.is_ignore_plate(plate_dict_for_kakiokoshi):
                    # サンプル料理の計量表を出力しない
                    plate_dict_for_kakiokoshi['is_mix_rice'] = False

                    # サンプルが初回の味噌汁の場合はフラグを次の味噌汁に引き継ぐ
                    if type(analyzed) in [AggMeasureMisoDevide, AggMeasureMiso]:
                        if analyzed.is_first:
                            store_first_miso_flag = True
                    continue

                plate_dict_for_kakiokoshi['work_name'] = analyzed.name

            adjst_result = analyzed.get_enge_adjust_status()
            if adjst_result == 1:
                # カレー・シチューの場合
                if enge_adjust_status == 0:
                    enge_adjust_timing = (analyzed.eating_day, analyzed.meal)
                    enge_adjust_status = 1
            try:
                if type(analyzed) is AggMeasureMisoDevide:
                    analyzed.items = self.miso_liquid_quantity
                    if store_first_miso_flag:
                        analyzed.is_first = True
                        store_first_miso_flag = False
                    message = analyzed.call_command(enge_adjust_status, order_manager)
                    aggregation.add_soup_quantity(message, analyzed.eating_day, analyzed.meal)

                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureSoupDevide:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureMiso:
                    analyzed.items = self.miso_liquid_quantity
                    if store_first_miso_flag:
                        analyzed.is_first = True
                        store_first_miso_flag = False
                    message = analyzed.call_command(enge_adjust_status, order_manager)
                    aggregation.add_soup_quantity(message, analyzed.eating_day, analyzed.meal)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureSoupFilling:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureSoupLiquid:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureLiquidSeasoning:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasureMixRice:
                    writer = MixRiceMeasureWriter(analyzed)
                    writer.write()
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = True
                    continue
                elif type(analyzed) is AggMeasureMixRiceParts:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = True
                    continue
                elif type(analyzed) is AggMeasurePlate:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateWithDensity:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateKoGram:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateKoGramPercent:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateKoGramDensity:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateGramGram:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                elif type(analyzed) is AggMeasurePlateWithAnotherUnit:
                    analyzed.call_command(enge_adjust_status, order_manager)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
                else:
                    analyzed.call_command(order_manager, enge_adjust_status)
                    if plate_dict_for_kakiokoshi:
                        plate_dict_for_kakiokoshi['is_mix_rice'] = False
                    continue
            except Exception as e:
                self.logger.error(analyzed.name)
                self.logger.info(traceback.format_exc())
                error_list.append(analyzed.name)

        del plate_generator
        del order_manager
        self.logger.info('メインループ終了')

        # 調理表の料理情報を保存
        CookingDirectionPlatesManager.save(analyzed_plates, cooking_day)

        # アレルギーの袋数を保存
        CookingDirectionPlatesManager.save_p7_allergen(analyzed_plates, cooking_day)

        self.logger.info('p7-allergen-end')

        # 味噌汁合計表の出力
        self.output_miso_total(new_dir_path, aggregation, cooking_day)
        self.logger.info('miso-output-end')

        shutil.make_archive(new_dir_path, 'zip', root_dir=new_dir_path)
        self.logger.info('zipped-end')

        if error_list:
            return ",".join(error_list)

    def output_miso_total(self, path, aggregation: MisoSoupAggregation, cooking_day: str):
        if not self.miso_items:
            return
        aggregation.item_quantity_list = self.miso_items

        # 味噌汁の合計表の出力
        miso_total_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/measure_miso_total.xlsx')  # 味噌汁用の合計テンプレ
        wb = excel.load_workbook(miso_total_template)
        worksheet = wb.worksheets[0]

        # 合計表の出力
        # 常食
        row = 4
        for name, value in self.miso_items:
            worksheet.cell(row, 2, name)
            worksheet.cell(row, 3, aggregation.get_j_item_gram_quantity_function(value))
            if row == 6:
                break
            else:
                row += 1

        # 常食(針刺し用)
        worksheet.cell(9, 3, aggregation.get_j_needle_gram_quantity_function())

        # 薄味
        if aggregation.u_total > 0:
            worksheet.cell(4, 5, aggregation.get_u_item_gram_quantity_function(0))
            worksheet.cell(5, 5, aggregation.get_u_item_gram_quantity_function(1))
            worksheet.cell(6, 5, aggregation.get_u_item_gram_quantity_function(2))

            # 薄味(針刺し用)
            worksheet.cell(9, 5, aggregation.get_u_needle_gram_quantity_function())

        total_file = os.path.join(path, cooking_day + '_全注文_味噌汁合計表.xlsx')
        wb.save(total_file)
        wb.close()
