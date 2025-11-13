import datetime as dt
import os

import openpyxl as excel

from openpyxl.drawing.image import Image

from django.conf import settings
from django.core.management.base import BaseCommand
from web_order.models import FoodPhoto, MonthlyMenu, EngeDirection

"""
    盛付指示書の作成

# 引数
    喫食日

# 献立DB
    Model: MonthlyMenu
    Model: FoodPhoto
"""


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('date', nargs='+', type=str)

    def handle(self, *args, **options):
        in_date = options['date'][0]

        setout_output_dir = os.path.join(settings.OUTPUT_DIR, 'setout')
        os.makedirs(setout_output_dir, exist_ok=True)  # 上書きOK
        output_file = os.path.join(setout_output_dir, '献立指示書_' + str(in_date) + '.xlsx')
        output_enge_file = os.path.join(setout_output_dir, '嚥下用献立指示書_' + str(in_date) + '.xlsx')

        xls_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/setout.xlsx')  # 指示書のテンプレート
        xls_enge_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/setout_enge.xlsx')  # 嚥下用指示書のテンプレート

        hot_text = '再加熱カートの場合は温めモード、\n\nスチコンの場合はスチーム88℃で15分加熱、\n湯煎の場合は沸騰後15分加熱し、開封して盛付け'
        cold_text = '再加熱カートの場合は冷蔵モード、加熱せず、開封して盛付け'
        no_text = ''

        book = excel.load_workbook(xls_template)

        def create_excel(queryset, sheetname):

            worksheet = book[sheetname]

            worksheet.cell(2, 5, in_date)

            i = 4
            j = 6
            p = 5

            for q in queryset:
                qs2 = FoodPhoto.objects.filter(food_name=q).first()

                worksheet.cell(i, 2, qs2.food_name)

                if qs2.hot_cool == '冷蔵':
                    worksheet.cell(i, 4, cold_text)
                elif qs2.hot_cool == '温め':
                    worksheet.cell(i, 4, hot_text)
                else:
                    worksheet.cell(i, 4, no_text)

                worksheet.cell(j, 4, qs2.direction)

                img_file = str(qs2.thumbnail)

                if (img_file):

                    img_dir = os.path.join(settings.MEDIA_ROOT, str(qs2.basic))

                    # 画像を選択
                    img_to_excel = excel.drawing.image.Image(img_dir)
                    # 指定の位置に画像を添付
                    worksheet.add_image(img_to_excel, 'C' + str(p))

                i = i + 5
                j = j + 5
                p = p + 5

        qs_b = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='朝食', option=False)
        qs_l = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='昼食', option=False)
        qs_d = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='夕食', option=False)
        qs_op_b = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='朝食', option=True)
        qs_op_l = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='昼食', option=True)
        qs_op_d = MonthlyMenu.objects.filter(eating_day=in_date, meal_name='夕食', option=True)

        create_excel(qs_b, '朝食')
        create_excel(qs_l, '昼食')
        create_excel(qs_d, '夕食')
        create_excel(qs_op_b, 'オプション・朝')
        create_excel(qs_op_l, 'オプション・昼')
        create_excel(qs_op_d, 'オプション・夕')

        book.save(output_file)

        # ------------------------------------------------------------------------------
        # 嚥下用の献立指示書の取得
        # ------------------------------------------------------------------------------
        enge_book = excel.load_workbook(xls_enge_template)
        enge_ws = enge_book['Sheet1']

        def get_rice_type(name):
            if '丼' in name:
                return 'other'
            elif 'カレーライス' in name:
                return 'curry'
            else:
                return None

        def get_noodle_type(name):
            if '焼きそば' in name:
                return ('yaki', True)
            elif 'うどん' in name:
                # ねぎは具とみなさない。他のものがあれば具とみなす(暫定)
                sp = name.split('・')

                if len(sp) == 1:
                    # 添えがない=うどんの麺のみ
                    return ('udon', False)
                for gu_name in sp[1:]:
                    # ねぎ以外の具が合ったら具ありとみなす
                    if gu_name != 'ねぎ':
                        return ('udon', True)

                # ねぎだけなら具なしとみなす
                return ('udon', False)
            else:
                return None

        def write_default_rice(ws, row, column, enge_type):
            if enge_type == 'ミキサー':
                ws.cell(row, column, f'全粥の{enge_type}食を作る。')
                ws.cell(row + 1, column, '丼の具は開封後、盛付け。')
            else:
                ws.cell(row, column, f'全粥の{enge_type}食を作る。丼の具は')
                ws.cell(row + 1, column, '開封後袋から出して、カットして盛付け。')

        def write_curry_rice(ws, row, column, enge_type, soe):
            if enge_type == 'ミキサー':
                ws.cell(row, column, f'全粥の{enge_type}食を作る。')
                if soe:
                    ws.cell(row + 1, column, f'カレールーは{soe}と開封後、盛付け。')
                else:
                    ws.cell(row + 1, column, 'カレールーは開封後、盛付け。')
            else:
                ws.cell(row, column, f'全粥の{enge_type}食を作る。丼の具は')
                if soe:
                    ws.cell(row + 1, column, f'カレールーと{soe}は、')
                else:
                    ws.cell(row + 1, column, 'カレールーは、')
                ws.cell(row + 2, column, '開封後袋から出して、カットして盛付け。')

        def write_noodle(ws, row, column, enge_type, noodle):
            if noodle[0] == 'yaki':
                noodle_name = '焼きそば'
            else:
                noodle_name = 'うどん'

            if enge_type == 'ミキサー':
                ws.cell(row, column, f'{noodle_name}の{enge_type}食を作る。')
                if noodle[1]:
                    ws.cell(row + 1, column, '具は開封後、盛付け。')
                else:
                    ws.cell(row + 1, column, 'スープは開封後、盛付け。')
            else:
                if noodle[1]:
                    ws.cell(row, column, f'{noodle_name}の{enge_type}食を作る。具は')
                    ws.cell(row + 1, column, '開封後袋から出して、カットして盛付け。')
                else:
                    ws.cell(row, column, f'{noodle_name}の{enge_type}食を作る。スープは')
                    ws.cell(row + 1, column, '開封後袋から出して、盛付け。')

        def write_enge_direcion(ws, row, column, enge_type, enge_direcion):
            if enge_direcion.direction_type == '2':
                # 盛付
                if enge_type == 'ミキサー':
                    ws.cell(row, column, '開封後、盛付け。')
                else:
                    ws.cell(row, column, '開封後袋から出して、カットして盛付け。')
                    ws.cell(row + 1, column, f'{enge_direcion.soe}は開封して盛付け。')
            elif enge_direcion.direction_type == '1':
                # かける
                if enge_type == 'ミキサー':
                    ws.cell(row, column, '開封後、盛付け。')
                else:
                    ws.cell(row, column, '開封後袋から出して、カットして盛付け。')
                    ws.cell(row + 1, column, f'{enge_direcion.soe}は開封して盛付け。')
                ws.cell(row + 1, column, f'{enge_direcion.soe}をかける。')
            else:
                # フリー入力
                if enge_type == 'ミキサー':
                    ws.cell(row, column, enge_direcion.direction_mixer_1)
                    ws.cell(row + 1, column, enge_direcion.direction_mixer_2)
                    ws.cell(row + 2, column, enge_direcion.direction_mixer_3)
                else:
                    ws.cell(row, column, enge_direcion.direction_other_1)
                    ws.cell(row + 1, column, enge_direcion.direction_other_2)
                    ws.cell(row + 2, column, enge_direcion.direction_other_3)

        def create_enge_excel(queryset, meal):
            if meal == '朝食':
                start_row = 5
                max_index = 2
            elif meal == '昼食':
                start_row = 15
                max_index = 3
            else:
                start_row = 30
                max_index = 3

            r = start_row
            for index, q in enumerate(queryset):
                if index == max_index:
                    break

                excel_food_name = q.food_name
                enge_ws.cell(r, 2, excel_food_name)

                rice_type = get_rice_type(excel_food_name)
                if rice_type:
                    # ご飯ものの対応
                    if rice_type == 'other':
                        write_default_rice(enge_ws, r+1, 3, 'ソフト')
                        write_default_rice(enge_ws, r + 1, 4, 'ミキサー')
                        write_default_rice(enge_ws, r + 1, 5, 'ゼリー')
                    if rice_type == 'curry':
                        sp = q.food_name.split('・')
                        soe = sp[1] if len(sp) > 1 else None

                        write_curry_rice(enge_ws, r + 1, 3, 'ソフト', soe)
                        write_curry_rice(enge_ws, r + 1, 4, 'ミキサー', soe)
                        write_curry_rice(enge_ws, r + 1, 5, 'ゼリー', soe)
                else:
                    noodle = get_noodle_type(excel_food_name)
                    if noodle:
                        write_noodle(enge_ws, r + 1, 3, 'ソフト', noodle)
                        write_noodle(enge_ws, r + 1, 4, 'ミキサー', noodle)
                        write_noodle(enge_ws, r + 1, 5, 'ゼリー', noodle)
                    else:
                        enge_direction_list = EngeDirection.objects.filter(food_name=excel_food_name)
                        if enge_direction_list:
                            enge = enge_direction_list.first()
                            write_enge_direcion(enge_ws, r + 1, 3, 'ソフト', enge)
                            write_enge_direcion(enge_ws, r + 1, 4, 'ミキサー', enge)
                            write_enge_direcion(enge_ws, r + 1, 5, 'ゼリー', enge)
                        else:
                            # 調理内容(ソフト)
                            enge_ws.cell(r+1, 3, '開封後袋から出して、カットして盛付け。')

                            # 調理内容(ミキサー)
                            enge_ws.cell(r+1, 4, '開封後、盛付け。')

                            # 調理内容(ゼリー)
                            enge_ws.cell(r+1, 5, '開封後袋から出して、カットして盛付け。')
                r += 5

        # 喫食日
        enge_ws.cell(1, 5, dt.datetime.strptime(in_date, '%Y-%m-%d'))

        create_enge_excel(qs_b, '朝食')
        create_enge_excel(qs_l, '昼食')
        create_enge_excel(qs_d, '夕食')

        enge_book.save(output_enge_file)
