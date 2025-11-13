from copy import copy
import datetime as dt
import logging
import os
import openpyxl as excel
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.utils import get_column_letter
import re

from django.conf import settings
from django.db.models import Sum

from web_order.contract import ContractManager
from web_order.models import Order, OrderEveryday, MenuMaster, AllergenMaster, UnitMaster

logger = logging.getLogger(__name__)


# 注文食数確認ページ用model
class AggregateOrder:
    def __init__(self, order, morning, lunch, dinner):
        self.model = order
        self.morning = morning
        self.lunch = lunch
        self.dinner = dinner


class BaseReferenceSoupCounter:
    def __init__(self):
        # カウンタは乾燥と冷凍で施設を分けてセル番地のリストを保管
        self.g_b_l = {"dry": [[], [], []], "cold": [[], [], []]}     # 具のみ 朝・昼
        self.g_l_d = {"dry": [[], [], []], "cold": [[], [], []]}     # 具のみ 昼・夕
        self.g_b_d = {"dry": [[], [], []], "cold": [[], [], []]}     # 具のみ 朝・夕

        self.sg_b_l = {"dry": [[], [], []], "cold": [[], [], []]}    # 汁と具 朝・昼
        self.sg_l_d = {"dry": [[], [], []], "cold": [[], [], []]}    # 汁と具 昼・夕
        self.sg_b_d = {"dry": [[], [], []], "cold": [[], [], []]}    # 汁と具 朝・夕

        self.g_b_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 具のみ 朝1回
        self.g_l_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 具のみ 昼1回
        self.g_d_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 具のみ 夕1回

        self.sg_b_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 汁と具 朝1回
        self.sg_l_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 汁と具 昼1回
        self.sg_d_1 = {"dry": [[], [], []], "cold": [[], [], []]}  # 汁と具 夕1回

        self.g_3 = {"dry": [[], [], []], "cold": [[], [], []]}       # 具のみ3回
        self.sg_3 = {"dry": [[], [], []], "cold": [[], [], []]}      # 汁と具3回
        self.s_none = {"dry": [[], [], []], "cold": [[], [], []]}    # 汁なし


class JoshokuSoupCounter(BaseReferenceSoupCounter):
    def add(self, unit: UnitMaster, meal: str, menu: str, coordinate: str, manager: ContractManager):
        # アレルギーも集計に含める

        if meal == '朝食':
            index = 0
        elif meal == '昼食':
            index = 1
        else:
            index = 2

        if unit.unit_number == 999:
            if '見本' in unit.unit_name:
                self.sg_3['cold'][index] += '2'
            else:
                self.sg_3['cold'][index] += [coordinate]
        elif '木沢・個食' in unit.unit_name:
            pass
        elif 'フリーズ' in unit.unit_name:
            pass
        else:
            user = unit.username
            contract = manager.get_user_contract(user).get_soup_contract_name(menu)
            if user.dry_cold_type == '乾燥':
                dry_cold_key = 'dry'
            else:
                dry_cold_key = 'cold'

            if contract == '汁と具　3回':
                self.sg_3[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　3回':
                self.g_3[dry_cold_key][index] += [coordinate]
            elif contract == '汁無し':
                self.s_none[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・昼':
                self.sg_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　昼・夕':
                self.sg_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・夕':
                self.sg_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・昼':
                self.g_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　昼・夕':
                self.g_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・夕':
                self.g_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　朝食':
                self.sg_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　昼食':
                self.sg_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　夕食':
                self.sg_d_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　朝食':
                self.g_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　昼食':
                self.g_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　夕食':
                self.g_d_1[dry_cold_key][index] += [coordinate]


class UsuajiSoupCounter(BaseReferenceSoupCounter):
    def add(self, unit: UnitMaster, meal: str, menu: str, coordinate: str, manager: ContractManager):
        none


class SoftSoupCounter(BaseReferenceSoupCounter):
    def add(self, unit: UnitMaster, meal: str, menu: str, coordinate: str, manager: ContractManager):
        # アレルギーも集計に含める

        if meal == '朝食':
            index = 0
        elif meal == '昼食':
            index = 1
        else:
            index = 2

        if unit.unit_number == 999:
            if '見本' in unit.unit_name:
                self.sg_3['cold'][index] += '2'
            elif ('針刺し' in unit.unit_name) or ('保存' in unit.unit_name):
                # 汁と具3回(冷凍にカウントする)
                self.sg_3['cold'][index] += [coordinate]
        else:
            user = unit.username
            contract = manager.get_user_contract(user).get_soup_contract_name(menu)
            if user.dry_cold_type == '乾燥':
                dry_cold_key = 'dry'
            else:
                dry_cold_key = 'cold'

            if contract == '汁と具　3回':
                self.sg_3[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　3回':
                self.g_3[dry_cold_key][index] += [coordinate]
            elif contract == '汁無し':
                self.s_none[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・昼':
                self.sg_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　昼・夕':
                self.sg_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・夕':
                self.sg_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・昼':
                self.g_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　昼・夕':
                self.g_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・夕':
                self.g_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　朝食':
                self.sg_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　昼食':
                self.sg_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　夕食':
                self.sg_d_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　朝食':
                self.g_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　昼食':
                self.g_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　夕食':
                self.g_d_1[dry_cold_key][index] += [coordinate]


class MixerSoupCounter(BaseReferenceSoupCounter):
    def add(self, unit: UnitMaster, meal: str, menu: str, coordinate: str, manager: ContractManager):
        # アレルギーも集計に含める

        if meal == '朝食':
            index = 0
        elif meal == '昼食':
            index = 1
        else:
            index = 2

        if unit.unit_number == 999:
            if '見本' in unit.unit_name:
                self.sg_3['cold'][index] += '2'
            elif ('針刺し' in unit.unit_name) or ('保存' in unit.unit_name):
                # 汁と具3回(冷凍にカウントする)
                self.sg_3['cold'][index] += [coordinate]
        else:
            user = unit.username
            contract = manager.get_user_contract(user).get_soup_contract_name(menu)
            if user.dry_cold_type == '乾燥':
                dry_cold_key = 'dry'
            else:
                dry_cold_key = 'cold'

            if contract == '汁と具　3回':
                self.sg_3[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　3回':
                self.g_3[dry_cold_key][index] += [coordinate]
            elif contract == '汁無し':
                self.s_none[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・昼':
                self.sg_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　昼・夕':
                self.sg_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・夕':
                self.sg_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・昼':
                self.g_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　昼・夕':
                self.g_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・夕':
                self.g_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　朝食':
                self.sg_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　昼食':
                self.sg_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　夕食':
                self.sg_d_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　朝食':
                self.g_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　昼食':
                self.g_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　夕食':
                self.g_d_1[dry_cold_key][index] += [coordinate]


class JellySoupCounter(BaseReferenceSoupCounter):
    def add(self, unit: UnitMaster, meal: str, menu: str, coordinate: str, manager: ContractManager):
        # アレルギーも集計に含める

        if meal == '朝食':
            index = 0
        elif meal == '昼食':
            index = 1
        else:
            index = 2

        if unit.unit_number == 999:
            if '見本' in unit.unit_name:
                self.sg_3['cold'][index] += '2'
            elif ('針刺し' in unit.unit_name) or ('保存' in unit.unit_name):
                # 汁と具3回(冷凍にカウントする)
                self.sg_3['cold'][index] += [coordinate]
        else:
            user = unit.username
            contract = manager.get_user_contract(user).get_soup_contract_name(menu)
            if user.dry_cold_type == '乾燥':
                dry_cold_key = 'dry'
            else:
                dry_cold_key = 'cold'

            if contract == '汁と具　3回':
                self.sg_3[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　3回':
                self.g_3[dry_cold_key][index] += [coordinate]
            elif contract == '汁無し':
                self.s_none[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・昼':
                self.sg_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　昼・夕':
                self.sg_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　2回　朝・夕':
                self.sg_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・昼':
                self.g_b_l[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　昼・夕':
                self.g_l_d[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　2回　朝・夕':
                self.g_b_d[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　朝食':
                self.sg_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　昼食':
                self.sg_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '汁具　1回　夕食':
                self.sg_d_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　朝食':
                self.g_b_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　昼食':
                self.g_l_1[dry_cold_key][index] += [coordinate]
            elif contract == '具のみ　1回　夕食':
                self.g_d_1[dry_cold_key][index] += [coordinate]


class ReferenceSoupCounter:
    def __init__(self):
        self.contract_manager = ContractManager()
        self.contract_manager.read_all()
        self.jsc = JoshokuSoupCounter()
        self.usc = UsuajiSoupCounter()
        self.ssc = SoftSoupCounter()
        self.msc = MixerSoupCounter()
        self.zsc = JellySoupCounter()

        self.freeze_total_cells = [[], [], []]
        self.freeze_dry_cells = [[], [], []]
        self.freeze_cold_cells = [[], [], []]

        self.koshoku_total_cells = [[], [], []]
        self.koshoku_dry_cells = [[], [], []]
        self.koshoku_cold_cells = [[], [], []]

    def add(self, unit: UnitMaster, meal: str, menu: str, cell):
        coordinate = cell.coordinate

        # フリーズ・個食の集計
        # 個食を別集計するのは、木沢・個食のみ。他の個食は常食にカウント
        if 'フリーズ' in unit.unit_name:
            self.add_freeze(unit, meal, coordinate)
            return
        elif '木沢・個食' in unit.unit_name:
            self.add_koshoku(unit, meal, coordinate)
            return

        if menu == '常食':
            self.jsc.add(unit, meal, menu, coordinate, self.contract_manager)
        elif menu == '薄味':
            self.usc.add(unit, meal, menu, coordinate, self.contract_manager)
        elif menu == 'ソフト':
            self.ssc.add(unit, meal, menu, coordinate, self.contract_manager)
        elif menu == 'ミキサー':
            self.msc.add(unit, meal, menu, coordinate, self.contract_manager)
        elif menu == 'ゼリー':
            self.zsc.add(unit, meal, menu, coordinate, self.contract_manager)

    def add_freeze(self, unit: UnitMaster, meal: str, coordinate: str):
        if meal == '朝食':
            self.freeze_total_cells[0] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.freeze_dry_cells[0] += [coordinate]
            else:
                self.freeze_cold_cells[0] += [coordinate]
        elif meal == '昼食':
            self.freeze_total_cells[1] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.freeze_dry_cells[1] += [coordinate]
            else:
                self.freeze_cold_cells[1] += [coordinate]
        elif meal == '夕食':
            self.freeze_total_cells[2] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.freeze_dry_cells[2] += [coordinate]
            else:
                self.freeze_cold_cells[2] += [coordinate]

    def add_koshoku(self, unit: UnitMaster, meal: str, coordinate: str):
        if meal == '朝食':
            self.koshoku_total_cells[0] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.koshoku_dry_cells[0] += [coordinate]
            else:
                self.koshoku_cold_cells[0] += [coordinate]
        elif meal == '昼食':
            self.koshoku_total_cells[1] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.koshoku_dry_cells[1] += [coordinate]
            else:
                self.koshoku_cold_cells[1] += [coordinate]
        elif meal == '夕食':
            self.koshoku_total_cells[2] += [coordinate]
            if unit.username.dry_cold_type == '乾燥':
                self.koshoku_dry_cells[2] += [coordinate]
            else:
                self.koshoku_cold_cells[2] += [coordinate]


class CookingProduceExporter:

    def __init__(self):
        self.soup_counter = ReferenceSoupCounter()

        excel_template = os.path.join(settings.STATICFILES_DIRS[0], 'excel/cook_product.xlsx')
        self.wb = excel.load_workbook(excel_template)
        self.cook_output_dir = os.path.join(settings.OUTPUT_DIR, settings.COOK_PRODUCE_DIR)
        self.thin_side = Side(style='thin')
        self.bold_side = Side(style='medium')
        self.top_border = Border(top=self.bold_side, left=self.thin_side, right=self.thin_side)
        self.top_border_no_left = Border(top=self.bold_side, right=self.thin_side)
        self.top_only_border = Border(top=self.bold_side)
        self.top_left_border = Border(top=self.bold_side, left=self.bold_side)
        self.top_right_border = Border(top=self.bold_side, right=self.bold_side)

        # 施設、献立種類毎の食数を出力する表の最大行
        # なぜか、settingsの値が反映されなかったので直接記述
        self.units_max_row = 69

    def fill_cell_color(self, row, name, menu, shift_col=0):
        fill = None
        if 'フリーズ' in name:
            fill = excel.styles.PatternFill(patternType='solid', fgColor='FFD966')
        elif '個食' in name:
            fill = excel.styles.PatternFill(patternType='solid', fgColor='33CC33')
        elif menu == '薄味':
            fill = excel.styles.PatternFill(patternType='solid', fgColor='BDD7EE')
        elif menu == 'ソフト':
            fill = excel.styles.PatternFill(patternType='solid', fgColor='FFCCFF')
        elif menu == 'ミキサー':
            fill = excel.styles.PatternFill(patternType='solid', fgColor='FFFF99')
        elif menu == 'ゼリー':
            fill = excel.styles.PatternFill(patternType='solid', fgColor='F8CBAD')
        else:
            fill = excel.styles.PatternFill(patternType='solid', fgColor='FFFFFF')

        if fill:
            row[0 + shift_col].fill = fill
            row[1 + shift_col].fill = fill
            row[2 + shift_col].fill = fill
            row[3 + shift_col].fill = fill
            row[4 + shift_col].fill = fill

    def get_converted_value(self, name, value):
        # valueがNoneでないことは、呼び出し元で確認する前提
        if '見本' in name:
            if value == 1:
                return (value, 0)
            else:
                converted = ''
                for x in range(value):
                    converted += '1+'
                return (converted[:-1], value)
        else:
            return (value, 0)

    def set_mihon_alignment(self, name, cells):
        if '見本' in name:
            for cell in cells:
                cell.alignment = Alignment(horizontal='right')

    def get_output_menu(self, name, menu):
        if ('フリーズ' in name) or ('個食' in name):
            return menu
        elif menu in ['常食', '薄味']:
            return ''
        else:
            return menu

    def append_sum_function(self, formatter, cell, column, shift_col):
        table_range_start = formatter.format(mincol=get_column_letter(column + shift_col), minrow=2)
        table_range_end = formatter.format(mincol=get_column_letter(column + shift_col), minrow=self.units_max_row)
        func = f'{cell.value}+SUM({table_range_start}:{table_range_end})'
        cell.value = func

    def append_mihon_count_to_function(self, cell, count):
        if count:
            func = f'{cell.value}+{count}'
            cell.value = func

    def get_font_for_unit(self, ws):
        default_font = ws['A2'].font
        return Font(name=default_font.name, color=default_font.color, size=default_font.size, bold=True)

    def pre_export(self, object_list, eating_day, from_date, to_date, start_meal, end_meal):
        ws = self.wb['order_list']

        ws.cell(1, 1, eating_day)  # 喫食日
        font_for_unit = self.get_font_for_unit(ws)

        logger.info('----- export start-----')

        order_iter = iter(object_list)
        order = next(order_iter)
        prev_number = None
        prev_name = None
        freeze_count = [0, 0, 0]
        koshoku_count = [0, 0, 0]
        mihon_count = [0, 0, 0]
        for row in ws.iter_rows(min_row=2, max_row=self.units_max_row):
            name = order.model.unit_name.calc_name or order.model.unit_name.unit_name
            menu = str(order.model.menu_name)
            row[1].value = self.get_output_menu(name, menu)
            if (prev_name == name) and (order.model.unit_name.unit_number != 999):
                pass
            else:
                row[0].value = f'{order.model.unit_name.unit_number}.{name}'
                row[0].font = font_for_unit
            if order.morning:
                value, raw = self.get_converted_value(name, order.morning)
                row[2].value = value
                mihon_count[0] += raw
            self.soup_counter.add(order.model.unit_name, '朝食', menu, row[2])
            if order.lunch:
                value, raw = self.get_converted_value(name, order.lunch)
                row[3].value = value
                mihon_count[1] += raw
            self.soup_counter.add(order.model.unit_name, '昼食', menu, row[3])
            if order.dinner:
                value, raw = self.get_converted_value(name, order.dinner)
                row[4].value = value
                mihon_count[2] += raw
            self.soup_counter.add(order.model.unit_name, '夕食', menu, row[4])
            self.set_mihon_alignment(name, [row[2], row[3], row[4]])

            # フリーズ・個食の集計
            # 個食を別集計するのは、木沢・個食のみ。他の個食は常食にカウント
            if 'フリーズ' in name:
                freeze_count[0] += order.morning
                freeze_count[1] += order.lunch
                freeze_count[2] += order.dinner
            elif '木沢・個食' in name:
                koshoku_count[0] += order.morning
                koshoku_count[1] += order.lunch
                koshoku_count[2] += order.dinner

            # 背景色の設定
            self.fill_cell_color(row, name, menu)

            # 施設の区切りの太線
            if prev_number != order.model.unit_name.unit_number:
                row[0].border = self.top_left_border
                row[1].border = self.top_border_no_left
                row[2].border = self.top_border
                row[3].border = self.top_border
                row[4].border = self.top_right_border
            try:
                prev_number = order.model.unit_name.unit_number
                prev_name = name
                order = next(order_iter)
            except StopIteration:
                order = None
                break

        shift_col = 0
        table_count = 1
        while order:
            # 表のコピー
            shift_col += 7
            cell_A1 = "{mincol}{minrow}"
            for col in range(1, 6):
                for row in range(1, self.units_max_row + 1):
                    # コピー元のセル番地をR1C1形式からA1形式に変換(例：column = 1 , row = 1 → A1)
                    copySrcCoord = cell_A1.format(mincol=get_column_letter(col), minrow=row + 151)

                    # コピー先のセル番地をR1C1形式からA1形式に変換
                    copyDstCoord = cell_A1.format(mincol=get_column_letter(col + shift_col), minrow=row)

                    # コピー先に値をコピー
                    if row == 1:
                        ws[copyDstCoord].value = ws[copySrcCoord].value

                    # コピー先に書式をコピー
                    if ws[copySrcCoord].has_style:
                        ws[copyDstCoord]._style = copy(ws[copySrcCoord]._style)

                max_row_cell = cell_A1.format(mincol=get_column_letter(col + shift_col), minrow=self.units_max_row + 1)
                ws[max_row_cell].border = self.top_only_border

            # 合計の計算式を修正
            # -朝食
            self.append_sum_function(cell_A1, ws["C70"], 3, shift_col)
            # -昼食
            self.append_sum_function(cell_A1, ws["D70"], 4, shift_col)
            # -夕食
            self.append_sum_function(cell_A1, ws["E70"], 5, shift_col)

            table_count += 1

            # 2目移行の表の記入
            prev_name = None    # 表の先頭行は、必ず呼び出し番号を表示する
            for index, row in enumerate(ws.iter_rows(min_row=2, max_row=self.units_max_row)):
                name = order.model.unit_name.calc_name or order.model.unit_name.unit_name
                menu = str(order.model.menu_name)
                row[1 + shift_col].value = self.get_output_menu(name, menu)
                if (prev_name == name) and (order.model.unit_name.unit_number != 999):
                    pass
                else:
                    row[0+shift_col].value = f'{order.model.unit_name.unit_number}.{name}'
                    row[0+shift_col].font = font_for_unit
                if order.morning:
                    value, raw = self.get_converted_value(name, order.morning)
                    row[2+shift_col].value = value
                    mihon_count[0] += raw
                self.soup_counter.add(order.model.unit_name, '朝食', menu, row[2+shift_col])
                if order.lunch:
                    value, raw = self.get_converted_value(name, order.lunch)
                    row[3+shift_col].value = value
                    mihon_count[1] += raw
                self.soup_counter.add(order.model.unit_name, '昼食', menu, row[3 + shift_col])
                if order.dinner:
                    value, raw = self.get_converted_value(name, order.dinner)
                    row[4 + shift_col].value = value
                    mihon_count[2] += raw
                self.soup_counter.add(order.model.unit_name, '夕食', menu, row[4 + shift_col])
                self.set_mihon_alignment(name, [row[2+shift_col], row[3+shift_col], row[4+shift_col]])

                # フリーズ・個食の集計
                if 'フリーズ' in name:
                    freeze_count[0] += order.morning
                    freeze_count[1] += order.lunch
                    freeze_count[2] += order.dinner
                elif '個食' in name:
                    koshoku_count[0] += order.morning
                    koshoku_count[1] += order.lunch
                    koshoku_count[2] += order.dinner

                # 背景色の設定
                self.fill_cell_color(row, name, menu, shift_col)

                # 施設の区切りの太線
                if prev_number != order.model.unit_name.unit_number:
                    row[0+shift_col].border = self.top_left_border
                    row[1+shift_col].border = self.top_border_no_left
                    row[2+shift_col].border = self.top_border
                    row[3+shift_col].border = self.top_border
                    row[4+shift_col].border = self.top_right_border
                try:
                    prev_number = order.model.unit_name.unit_number
                    prev_name = name
                    order = next(order_iter)
                    if index >= 65:
                        break
                except StopIteration:
                    order = None
                    break

        # 見本は2件以上の場合、数字でなくなるので、ここで補正
        self.append_mihon_count_to_function(ws["C70"], mihon_count[0])
        self.append_mihon_count_to_function(ws["D70"], mihon_count[1])
        self.append_mihon_count_to_function(ws["E70"], mihon_count[2])

        if table_count > 2:
            print_max_col = 12 + 7 * (table_count - 2)
            print_area_end = cell_A1.format(mincol=get_column_letter(print_max_col), minrow=145)
            ws.print_area = f'A1:{print_area_end}'

        # アレルギーの出力は不要

        # 汁食数の出力
        self._write_soup_counter(ws, eating_day)

        # フリーズ数の出力
        self._write_freeze_counter(ws)

        # 個食数の出力
        self._write_koshoku_counter(ws)


    def _write_freeze_counter(self, ws):
        # 冷凍
        ws['I89'] = self._get_aggregate_function(self.soup_counter.freeze_cold_cells[0])
        ws['K89'] = self._get_aggregate_function(self.soup_counter.freeze_cold_cells[1])
        ws['M89'] = self._get_aggregate_function(self.soup_counter.freeze_cold_cells[2])

        # 乾燥
        ws['J89'] = self._get_aggregate_function(self.soup_counter.freeze_dry_cells[0])
        ws['L89'] = self._get_aggregate_function(self.soup_counter.freeze_dry_cells[1])
        ws['N89'] = self._get_aggregate_function(self.soup_counter.freeze_dry_cells[2])

    def _write_koshoku_counter(self, ws):
        # 冷凍
        ws['I144'] = self._get_aggregate_function(self.soup_counter.koshoku_cold_cells[0])
        ws['K144'] = self._get_aggregate_function(self.soup_counter.koshoku_cold_cells[1])
        ws['M144'] = self._get_aggregate_function(self.soup_counter.koshoku_cold_cells[0])

        # 乾燥
        ws['J144'] = self._get_aggregate_function(self.soup_counter.koshoku_dry_cells[0])
        ws['L144'] = self._get_aggregate_function(self.soup_counter.koshoku_dry_cells[1])
        ws['N144'] = self._get_aggregate_function(self.soup_counter.koshoku_dry_cells[2])


    def _is_blank_row(self, ws, coordinate_list):
        for coordinate in coordinate_list:
            if coordinate == '2':
                # 見本の場合
                return False

            if ws[coordinate].value != "":
                # 0は出力しないので、何かが入っていれば空でない
                return False

        return True

    def _get_aggregate_function(self, coordinate_list):
        if coordinate_list:
            return f"={'+'.join(coordinate_list)}"
        else:
            return ""

    def _write_soup_counter_detail(self, ws, row, arry_dict):
        # 冷凍の入力
        ws[f'I{row}'] = self._get_aggregate_function(arry_dict['cold'][0])
        ws[f'K{row}'] = self._get_aggregate_function(arry_dict['cold'][1])
        ws[f'M{row}'] = self._get_aggregate_function(arry_dict['cold'][2])

        # 乾燥の入力
        ws[f'J{row}'] = self._get_aggregate_function(arry_dict['dry'][0])
        ws[f'L{row}'] = self._get_aggregate_function(arry_dict['dry'][1])
        ws[f'N{row}'] = self._get_aggregate_function(arry_dict['dry'][2])
        is_blank = self._is_blank_row(ws, arry_dict['cold'][0]) and \
                   self._is_blank_row(ws, arry_dict['cold'][1]) and \
                   self._is_blank_row(ws, arry_dict['cold'][2]) and \
                   self._is_blank_row(ws, arry_dict['dry'][0]) and \
                   self._is_blank_row(ws, arry_dict['dry'][1]) and \
                   self._is_blank_row(ws, arry_dict['dry'][2])
        if is_blank:
            ws.row_dimensions[row].hidden = True

    def _write_soup_counter(self, ws, eating_day):
        t_date = dt.datetime(2023, 1, 31).date()
        if self.soup_counter:
            jsc = self.soup_counter.jsc
            jsc_row = 73

            # 常食（具のみ朝・昼）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_b_l)
            jsc_row += 1

            # 常食（具のみ昼・夕）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_l_d)
            jsc_row += 1

            # 常食（具のみ朝・夕）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_b_d)
            jsc_row += 1

            # 常食（汁と具朝・昼）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_b_l)
            jsc_row += 1

            # 常食（汁と具昼・夕）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_l_d)
            jsc_row += 1

            # 常食（汁と具朝・夕）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_b_d)
            jsc_row += 1

            # 常食（具のみ朝1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_b_1)
            jsc_row += 1

            # 常食（具のみ昼1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_l_1)
            jsc_row += 1

            # 常食（具のみ夕1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_d_1)
            jsc_row += 1

            # 常食（汁具朝1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_b_1)
            jsc_row += 1

            # 常食（汁具昼1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_l_1)
            jsc_row += 1

            # 常食（汁具夕1回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_d_1)
            jsc_row += 1

            # 常食（具のみ3回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.g_3)
            jsc_row += 1

            # 常食（汁と具3回）
            self._write_soup_counter_detail(ws, jsc_row, jsc.sg_3)
            jsc_row += 1

            # 常食(汁無し）
            self._write_soup_counter_detail(ws, jsc_row, jsc.s_none)
            jsc_row += 1

            if eating_day < t_date:
                usc = self.soup_counter.usc

            ssc = self.soup_counter.ssc
            ssc_row = 92

            # ソフト（具のみ朝・昼）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_b_l)
            ssc_row += 1

            # ソフト（具のみ昼・夕）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_l_d)
            ssc_row += 1

            # ソフト（具のみ朝・夕）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_b_d)
            ssc_row += 1

            # ソフト（汁と具朝・昼）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_b_l)
            ssc_row += 1

            # ソフト（汁と具昼・夕）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_l_d)
            ssc_row += 1

            # ソフト（汁と具朝・夕）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_b_d)
            ssc_row += 1

            # ソフト（具のみ朝1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_b_1)
            ssc_row += 1

            # ソフト（具のみ昼1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_l_1)
            ssc_row += 1

            # ソフト（具のみ夕1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_d_1)
            ssc_row += 1

            # ソフト（汁具朝1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_b_1)
            ssc_row += 1

            # ソフト（汁具昼1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_l_1)
            ssc_row += 1

            # ソフト（汁具夕1回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_d_1)
            ssc_row += 1

            # ソフト（具のみ3回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.g_3)
            ssc_row += 1

            # ソフト（汁と具3回）
            self._write_soup_counter_detail(ws, ssc_row, ssc.sg_3)
            ssc_row += 1

            # ソフト(汁無し）
            self._write_soup_counter_detail(ws, ssc_row, ssc.s_none)
            ssc_row += 1

            msc = self.soup_counter.msc
            msc_row = 109

            # ミキサー（具のみ朝・昼）
            self._write_soup_counter_detail(ws, msc_row, msc.g_b_l)
            msc_row += 1

            # ミキサー（具のみ昼・夕）
            self._write_soup_counter_detail(ws, msc_row, msc.g_l_d)
            msc_row += 1

            # ミキサー（具のみ朝・夕）
            self._write_soup_counter_detail(ws, msc_row, msc.g_b_d)
            msc_row += 1

            # ミキサー（汁と具朝・昼）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_b_l)
            msc_row += 1

            # ミキサー（汁と具昼・夕）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_l_d)
            msc_row += 1

            # ミキサー（汁と具朝・夕）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_b_d)
            msc_row += 1

            # ミキサー（具のみ朝1回）
            self._write_soup_counter_detail(ws, msc_row, msc.g_b_1)
            msc_row += 1

            # ミキサー（具のみ昼1回）
            self._write_soup_counter_detail(ws, msc_row, msc.g_l_1)
            msc_row += 1

            # ミキサー（具のみ夕1回）
            self._write_soup_counter_detail(ws, msc_row, msc.g_d_1)
            msc_row += 1

            # ミキサー（汁具朝1回）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_b_1)
            msc_row += 1

            # ミキサー（汁具昼1回）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_l_1)
            msc_row += 1

            # ミキサー（汁具夕1回）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_d_1)
            msc_row += 1

            # ミキサー（具のみ3回）
            self._write_soup_counter_detail(ws, msc_row, msc.g_3)
            msc_row += 1

            # ミキサー（汁と具3回）
            self._write_soup_counter_detail(ws, msc_row, msc.sg_3)
            msc_row += 1

            # ミキサー(汁無し）
            self._write_soup_counter_detail(ws, msc_row, msc.s_none)
            msc_row += 1

            zsc = self.soup_counter.zsc
            zsc_row = 126

            # ゼリー（具のみ朝・昼）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_b_l)
            zsc_row += 1

            # ゼリー（具のみ昼・夕）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_l_d)
            zsc_row += 1

            # ゼリー（具のみ朝・夕）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_b_d)
            zsc_row += 1

            # ゼリー（汁と具朝・昼）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_b_l)
            zsc_row += 1

            # ゼリー（汁と具昼・夕）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_l_d)
            zsc_row += 1

            # ゼリー（汁と具朝・夕）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_b_d)
            zsc_row += 1

            # ゼリー（具のみ朝1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_b_1)
            zsc_row += 1

            # ゼリー（具のみ昼1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_l_1)
            zsc_row += 1

            # ゼリー（具のみ夕1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_d_1)
            zsc_row += 1

            # ゼリー（汁具朝1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_b_1)
            zsc_row += 1

            # ゼリー（汁具昼1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_l_1)
            zsc_row += 1

            # ゼリー（汁具夕1回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_d_1)
            zsc_row += 1

            # ゼリー（具のみ3回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.g_3)
            zsc_row += 1

            # ゼリー（汁と具3回）
            self._write_soup_counter_detail(ws, zsc_row, zsc.sg_3)
            zsc_row += 1

            # ゼリー(汁無し）
            self._write_soup_counter_detail(ws, zsc_row, zsc.s_none)
            zsc_row += 1

    def save(self, eating_day):
        export_file = os.path.join(self.cook_output_dir, '製造作成表_' + str(eating_day) + '.xlsx')
        self.wb.save(export_file)
        logger.info('----- export end-----')
